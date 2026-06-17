"""Export agent progress report to Excel for presentation."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parents[1] / "docs" / "reports" / "agent-progress-report-20260610.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=11, color="1F4E79")
BODY_FONT = Font(size=10)
THIN = Side(style="thin", color="B4C6E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row: int, ncol: int) -> None:
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_table(
    ws,
    start_row: int,
    headers: list[str],
    rows: list[list[str]],
    col_widths: list[int] | None = None,
) -> int:
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    style_header_row(ws, start_row, len(headers))
    r = start_row + 1
    for row_data in rows:
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = BODY_FONT
            c.alignment = WRAP
            c.border = BORDER
        r += 1
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return r


def sheet_overview(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "总览"
    ws["A1"] = "固定资产质检 Agent 开发进度汇报"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:D1")

    meta = [
        ("汇报日期", "2026-06-10"),
        ("项目阶段", "M2a（Agent P1）进行中"),
        ("数据来源", "开发进度汇总20260610.txt、代码库、handoff/latest.md"),
        ("适用读者", "审计/质检业务人员、项目管理人员"),
    ]
    row = 3
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = SUBTITLE_FONT
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="项目定位").font = TITLE_FONT
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=(
            "固定资产质检 Agent 用于自动完成底稿中的重复性核对与可结构化检查，"
            "输出质检报告与带批注的底稿副本，让质检人员把更多时间用于高风险事项识别、"
            "重大审计判断和风险管理。"
        ),
    ).alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 48

    row += 2
    write_table(
        ws,
        row,
        ["交付物", "说明", "当前状态"],
        [
            ["质检报告", "findings 清单、严重级别、程序/资产维度汇总", "JSON + HTML 已可用；正式 Excel 版待完善"],
            ["底稿标注", "在原底稿副本上批注/高亮问题位置", "首版已通（双 Comments 表 + 单元格批注）"],
        ],
        [18, 42, 36],
    )

    row += 5
    ws.cell(row=row, column=1, value="端到端流水线").font = TITLE_FONT
    row += 1
    for line in [
        "上传/指定底稿 (.xlsx)",
        "  → ingest（读取并结构化）",
        "  → rules（确定性规则检查）",
        "  → LLM（可选，语义辅助复核）",
        "  → report（JSON + HTML + 标注副本）",
    ]:
        ws.cell(row=row, column=1, value=line).font = BODY_FONT
        row += 1

    row += 1
    write_table(
        ws,
        row,
        ["使用方式", "命令/入口", "说明"],
        [
            ["图形界面（推荐）", "fa-qc-ui 或 启动质检界面.bat", "选文件 → 开始质检 → 下载报告与标注底稿"],
            ["命令行", "fa-qc-run <底稿.xlsx>", "适合脚本与 CI；FAIL 时退出码 3"],
            ["纯规则验收", "fa-qc-run（默认不启 LLM）", "团队基线验收推荐"],
        ],
        [20, 28, 40],
    )

    row += 5
    ws.cell(row=row, column=1, value="核心结论").font = TITLE_FONT
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=(
            "Agent 当前可支撑「汇总 → Lead → K.01 → FA list → K.02.1 新增」主链路质检；"
            "K.02.2 处置、K.03 折旧为下一阶段重点。"
        ),
    ).alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 36


def sheet_coverage(wb: Workbook) -> None:
    ws = wb.create_sheet("程序覆盖")
    write_table(
        ws,
        1,
        ["程序/模块", "读取(ingest)", "规则检查(rules)", "报告摘录", "LLM辅助", "备注"],
        [
            ["汇总页 / PSP", "已完成", "已完成 AE-003 等", "已完成", "拒绝理由语义复核", ""],
            ["K.00 Lead", "已完成 6 块分区域", "已完成 18 条规则", "已完成", "预期/波动/调整汇总", ""],
            ["K.01 后推表", "已完成 六区块+TB/表4", "已完成 7 条规则", "已完成", "Notes 充分性", ""],
            ["FA list", "已完成 字段映射", "已完成 6 条规则", "已完成+标注", "—", ""],
            ["K.02.1 新增测试", "已完成 清单+K.02.1+K.02.1a", "已完成 8 条+门控", "已完成", "异常说明语义", "B 公司样本匹配已验证"],
            ["K.02.2 处置测试", "进行中 P0", "仅程序包门控", "未做", "—", "下一阶段重点"],
            ["K.03 折旧测试", "PSP/TOD 识别部分", "未做", "未做", "—", "规划阶段"],
        ],
        [18, 22, 22, 14, 18, 24],
    )


def sheet_checkpoints(wb: Workbook) -> None:
    ws = wb.create_sheet("检查点明细")
    rows = [
        ["汇总页 / PSP", "AE-003", "程序是否执行、sheet 是否存在", "psp_completion"],
        ["汇总页 / PSP", "AE-003", "不执行/拒绝理由是否明显空泛", "规则+LLM"],
        ["汇总页 / PSP", "AE-003", "汇总勾选与底稿证据是否冲突", "psp_completion"],
        ["K.00 Lead", "LEAD-001等", "基础信息：客户、期间、TE/SAD、准则、币种", "18 条 Lead 规则"],
        ["K.00 Lead", "LEAD-010", "Lead 期末 vs K.01 CHECK/期末合计", "lead_rollforward_tb_reconciliation"],
        ["K.00 Lead", "—", "预期分析、异常波动、Notes、Check with A3", "Lead 规则+LLM"],
        ["K.00 Lead", "—", "调整汇总表一致性、借贷方向", "规则+LLM（B 案例校准中）"],
        ["K.01 后推", "GL-002", "表3 FA list 与后推 check；超 SAD 查 Notes", "rollforward_fa_list_reconciliation"],
        ["K.01 后推", "GL-008", "TB 差异超 SAD；无 Note 标识 FAIL", "rollforward_difference_over_sad"],
        ["K.01 后推", "GL-004", "表4 折旧费用与利润表/TB 核对", "rollforward_depreciation_pl_reconciliation"],
        ["K.01 后推", "GL-005/006/007", "异常金额：累折>原值、净值为负等", "rollforward_abnormal_amounts"],
        ["K.01 后推", "—", "表1 矩阵购置/处置行按类别审定列汇总", "ingest 已修复重复加总"],
        ["FA list", "FA-RC 系列", "必填字段、编号唯一、金额非负、净值勾稽", "6 条 fa_list 规则"],
        ["FA list", "—", "累折负数列示适配；主 Comments 同类汇总", "ingest+report"],
        ["K.02.1 新增", "—", "三表程序包识别与执行路径门控", "addition/disposal_test_package"],
        ["K.02.1 新增", "—", "清单字段、同质性、购置 vs K.01", "addition 规则集"],
        ["K.02.1 新增", "—", "K.02.1a TE/CRA、样本池、选样 vs 实测匹配", "addition_sampling_output 等"],
        ["输出交付", "—", "JSON / HTML / *_qc_annotated.xlsx", "report 模块"],
        ["输出交付", "—", "双 Comments 表 + 单元格批注 + 外链兼容", "export_annotated_workbook"],
    ]
    write_table(
        ws,
        1,
        ["程序", "规则编码", "检查内容", "实现说明"],
        rows,
        [16, 14, 42, 28],
    )


def sheet_llm(wb: Workbook) -> None:
    ws = wb.create_sheet("LLM边界")
    write_table(
        ws,
        1,
        ["场景", "LLM 作用", "LLM 不做什么"],
        [
            ["汇总页 PSP 拒绝理由", "判断理由是否空泛、是否结合上下文", "不决定是否应执行程序"],
            ["Lead 预期/波动说明", "判断叙述是否与已读事实冲突", "不重新计算 Lead/K.01 金额"],
            ["Lead 调整汇总表", "辅助理解借贷方向、跨科目调整", "不替代确定性勾稽 FAIL"],
            ["K.01 Notes 充分性", "TB/表3/表4 差异说明是否回应问题", "不与确定性 Notes 规则重复输出"],
            ["K.02.1 异常说明", "跨表叙述、拒绝执行说明是否充分", "不判断样本量/TE/CRA"],
            ["报告润色（层4）", "文字摘要", "优先级最低；不改变 severity"],
        ],
        [22, 36, 36],
    )
    r = 10
    ws.cell(row=r, column=1, value="LLM 硬边界（产品原则）").font = TITLE_FONT
    r += 1
    principles = [
        "金额勾稽、TE/CRA、样本匹配、字段缺失、唯一性 → 仅 rules 判定",
        "LLM 不得将规则 FAIL 改为 PASS",
        "LLM 证据不足时 → NEED_REVIEW 或不输出，不编造原因",
        "默认关闭（FA_QC_LLM_ENABLED=false）；不开 LLM 时规则仍可完整运行",
        "sample_selection 不再作为 finding；抽样结论识别降为诊断信息",
        "确定性规则已报同类 Notes 时，LLM 去重不再重复生成",
    ]
    for p in principles:
        ws.cell(row=r, column=1, value=f"• {p}").font = BODY_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
    ws.column_dimensions["A"].width = 90


def sheet_pain_points(wb: Workbook) -> None:
    ws = wb.create_sheet("卡点与解决")
    ws["A1"] = "核心共识：先读对底稿结构，再让规则判；LLM 只能补语义，不能替代取数和勾稽。"
    ws["A1"].font = SUBTITLE_FONT
    ws.merge_cells("A1:D1")
    write_table(
        ws,
        3,
        ["阶段/模块", "卡点", "表现或原因", "解决办法"],
        [
            ["K.01", "表结构不统一", "表1/2/3/4/TB/Notes 易误读区域", "六区块识别、表2/3/4 专门读取"],
            ["K.01", "表1 矩阵重复加总", "B 公司购置 86.6 万误读为 173.3 万", "类别审定列汇总；总计不重复加总"],
            ["K.01", "Lead↔K.01 定位错", "差异应看 CHECK 列却落到错误行", "优先 K.01 CHECK；读不到再退回"],
            ["K.01", "Notes 充分性", "有 Notes ≠ 说明充分", "规则判有没有；LLM 判够不够；去重"],
            ["汇总页", "sheet 识别不稳定", "「汇总 」带空格被误分类", "名称命中汇总时优先作 summary"],
            ["汇总页", "拒绝理由判断", "过严或过松", "规则识别空泛；LLM 结合上下文"],
            ["汇总页", "程序包不完整", "缺 K.02.1a 不应机械 FAIL", "执行路径门控 waived/documented_limited"],
            ["Lead", "多信息块版式变体", "CRA 区有无、波动来源不同", "6 块 LeadSheetDataset；简版支持"],
            ["Lead", "调整汇总黑箱", "只看到「有差异」无取数路径", "拆开主表/调整/借贷/跨科目"],
            ["K.02.1", "SOP 区干扰", "右侧说明文字含「差异」等", "限定左侧业务区；金额须取数值"],
            ["K.02.1", "LLM 样本误报", "底稿已有说明仍报样本选择不足", "sample_selection 不再作 finding"],
            ["K.02.1", "购置勾稽误判", "表面 K.02 差异，根因 K.01 取数", "回 ingest 修 K.01，不在规则硬调"],
            ["FA list", "Comments 不透明", "只显示规则码黑箱标题", "同类计数+代表性说明+明细索引"],
            ["FA list", "累折负数", "净值勾稽批量误判", "原值-abs(累折)-abs(减值)"],
            ["标注", "批注落点", "合并单元格/外链破坏", "锚定左上角；OOXML 原位注入"],
            ["方法论", "单测≠UI正确", "cmts/批注问题单测难覆盖", "以 UI 下载 JSON/HTML/标注回归"],
            ["方法论", "有效修改顺序", "盲目改规则", "下载→定位→回底稿→判层级→最小修复→复测"],
        ],
        [14, 22, 38, 38],
    )


def sheet_gaps_plan(wb: Workbook) -> None:
    ws = wb.create_sheet("差距与计划")
    write_table(
        ws,
        1,
        ["能力", "状态", "说明"],
        [
            ["全 checklist 覆盖", "进行中 约30%+", "46 条已映射；大量仍为 planned"],
            ["K.02.2 处置测试", "P0 待开发", "门控已有；清单/disposal_common/K.01 勾稽待做"],
            ["K.03 折旧测试", "规划阶段", "SAP/TOD/政策复核规则未落地"],
            ["正式 Excel 质检报告", "未做", "当前 JSON + HTML + 标注副本"],
            ["--llm-rules / --llm-checklist", "规划 M3c", "部分 LLM 已挂 pipeline"],
            ["案例库全量回归", "进行中", "B–G 部分回归；42MB 大文件待优化"],
            ["标注精度", "进行中", "无行号 finding、FA 合并粒度待确认"],
        ],
        [28, 14, 50],
    )
    r = 12
    ws.cell(row=r, column=1, value="下一步开发计划（建议顺序）").font = TITLE_FONT
    r += 1
    plans = [
        "1. K.02.2 处置 ingest P0（DT-B 字段 → DT-C 出售+报废净值 → DT-D 与 K.01 勾稽）",
        "2. ingest 修复：案例库路由脚本 bug、addition_method 映射",
        "3. K.01 M2b：表3 模板变体、>TE 路由、Notes 充分性规则化",
        "4. K.03 折旧：SAP/TOD 识别与基础规则",
        "5. report：独立 Excel 质检报告、标注 Cell Ref. 优化",
        "",
        "研发顺序共识：ingest 先读对 → report 先展示 → rules 再判对 → 案例库回归校准",
    ]
    for p in plans:
        ws.cell(row=r, column=1, value=p).font = BODY_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
    ws.column_dimensions["A"].width = 90


def sheet_metrics(wb: Workbook) -> None:
    ws = wb.create_sheet("量化指标")
    write_table(
        ws,
        1,
        ["指标", "数值/状态", "说明"],
        [
            ["已实现规则（注册表）", "46 条", "含 Lead/K.01/FA list/K.02/汇总等"],
            ["自动化测试", "416 条", "ingest + rules + report + llm"],
            ["已通程序链路", "5 段", "汇总 + Lead + K.01 + FA list + K.02.1"],
            ["必交付：底稿标注", "首版已通", "双 Comments + 单元格批注"],
            ["必交付：质检报告", "JSON/HTML 已通", "面向业务的 Excel 报告待做"],
            ["案例库实测（E 锂原 0603）", "405 → 14 issues", "ingest/规则修复后误报大幅下降"],
            ["结论枚举", "4 种", "PASS / WARN / FAIL / NEED_REVIEW"],
        ],
        [28, 22, 40],
    )
    r = 12
    ws.cell(row=r, column=1, value="界面演示流程（汇报可用）").font = TITLE_FONT
    r += 1
    demo = [
        "1. 启动 fa-qc-ui 或 启动质检界面.bat",
        "2. 上传案例底稿（workbook_with_lead.xlsx 或案例库 B/E）",
        "3. 展示问题清单、人工核对 HTML、K.02/K.01/Lead 摘录",
        "4. 下载 JSON、*_qc_annotated.xlsx，打开 Comments 与单元格批注",
    ]
    for d in demo:
        ws.cell(row=r, column=1, value=d).font = BODY_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
    ws.column_dimensions["A"].width = 90


def main() -> Path:
    wb = Workbook()
    sheet_overview(wb)
    sheet_coverage(wb)
    sheet_checkpoints(wb)
    sheet_llm(wb)
    sheet_pain_points(wb)
    sheet_gaps_plan(wb)
    sheet_metrics(wb)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    path = main()
    print(f"Wrote {path}")
