"""导出 Plan B 研发四层汇总（尝试阶段）为 Excel。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = ROOT / "artifacts" / "plan_b_rd_summary.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")


def _write_table(ws, start_row: int, headers: list[str], rows: list[list[str]]) -> int:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    r = start_row + 1
    for row in rows:
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val).alignment = WRAP
        r += 1
    return r


def _autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(
            max_width,
            max((len(str(c.value or "")) for c in col), default=8) + 2,
        )
        ws.column_dimensions[letter].width = width


def build_workbook() -> Workbook:
    wb = Workbook()

    # --- 封面 / 说明 ---
    ws0 = wb.active
    ws0.title = "说明"
    ws0["A1"] = "固定资产质检 Agent — Plan B 研发汇总（尝试阶段）"
    ws0["A1"].font = TITLE_FONT
    lines = [
        "",
        "阶段定位：尝试验证；与团队内路径 A 并行探索，尚未确定最终技术路线。",
        "Plan B 思路：L1 认表读数 → L2 按 sheet 加规则 → L3 报告与标注 → L4 可选 LLM。",
        "终态对齐：质检报告 + 底稿标注副本（不覆盖原件）。",
        "",
        "图例：✅ 已有  ⏳ 部分  ❌ 尚未",
        "生成方式：python scripts/export_plan_b_summary_xlsx.py",
    ]
    for i, line in enumerate(lines, 2):
        ws0.cell(row=i, column=1, value=line).alignment = WRAP
    ws0.column_dimensions["A"].width = 90

    # --- 四层职责 ---
    ws1 = wb.create_sheet("四层结构")
    _write_table(
        ws1,
        1,
        ["层级", "名称", "这一层解决什么"],
        [
            ["L1", "底稿识别与读数", "每张 sheet 是什么？关键字段读在哪里？"],
            ["L2", "规则检查", "对照 checklist：通过 / 提醒 / 不通过 / 待复核"],
            ["L3", "报告与标注", "JSON、界面、HTML、带批注 Excel"],
            ["L4", "大模型增强（可选）", "文字与模糊匹配补充；不改 L2 硬性结论"],
        ],
    )
    _autosize(ws1)

    # --- L1 ---
    ws2 = wb.create_sheet("L1_识别读数")
    _write_table(
        ws2,
        1,
        ["工作表/能力", "已实现", "未实现/待加强"],
        [
            ["汇总", "表名+表头；SWP/四列简版；程序行解析", "—"],
            ["K.00 Lead", "6 块锚点；简版无 CRA", "—"],
            ["K.01 后推", "六区块；L1 金额列绑定", "表2/3 与 FA list 跨表读数未闭环"],
            ["FA list", "表名变体；同义词映射；防误映射", "案例库全量 diagnose 待更新"],
            ["新增/处置清单", "能分类、能读行", "未作质检主线深度打磨"],
            ["K.03 SAP/TOD/政策", "能分类", "无专用深度解析供规则使用"],
            ["K.02 测试页", "部分未分类", "识别不稳定"],
            ["整本结构", "fa-qc-diagnose；案例回归脚本", "42MB+ 跳过；性能优化"],
        ],
    )
    _autosize(ws2)

    # --- L2 ---
    ws3 = wb.create_sheet("L2_规则检查")
    _write_table(
        ws3,
        1,
        ["工作表", "已实现检查点（概括）", "条数级", "未实现/待做"],
        [
            [
                "汇总",
                "程序表可读；已执行→程序页；不执行→理由；合并单元格；K.03 二选一；模板齐全；空表提醒",
                "1 套（多类检查）",
                "Canvas PM/TE/SAD；与 K.01 >TE 路由",
            ],
            [
                "K.00 Lead",
                "必填；TT/GAM；预期/波动；主表/A3；与 K.01 勾稽；波动调查；调整一致；AE-001/002 摘录",
                "约17条+2条LLM",
                "ARP 三触发；波动说明金额一致等",
            ],
            ["K.01 后推", "存在；表1可解析；列完整；异常金额", "3条 P0", "表2/3↔FA list；超SAD；Notes/TE"],
            ["FA list", "必填；唯一；非负；勾稽；寿命；残值率", "6条", "非当前扩展主线"],
            ["新增/处置清单", "—", "0", "K.02 字段与后推勾稽"],
            ["K.03 折旧", "—", "0", "SAP/TOD/政策/重算"],
            ["全 checklist", "约30条已实现", "—", "其余 REVIEW 或规划"],
        ],
    )
    _autosize(ws3)

    # --- L3 ---
    ws4 = wb.create_sheet("L3_报告标注")
    _write_table(
        ws4,
        1,
        ["交付物", "已实现", "未实现/待做"],
        [
            ["JSON 质检报告", "findings；Lead/K.01/汇总专块", "独立 Excel 业务报告"],
            ["命令行 fa-qc-run", "整本跑通；FAIL 退出码", "案例库端到端 CI 门禁"],
            ["本地界面 fa-qc-ui", "上传；分程序页签；下载", "页签体验可再统一"],
            ["人工核对 HTML", "AE-001/002 摘录", "—"],
            ["底稿标注", "*_qc_annotated.xlsx；双 Comments", "无行号项；FA 合并粒度待确认"],
        ],
    )
    _autosize(ws4)

    # --- L4 ---
    ws5 = wb.create_sheet("L4_大模型")
    _write_table(
        ws5,
        1,
        ["能力", "已实现", "未实现/待做"],
        [
            ["基础设施", "API 配置；.env；脱敏", "—"],
            ["汇总语义", "不执行理由；程序页模糊匹配（--llm）", "独立 --llm-rules"],
            ["Lead 语义", "预期分析；波动说明（--llm）", "同上"],
            ["报告叙述", "层4文字摘要", "—"],
            ["ingest 映射", "—", "--llm-map"],
            ["checklist 逐条", "—", "--llm-checklist"],
            ["原则", "L2 severity 不被 LLM 推翻", "tests/llm 体系化回归"],
        ],
    )
    _autosize(ws5)

    # --- 程序总览 ---
    ws6 = wb.create_sheet("程序总览")
    _write_table(
        ws6,
        1,
        ["程序/Sheet", "L1识别", "L2自动查", "L3专块或标注"],
        [
            ["汇总", "✅", "✅", "✅ 汇总页(PSP)"],
            ["K.00 Lead", "✅", "✅", "✅ Lead(K.00)"],
            ["K.01 后推", "✅", "✅ P0三条", "✅ K.01"],
            ["FA list", "✅", "✅", "✅ findings+Comments"],
            ["新增/处置清单", "⏳", "❌", "❌"],
            ["K.02/K.03", "⏳", "❌", "❌"],
        ],
    )
    _autosize(ws6)

    # --- 进度一句话 ---
    ws7 = wb.create_sheet("进度摘要")
    _write_table(
        ws7,
        1,
        ["层级", "进度一句话"],
        [
            ["L1", "四类主表识别读数已通；清单/K.03 仅结构识别；大文件待加强"],
            ["L2", "汇总+Lead+K.01 P0+FA list 已上线；K.02/K.03 与 K.01 跨表未做"],
            ["L3", "JSON+UI+标注副本已通；独立 Excel 报告与标注细粒度待完善"],
            ["L4", "语义与摘要可演示；与规则分层的产品化未做完"],
        ],
    )
    _autosize(ws7)

    return wb


def main() -> None:
    out = DEFAULT_OUT
    out.parent.mkdir(parents=True, exist_ok=True)
    build_workbook().save(out)
    print(f"已写入: {out}")


if __name__ == "__main__":
    main()
