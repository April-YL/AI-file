"""生成 Lead 调整汇总 LLM 回归用脱敏 fixture（英文双列 / 跨科目 AA#）。"""

from __future__ import annotations

from pathlib import Path

import openpyxl

FIXTURES = Path(__file__).resolve().parent.parent / "tests" / "fixtures"


def _write_minimal_lead_blocks(ws) -> None:
    """满足 detect_lead_blocks 所需锚点的最小 Lead 结构。"""
    ws["B1"] = "客户名称"
    ws["C1"] = "FA-TEST-CLIENT"
    ws["B2"] = "期末"
    ws["C2"] = "2025-12-31"
    ws["B3"] = "可容忍误差"
    ws["C3"] = "100000"
    ws["B4"] = "名义金额"
    ws["C4"] = "50000"
    ws["B6"] = "认定"
    ws["B7"] = "计价和分摊"
    ws["C7"] = "Low"
    ws["D7"] = "80000"
    ws["B20"] = "账户变更"
    ws["B21"] = "新增"
    ws["C21"] = "预计因产线扩建增加。"
    ws["B48"] = "总账科目编码"
    ws["C48"] = "科目名称"
    ws["D48"] = "索引号"
    ws["E48"] = "期末账面数"
    ws["C50"] = "累计折旧"
    ws["D50"] = "K.01"
    ws["E50"] = 200000
    ws["B58"] = "波动说明"
    ws["B59"] = "无异常波动"


def build_en_debit_credit(path: Path) -> None:
    """英文 Dr/Cr 双列 + 中文块标题；direct PPE 行与引导表审计调整列勾稽。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    _write_minimal_lead_blocks(ws)
    ws["K48"] = "审计调整数"
    ws["C49"] = "原值"
    ws["D49"] = "K.01"
    ws["E49"] = 1000000
    ws["K49"] = 500000
    ws["B64"] = "调整汇总表 Adjustment Summary (delete if N/A)"
    ws["B65"] = "调整类型 Adjustment Type"
    ws["C65"] = "Ref"
    ws["D65"] = "Account"
    ws["E65"] = "Dr"
    ws["F65"] = "Cr"
    ws["B66"] = "未更正审计调整"
    ws["C66"] = "AA1"
    ws["D66"] = "PPE Cost / 原值"
    ws["E66"] = 500000
    ws["F66"] = 0
    wb.save(path)
    wb.close()


def build_cross_account_aa(path: Path) -> None:
    """跨科目 AA#：汇总表为费用科目间接调整，引导主表无 PPE 调整列金额。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    _write_minimal_lead_blocks(ws)
    ws["C49"] = "原值"
    ws["D49"] = "K.01"
    ws["E49"] = 1000000
    ws["B64"] = "调整汇总表（如不适用请删除）"
    ws["B65"] = "调整类型"
    ws["C65"] = "Ref"
    ws["D65"] = "Account"
    ws["E65"] = "Amount"
    ws["F65"] = "Description"
    ws["B66"] = "未更正审计调整"
    ws["C66"] = "AA3"
    ws["D66"] = "SG&A expense / 管理费用"
    ws["E66"] = 250000
    ws["F66"] = "Reclass to PPE cost; see K.02.1 / 影响固定资产原值，Refer K.02.1"
    wb.save(path)
    wb.close()


def main() -> None:
    FIXTURES.mkdir(parents=True, exist_ok=True)
    build_en_debit_credit(FIXTURES / "lead_adjustment_en_debit_credit.xlsx")
    build_cross_account_aa(FIXTURES / "lead_adjustment_cross_account_aa.xlsx")
    print("Wrote:", FIXTURES / "lead_adjustment_en_debit_credit.xlsx")
    print("Wrote:", FIXTURES / "lead_adjustment_cross_account_aa.xlsx")


if __name__ == "__main__":
    main()
