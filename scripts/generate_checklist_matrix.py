from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_PATH = (
    Path(__file__).resolve().parents[1]
    / "docs"
    / "planning"
    / "checklist-four-layer-decision-matrix.xlsx"
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GROUP_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


MATRIX_HEADERS = [
    "程序",
    "dict_code",
    "rule_id",
    "检查点",
    "目标结论",
    "ingest-coding",
    "ingest-LLM",
    "rules-coding",
    "rules-LLM",
    "report",
    "annotated workbook",
    "PASS 条件",
    "WARN 条件",
    "FAIL 条件",
    "NEED_REVIEW 条件",
    "测试用例",
    "当前状态",
]


def row(
    procedure,
    dict_code,
    rule_id,
    checkpoint,
    target,
    ingest_coding,
    ingest_llm,
    rules_coding,
    rules_llm,
    report,
    annotation,
    pass_condition,
    warn_condition,
    fail_condition,
    need_review_condition,
    tests,
    status,
):
    return {
        "程序": procedure,
        "dict_code": dict_code,
        "rule_id": rule_id,
        "检查点": checkpoint,
        "目标结论": target,
        "ingest-coding": ingest_coding,
        "ingest-LLM": ingest_llm,
        "rules-coding": rules_coding,
        "rules-LLM": rules_llm,
        "report": report,
        "annotated workbook": annotation,
        "PASS 条件": pass_condition,
        "WARN 条件": warn_condition,
        "FAIL 条件": fail_condition,
        "NEED_REVIEW 条件": need_review_condition,
        "测试用例": tests,
        "当前状态": status,
    }


MATRIX_ROWS = [
    row(
        "交付完成度",
        "DL-001",
        "first_delivery_standard.psp_done",
        "首次交付：PSP 是否完成",
        "PASS/WARN/NEED_REVIEW",
        "读取 UI 输入的交付阶段；仅当阶段=首次交付时启用。读取汇总页 PSP 状态、拒绝理由、相关 finding。",
        "字段标题不标准时，识别“PSP/计划程序/拒绝理由”等语义字段。",
        "PSP 已执行或拒绝理由充分为 PASS；信息不完整为 WARN；完全读不到为 NEED_REVIEW。",
        "判断拒绝理由是否为实质理由；不得单独把规则 FAIL 改为 PASS。",
        "归入“交付完成度/DL-001”，显示首次交付口径。",
        "标注汇总页 PSP 状态或拒绝理由单元格。",
        "阶段=首次交付，且 PSP 已完成或拒绝理由充分。",
        "PSP 信息不完整但有执行痕迹。",
        "窄版不建议直接 FAIL，除非后续 checklist 明确首次交付必须完成。",
        "读不到 PSP 区域或拒绝理由需要人工判断。",
        "首次交付、有 PSP；无 PSP；拒绝理由为空。",
        "待实现/规则草案",
    ),
    row(
        "交付完成度",
        "DL-001",
        "first_delivery_standard.tod_cutoff_evidence",
        "首次交付：TOD/cutoff 支持性证据是否已纳入底稿范围",
        "PASS/WARN/NEED_REVIEW",
        "读取 sheet 名、索引、样本状态、附件/证据标识，识别 TOD、cutoff、新增/处置相关程序。",
        "识别“截止测试”“控制测试”“新增抽样支持”等非标准命名。",
        "发现相关程序且证据状态可追踪为 PASS；证据状态缺失为 WARN；无法定位为 NEED_REVIEW。",
        "判断文字说明是否明确证据尚未获取、由谁提供、预计完成时间。",
        "报告说明缺口属于首次交付范围，不直接等同整体交付失败。",
        "标注相关样本状态列或程序行。",
        "相关程序已识别，证据状态完整。",
        "证据状态不完整或仅部分样本有状态。",
        "窄版不直接 FAIL。",
        "找不到 TOD/cutoff 区域或证据口径无法判断。",
        "标准 sheet 名、非标准 sheet 名、无证据状态。",
        "待实现/需样例",
    ),
    row(
        "交付完成度",
        "DL-001",
        "first_delivery_standard.risk_adjustment_response",
        "首次交付：重大错报风险和调整事项是否已有应对说明",
        "PASS/WARN/NEED_REVIEW",
        "读取风险提示、调整事项 finding、汇总页说明、审计调整表索引。",
        "识别“重大风险”“待调整”“管理层确认中”等文本含义。",
        "无相关事项或已有应对说明为 PASS；存在事项但说明不完整为 WARN；业务背景不足为 NEED_REVIEW。",
        "判断说明是否覆盖风险、影响金额、责任人、后续动作。",
        "报告列示需质检人员重点复核的风险和调整事项。",
        "标注风险说明区、调整事项表或 finding 来源位置。",
        "风险/调整事项均有清晰应对或无相关事项。",
        "有事项但闭环说明不完整。",
        "窄版不直接 FAIL。",
        "需要项目背景或外部沟通记录才能判断。",
        "无风险、有风险已说明、有风险未说明。",
        "待实现/LLM 设计",
    ),
    row(
        "交付完成度",
        "DL-002",
        "final_delivery_standard.comments_cleared",
        "整体交付：Comments 是否全部清理",
        "PASS/FAIL/NEED_REVIEW",
        "读取全 workbook 批注、Comments/Review notes 等 sheet，统计未关闭、未回复、open 状态。",
        "识别非标准 Comments 表中 open/closed/已解决/未解决的语义。",
        "开放 comments 数=0 为 PASS；>0 为 FAIL；无法读取状态为 NEED_REVIEW。",
        "解释 comment 文本属于开放问题还是历史记录，但不得覆盖明确 open 状态。",
        "报告显示开放 comments 数量和位置。",
        "标注 Comments 状态列或原批注单元格。",
        "开放 comments 数=0。",
        "不适用。",
        "开放 comments 数>0。",
        "Comments 表结构异常或状态字段无法识别。",
        "无 comments、open comments、异常 comments 表。",
        "已实现窄版",
    ),
    row(
        "交付完成度",
        "DL-002",
        "final_delivery_standard.supporting_samples_checked",
        "整体交付：未获取并检查支持性样本数是否为零",
        "PASS/FAIL/NEED_REVIEW",
        "读取样本状态、支持性文件状态、finding 中“未获取/未检查/缺证据”的计数。",
        "识别“待客户提供”“未核”“N/A”等状态列语义。",
        "未获取或未检查样本数=0 为 PASS；>0 为 FAIL；无法识别样本/状态列为 NEED_REVIEW。",
        "判断状态说明是否实质表示未完成，不得将缺证据改为 PASS。",
        "报告显示未完成样本数量、sheet、行号、样本编号。",
        "标注样本状态列或证据索引列。",
        "所有支持性样本已获取且已检查。",
        "不适用。",
        "存在未获取或未检查支持性样本。",
        "样本清单或状态口径无法识别。",
        "全部完成、1 个未获取、状态列缺失。",
        "已实现窄版",
    ),
    row(
        "交付完成度",
        "DL-002",
        "final_delivery_standard.workpaper_procedures_completed",
        "整体交付：底稿程序是否全部完成",
        "PASS/FAIL/NEED_REVIEW",
        "读取程序完成标识、checkbox、状态列、finding 中程序未完成信息。",
        "识别“进行中”“待复核”“未执行”等非标准状态文本。",
        "全部程序完成为 PASS；存在未完成程序为 FAIL；找不到状态为 NEED_REVIEW。",
        "识别说明是备注还是实际未完成状态，不覆盖明确未完成。",
        "报告显示未完成程序清单和位置。",
        "标注程序行或完成状态单元格。",
        "所有适用程序均完成。",
        "不适用。",
        "存在适用程序未完成。",
        "程序区无法定位或状态字段无法识别。",
        "全部完成、未完成程序、无状态列。",
        "已实现窄版",
    ),
    row(
        "汇总页 PSP",
        "AE-003",
        "psp_completion",
        "汇总页：应执行 PSP 是否有对应工作表",
        "PASS/WARN/FAIL/NEED_REVIEW",
        "读取汇总页程序主表，绑定程序页、执行、不执行原因、注意事项列；把执行=是的程序页与 workbook sheet 名做规范化/模糊匹配。",
        "程序页标题不标准时，辅助识别“新增测试/折旧测试/后推”等语义相近 sheet。",
        "已执行且能匹配目标 sheet 为 PASS；目标 sheet 过空为 WARN；无匹配为 FAIL；弱匹配为 NEED_REVIEW。",
        "仅解释弱匹配或拒绝理由，不得忽略明确缺失 sheet。",
        "输出 summary_sheet_section，列示 PSP 表、列绑定、匹配结果和 findings。",
        "标注汇总页程序行、程序页索引列、执行列。",
        "执行=是且对应 sheet 存在并有基本内容。",
        "目标 sheet 存在但内容密度过低。",
        "执行=是但工作簿无对应 sheet。",
        "只能弱匹配或列绑定置信度不足。",
        "summary_sheet、psp_completion 单测；workbook_psp_demo。",
        "已实现",
    ),
    row(
        "汇总页 PSP",
        "AE-003",
        "summary_execution_refusal_reason",
        "汇总页：拒绝执行是否有理由",
        "PASS/WARN/NEED_REVIEW",
        "读取执行列为否/不执行的程序行及 H 列不执行原因。",
        "判断拒绝理由是否为空泛，如“无需执行”但未说明业务原因。",
        "有明确拒绝理由为 PASS；理由空白为 WARN；理由是否充分无法结构化时 NEED_REVIEW。",
        "识别理由是否对应 Lead/K.01/K.02/K.03 可见事实，如低于 SAD、无本期新增。",
        "报告列示拒绝执行程序和理由。",
        "标注不执行原因单元格。",
        "拒绝执行且理由具体、可追溯。",
        "理由简短或缺少金额/业务依据。",
        "不直接 FAIL，除非后续规定必填拒绝理由。",
        "需要结合项目背景判断理由充分性。",
        "拒绝理由为空、低于 SAD、其他底稿执行三类。",
        "部分实现/待增强 LLM",
    ),
    row(
        "K.00 Lead",
        "LEAD-001",
        "lead_required_fields",
        "Lead：基础信息必填",
        "PASS/FAIL/WARN",
        "读取 Lead 六块中的基础信息字段：客户名称、期末、分析日期、TE、SAD、GAAP、币种。",
        "字段标签变体较大时，辅助识别同义标题。",
        "必填字段存在且非空为 PASS；核心字段缺失为 FAIL；日期解析失败为 WARN。",
        "不得用外部推测值补足底稿空白。",
        "Lead section 汇总字段缺口。",
        "标注基础信息块缺失字段。",
        "基础信息完整且可解析。",
        "字段存在但格式无法解析。",
        "核心必填字段为空。",
        "无法定位 Lead 基础信息块。",
        "test_lead_sheet、workbook_with_lead。",
        "已实现/需持续校准",
    ),
    row(
        "K.00 Lead",
        "LEAD-002",
        "lead_analysis_date_after_period_end",
        "Lead：分析日期不得早于期末",
        "PASS/FAIL/WARN",
        "读取 period_end 与 analysis_date，解析日期。",
        "识别“据实日期/分析日期”等标签变体。",
        "分析日期>=期末为 PASS；早于期末为 FAIL；日期不可解析为 WARN。",
        "不得因文字说明把明确早于期末改为 PASS。",
        "报告显示两个日期和差异。",
        "标注分析日期单元格。",
        "分析日期不早于期末。",
        "日期格式不清。",
        "分析日期早于期末。",
        "读不到日期字段。",
        "早于、等于、晚于、不可解析日期。",
        "已实现",
    ),
    row(
        "K.00 Lead",
        "AE-001",
        "materiality_consistency",
        "Lead：TE/SAD 与外部最终版一致性摘录",
        "NEED_REVIEW/WARN",
        "读取 Lead 中 PM/TE/SAD，当前无 Canvas/A3 外部输入时只摘录。",
        "辅助解释 TE/SAD 最终化说明是否存在。",
        "有值则输出 NEED_REVIEW 供人工比对；缺关键值可 WARN。",
        "LLM 只说明需比对的口径，不替代 Canvas/A3。",
        "manual_review_sections 展示重要性摘录。",
        "标注 TE/SAD 区域。",
        "接入外部输入后与最终版一致。",
        "底稿有值但尚待人工比对。",
        "当前无外部输入时不直接 FAIL。",
        "缺外部最终版或字段无法定位。",
        "Lead 摘录回归。",
        "已实现摘录/外部比对待接入",
    ),
    row(
        "K.00 Lead",
        "AE-002",
        "risk_threshold_consistency",
        "Lead：CRA/TT 与 GAM 区间及项目 CRA 摘录",
        "PASS/WARN/NEED_REVIEW",
        "读取 CRA/TT 块、各认定 TT、整体 TT；简版无 CRA 时识别 layout_variant。",
        "辅助识别 Minimal/Low/Moderate/High 与中文风险等级。",
        "TT 在 GAM 区间且整体 TT=min(非零 TT) 为 PASS；偏离区间为 WARN；缺项目 CRA 表为 NEED_REVIEW。",
        "不得替代项目组 CRA 判断。",
        "报告展示 CRA/TT 摘录和需人工比对项。",
        "标注 CRA/TT 块。",
        "CRA/TT 块完整且数值关系合理。",
        "TT 偏离 GAM 区间或整体 TT 异常。",
        "当前多为 WARN/NEED_REVIEW，不轻易 FAIL。",
        "无外部 CRA 表或简版 Lead。",
        "标准版、简版 Lead 回归。",
        "已实现摘录/部分规则待增强",
    ),
    row(
        "K.00 Lead",
        "LEAD-014",
        "lead_expectation_basis_present",
        "Lead：预期分析是否有判断依据",
        "PASS/WARN/NEED_REVIEW",
        "读取预期分析块和实际变动方向。",
        "判断说明是否只是结论，还是包含业务原因和判断依据。",
        "有明确业务原因和方向为 PASS；仅短结论为 WARN；语义不清为 NEED_REVIEW。",
        "不得要求每个账户都长篇说明，但应识别明显空泛描述。",
        "报告列示预期分析不足项。",
        "标注预期分析说明区。",
        "预期与实际变动方向基本一致且有依据。",
        "说明过短或缺依据。",
        "不直接 FAIL。",
        "需要业务背景判断充分性。",
        "无预期、简短预期、充分预期。",
        "已实现/LLM 已收窄",
    ),
    row(
        "K.00 Lead",
        "LEAD-016",
        "lead_fluctuation_notes_refs",
        "Lead：主表 Notes 与波动说明编号一致",
        "PASS/WARN/FAIL/NEED_REVIEW",
        "读取 Lead 主表 Notes 编号和波动说明区编号。",
        "辅助识别非标准编号，如 NB1、Note A、说明1。",
        "主表触发调查且说明区有对应编号为 PASS；缺对应说明为 FAIL/WARN；编号弱匹配为 NEED_REVIEW。",
        "不得用远处无关说明替代对应 Notes。",
        "报告列示缺失或弱匹配编号。",
        "标注主表 Notes 单元格和说明区。",
        "编号双向匹配。",
        "编号格式异常但可能对应。",
        "触发调查但无对应说明。",
        "编号无法识别。",
        "Notes 编号匹配/缺失/弱匹配。",
        "已实现/待加强",
    ),
    row(
        "K.00 Lead",
        "LEAD-017",
        "lead_adjustment_internal_consistency",
        "Lead：调整事项汇总表与主表调整列一致",
        "PASS/WARN/FAIL/NEED_REVIEW",
        "读取 Lead 主表调整列、调整事项汇总表、调整说明文本。",
        "识别“本年度不涉及审计调整”等结论性文字，避免误当调整明细。",
        "有调整且主表/汇总表一致为 PASS；金额不一致为 FAIL/WARN；复杂借贷或跨科目为 NEED_REVIEW。",
        "LLM 用于判断复杂调整是否恰当，不替代金额勾稽。",
        "报告显示调整事项不一致或待复核。",
        "标注调整列和调整汇总表。",
        "主表调整列与汇总表一致。",
        "格式复杂或说明不完整。",
        "明确金额不一致或缺少来源。",
        "复杂调整需人工判断。",
        "无调整、有调整一致、有调整不一致。",
        "已实现/LLM 辅助待增强",
    ),
    row(
        "K.01",
        "GL-006",
        "rollforward_exists",
        "K.01：后推明细表存在且可识别",
        "PASS/FAIL/NEED_REVIEW",
        "识别 K.01 sheet 和表1 BKD 主矩阵；保存六区块 presence/confidence。",
        "辅助识别非标准 sheet 名，如 Agree SL to GL、BKD、后推。",
        "K.01 存在且表1可解析为 PASS；缺失为 FAIL；弱识别为 NEED_REVIEW。",
        "不得用普通 FA list 误替代后推表。",
        "rollforward_sheet_section 展示区块识别。",
        "标注 K.01 表1锚点或缺失提示。",
        "K.01 sheet 与表1均识别。",
        "不适用。",
        "应执行但无 K.01/表1。",
        "命名或布局弱识别。",
        "K.01 标准/变体/缺失。",
        "已实现",
    ),
    row(
        "K.01",
        "GL-007",
        "rollforward_columns_complete",
        "K.01：金额口径和期初/变动/期末列完整",
        "PASS/FAIL/NEED_REVIEW",
        "读取原值、累计折旧、减值准备、净值四口径，以及期初、本期变动、期末列。",
        "辅助识别“累折/账面价值/审定数”等标题变体。",
        "核心口径和列完整为 PASS；缺核心列为 FAIL；表头冲突为 NEED_REVIEW。",
        "不得把说明文字当金额列。",
        "报告列示缺失列和绑定置信度。",
        "标注缺失或冲突列位置。",
        "四口径和关键期间列完整。",
        "不适用。",
        "缺核心金额列。",
        "表头冲突或布局无法判定。",
        "完整列、缺列、表头冲突。",
        "已实现",
    ),
    row(
        "K.01",
        "GL-005",
        "rollforward_abnormal_amounts",
        "K.01：异常金额关系",
        "PASS/FAIL/WARN",
        "读取表1合计与明细金额，检查累计折旧>原值、净值为负等。",
        "不需要 LLM。",
        "金额关系正常为 PASS；明确异常为 FAIL；轻微格式或尾差为 WARN。",
        "LLM 不参与覆盖金额勾稽。",
        "报告显示异常金额、sheet、行列。",
        "标注异常金额单元格。",
        "金额关系合理。",
        "可能为格式/尾差。",
        "累计折旧>原值、净值为负等明确异常。",
        "金额不可解析。",
        "异常金额、负净值、不可解析金额。",
        "已实现",
    ),
    row(
        "K.01",
        "GL-002",
        "rollforward_fa_list_reconciliation",
        "K.01：FA list/表3 与表1勾稽",
        "PASS/WARN/FAIL/NEED_REVIEW",
        "读取表2/表3 check、FA list 合计和 K.01 表1审定数。",
        "识别表3说明或 Notes 是否解释差异。",
        "差异=0或不超过 SAD 为 PASS；超过 SAD 且无 Notes 为 FAIL；表3不可读为 NEED_REVIEW。",
        "不得把远处无关 Notes 当作表3差异说明。",
        "报告显示差异金额和 SAD 口径。",
        "标注表3 check 差异和 Notes。",
        "差异为 0 或不超过 SAD。",
        "差异存在但已说明或未超 SAD。",
        "超过 SAD 且无说明。",
        "表3或 FA list 合计不可读。",
        "表3差异、SAD、Notes 变体。",
        "已实现/变体增强中",
    ),
    row(
        "K.01",
        "GL-008",
        "rollforward_difference_over_sad",
        "K.01：TB check 差异超过 SAD 是否有说明",
        "PASS/FAIL/NEED_REVIEW",
        "读取 TB/试算表核对区、差异金额、相邻 Note/NB 标识。",
        "判断 Notes 是否说明差异原因，但不得覆盖无 Note 的明确差异。",
        "差异未超 SAD 为 PASS；超过 SAD 且无 Notes 为 FAIL；超过 SAD 且有 Notes 为 NEED_REVIEW。",
        "LLM 可评价 Notes 充分性。",
        "报告列示所有超过 SAD 的差异单元格。",
        "标注 TB check 差异行和 Notes。",
        "差异不超过 SAD。",
        "不适用。",
        "超过 SAD 且无 Notes。",
        "超过 SAD 但 Notes 充分性需判断。",
        "多行差异、无 Note、有 Note。",
        "已实现",
    ),
    row(
        "K.01",
        "GL-004",
        "rollforward_depreciation_pl_reconciliation",
        "K.01：表4折旧费用与利润表核对",
        "PASS/FAIL/NEED_REVIEW",
        "读取表4折旧费用、利润表科目核对金额、差异和 Notes。",
        "判断差异说明是否为折旧费用核对，而非 TB 差异说明错配。",
        "差异为 0 或不超过 SAD 为 PASS；超过 SAD 且无 Notes 为 FAIL；有 Notes 为 NEED_REVIEW。",
        "LLM 仅判断说明充分性。",
        "报告显示折旧费用差异和定位。",
        "标注表4差异单元格和说明。",
        "差异为 0 或不超过 SAD。",
        "不适用。",
        "超过 SAD 且无说明。",
        "有说明但充分性需判断。",
        "表4差异、Notes 错配。",
        "已实现",
    ),
    row(
        "K.01",
        "GL-009",
        "rollforward_te_program_routing",
        "K.01：超过 TE 的变动是否路由到 K.02/K.03",
        "PASS/WARN/NEED_REVIEW",
        "读取 K.01 购置/处置/折旧等变动金额、Lead TE、汇总页执行状态。",
        "判断拒绝执行理由是否与金额、性质和其他底稿一致。",
        "超过 TE 且相关程序已执行为 PASS；未执行但有合理说明为 WARN/NEED_REVIEW；无说明为 WARN。",
        "LLM 识别说明是否实质支持拒绝程序。",
        "报告列示需补做或需说明的程序。",
        "标注 K.01 变动行和汇总页程序行。",
        "重大变动已路由到相应程序。",
        "未执行但存在说明。",
        "当前先不直接 FAIL。",
        "TE 或程序执行状态无法读取。",
        "购置/处置/折旧超过 TE 与汇总页路由。",
        "规划中",
    ),
    row(
        "K.02 新增/处置",
        "AD-001",
        "addition_test_package_complete",
        "K.02.1：新增测试三表程序包完整",
        "PASS/WARN/NEED_REVIEW",
        "识别新增清单、K.02.1 新增测试、K.02.1a 新增选样输出；结合汇总页执行/拒绝路径。",
        "识别“细节测试”“新增抽样输出”等名称变体和测试页 waiver 说明。",
        "程序包完整为 PASS；不完整但有拒绝/受限说明为 WARN；执行路径不清为 NEED_REVIEW。",
        "LLM 判断拒绝说明是否实质合理。",
        "报告展示新增程序包存在性和执行路径。",
        "标注缺失 sheet 或测试页说明。",
        "三表齐全或有合理受限路径。",
        "程序包不完整但有说明。",
        "当前概略版不直接 FAIL。",
        "汇总页和测试页说明冲突。",
        "三表齐全、汇总页拒绝、测试页 waiver。",
        "已实现部分/概略补充",
    ),
    row(
        "K.02 新增/处置",
        "AD-002",
        "addition_required_fields",
        "K.02.1：新增清单字段完整",
        "PASS/FAIL/NEED_REVIEW",
        "读取新增清单字段：类别、编号、名称、入账开始日期、原值、新增方式；过滤小计/合计行。",
        "辅助识别“卡片编码/新增原值/期末原值”等字段语义。",
        "必需字段完整为 PASS；缺核心字段为 FAIL；金额列口径冲突为 NEED_REVIEW。",
        "不得将单据编号误映射为资产编号。",
        "报告展示字段映射、明细行数、购置总体金额。",
        "标注缺失字段表头或异常明细行。",
        "必需字段完整且金额口径清晰。",
        "不适用。",
        "缺核心字段。",
        "金额列或资产编号字段无法确认。",
        "B/E/G/I/H 案例新增清单字段映射。",
        "已实现/持续校准",
    ),
    row(
        "K.02 新增/处置",
        "AD-003",
        "addition_rollforward_reconciliation",
        "K.02.1：购置类新增清单与 K.01 购置行勾稽",
        "PASS/WARN/NEED_REVIEW",
        "读取新增清单购置/外购总体金额和 K.01 表1购置行审定合计。",
        "判断新增方式文本是否属于购置、在建转入、企业合并、调拨等不同总体。",
        "两侧一致为 PASS；不一致为 WARN；任一侧读不到为 NEED_REVIEW。",
        "LLM 只解释非购置总体，不替代金额核对。",
        "报告展示清单侧、K.01侧、差异和 SAD。",
        "标注新增方式列、金额列、K.01购置行。",
        "购置类新增金额与 K.01 一致。",
        "存在差异或差异超过 SAD 需调查。",
        "当前按 AUTO_WARN，不直接 FAIL。",
        "新增方式或 K.01 购置行无法识别。",
        "购置一致、不一致、非购置混入。",
        "已实现/口径校准中",
    ),
    row(
        "K.02 新增/处置",
        "AD-004",
        "addition_population_homogeneity",
        "K.02.1：新增总体同质性",
        "NEED_REVIEW/WARN/PASS",
        "读取新增方式分布，区分购置、在建工程转入、企业合并、内部划转、重分类等。",
        "判断非购置新增是否应另设总体或索引其他程序。",
        "全部为购置或已分总体为 PASS；混入非购置为 NEED_REVIEW/WARN。",
        "不得将非购置新增直接纳入购置总体。",
        "报告展示新增方式分布。",
        "标注新增方式列和非购置行。",
        "总体同质或已清楚分组。",
        "存在非购置但有说明。",
        "不直接 FAIL。",
        "需业务判断是否另设总体。",
        "购置、在建转入、企业合并混合。",
        "已实现提示/概略补充",
    ),
    row(
        "K.02 新增/处置",
        "DT-001",
        "disposal_required_fields",
        "K.02.2：处置清单字段完整",
        "PASS/FAIL/NEED_REVIEW",
        "读取处置清单字段：类别、编号、名称、原值、累计折旧、减值、净值、处置日期、减少方式。",
        "识别“减少日期/业务日期/减少方式/变动方式”等语义字段。",
        "字段完整为 PASS；缺核心字段为 FAIL；单据编号误映射风险为 NEED_REVIEW。",
        "不得把业务单据编号自动作为资产编号。",
        "报告展示字段映射和处置净值总体。",
        "标注缺失字段和疑似误映射列。",
        "必需字段完整。",
        "不适用。",
        "缺核心字段。",
        "字段语义或 sheet 类型不清。",
        "处置清单字段完整/缺失/误映射。",
        "待实现/概略补充",
    ),
    row(
        "K.02 新增/处置",
        "DT-002",
        "disposal_rollforward_reconciliation",
        "K.02.2：处置净值与 K.01 处置行勾稽",
        "PASS/WARN/NEED_REVIEW",
        "读取出售+报废处置净值合计，与 K.01 处置/减少行核对。",
        "判断减少方式是否属于出售、报废、其他减少，避免口径混淆。",
        "两侧一致为 PASS；不一致为 WARN；任一侧读不到为 NEED_REVIEW。",
        "LLM 只辅助解释减少方式，不替代金额勾稽。",
        "报告展示处置清单、K.01侧、差异和 SAD。",
        "标注处置净值列、减少方式列、K.01处置行。",
        "出售+报废净值与 K.01 一致。",
        "存在差异或差异超过 SAD。",
        "当前按 AUTO_WARN，不直接 FAIL。",
        "处置方式或 K.01 处置行无法识别。",
        "出售/报废、其他减少、差异超过 SAD。",
        "待实现/概略补充",
    ),
    row(
        "K.02 新增/处置",
        "DT-003",
        "disposal_date_reasonable",
        "K.02.2：处置日期不晚于期末",
        "PASS/WARN/NEED_REVIEW",
        "读取处置日期/减少日期/业务日期和期末日期。",
        "识别日期字段标题和文本日期格式。",
        "处置日期<=期末为 PASS；晚于期末为 WARN；日期不可解析为 NEED_REVIEW。",
        "不得用备注解释覆盖明显晚于期末的日期。",
        "报告列示异常日期资产。",
        "标注处置日期单元格。",
        "日期不晚于期末。",
        "处置日期晚于期末。",
        "当前 checklist 为 AUTO_WARN。",
        "日期或期末无法读取。",
        "正常日期、晚于期末、不可解析日期。",
        "待实现/概略补充",
    ),
    row(
        "K.03 折旧",
        "DEP-001",
        "depreciation_required_fields",
        "K.03.2：折旧测试清单字段完整",
        "PASS/FAIL/NEED_REVIEW",
        "读取类别、编号、名称、入账开始日期、使用寿命、残值率、原值、累计折旧、净值、本期计提折旧。",
        "识别 By item 与普通 TOD 字段差异。",
        "必需字段完整为 PASS；缺核心字段为 FAIL；TOD 样本表不等同全量清单时 NEED_REVIEW。",
        "不得把测试程序说明误当字段。",
        "报告展示折旧字段映射和缺口。",
        "标注缺失表头或字段冲突。",
        "字段完整且表类型清楚。",
        "不适用。",
        "缺核心折旧字段。",
        "普通 TOD 与 By item 口径无法区分。",
        "标准 by item、普通 TOD、缺字段。",
        "待实现/概略补充",
    ),
    row(
        "K.03 折旧",
        "DEP-002",
        "depreciation_rollforward_reconciliation",
        "K.03.2：本期计提折旧与 K.01 后推核对",
        "PASS/WARN/NEED_REVIEW",
        "读取折旧测试本期计提合计和 K.01 表1/表4折旧金额。",
        "识别折旧费用与制造费用/成本分摊等说明。",
        "一致为 PASS；存在差异为 WARN；任一侧读不到为 NEED_REVIEW。",
        "LLM 仅解释差异说明，不替代金额核对。",
        "报告展示折旧测试侧、K.01侧、差异。",
        "标注折旧合计、K.01折旧行/表4。",
        "本期计提折旧与 K.01 一致。",
        "存在差异或需说明。",
        "当前 checklist 为 AUTO_WARN。",
        "折旧金额或 K.01 折旧行无法识别。",
        "一致、差异、K.01缺失。",
        "待实现/概略补充",
    ),
    row(
        "K.03 折旧",
        "DEP-003",
        "depreciation_recalculation_difference",
        "K.03.2：折旧重新计算差异",
        "PASS/WARN/NEED_REVIEW",
        "读取账面折旧、重算折旧、差异、容差/阈值。",
        "判断复杂情形说明，如新增/处置日期、提足折旧、减值后折旧。",
        "差异在容差内为 PASS；超过容差为 WARN；公式或口径不可读为 NEED_REVIEW。",
        "不得由 LLM 代替重新计算公式。",
        "报告列示差异超过容差的样本。",
        "标注差异列和相关折旧参数。",
        "重算差异在容差内。",
        "差异超过容差。",
        "当前 checklist 为 AUTO_WARN。",
        "公式、参数或容差无法读取。",
        "正常差异、超容差、提足折旧、处置资产。",
        "待实现/概略补充",
    ),
    row(
        "K.03 折旧",
        "SAP-001",
        "sap_cra_consistency",
        "K.03.1：SAP CRA 与 Lead 认定一致",
        "PASS/NEED_REVIEW/WARN",
        "读取 K.03.1 SAP 的 CRA/精确度选择和 Lead 计价/计量认定 CRA。",
        "识别 Minimal/Low/Moderate/High 与中文等级。",
        "两侧一致为 PASS；缺任一侧为 NEED_REVIEW；明显不一致为 WARN。",
        "不得替代项目组风险评估。",
        "报告展示 SAP 与 Lead CRA 对照。",
        "标注 SAP CRA 字段和 Lead CRA 行。",
        "SAP CRA 与 Lead 相关认定一致。",
        "明显不一致。",
        "不直接 FAIL。",
        "缺 SAP 或 Lead CRA。",
        "标准 SAP、无 SAP、CRA 不一致。",
        "待实现/概略补充",
    ),
    row(
        "K.03 折旧",
        "POL-001",
        "depreciation_policy_complete",
        "K.03.3：折旧政策三要素完整",
        "PASS/WARN/NEED_REVIEW",
        "读取折旧方法、使用寿命、预计净残值/残值率政策说明。",
        "判断政策说明是否覆盖三要素且不是仅复制标题。",
        "三要素完整为 PASS；缺要素为 WARN；合理性需判断为 NEED_REVIEW。",
        "LLM 只能辅助识别政策文本完整性，不作准则最终判断。",
        "报告列示政策缺失要素。",
        "标注政策复核区域。",
        "方法、寿命、残值三要素均清晰。",
        "缺少某一要素或说明过短。",
        "不直接 FAIL。",
        "需结合会计准则和业务模式判断。",
        "完整政策、缺残值、政策变更。",
        "待实现/概略补充",
    ),
    row(
        "K.03 折旧",
        "POL-002",
        "depreciation_policy_change_reason",
        "K.03.3：折旧政策与上期不一致时是否说明依据",
        "PASS/WARN/NEED_REVIEW",
        "读取本期政策和上期政策/说明；如无上期输入则摘录。",
        "判断变更说明是否包含原因、影响和依据。",
        "一致或变更有说明为 PASS；变更说明不足为 WARN；缺上期政策为 NEED_REVIEW。",
        "不得替代政策变更合理性判断。",
        "报告列示政策变更待复核项。",
        "标注政策变更说明区域。",
        "政策一致，或变更依据清楚。",
        "变更说明不充分。",
        "不直接 FAIL。",
        "缺上期政策或需业务判断。",
        "政策一致、变更有说明、缺上期。",
        "待实现/概略补充",
    ),
]


def add_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for item in rows:
        ws.append(item)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = WRAP
            cell.border = BORDER

    for cell in ws["A"][1:]:
        cell.fill = GROUP_FILL
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for cell in ws[get_column_letter(idx)]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 44))
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(max_len + 4, 48))

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 30 if row_idx == 1 else 66


def overview_rows():
    return [
        ["交付完成度", "区分首次交付与整体交付，分别判断 DL-001/DL-002。", "已填具体明细。", "整体交付三项已实现窄版；首次交付待实现。", "读取交付阶段、Comments、样本状态、程序完成状态、finding 汇总。", "解释文本性证据、重大风险应对、调整事项闭环。", "需补首次交付样例和 LLM prompt。"],
        ["汇总页 PSP", "程序目录、执行状态、拒绝理由、程序页索引。", "已填具体明细。", "AE-003 已实现，拒绝理由语义待增强。", "读取汇总页主表、列绑定、sheet 匹配、目标 sheet 内容密度。", "判断拒绝理由是否实质合理。", "需加强拒绝理由与 Lead/K.01/K.02/K.03 的交叉判断。"],
        ["K.00 Lead", "基础信息、CRA/TT、预期分析、两期引导主表、波动说明、调整事项。", "已填具体明细。", "已有 Lead ingest、规则和报告块；部分规则需增强。", "读取 Lead 六块、日期、TE/SAD、CRA/TT、Notes、调整表。", "判断预期/波动说明/调整事项语义充分性。", "外部 A3/Canvas/CRA 表尚未接入。"],
        ["K.01", "后推表存在性、列完整、异常金额、FA list/TB/折旧费用勾稽、TE 路由。", "已填具体明细。", "P0 和部分 M2b 已实现。", "读取 K.01 六区块、表1矩阵、表3、TB check、表4。", "判断 Notes 充分性和程序路由理由。", "GL-002 变体、TE 路由和 Notes 充分性仍需增强。"],
        ["K.02 新增/处置", "新增/处置三表程序包、清单字段、与 K.01 勾稽、总体同质性、日期合理性。", "已按当前规划概略补充。", "新增侧部分已实现；处置侧待做。", "读取新增/处置清单、测试页、选样输出、K.01 变动行。", "判断拒绝说明、总体同质性、证据充分性。", "处置 ingest 与勾稽规则仍是后续重点。"],
        ["K.03 折旧", "SAP、折旧 TOD/By item、折旧政策复核。", "已按当前规划概略补充。", "规划中。", "读取折旧测试字段、重算差异、K.01 折旧行、SAP CRA、政策文本。", "判断复杂折旧说明和政策合理性。", "需先实现折旧字段读取和 K.01 勾稽。"],
    ]


def matrix_sheet_rows():
    return [[r[h] for h in MATRIX_HEADERS] for r in MATRIX_ROWS]


def ingest_rows():
    headers = [
        "dict_code",
        "rule_id",
        "检查点",
        "来源 sheet / workbook 区域",
        "来源字段/对象",
        "coding 识别方法",
        "LLM 识别方法",
        "读不到怎么办",
        "当前实现",
    ]
    rows = []
    for r in MATRIX_ROWS:
        rows.append(
            [
                r["dict_code"],
                r["rule_id"],
                r["检查点"],
                r["程序"],
                "见主矩阵 ingest-coding。",
                r["ingest-coding"],
                r["ingest-LLM"],
                r["NEED_REVIEW 条件"],
                r["当前状态"],
            ]
        )
    return headers, rows


def rules_rows():
    headers = [
        "dict_code",
        "rule_id",
        "检查点",
        "coding 确定性规则",
        "阈值/口径",
        "LLM 规则辅助",
        "禁止事项",
        "输出结论映射",
        "当前实现",
    ]
    rows = []
    for r in MATRIX_ROWS:
        rows.append(
            [
                r["dict_code"],
                r["rule_id"],
                r["检查点"],
                r["rules-coding"],
                f"PASS：{r['PASS 条件']}；WARN：{r['WARN 条件']}；FAIL：{r['FAIL 条件']}。",
                r["rules-LLM"],
                "LLM 不得覆盖确定性金额勾稽、缺字段、缺 sheet、明确 open 状态等 coding 结论。",
                r["目标结论"],
                r["当前状态"],
            ]
        )
    return headers, rows


def llm_rows():
    headers = [
        "dict_code",
        "rule_id",
        "检查点",
        "LLM 参与场景",
        "LLM 输入",
        "Prompt 任务",
        "禁止事项",
        "输出 JSON 建议",
        "结果如何使用",
    ]
    rows = []
    for r in MATRIX_ROWS:
        rows.append(
            [
                r["dict_code"],
                r["rule_id"],
                r["检查点"],
                r["ingest-LLM"],
                "相关 sheet 摘录、字段绑定、规则初判、定位信息。",
                r["rules-LLM"],
                "不得把 FAIL 改为 PASS；不得推断底稿外证据；不得替代项目组审计判断。",
                "{status_hint, evidence_quote, reason, confidence, missing_parts}",
                "低置信度转 NEED_REVIEW；高置信度仅补充理由、建议和人工复核提示。",
            ]
        )
    return headers, rows


def report_rows():
    headers = [
        "dict_code",
        "rule_id",
        "检查点",
        "report 字段/分组",
        "message 口径",
        "suggestion 口径",
        "Comments sheet",
        "单元格批注/定位",
        "用户复核动作",
    ]
    rows = []
    for r in MATRIX_ROWS:
        rows.append(
            [
                r["dict_code"],
                r["rule_id"],
                r["检查点"],
                f"category={r['程序']}, code={r['dict_code']}, rule_id={r['rule_id']}",
                r["report"],
                "按 PASS/WARN/FAIL/NEED_REVIEW 给出补充底稿、清理问题或人工复核建议。",
                "汇总 finding、来源 sheet、行列、证据摘录、规则依据。",
                r["annotated workbook"],
                "按标注位置回到底稿核对并补充/修正。",
            ]
        )
    return headers, rows


def status_rows():
    headers = [
        "程序",
        "dict_code",
        "rule_id",
        "检查点",
        "ingest 状态",
        "rules 状态",
        "report 状态",
        "llm 状态",
        "测试状态",
        "下一步",
    ]
    rows = []
    for r in MATRIX_ROWS:
        status = r["当前状态"]
        implemented = "已实现" in status
        rows.append(
            [
                r["程序"],
                r["dict_code"],
                r["rule_id"],
                r["检查点"],
                "已实现/校准中" if implemented else "待实现或规划中",
                "已实现/校准中" if implemented else "待实现或规划中",
                "可复用现有 finding/report/annotation 框架",
                "待接入或仅作语义辅助",
                r["测试用例"],
                "优先补脱敏 fixture，再实现确定性规则，最后接 LLM 语义辅助。",
            ]
        )
    return headers, rows


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    add_sheet(
        wb,
        "总览",
        ["模块", "内容", "当前填充范围", "自动化状态", "Coding 重点", "LLM 重点", "当前缺口"],
        overview_rows(),
    )
    add_sheet(wb, "Checklist四层判定矩阵", MATRIX_HEADERS, matrix_sheet_rows())

    headers, rows = ingest_rows()
    add_sheet(wb, "Ingest取证设计", headers, rows)

    headers, rows = rules_rows()
    add_sheet(wb, "Rules判定设计", headers, rows)

    headers, rows = llm_rows()
    add_sheet(wb, "LLM Prompt设计", headers, rows)

    headers, rows = report_rows()
    add_sheet(wb, "Report标注设计", headers, rows)

    headers, rows = status_rows()
    add_sheet(wb, "状态与缺口", headers, rows)

    return wb


def main():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUTPUT_PATH)

    check = load_workbook(OUTPUT_PATH, read_only=True)
    print(OUTPUT_PATH)
    print("|".join(check.sheetnames))
    for ws in check.worksheets:
        print(f"{ws.title}:{ws.max_row}x{ws.max_column}")


if __name__ == "__main__":
    main()
