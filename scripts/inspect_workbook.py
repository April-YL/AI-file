from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Inspect an Excel workbook without modifying it.")
    parser.add_argument("--path", required=True, help="Workbook path")
    parser.add_argument("--sheet-pattern", default="", help="Only show sheets whose name contains this text")
    parser.add_argument("--max-rows", type=int, default=40, help="Max rows to print per sheet")
    parser.add_argument("--max-cols", type=int, default=20, help="Max columns to print per row")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    path = Path(args.path)

    if not path.exists():
        print(f"file not found: {path}")
        return 2

    wb = openpyxl.load_workbook(path, data_only=False)
    print(f"workbook: {path}")
    print("sheets:", wb.sheetnames)

    needle = args.sheet_pattern.lower().strip()
    for name in wb.sheetnames:
        if needle and needle not in name.lower():
            continue

        ws = wb[name]
        print(f"\n--- sheet: {name} ---")
        print(f"dims: rows={ws.max_row}, cols={ws.max_column}")
        merged = [str(rng) for rng in ws.merged_cells.ranges]
        print("merged:", merged[:20] if merged else [])

        for r in range(1, min(ws.max_row, args.max_rows) + 1):
            values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, args.max_cols) + 1)]
            if any(v not in (None, "") for v in values):
                print(r, values)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
