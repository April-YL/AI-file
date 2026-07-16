"""LLM ingest review helpers.

This module provides the project-level ingest review entrypoint. Program-level
profiles such as K.01 can add precision, but all core workpaper objects share
the same missing-sheet discovery and output guardrails.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Any

import openpyxl

from ingest.disposal_test_sheet import (
    DisposalExecutionPathDataset,
    DisposalSampleOutputDataset,
    DisposalTestSheetDataset,
)
from ingest.lead_sheet import LeadSheetDataset
from ingest.models import ResolutionStatus, SheetKind
from ingest.records import FaListDataset
from ingest.rollforward_sheet import K01_SECTION_IDS, RollforwardSheetDataset
from ingest.sheet_classifier import classify_sheet, score_by_name
from ingest.workbook_reader import read_worksheet_rows
from llm.client import LlmClientError, chat_completion_json
from llm.config import LlmConfig
from llm.router import LlmCapability, LlmRouter
from llm.ingest_profiles import (
    K01_PROFILE_HINT,
    K021_ADDITION_PROFILE_HINT,
    K022_DISPOSAL_PROFILE_HINT,
    LEAD_PROFILE_HINT,
    PROGRAM_PROFILE_HINT,
    SUMMARY_PROFILE_HINT,
)

SYSTEM_PROMPT = """你是固定资产审计底稿的资深质检 CPA，负责复核 Agent 的 ingest 读取结果是否可靠。

你的任务不是重新读取整本 Excel，也不是计算金额，而是基于 coding 已经提供的结构化读取结果、置信度、冲突信息、候选 sheet 预览和局部文本，判断是否存在以下风险：

1. 漏读：Agent 未识别到某个核心 sheet、模块、字段、Notes，但候选内容中疑似存在。
2. 错读：Agent 已读取某个 sheet、模块或金额，但读取结果与底稿预览、锚点顺序或业务结构不自洽。
3. 错分：sheet 类型、区块类型或 Notes 归属可能错误，例如把表4折旧差异当作 TB 差异。
4. 错映射：字段映射可能错误，例如把“单据编号”误当固定资产编号。
5. 低置信度：coding 已提示置信度低、锚点重复、表头冲突、区块边界异常，需要人工复核。

你必须遵守：

1. 不得编造输入中没有的 sheet、字段、金额、行号、单元格或证据。
2. 不得直接给出金额勾稽结论，不判断差异是否超过 SAD，不判断规则 PASS/FAIL。
3. 不得将低置信度读取直接改成高置信度。
4. 不得推翻 coding 规则已形成的确定性结论；只能提示“读取层可能存在风险”。
5. 如果证据不足，返回 unclear，并说明需要人工打开底稿核对。
6. 如果发现疑似漏读或错读，只提出候选 sheet、候选模块、风险原因和建议动作。
7. 最终输出必须是 JSON，不要输出 markdown。

判断口径：

- likely_ok：读取结果与候选预览、锚点、字段含义基本自洽，未见明显漏读或错读风险。
- suspicious：存在较明确的漏读、错读、错分、错映射风险，建议二次 deterministic ingest 或人工复核。
- unclear：输入信息不足，无法判断是否读对，应人工复核。
- not_found：针对漏读发现任务，候选内容中未见目标对象的明显证据。

注意：
coding 是事实取数层；你是读取结果复核层。你只能输出复核建议，不生成最终审计结论。"""

@dataclass(frozen=True)
class ExpectedIngestObject:
    procedure_code: str
    object_name: str
    object_type: str
    sheet_kinds: tuple[SheetKind, ...]
    why_expected: str
    profile_hint: str = PROGRAM_PROFILE_HINT


EXPECTED_INGEST_OBJECTS: tuple[ExpectedIngestObject, ...] = (
    ExpectedIngestObject(
        procedure_code="SUMMARY",
        object_name="汇总",
        object_type="sheet",
        sheet_kinds=(SheetKind.SUMMARY,),
        why_expected="汇总页用于列示固定资产 PSP 是否执行、不执行理由和程序页索引。",
        profile_hint=SUMMARY_PROFILE_HINT,
    ),
    ExpectedIngestObject(
        procedure_code="K.00",
        object_name="K.00 Lead Sheet",
        object_type="sheet",
        sheet_kinds=(SheetKind.LEAD,),
        why_expected="Lead 页用于读取基础信息、TE/SAD、CRA/TT、预期分析和波动说明。",
        profile_hint=LEAD_PROFILE_HINT,
    ),
    ExpectedIngestObject(
        procedure_code="K.01",
        object_name="K.01 Agree SL to GL",
        object_type="sheet",
        sheet_kinds=(SheetKind.ROLLFORWARD,),
        why_expected="K.01 后推表用于明细账/后推明细表与总账、资产清单和试算表核对。",
        profile_hint=K01_PROFILE_HINT,
    ),
    ExpectedIngestObject(
        procedure_code="FA_LIST",
        object_name="FA list",
        object_type="sheet",
        sheet_kinds=(SheetKind.FA_LIST,),
        why_expected="FA list 是资产级字段、金额关系和折旧测试的基础明细清单。",
    ),
    ExpectedIngestObject(
        procedure_code="K.02.1",
        object_name="新增清单",
        object_type="sheet",
        sheet_kinds=(SheetKind.ADDITION_LIST,),
        why_expected="新增清单用于确定新增测试总体、字段完整性和与 K.01 购置金额勾稽。",
        profile_hint=K021_ADDITION_PROFILE_HINT,
    ),
    ExpectedIngestObject(
        procedure_code="K.02.1",
        object_name="K.02.1 新增测试",
        object_type="sheet",
        sheet_kinds=(SheetKind.ADDITION_TEST,),
        why_expected="K.02.1 新增测试页用于记录执行路径、总体金额、样本测试和异常说明。",
        profile_hint=K021_ADDITION_PROFILE_HINT,
    ),
    ExpectedIngestObject(
        procedure_code="K.02.1a",
        object_name="K.02.1a 新增选样输出",
        object_type="sheet",
        sheet_kinds=(SheetKind.ADDITION_SAMPLE_OUTPUT,),
        why_expected="K.02.1a 选样输出用于核对样本池、抽样参数和已选取样本。",
        profile_hint=K021_ADDITION_PROFILE_HINT,
    ),
    ExpectedIngestObject(
        procedure_code="K.02.2",
        object_name="处置清单",
        object_type="sheet",
        sheet_kinds=(SheetKind.DISPOSAL_LIST,),
        why_expected="处置清单用于确定处置测试总体、处置净值和与 K.01 处置金额勾稽。",
        profile_hint=K022_DISPOSAL_PROFILE_HINT,
    ),
    ExpectedIngestObject(
        procedure_code="K.02.2",
        object_name="K.02.2 处置测试",
        object_type="sheet",
        sheet_kinds=(SheetKind.DISPOSAL_TEST,),
        why_expected="K.02.2 处置测试页用于记录执行路径、总体金额、样本测试和异常说明。",
        profile_hint=K022_DISPOSAL_PROFILE_HINT,
    ),
    ExpectedIngestObject(
        procedure_code="K.02.2a",
        object_name="K.02.2a 处置选样输出",
        object_type="sheet",
        sheet_kinds=(SheetKind.DISPOSAL_SAMPLE_OUTPUT,),
        why_expected="K.02.2a 选样输出用于核对处置测试样本池、抽样参数和已选取样本。",
        profile_hint=K022_DISPOSAL_PROFILE_HINT,
    ),
    ExpectedIngestObject(
        procedure_code="K.03.1",
        object_name="K.03.1 SAP",
        object_type="sheet",
        sheet_kinds=(SheetKind.SAP,),
        why_expected="K.03.1 SAP 用于折旧实质性分析程序和是否需要 TOD 的判断依据。",
    ),
    ExpectedIngestObject(
        procedure_code="K.03.2",
        object_name="K.03.2 折旧测试",
        object_type="sheet",
        sheet_kinds=(SheetKind.DEPRECIATION_TOD, SheetKind.DEPRECIATION_TOD_SAMPLE),
        why_expected="K.03.2 折旧测试用于重新计算折旧或执行 by item/TOD 详细测试。",
    ),
    ExpectedIngestObject(
        procedure_code="K.03.3",
        object_name="K.03.3 折旧政策复核",
        object_type="sheet",
        sheet_kinds=(SheetKind.DEPRECIATION_POLICY,),
        why_expected="K.03.3 折旧政策复核用于检查折旧方法、使用寿命和残值率政策。",
    ),
)

USER_TEMPLATE = """请复核以下固定资产底稿 ingest 读取结果。

复核目标：
{review_target}

复核类型：
{review_type}

请返回 JSON：
{{
  "assessment": "likely_ok|suspicious|unclear|not_found",
  "risk_level": "high|medium|low",
  "risk_area": "sheet_classification|section_boundary|field_mapping|amount_anchor|notes_location|missing_sheet|missing_module|other",
  "suspected_object": "",
  "candidate_sheet": "",
  "candidate_rows": [],
  "evidence_anchors": [],
  "rationale": "",
  "suggested_action": "",
  "should_retry_deterministic_ingest": true,
  "manual_review_focus": ""
}}

输入数据：
{payload}
"""

ASSESSMENTS = {"likely_ok", "suspicious", "unclear", "not_found"}
RISK_LEVELS = {"high", "medium", "low"}
RISK_AREAS = {
    "sheet_classification",
    "section_boundary",
    "field_mapping",
    "amount_anchor",
    "notes_location",
    "missing_sheet",
    "missing_module",
    "other",
}
REVIEW_TYPES = {
    "read_result_review",
    "missing_object_discovery",
    "field_mapping_review",
    "notes_location_review",
    "section_boundary_review",
}
FORBIDDEN_OUTPUT_KEYS = {
    "amount",
    "amounts",
    "difference",
    "severity",
    "pass_fail",
    "rule_severity",
}

IDENTIFICATION_SYSTEM_PROMPT = """You review deterministic field candidates for an audit workpaper.
Choose only from the supplied candidate column numbers. Do not create mappings, calculate audit
rules, or infer PASS/FAIL. A suggestion is advisory and will be revalidated by code. Return JSON:
{"selections":[{"standard_field":"", "column":1, "confidence":0.0, "reason":""}]}.
If evidence does not distinguish candidates, omit that field."""


def run_field_identification_fallback(
    config: LlmConfig,
    dataset: FaListDataset,
    *,
    router: LlmRouter | None = None,
) -> dict[str, int]:
    """Ask identification LLM to choose only among existing ambiguous candidates."""
    llm_router = router or LlmRouter(config)
    unresolved = [
        decision
        for decision in dataset.field_resolutions.values()
        if decision.status == ResolutionStatus.AMBIGUOUS
        and decision.reorganization_count < 1
    ]
    if not unresolved or not llm_router.is_enabled(LlmCapability.IDENTIFICATION):
        return {}
    payload = {
        "source_sheet": dataset.source_sheet,
        "fields": [
            {
                "standard_field": decision.standard_field,
                "candidates": [
                    {
                        "column": candidate.column_index,
                        "header": candidate.source_header[:80],
                        "evidence": [
                            {
                                "type": item.evidence_type.value,
                                "description": item.description[:160],
                            }
                            for item in candidate.evidence
                        ],
                        "negative_evidence": [
                            item.description[:160] for item in candidate.negative_evidence
                        ],
                    }
                    for candidate in decision.candidates
                ],
            }
            for decision in unresolved
        ],
    }
    try:
        raw = llm_router.complete_json(
            capability=LlmCapability.IDENTIFICATION,
            task="field_candidate_selection_v1",
            system=IDENTIFICATION_SYSTEM_PROMPT,
            user=json.dumps(payload, ensure_ascii=False),
            client=chat_completion_json,
        )
    except LlmClientError:
        return {}
    selections: dict[str, int] = {}
    allowed = {
        decision.standard_field: {
            candidate.column_index
            for candidate in decision.candidates
            if not candidate.negative_evidence
            and len({item.evidence_type for item in candidate.evidence}) >= 2
        }
        for decision in unresolved
    }
    items = raw.get("selections") if isinstance(raw, dict) else None
    if not isinstance(items, list):
        return {}
    for item in items:
        if not isinstance(item, dict):
            continue
        field_name = str(item.get("standard_field") or "")
        try:
            column = int(item.get("column"))
            confidence = float(item.get("confidence", 0))
        except (TypeError, ValueError):
            continue
        if column in allowed.get(field_name, set()) and 0 <= confidence <= 1:
            selections[field_name] = column
    return selections


@dataclass(frozen=True)
class IngestReviewCandidatePreview:
    sheet_name: str
    name_score: float | None = None
    content_score: float | None = None
    preview_lines: list[dict[str, Any]] = field(default_factory=list)
    anchor_hits: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "name_score": self.name_score,
            "content_score": self.content_score,
            "preview_lines": self.preview_lines,
            "anchor_hits": self.anchor_hits,
        }


@dataclass(frozen=True)
class IngestReviewPayload:
    review_target: str
    review_type: str
    coding_result: dict[str, Any]
    expected_object: dict[str, Any]
    candidate_previews: list[IngestReviewCandidatePreview]
    deterministic_findings: list[dict[str, Any]] = field(default_factory=list)
    question: str = ""
    program_profile_hint: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "review_target": self.review_target,
            "review_type": self.review_type,
            "program_profile_hint": self.program_profile_hint,
            "coding_result": self.coding_result,
            "expected_object": self.expected_object,
            "candidate_previews": [c.to_dict() for c in self.candidate_previews],
            "deterministic_findings": self.deterministic_findings,
            "question": self.question,
        }


@dataclass(frozen=True)
class IngestReviewResult:
    assessment: str
    risk_level: str
    risk_area: str
    suspected_object: str = ""
    candidate_sheet: str = ""
    candidate_rows: list[int] = field(default_factory=list)
    evidence_anchors: list[str] = field(default_factory=list)
    rationale: str = ""
    suggested_action: str = ""
    should_retry_deterministic_ingest: bool = False
    manual_review_focus: str = ""
    procedure_code: str = ""
    source_sheet: str = ""
    review_type: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "assessment": self.assessment,
            "risk_level": self.risk_level,
            "risk_area": self.risk_area,
            "suspected_object": self.suspected_object,
            "candidate_sheet": self.candidate_sheet,
            "candidate_rows": self.candidate_rows,
            "evidence_anchors": self.evidence_anchors,
            "rationale": self.rationale,
            "suggested_action": self.suggested_action,
            "should_retry_deterministic_ingest": self.should_retry_deterministic_ingest,
            "manual_review_focus": self.manual_review_focus,
            "procedure_code": self.procedure_code,
            "source_sheet": self.source_sheet,
            "review_type": self.review_type,
        }


def build_ingest_review_user_prompt(payload: IngestReviewPayload) -> str:
    """Build the user prompt for one ingest review task."""
    review_type = payload.review_type
    if review_type not in REVIEW_TYPES:
        review_type = "read_result_review"
    return USER_TEMPLATE.format(
        review_target=payload.review_target,
        review_type=review_type,
        payload=json.dumps(payload.to_dict(), ensure_ascii=False, indent=2),
    )


def run_ingest_review(
    config: LlmConfig,
    payload: IngestReviewPayload,
    *,
    router: LlmRouter | None = None,
) -> tuple[IngestReviewResult | None, dict[str, Any] | None]:
    """Run LLM ingest review and validate the returned JSON strictly."""
    if not config.enabled:
        return None, None

    user = build_ingest_review_user_prompt(payload)
    try:
        raw = (router or LlmRouter(config)).complete_json(
            capability=LlmCapability.RULE_REVIEW,
            task="legacy_ingest_advisory",
            system=SYSTEM_PROMPT,
            user=user,
            client=chat_completion_json,
        )
    except LlmClientError:
        return None, None
    result = parse_ingest_review_result(raw, payload)
    if result is None:
        return None, raw
    return result, raw


def run_workbook_ingest_reviews(
    config: LlmConfig,
    *,
    lead: LeadSheetDataset | None = None,
    rollforward: RollforwardSheetDataset | None = None,
    disposal_test: DisposalTestSheetDataset | None = None,
    disposal_sample_output: DisposalSampleOutputDataset | None = None,
    disposal_execution_path: DisposalExecutionPathDataset | None = None,
    workbook_path: str | None = None,
    workbook_sheet_titles: list[str] | None = None,
    recognized_sheet_kinds: dict[str, bool] | None = None,
    router: LlmRouter | None = None,
) -> list[IngestReviewResult]:
    """Run project-level ingest reviews that are ready for pipeline use."""
    if not config.enabled:
        return []

    payloads: list[IngestReviewPayload] = []
    recognized = recognized_sheet_kinds or {}
    for expected in EXPECTED_INGEST_OBJECTS:
        if _expected_object_recognized(expected, recognized):
            continue
        payload = build_missing_ingest_object_review_payload(
            expected,
            workbook_path=workbook_path,
            workbook_sheet_titles=workbook_sheet_titles,
        )
        if payload is not None:
            payloads.append(payload)

    if rollforward is not None and should_review_k01_ingest(rollforward):
        payloads.append(
            build_k01_ingest_review_payload(
                rollforward,
                workbook_path=workbook_path,
            )
        )

    if should_review_lead_ingest(lead):
        payloads.append(
            build_lead_ingest_review_payload(
                lead,
                workbook_path=workbook_path,
            )
        )

    if should_review_k022_disposal_ingest(
        disposal_test=disposal_test,
        disposal_sample_output=disposal_sample_output,
        disposal_execution_path=disposal_execution_path,
    ):
        payloads.append(
            build_k022_disposal_ingest_review_payload(
                disposal_test=disposal_test,
                disposal_sample_output=disposal_sample_output,
                disposal_execution_path=disposal_execution_path,
                workbook_path=workbook_path,
            )
        )

    results: list[IngestReviewResult] = []
    for payload in payloads:
        result, _raw = run_ingest_review(config, payload, router=router)
        if result is not None and result.assessment in {"suspicious", "unclear"}:
            procedure = str(payload.expected_object.get("procedure") or "")
            results.append(
                replace(
                    result,
                    procedure_code=procedure,
                    source_sheet=result.candidate_sheet,
                    review_type=f"{procedure.lower().replace('.', '').replace('_', '')}_ingest_review",
                )
            )
    return results


def should_review_k01_ingest(rollforward: RollforwardSheetDataset | None) -> bool:
    """Whether K.01 ingest looks unstable enough to ask LLM for review."""
    if rollforward is None:
        return False
    return bool(
        rollforward.recognition_confidence < 0.65
        or rollforward.section_conflicts
        or _missing_important_k01_sections(rollforward)
    )


def should_review_lead_ingest(lead: LeadSheetDataset | None) -> bool:
    """Ask the LLM only when deterministic Lead ingest is not rule-usable."""
    return bool(lead is not None and lead.usable_for_rules is False)


def build_lead_ingest_review_payload(
    lead: LeadSheetDataset,
    *,
    workbook_path: str | None = None,
) -> IngestReviewPayload:
    """Build a compact Lead read-result review without asking for amounts."""
    rows = _load_sheet_rows(workbook_path, lead.source_sheet)
    preview_lines = _generic_preview_lines(rows)
    return IngestReviewPayload(
        review_target="K.00 Lead Sheet 读取结果复核",
        review_type="read_result_review",
        program_profile_hint=LEAD_PROFILE_HINT,
        coding_result={
            "classified_sheet": lead.source_sheet,
            "usable_for_rules": lead.usable_for_rules,
            "recognized_blocks": [
                {
                    "kind": block.kind.value,
                    "anchor_row": block.anchor_row,
                    "start_row": block.start_row,
                    "end_row": block.end_row,
                    "confidence": block.confidence,
                }
                for block in lead.blocks
            ],
            "movement_labels": [row.account_label for row in lead.movement_rows[:8]],
            "movement_column_roles": [binding.role for binding in lead.movement_bindings],
            "check_with_a3_detected": lead.check_with_a3 is not None,
            "fluctuation_notes_detected": bool(lead.fluctuation_notes),
            "ingest_notes": list(lead.notes or [])[:20],
        },
        expected_object={
            "procedure": "K.00",
            "object_type": "module",
            "object_name": "Lead movement table and Notes",
        },
        candidate_previews=[
            IngestReviewCandidatePreview(
                sheet_name=lead.source_sheet,
                preview_lines=preview_lines[:30],
                anchor_hits=_anchor_hits_from_preview(preview_lines[:30]),
            )
        ],
        question=(
            "请判断 Lead 主表的账户行、两期列、Check with A3/Diff 和 Notes "
            "是否可能漏读、错分或错映射；不要计算金额或判断规则结论。"
        ),
    )


def should_review_k022_disposal_ingest(
    *,
    disposal_test: DisposalTestSheetDataset | None = None,
    disposal_sample_output: DisposalSampleOutputDataset | None = None,
    disposal_execution_path: DisposalExecutionPathDataset | None = None,
) -> bool:
    """Whether K.02.2 disposal ingest looks unstable enough for LLM review."""
    if not (disposal_test or disposal_sample_output or disposal_execution_path):
        return False

    if disposal_execution_path is not None:
        if disposal_execution_path.recognition_confidence < 0.65:
            return True
        if disposal_execution_path.missing_components:
            return True
        if disposal_execution_path.notes:
            return True

    if disposal_test is not None:
        if disposal_test.recognition_confidence < 0.65:
            return True
        if disposal_test.notes:
            return True
        if disposal_test.usable_for_rules is False:
            return True
        if _has_unrecognized_modules(disposal_test.module_assessments):
            return True

    if disposal_sample_output is not None:
        if disposal_sample_output.recognition_confidence < 0.65:
            return True
        if disposal_sample_output.notes:
            return True
        if disposal_sample_output.usable_for_rules is False:
            return True
        if _has_unrecognized_modules(disposal_sample_output.module_assessments):
            return True

    return False


def build_missing_ingest_object_review_payload(
    expected: ExpectedIngestObject,
    *,
    workbook_path: str | None = None,
    workbook_sheet_titles: list[str] | None = None,
) -> IngestReviewPayload | None:
    """Build a compact payload when deterministic ingest missed one object."""
    candidate_previews = _build_missing_object_candidate_previews(
        expected,
        workbook_path=workbook_path,
        workbook_sheet_titles=workbook_sheet_titles,
    )
    if not candidate_previews:
        return None
    return IngestReviewPayload(
        review_target=f"{expected.object_name} 疑似漏读发现",
        review_type="missing_object_discovery",
        program_profile_hint=expected.profile_hint,
        coding_result={
            "classified_sheet": "",
            "recognized_sections": [],
            "missing_sections": [expected.object_name],
            "recognition_confidence": 0.0,
            "section_conflicts": [],
            "ingest_notes": [f"{expected.object_name}_not_recognized"],
        },
        expected_object={
            "procedure": expected.procedure_code,
            "object_type": expected.object_type,
            "object_name": expected.object_name,
            "sheet_kinds": [kind.value for kind in expected.sheet_kinds],
            "why_expected": expected.why_expected,
        },
        candidate_previews=candidate_previews,
        deterministic_findings=[],
        question=(
            f"deterministic ingest 未识别到 {expected.object_name}。请基于候选 sheet 名称、"
            "局部预览、表头和锚点判断是否存在疑似漏读；只提出候选位置，不计算金额。"
        ),
    )


def build_missing_k01_ingest_review_payload(
    *,
    workbook_path: str | None = None,
    workbook_sheet_titles: list[str] | None = None,
) -> IngestReviewPayload | None:
    """Compatibility wrapper for K.01 missing-sheet tests and callers."""
    expected = next(
        obj for obj in EXPECTED_INGEST_OBJECTS if obj.sheet_kinds == (SheetKind.ROLLFORWARD,)
    )
    return build_missing_ingest_object_review_payload(
        expected,
        workbook_path=workbook_path,
        workbook_sheet_titles=workbook_sheet_titles,
    )


def build_k01_ingest_review_payload(
    rollforward: RollforwardSheetDataset,
    *,
    workbook_path: str | None = None,
) -> IngestReviewPayload:
    """Build a compact K.01 ingest-review payload.

    The payload intentionally includes only coding results, confidence/conflicts,
    and small sheet previews. It does not ask the LLM to calculate amounts.
    """
    missing_sections = [
        sid for sid in K01_SECTION_IDS if not rollforward.section_presence.get(sid)
    ]
    review_type = (
        "missing_object_discovery"
        if missing_sections
        else "section_boundary_review"
    )
    expected = _k01_expected_object(missing_sections)
    return IngestReviewPayload(
        review_target="K.01 后推表读取结果复核",
        review_type=review_type,
        program_profile_hint=K01_PROFILE_HINT,
        coding_result={
            "classified_sheet": rollforward.source_sheet,
            "layout_profile": rollforward.layout_profile.value,
            "recognized_sections": [
                sid
                for sid in K01_SECTION_IDS
                if rollforward.section_presence.get(sid)
            ],
            "missing_sections": missing_sections,
            "recognition_confidence": rollforward.recognition_confidence,
            "section_conflicts": list(rollforward.section_conflicts[:12]),
            "section_regions": {
                sid: {
                    "anchor_row": region.anchor_row,
                    "start_row": region.start_row,
                    "end_row": region.end_row,
                    "evidence": list(region.evidence),
                }
                for sid, region in rollforward.section_regions.items()
            },
            "tb_reconciliation_confidence": rollforward.tb_reconciliation_confidence,
            "table3_check_row": rollforward.table3_check_row,
            "table3_notes_row": rollforward.table3_notes_row,
            "table4_difference_row": rollforward.table4_difference_row,
            "table4_notes_row": rollforward.table4_notes_row,
            "tb_difference_row": rollforward.tb_difference_row,
            "tb_notes_row": rollforward.tb_notes_row,
            "ingest_notes": list(rollforward.notes or [])[:20],
        },
        expected_object=expected,
        candidate_previews=[
            _build_k01_candidate_preview(
                rollforward,
                workbook_path=workbook_path,
            )
        ],
        deterministic_findings=[],
        question=(
            "请判断 K.01 后推表读取结果是否存在漏读、错读、错分专题或 Notes 归属风险；"
            "只提出候选位置和人工复核重点，不计算金额、不判断是否超过 SAD。"
        ),
    )


def build_k022_disposal_ingest_review_payload(
    *,
    disposal_test: DisposalTestSheetDataset | None = None,
    disposal_sample_output: DisposalSampleOutputDataset | None = None,
    disposal_execution_path: DisposalExecutionPathDataset | None = None,
    workbook_path: str | None = None,
) -> IngestReviewPayload:
    """Build a compact K.02.2 disposal ingest-review payload.

    This is a read-result review for the sixth ingest layer. It sends only
    recognition metadata, module evidence and small sheet previews; the LLM must
    not calculate disposal amounts or judge sample matching.
    """
    expected = _k022_expected_object()
    previews = _build_k022_candidate_previews(
        disposal_test=disposal_test,
        disposal_sample_output=disposal_sample_output,
        disposal_execution_path=disposal_execution_path,
        workbook_path=workbook_path,
    )
    return IngestReviewPayload(
        review_target="K.02.2 处置测试读取结果复核",
        review_type="read_result_review",
        program_profile_hint=K022_DISPOSAL_PROFILE_HINT,
        coding_result={
            "classified_sheet": {
                "disposal_test": disposal_test.source_sheet if disposal_test else None,
                "disposal_sample_output": (
                    disposal_sample_output.source_sheet
                    if disposal_sample_output
                    else None
                ),
            },
            "recognized_sections": _k022_recognized_modules(
                disposal_test=disposal_test,
                disposal_sample_output=disposal_sample_output,
            ),
            "missing_sections": _k022_missing_modules(
                disposal_test=disposal_test,
                disposal_sample_output=disposal_sample_output,
                disposal_execution_path=disposal_execution_path,
            ),
            "recognition_confidence": _min_confidence(
                disposal_test.recognition_confidence if disposal_test else None,
                disposal_sample_output.recognition_confidence
                if disposal_sample_output
                else None,
                disposal_execution_path.recognition_confidence
                if disposal_execution_path
                else None,
            ),
            "execution_path": (
                disposal_execution_path.to_dict()
                if disposal_execution_path
                else None
            ),
            "disposal_test": _disposal_test_excerpt(disposal_test),
            "disposal_sample_output": _disposal_sample_output_excerpt(
                disposal_sample_output
            ),
            "ingest_notes": _k022_ingest_notes(
                disposal_test=disposal_test,
                disposal_sample_output=disposal_sample_output,
                disposal_execution_path=disposal_execution_path,
            ),
        },
        expected_object=expected,
        candidate_previews=previews,
        deterministic_findings=[],
        question=(
            "请判断 K.02.2 处置测试读取结果是否存在 sheet 错分、模块漏读、"
            "执行路径误读或样本/净值锚点错位风险；只提出候选位置和人工复核重点，"
            "不计算金额、不判断样本是否匹配。"
        ),
    )


def parse_ingest_review_result(
    raw: dict[str, Any],
    payload: IngestReviewPayload,
) -> IngestReviewResult | None:
    """Validate and normalize an LLM ingest review response.

    Invalid or hallucinated responses are discarded instead of being downgraded,
    because this layer is meant to reduce ingest risk, not introduce new noise.
    """
    if not isinstance(raw, dict):
        return None
    if _contains_forbidden_output_keys(raw):
        return None

    assessment = str(raw.get("assessment", "")).strip().lower()
    if assessment not in ASSESSMENTS:
        return None

    risk_level = str(raw.get("risk_level", "low")).strip().lower() or "low"
    if risk_level not in RISK_LEVELS:
        return None

    risk_area = str(raw.get("risk_area", "other")).strip().lower() or "other"
    if risk_area not in RISK_AREAS:
        return None

    candidate_sheet = str(raw.get("candidate_sheet", "") or "").strip()
    candidate_sheets = {c.sheet_name for c in payload.candidate_previews}
    if candidate_sheet and candidate_sheet not in candidate_sheets:
        return None

    preview_rows = _allowed_preview_rows(payload, candidate_sheet=candidate_sheet)
    candidate_rows = _parse_int_list(raw.get("candidate_rows"))
    if any(row not in preview_rows for row in candidate_rows):
        return None

    evidence_anchors = _parse_str_list(raw.get("evidence_anchors"))
    allowed_anchors = _allowed_anchors(payload)
    if any(anchor not in allowed_anchors for anchor in evidence_anchors):
        return None

    if assessment == "suspicious" and not (
        candidate_sheet or candidate_rows or evidence_anchors
    ):
        return None

    return IngestReviewResult(
        assessment=assessment,
        risk_level=risk_level,
        risk_area=risk_area,
        suspected_object=str(raw.get("suspected_object", "") or "").strip(),
        candidate_sheet=candidate_sheet,
        candidate_rows=candidate_rows,
        evidence_anchors=evidence_anchors,
        rationale=str(raw.get("rationale", "") or "").strip(),
        suggested_action=str(raw.get("suggested_action", "") or "").strip(),
        should_retry_deterministic_ingest=bool(
            raw.get("should_retry_deterministic_ingest", False)
        ),
        manual_review_focus=str(raw.get("manual_review_focus", "") or "").strip(),
    )


def _missing_important_k01_sections(rollforward: RollforwardSheetDataset) -> bool:
    important = {
        "b1_bkd_main_table",
        "b4_table3_check_with_table1",
        "b5_table4_depreciation_pl",
    }
    return any(not rollforward.section_presence.get(sid) for sid in important)


def _has_unrecognized_modules(modules: list[Any]) -> bool:
    for module in modules:
        status = str(getattr(module, "status", "") or "").strip().lower()
        if status and status != "recognized":
            return True
    return False


def _min_confidence(*values: float | None) -> float | None:
    numeric = [float(v) for v in values if v is not None]
    if not numeric:
        return None
    return round(min(numeric), 3)


def _k022_recognized_modules(
    *,
    disposal_test: DisposalTestSheetDataset | None,
    disposal_sample_output: DisposalSampleOutputDataset | None,
) -> list[str]:
    out: list[str] = []
    for prefix, modules in (
        ("disposal_test", disposal_test.module_assessments if disposal_test else []),
        (
            "disposal_sample_output",
            disposal_sample_output.module_assessments
            if disposal_sample_output
            else [],
        ),
    ):
        for module in modules:
            if str(getattr(module, "status", "") or "").strip().lower() == "recognized":
                out.append(f"{prefix}:{module.module_key}")
    return out


def _k022_missing_modules(
    *,
    disposal_test: DisposalTestSheetDataset | None,
    disposal_sample_output: DisposalSampleOutputDataset | None,
    disposal_execution_path: DisposalExecutionPathDataset | None,
) -> list[str]:
    missing: list[str] = []
    if disposal_execution_path is not None:
        missing.extend(disposal_execution_path.missing_components)
    for prefix, modules in (
        ("disposal_test", disposal_test.module_assessments if disposal_test else []),
        (
            "disposal_sample_output",
            disposal_sample_output.module_assessments
            if disposal_sample_output
            else [],
        ),
    ):
        for module in modules:
            status = str(getattr(module, "status", "") or "").strip().lower()
            if status and status != "recognized":
                missing.append(f"{prefix}:{module.module_key}:{status}")
    return missing[:20]


def _k022_ingest_notes(
    *,
    disposal_test: DisposalTestSheetDataset | None,
    disposal_sample_output: DisposalSampleOutputDataset | None,
    disposal_execution_path: DisposalExecutionPathDataset | None,
) -> list[str]:
    notes: list[str] = []
    if disposal_execution_path is not None:
        notes.extend(disposal_execution_path.notes)
    if disposal_test is not None:
        notes.extend(disposal_test.notes)
        if disposal_test.usable_for_rules is False:
            notes.append("disposal_test_not_usable_for_rules")
    if disposal_sample_output is not None:
        notes.extend(disposal_sample_output.notes)
        if disposal_sample_output.usable_for_rules is False:
            notes.append("disposal_sample_output_not_usable_for_rules")
    return [str(note) for note in notes if str(note).strip()][:20]


def _disposal_test_excerpt(
    disposal_test: DisposalTestSheetDataset | None,
) -> dict[str, Any] | None:
    if disposal_test is None:
        return None
    return {
        "source_sheet": disposal_test.source_sheet,
        "recognition_confidence": disposal_test.recognition_confidence,
        "usable_for_rules": disposal_test.usable_for_rules,
        "module_assessments": [
            module.to_dict() for module in disposal_test.module_assessments[:12]
        ],
        "waiver_note_rows": disposal_test.waiver_note_rows[:8],
        "amount_anchor_keys": sorted(disposal_test.amounts.keys())[:20],
        "tested_sample_count": len(disposal_test.tested_samples),
        "tested_sample_rows": [
            {
                "source_row": row.source_row,
                "sample_type": row.sample_type,
                "asset_id_present": bool(row.asset_id),
                "net_value_present": bool(row.net_value),
                "disposal_method_present": bool(row.disposal_method),
                "evidence_description_present": bool(row.evidence_description),
            }
            for row in disposal_test.tested_samples[:12]
        ],
        "notes": list(disposal_test.notes[:20]),
    }


def _disposal_sample_output_excerpt(
    sample_output: DisposalSampleOutputDataset | None,
) -> dict[str, Any] | None:
    if sample_output is None:
        return None
    return {
        "source_sheet": sample_output.source_sheet,
        "recognition_confidence": sample_output.recognition_confidence,
        "usable_for_rules": sample_output.usable_for_rules,
        "module_assessments": [
            module.to_dict() for module in sample_output.module_assessments[:12]
        ],
        "parameter_keys": sorted(sample_output.parameters.keys())[:20],
        "amount_anchor_keys": sorted(sample_output.amounts.keys())[:20],
        "selected_sample_count": len(sample_output.selected_samples),
        "selected_sample_rows": [
            {
                "source_row": row.source_row,
                "sample_type": row.sample_type,
                "asset_id_present": bool(row.asset_id),
                "net_value_present": bool(row.net_value),
                "disposal_method_present": bool(row.disposal_method),
            }
            for row in sample_output.selected_samples[:12]
        ],
        "notes": list(sample_output.notes[:20]),
    }


def _k01_expected_object(missing_sections: list[str]) -> dict[str, str]:
    object_name = missing_sections[0] if missing_sections else "K.01 section boundary"
    return {
        "procedure": "K.01",
        "object_type": "module" if missing_sections else "section_boundary",
        "object_name": object_name,
        "why_expected": (
            "K.01 后推表通常包含表1、TB区、表2、表3、表4和 Notes；"
            "低置信度或区块冲突会影响后续规则判断。"
        ),
    }


def _k022_expected_object() -> dict[str, str]:
    return {
        "procedure": "K.02.2",
        "object_type": "module",
        "object_name": "K.02.2 disposal ingest result",
        "why_expected": (
            "K.02.2 处置测试通常需要结合处置清单、K.02.2 测试页、"
            "K.02.2a 选样输出和汇总页执行路径判断是否读对。"
        ),
    }


def _build_k01_candidate_preview(
    rollforward: RollforwardSheetDataset,
    *,
    workbook_path: str | None,
) -> IngestReviewCandidatePreview:
    rows = _load_sheet_rows(workbook_path, rollforward.source_sheet)
    preview_lines = _k01_preview_lines(rows, rollforward)
    return IngestReviewCandidatePreview(
        sheet_name=rollforward.source_sheet,
        name_score=None,
        content_score=None,
        preview_lines=preview_lines,
        anchor_hits=_anchor_hits_from_preview(preview_lines),
    )


def _build_k022_candidate_previews(
    *,
    disposal_test: DisposalTestSheetDataset | None,
    disposal_sample_output: DisposalSampleOutputDataset | None,
    disposal_execution_path: DisposalExecutionPathDataset | None,
    workbook_path: str | None,
) -> list[IngestReviewCandidatePreview]:
    previews: list[IngestReviewCandidatePreview] = []
    seen: set[str] = set()

    for sheet_name, fallback_lines in (
        (
            disposal_test.source_sheet if disposal_test else None,
            _k022_disposal_test_fallback_lines(disposal_test),
        ),
        (
            disposal_sample_output.source_sheet if disposal_sample_output else None,
            _k022_sample_output_fallback_lines(disposal_sample_output),
        ),
    ):
        if not sheet_name or sheet_name in seen:
            continue
        rows = _load_sheet_rows(workbook_path, sheet_name)
        preview_lines = _generic_preview_lines(rows) if rows else fallback_lines
        previews.append(
            IngestReviewCandidatePreview(
                sheet_name=sheet_name,
                name_score=None,
                content_score=None,
                preview_lines=preview_lines[:30],
                anchor_hits=_anchor_hits_from_preview(preview_lines[:30]),
            )
        )
        seen.add(sheet_name)

    for sheet_name in (
        disposal_execution_path.disposal_list_sheet
        if disposal_execution_path
        else None,
        disposal_execution_path.disposal_test_sheet
        if disposal_execution_path
        else None,
        disposal_execution_path.disposal_sample_output_sheet
        if disposal_execution_path
        else None,
    ):
        if not sheet_name or sheet_name in seen:
            continue
        rows = _load_sheet_rows(workbook_path, sheet_name)
        if not rows:
            continue
        preview_lines = _generic_preview_lines(rows)
        previews.append(
            IngestReviewCandidatePreview(
                sheet_name=sheet_name,
                name_score=None,
                content_score=None,
                preview_lines=preview_lines[:20],
                anchor_hits=_anchor_hits_from_preview(preview_lines[:20]),
            )
        )
        seen.add(sheet_name)

    return [preview for preview in previews if preview.preview_lines]


def _k022_disposal_test_fallback_lines(
    disposal_test: DisposalTestSheetDataset | None,
) -> list[dict[str, Any]]:
    if disposal_test is None:
        return []
    lines: list[dict[str, Any]] = []
    for module in disposal_test.module_assessments[:12]:
        evidence = " | ".join(str(e) for e in module.evidence if str(e).strip())
        text = f"{module.module_name} | {module.status}"
        if evidence:
            text += f" | {evidence}"
        lines.append({"row": 1 + len(lines), "text": text})
    for row in disposal_test.tested_samples[:8]:
        parts = [
            "处置测试样本",
            row.sample_type,
            row.asset_id,
            row.asset_name,
            row.net_value,
            row.disposal_method,
        ]
        text = " | ".join(str(p) for p in parts if p)
        if text:
            lines.append({"row": row.source_row, "text": text})
    for row_no in disposal_test.waiver_note_rows[:5]:
        lines.append({"row": row_no, "text": "K.02.2 处置测试 waiver note"})
    return lines


def _k022_sample_output_fallback_lines(
    sample_output: DisposalSampleOutputDataset | None,
) -> list[dict[str, Any]]:
    if sample_output is None:
        return []
    lines: list[dict[str, Any]] = []
    for module in sample_output.module_assessments[:12]:
        evidence = " | ".join(str(e) for e in module.evidence if str(e).strip())
        text = f"{module.module_name} | {module.status}"
        if evidence:
            text += f" | {evidence}"
        lines.append({"row": 1 + len(lines), "text": text})
    for row in sample_output.selected_samples[:8]:
        parts = [
            "处置选样输出",
            row.sample_type,
            row.asset_id,
            row.asset_name,
            row.net_value,
            row.disposal_method,
        ]
        text = " | ".join(str(p) for p in parts if p)
        if text:
            lines.append({"row": row.source_row, "text": text})
    return lines


def _load_sheet_rows(
    workbook_path: str | None,
    sheet_name: str,
) -> list[tuple[Any, ...]]:
    if not workbook_path:
        return []
    path = Path(workbook_path)
    if not path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    try:
        if sheet_name not in wb.sheetnames:
            return []
        return read_worksheet_rows(wb[sheet_name], max_rows=140)
    finally:
        wb.close()


def _k01_preview_lines(
    rows: list[tuple[Any, ...]],
    rollforward: RollforwardSheetDataset,
) -> list[dict[str, Any]]:
    selected_rows = _k01_candidate_row_numbers(rollforward)
    if not selected_rows and rows:
        selected_rows = list(range(1, min(len(rows), 20) + 1))

    out: list[dict[str, Any]] = []
    for row_no in selected_rows:
        if row_no < 1 or row_no > len(rows):
            continue
        row = rows[row_no - 1]
        cells = [
            str(v).strip()
            for v in row[:16]
            if v is not None and str(v).strip()
        ]
        if not cells:
            continue
        out.append({"row": row_no, "text": " | ".join(cells)})
        if len(out) >= 30:
            break

    if out:
        return out

    # Fallback: keep region evidence rows available for validation even if the
    # workbook file is unavailable during unit tests or API-only runs.
    for sid, region in rollforward.section_regions.items():
        if region.anchor_row:
            evidence = " | ".join(str(e) for e in region.evidence if str(e).strip())
            out.append({"row": region.anchor_row, "text": f"{sid} | {evidence}"})
    return out[:30]


def _k01_candidate_row_numbers(rollforward: RollforwardSheetDataset) -> list[int]:
    rows: set[int] = set()
    for region in rollforward.section_regions.values():
        if region.anchor_row:
            rows.update(range(max(1, region.anchor_row - 2), region.anchor_row + 4))
    for row in (
        rollforward.table3_check_row,
        rollforward.table3_notes_row,
        rollforward.table4_difference_row,
        rollforward.table4_notes_row,
        rollforward.tb_difference_row,
        rollforward.tb_notes_row,
    ):
        if row:
            rows.update(range(max(1, row - 2), row + 4))
    return sorted(rows)


def _expected_object_recognized(
    expected: ExpectedIngestObject,
    recognized_sheet_kinds: dict[str, bool],
) -> bool:
    return any(recognized_sheet_kinds.get(kind.value, False) for kind in expected.sheet_kinds)


def _build_missing_object_candidate_previews(
    expected: ExpectedIngestObject,
    *,
    workbook_path: str | None,
    workbook_sheet_titles: list[str] | None,
) -> list[IngestReviewCandidatePreview]:
    return _build_candidate_previews_for_kinds(
        expected.sheet_kinds,
        workbook_path=workbook_path,
        workbook_sheet_titles=workbook_sheet_titles,
    )


def _build_missing_k01_candidate_previews(
    *,
    workbook_path: str | None,
    workbook_sheet_titles: list[str] | None,
) -> list[IngestReviewCandidatePreview]:
    return _build_candidate_previews_for_kinds(
        (SheetKind.ROLLFORWARD,),
        workbook_path=workbook_path,
        workbook_sheet_titles=workbook_sheet_titles,
    )


def _build_candidate_previews_for_kinds(
    target_kinds: tuple[SheetKind, ...],
    *,
    workbook_path: str | None,
    workbook_sheet_titles: list[str] | None,
) -> list[IngestReviewCandidatePreview]:
    rows_by_sheet = _load_candidate_sheet_rows(workbook_path)
    titles = workbook_sheet_titles or list(rows_by_sheet.keys())
    candidates: list[tuple[float, IngestReviewCandidatePreview]] = []
    for title in titles:
        rows = rows_by_sheet.get(title, [])
        name_kind, name_score, _name_hint = score_by_name(title)
        target_name_score = max(
            name_score if name_kind in target_kinds else 0.0,
            _target_kind_name_score(title, target_kinds),
        )
        content_kind = SheetKind.UNCLASSIFIED
        content_score = 0.0
        if rows:
            content_kind, confidence, name_component, content_component, _hint, _header = (
                classify_sheet(title, rows)
            )
            if content_kind in target_kinds:
                content_score = max(content_component, confidence)
                target_name_score = max(target_name_score, name_component)
        preview_lines = _generic_preview_lines(rows)
        anchor_hits = _anchor_hits_from_preview(preview_lines)
        anchor_score = (
            min(0.4, 0.08 * sum(len(h.get("anchors", [])) for h in anchor_hits))
            if name_kind in target_kinds or content_kind in target_kinds
            else 0.0
        )
        kind_bonus = 0.35 if name_kind in target_kinds or content_kind in target_kinds else 0.0
        score = max(target_name_score, content_score) + anchor_score + kind_bonus
        if score < 0.55:
            continue
        candidates.append(
            (
                score,
                IngestReviewCandidatePreview(
                    sheet_name=title,
                    name_score=round(target_name_score, 3),
                    content_score=round(content_score, 3),
                    preview_lines=preview_lines,
                    anchor_hits=anchor_hits,
                ),
            )
        )
    candidates.sort(key=lambda item: item[0], reverse=True)
    return [preview for _score, preview in candidates[:5]]


def _target_kind_name_score(sheet_name: str, target_kinds: tuple[SheetKind, ...]) -> float:
    """Loose name similarity for LLM discovery only, not deterministic ingest."""
    raw = sheet_name.strip().lower()
    n = re.sub(r"\s+", "", raw)
    score = 0.0
    for kind in target_kinds:
        if kind == SheetKind.SUMMARY and "汇总" in sheet_name:
            score = max(score, 0.9)
        elif kind == SheetKind.LEAD and ("k.00" in n or "k00" in n or "lead" in n):
            score = max(score, 0.9)
        elif kind == SheetKind.ROLLFORWARD and (
            "k.01" in n or "k01" in n or "sl-gl" in raw or "agree" in n and "gl" in n or "后推" in sheet_name
        ):
            score = max(score, 0.9)
        elif kind == SheetKind.FA_LIST and ("falist" in n or "fa-list" in raw or "资产清单" in sheet_name):
            score = max(score, 0.88)
        elif kind == SheetKind.ADDITION_LIST and ("新增清单" in sheet_name or "k.02.1b" in n or "k021b" in n):
            score = max(score, 0.88)
        elif kind == SheetKind.ADDITION_TEST and ("k.02.1" in n or "k021" in n or "新增测试" in sheet_name):
            score = max(score, 0.86)
        elif kind == SheetKind.ADDITION_SAMPLE_OUTPUT and (
            "k.02.1a" in n or "k021a" in n or ("新增" in sheet_name and "选样" in sheet_name)
        ):
            score = max(score, 0.88)
        elif kind == SheetKind.DISPOSAL_LIST and (
            "处置清单" in sheet_name or "减少清单" in sheet_name or "k.02.2b" in n or "k022b" in n
        ):
            score = max(score, 0.88)
        elif kind == SheetKind.DISPOSAL_TEST and (
            "k.02.2" in n or "k022" in n or "处置测试" in sheet_name or "减少测试" in sheet_name
        ):
            score = max(score, 0.86)
        elif kind == SheetKind.DISPOSAL_SAMPLE_OUTPUT and (
            "k.02.2a" in n or "k022a" in n or ("处置" in sheet_name and "选样" in sheet_name)
        ):
            score = max(score, 0.88)
        elif kind == SheetKind.SAP and ("k.03.1" in n or "k031" in n or "sap" in n):
            score = max(score, 0.88)
        elif kind in {SheetKind.DEPRECIATION_TOD, SheetKind.DEPRECIATION_TOD_SAMPLE} and (
            "k.03.2" in n or "k032" in n or "折旧测试" in sheet_name
        ):
            score = max(score, 0.88)
        elif kind == SheetKind.DEPRECIATION_POLICY and ("k.03.3" in n or "k033" in n or "折旧政策" in sheet_name):
            score = max(score, 0.88)
    return score


def _load_candidate_sheet_rows(workbook_path: str | None) -> dict[str, list[tuple[Any, ...]]]:
    if not workbook_path:
        return {}
    path = Path(workbook_path)
    if not path.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    try:
        return {
            ws.title: read_worksheet_rows(ws, max_rows=80, max_col=24)
            for ws in wb.worksheets
        }
    finally:
        wb.close()


def _generic_preview_lines(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for idx, row in enumerate(rows[:40], start=1):
        cells = [
            str(v).strip()
            for v in row[:16]
            if v is not None and str(v).strip()
        ]
        if not cells:
            continue
        text = " | ".join(cells)
        anchors = _anchor_tokens_from_text(text)
        if anchors or len(out) < 10:
            out.append({"row": idx, "text": text})
        if len(out) >= 20:
            break
    return out


def _anchor_hits_from_preview(lines: list[dict[str, Any]]) -> list[dict[str, Any]]:
    hits: list[dict[str, Any]] = []
    for line in lines:
        row = _as_int(line.get("row"))
        text = str(line.get("text", "") or "")
        anchors = sorted(_anchor_tokens_from_text(text))
        if row is not None and anchors:
            hits.append({"row": row, "anchors": anchors})
    return hits


def _contains_forbidden_output_keys(value: Any) -> bool:
    if isinstance(value, dict):
        for key, nested in value.items():
            if str(key).strip().lower() in FORBIDDEN_OUTPUT_KEYS:
                return True
            if _contains_forbidden_output_keys(nested):
                return True
    elif isinstance(value, list):
        return any(_contains_forbidden_output_keys(item) for item in value)
    return False


def _allowed_preview_rows(
    payload: IngestReviewPayload,
    *,
    candidate_sheet: str,
) -> set[int]:
    rows: set[int] = set()
    for preview in payload.candidate_previews:
        if candidate_sheet and preview.sheet_name != candidate_sheet:
            continue
        for line in preview.preview_lines:
            row = _as_int(line.get("row"))
            if row is not None:
                rows.add(row)
        for hit in preview.anchor_hits:
            row = _as_int(hit.get("row"))
            if row is not None:
                rows.add(row)
    return rows


def _allowed_anchors(payload: IngestReviewPayload) -> set[str]:
    anchors: set[str] = set()
    for preview in payload.candidate_previews:
        for hit in preview.anchor_hits:
            raw_anchors = hit.get("anchors")
            if isinstance(raw_anchors, list):
                anchors.update(str(a).strip() for a in raw_anchors if str(a).strip())
        for line in preview.preview_lines:
            text = str(line.get("text", "") or "")
            for token in _anchor_tokens_from_text(text):
                anchors.add(token)
    return anchors


def _anchor_tokens_from_text(text: str) -> set[str]:
    """Return conservative, exact text fragments that may be cited as evidence."""
    tokens: set[str] = set()
    for token in (
        "汇总",
        "程序页",
        "是否执行",
        "不执行",
        "Lead Sheet",
        "TE",
        "SAD",
        "CRA",
        "TT",
        "预期分析",
        "异常波动",
        "表1",
        "表2",
        "表3",
        "表4",
        "表2 check with 表1",
        "固定资产类别",
        "固定资产编号",
        "固定资产名称",
        "资产编号",
        "资产名称",
        "原值",
        "累计折旧",
        "净值",
        "入账开始日期",
        "使用寿命",
        "残值率",
        "新增清单",
        "新增方式",
        "新增测试",
        "选样输出",
        "抽样",
        "样本池",
        "已选取样本",
        "处置清单",
        "减少清单",
        "处置测试",
        "处置日期",
        "处置方式",
        "减少方式",
        "报废",
        "出售",
        "处置选样输出",
        "处置测试样本",
        "处置选样",
        "样本类型",
        "SAP",
        "折旧测试",
        "折旧政策",
        "本期计提折旧",
        "年初余额",
        "年末余额",
        "审定数",
        "TB-原值",
        "TB-累计折旧",
        "试算表",
        "差异",
        "折旧费用与利润表",
        "Notes",
    ):
        if token in text:
            tokens.add(token)
    return tokens


def _parse_int_list(value: Any) -> list[int]:
    if not isinstance(value, list):
        return []
    out: list[int] = []
    for item in value:
        row = _as_int(item)
        if row is not None:
            out.append(row)
    return out


def _parse_str_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return [str(item).strip() for item in value if str(item).strip()]


def _as_int(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None
