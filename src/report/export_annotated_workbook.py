"""将 findings 写回底稿副本：主汇总 sheet + FA list 明细 sheet + 单元格批注。"""

from __future__ import annotations

import shutil
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from report.ooxml_workbook import (
    build_worksheet_xml,
    inject_cell_comments,
    inject_worksheets_at_front,
    workbook_has_external_links,
)
from report.summary import QcReport, worst_severity
from rules.models import QcIssue, Severity

COMMENTS_SHEET_NAME = "Comments【归档前删除】"
FA_LIST_COMMENTS_SHEET_NAME = "Comments【FA list】"
LOCATOR_SHEET_NAME = "QC_Locator"
LLM_INGEST_REVIEW_SHEET_NAME = "LLM识别复核【归档前删除】"
_AGENT_REF_HEADER = "Agent 参考（质检建议）"
_REVIEW_SOURCE_HEADER = "判断来源"
_COMMENT_HEADERS = (
    "EY Ref.",
    "Tab Ref.",
    "Cell Ref.",
    "Question/Comment",
    "Answer/Comment",
    "Closed?",
    _AGENT_REF_HEADER,
    _REVIEW_SOURCE_HEADER,
)
_LOCATOR_HEADERS = (
    "EY Ref.",
    "Severity",
    "Rule",
    "Tab Ref.",
    "Cell Ref.",
    "Question/Comment",
    _AGENT_REF_HEADER,
    "Navigate",
    _REVIEW_SOURCE_HEADER,
)
_LLM_INGEST_REVIEW_HEADERS = (
    "序号",
    "程序",
    "复核类型",
    "LLM 判断",
    "风险级别",
    "风险区域",
    "候选 Tab",
    "候选行",
    "锚点证据",
    "判断依据",
    "建议动作",
    "人工复核重点",
    "说明",
)

_DEFAULT_COMMENT_COL = 2
_AUTHOR = "FA-QC"
_FILL_FAIL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_WARN = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_NR = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

_SEV_RANK = {Severity.FAIL: 0, Severity.WARN: 1, Severity.NEED_REVIEW: 2}
_SHORT_TITLE_BY_CODE_FIELD: dict[tuple[str, str], str] = {
    ("AE-003", "execution_status_consistency"): "汇总勾选与底稿证据不一致（K.03.2/TOD）",
    ("AE-003", "waiver_reason"): "不执行理由不充分（需补风险/阈值/替代程序）",
    ("AE-003", "execution_status"): "执行状态未填写或不明确",
    ("LEAD-010", ""): "Lead 与 K.01 期末数不一致",
    ("FA-RC-003", "net_value"): "净值勾稽不一致（原值-累折-减值≠净值）",
}


def annotated_workbook_path(input_path: str | Path) -> Path:
    p = Path(input_path)
    return p.with_name(f"{p.stem}_qc_annotated.xlsx")


def _cell_ref_a1(row: int | None, col: int = _DEFAULT_COMMENT_COL) -> str:
    if not row or row < 1:
        return ""
    return f"${get_column_letter(col)}${row}"


def _cell_ref_plain(row: int | None, col: int = _DEFAULT_COMMENT_COL) -> str:
    if not row or row < 1:
        return ""
    return f"{get_column_letter(col)}{row}"


def _quote_sheet_for_location(sheet: str) -> str:
    escaped = sheet.replace("'", "''")
    return f"'{escaped}'"


def _cell_location(sheet: str | None, row: int | None, col: int | None = None) -> str | None:
    if not sheet or not row or row < 1:
        return None
    return f"{_quote_sheet_for_location(sheet)}!{_cell_ref_plain(row, _issue_col(col))}"


def _issue_col(col: int | None) -> int:
    return col if isinstance(col, int) and col > 0 else _DEFAULT_COMMENT_COL


def _issue_comment_text(issue: QcIssue) -> str:
    code = issue.dict_rule_code or issue.rule_id
    lines = [f"[{issue.severity.value}] {code}"]
    addition_field_labels = {
        "waiver_reason": "不执行理由",
        "sample_selection": "样本选择依据",
        "exception_summary": "异常结论",
        "special_addition_source": "特殊新增来源",
        "cross_sheet_explanation": "跨表勾稽说明",
    }
    if issue.field in addition_field_labels:
        issue.field = addition_field_labels[issue.field]
    lines.append(f"判断来源: {_review_source_text(issue)}")
    if issue.field:
        lines.append(f"字段: {issue.field}")
    lines.append(issue.message)
    if issue.suggestion:
        lines.append(f"建议: {issue.suggestion}")
    return "\n".join(lines)


def _question_text(issue: QcIssue) -> str:
    """Question/Comment：仅质检问题；建议见最后一列。"""
    code = issue.dict_rule_code or issue.rule_id
    short = _short_title_for_issue(issue) or _compact_issue_message(issue.message)
    return f"[{issue.severity.value}] {code} {short}"


def _agent_suggestion(issue: QcIssue) -> str | None:
    return issue.suggestion or None


def _review_source_text(issue: QcIssue) -> str:
    source = (issue.review_source or "规则判断").strip()
    if issue.llm_review_type:
        return f"{source}（{issue.llm_review_type}）"
    return source


def _answer_for_preparer() -> None:
    """Answer/Comment 由底稿 prepare 根据 review comments 回复，Agent 不预填。"""
    return None


def _finding_issues(report: QcReport) -> list[QcIssue]:
    return [i for i in report.issues if i.severity != Severity.PASS]


def split_fa_list_issues(issues: list[QcIssue]) -> tuple[list[QcIssue], list[QcIssue]]:
    fa_list: list[QcIssue] = []
    other: list[QcIssue] = []
    for issue in issues:
        if issue.procedure_code == "FA_LIST":
            fa_list.append(issue)
        else:
            other.append(issue)
    return fa_list, other


def _aggregate_fa_list_issues(
    issues: list[QcIssue],
) -> list[tuple[QcIssue, int, list[str]]]:
    """按 (rule_id, field, severity) 合并 FA list 共性问题。"""
    buckets: dict[tuple[str, str, Severity], list[QcIssue]] = defaultdict(list)
    for issue in issues:
        key = (issue.rule_id, issue.field or "", issue.severity)
        buckets[key].append(issue)
    merged: list[tuple[QcIssue, int, list[str]]] = []
    for group in buckets.values():
        rep = group[0]
        sheets = sorted({i.source_sheet for i in group if i.source_sheet})
        merged.append((rep, len(group), sheets))
    merged.sort(
        key=lambda x: (
            _SEV_RANK.get(x[0].severity, 9),
            x[0].rule_id,
            x[0].field or "",
        )
    )
    return merged


def _fa_list_summary_question(rep: QcIssue, count: int) -> str:
    code = rep.dict_rule_code or rep.rule_id
    field_part = f"\uff08\u5b57\u6bb5\uff1a{rep.field}\uff09" if rep.field else ""
    detail = _compact_issue_message(rep.message)
    if detail:
        return (
            f"[{rep.severity.value}] {code}{field_part} \u2014 \u5171 {count} \u6761\u540c\u7c7b\u95ee\u9898\uff1b"
            f"\u4ee3\u8868\u6027\u95ee\u9898\uff1a{detail}\u3002\u8be6\u89c1\u300a{FA_LIST_COMMENTS_SHEET_NAME}\u300b"
        )
    return f"[{rep.severity.value}] {code}{field_part} \u2014 \u5171 {count} \u6761\u540c\u7c7b\u95ee\u9898\uff0c\u8be6\u89c1\u300a{FA_LIST_COMMENTS_SHEET_NAME}\u300b"


def build_main_comments_rows(
    other_issues: list[QcIssue],
    fa_list_issues: list[QcIssue],
) -> list[tuple]:
    """主汇总表：其他程序逐条 + FA list 仅共性合并行。"""
    rows: list[tuple] = []
    ey = 0

    for issue in sorted(
        other_issues,
        key=lambda i: (
            _sheet_group_rank(i),
            (i.source_sheet or "").strip(),
            _SEV_RANK.get(i.severity, 9),
            i.source_row or 0,
        ),
    ):
        ey += 1
        rows.append(
            (
                ey,
                issue.source_sheet or "—",
                _cell_ref_a1(issue.source_row, _issue_col(issue.source_col)),
                _question_text(issue),
                _answer_for_preparer(),
                "No",
                _agent_suggestion(issue),
                _review_source_text(issue),
            )
        )

    for rep, count, sheets in _aggregate_fa_list_issues(fa_list_issues):
        ey += 1
        if not sheets:
            tab = "FA list"
        elif len(sheets) == 1:
            tab = sheets[0]
        else:
            tab = f"{sheets[0]} 等（{len(sheets)} 张表）"
        rows.append(
            (
                ey,
                tab or "FA list",
                f"见附表 {FA_LIST_COMMENTS_SHEET_NAME}",
                _fa_list_summary_question(rep, count),
                _answer_for_preparer(),
                "No",
                _agent_suggestion(rep),
                _review_source_text(rep),
            )
        )

    return rows


def build_main_comments_hyperlinks(
    other_issues: list[QcIssue],
    fa_list_issues: list[QcIssue],
) -> dict[tuple[int, int], str]:
    """主 Comments 表 Cell Ref. 列内部跳转链接。"""
    links: dict[tuple[int, int], str] = {}
    row_idx = 2
    for issue in sorted(
        other_issues,
        key=lambda i: (
            _sheet_group_rank(i),
            (i.source_sheet or "").strip(),
            _SEV_RANK.get(i.severity, 9),
            i.source_row or 0,
        ),
    ):
        loc = _cell_location(issue.source_sheet, issue.source_row, issue.source_col)
        if loc:
            links[(row_idx, 3)] = loc
        row_idx += 1
    for _rep, _count, _sheets in _aggregate_fa_list_issues(fa_list_issues):
        row_idx += 1
    return links


def _compact_issue_message(message: str | None) -> str:
    text = (message or "").strip()
    if not text:
        return "发现问题，请复核。"
    text = " ".join(text.split())
    for marker in ("；模型提示", ";模型提示", "模型提示：", "模型提示"):
        if marker in text:
            text = text.split(marker, 1)[0].strip()
            break
    # 压缩超长的程序名包裹文本，保留问题本质。
    text = text.replace("程序「", "程序(").replace("」", ")")
    return text


def _short_title_for_issue(issue: QcIssue) -> str | None:
    code = (issue.dict_rule_code or issue.rule_id or "").strip().upper()
    field = (issue.field or "").strip().lower()
    if not code:
        return None
    return _SHORT_TITLE_BY_CODE_FIELD.get((code, field)) or _SHORT_TITLE_BY_CODE_FIELD.get((code, ""))


def _sheet_group_rank(issue: QcIssue) -> int:
    sheet = (issue.source_sheet or "").strip().lower()
    proc = (issue.procedure_code or "").strip().upper()
    if proc == "SUMMARY" or "汇总" in sheet:
        return 0
    if proc == "K.00" or "lead" in sheet:
        return 1
    return 2


def build_fa_list_detail_rows(fa_list_issues: list[QcIssue]) -> list[tuple]:
    sorted_issues = sorted(
        fa_list_issues,
        key=lambda i: (_SEV_RANK.get(i.severity, 9), i.source_sheet or "", i.source_row or 0),
    )
    rows: list[tuple] = []
    for idx, issue in enumerate(sorted_issues, start=1):
        rows.append(
            (
                idx,
                issue.source_sheet or "—",
                _cell_ref_a1(issue.source_row, _issue_col(issue.source_col)),
                _question_text(issue),
                _answer_for_preparer(),
                "No",
                _agent_suggestion(issue),
                _review_source_text(issue),
            )
        )
    return rows


def build_fa_list_detail_hyperlinks(
    fa_list_issues: list[QcIssue],
) -> dict[tuple[int, int], str]:
    sorted_issues = sorted(
        fa_list_issues,
        key=lambda i: (_SEV_RANK.get(i.severity, 9), i.source_sheet or "", i.source_row or 0),
    )
    links: dict[tuple[int, int], str] = {}
    for idx, issue in enumerate(sorted_issues, start=2):
        loc = _cell_location(issue.source_sheet, issue.source_row, issue.source_col)
        if loc:
            links[(idx, 3)] = loc
    return links


def build_locator_rows(issues: list[QcIssue]) -> list[tuple]:
    """导航定位表：列出全部 findings 的定位信息，便于快速检索。"""
    sorted_issues = sorted(
        issues,
        key=lambda i: (
            _sheet_group_rank(i),
            (i.source_sheet or "").strip(),
            _SEV_RANK.get(i.severity, 9),
            i.source_row or 0,
            i.rule_id,
        ),
    )
    rows: list[tuple] = []
    for idx, issue in enumerate(sorted_issues, start=1):
        code = issue.dict_rule_code or issue.rule_id
        tab = issue.source_sheet or "—"
        cell_ref = _cell_ref_a1(issue.source_row, _issue_col(issue.source_col))
        navigate = ""
        if tab != "—":
            navigate = tab if not cell_ref else f"{tab}!{cell_ref.replace('$', '')}"
        rows.append(
            (
                idx,
                issue.severity.value,
                code,
                tab,
                cell_ref,
                _question_text(issue),
                _agent_suggestion(issue),
                navigate,
                _review_source_text(issue),
            )
        )
    return rows


def build_locator_hyperlinks(issues: list[QcIssue]) -> dict[tuple[int, int], str]:
    sorted_issues = sorted(
        issues,
        key=lambda i: (
            _sheet_group_rank(i),
            (i.source_sheet or "").strip(),
            _SEV_RANK.get(i.severity, 9),
            i.source_row or 0,
            i.rule_id,
        ),
    )
    links: dict[tuple[int, int], str] = {}
    for idx, issue in enumerate(sorted_issues, start=2):
        loc = _cell_location(issue.source_sheet, issue.source_row, issue.source_col)
        if loc:
            links[(idx, 5)] = loc
            links[(idx, 8)] = loc
    return links


def build_llm_ingest_review_rows(report: QcReport) -> list[tuple]:
    """LLM 读取层复核提示，独立于业务 findings 展示。"""
    section = report.ingest_review_section or {}
    reviews = section.get("reviews") or []
    rows: list[tuple] = []
    if not isinstance(reviews, list):
        return rows
    for idx, review in enumerate(reviews, start=1):
        if not isinstance(review, dict):
            continue
        rows.append(
            (
                idx,
                review.get("procedure_code") or "",
                review.get("review_type") or "",
                review.get("assessment") or "",
                review.get("risk_level") or "",
                review.get("risk_area") or "",
                review.get("candidate_sheet") or review.get("source_sheet") or "",
                _join_list(review.get("candidate_rows")),
                _join_list(review.get("evidence_anchors")),
                review.get("rationale") or "",
                review.get("suggested_action") or "",
                review.get("manual_review_focus") or "",
                review.get("note") or "读取结果复核提示，不等同于业务规则 finding。",
            )
        )
    return rows


def build_llm_ingest_review_hyperlinks(report: QcReport) -> dict[tuple[int, int], str]:
    section = report.ingest_review_section or {}
    reviews = section.get("reviews") or []
    links: dict[tuple[int, int], str] = {}
    if not isinstance(reviews, list):
        return links
    for row_idx, review in enumerate(reviews, start=2):
        if not isinstance(review, dict):
            continue
        sheet = str(review.get("candidate_sheet") or review.get("source_sheet") or "").strip()
        rows = review.get("candidate_rows")
        first_row = rows[0] if isinstance(rows, list) and rows else None
        loc = _cell_location(sheet, first_row)
        if loc:
            links[(row_idx, 7)] = loc
            links[(row_idx, 8)] = loc
    return links


def _join_list(value) -> str:
    if isinstance(value, list):
        return ", ".join(str(v) for v in value if str(v).strip())
    if value is None:
        return ""
    return str(value)


def build_comments_rows(issues: list[QcIssue]) -> list[tuple]:
    """兼容旧接口：等同主表行（含 FA 合并逻辑）。"""
    fa, other = split_fa_list_issues(issues)
    return build_main_comments_rows(other, fa)


def _fill_for_severity(sev: Severity) -> PatternFill | None:
    if sev == Severity.FAIL:
        return _FILL_FAIL
    if sev == Severity.WARN:
        return _FILL_WARN
    if sev == Severity.NEED_REVIEW:
        return _FILL_NR
    return None


def _apply_cell_annotations(wb: openpyxl.Workbook, issues: list[QcIssue]) -> int:
    count = 0
    by_sheet: dict[str, list[QcIssue]] = {}
    for issue in issues:
        sheet = (issue.source_sheet or "").strip()
        if not sheet or sheet not in wb.sheetnames:
            continue
        by_sheet.setdefault(sheet, []).append(issue)

    for sheet_name, sheet_issues in by_sheet.items():
        ws = wb[sheet_name]
        for issue in sheet_issues:
            row = issue.source_row
            if not row or row < 1:
                continue
            anchor_row, anchor_col = _annotation_cell_position(ws, row, _issue_col(issue.source_col))
            cell = ws.cell(row=anchor_row, column=anchor_col)
            text = _issue_comment_text(issue)
            if cell.comment:
                cell.comment = Comment(f"{cell.comment.text}\n\n{text}", _AUTHOR)
            else:
                cell.comment = Comment(text, _AUTHOR)
            fill = _fill_for_severity(issue.severity)
            if fill:
                cell.fill = fill
            count += 1
    return count


def _annotation_cell_position(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    col: int = _DEFAULT_COMMENT_COL,
) -> tuple[int, int]:
    coord = f"{get_column_letter(col)}{row}"
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            return merged_range.min_row, merged_range.min_col
    return row, col


def _annotation_cell_ref(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int | None,
    col: int = _DEFAULT_COMMENT_COL,
) -> str:
    if not row or row < 1:
        return ""
    anchor_row, anchor_col = _annotation_cell_position(ws, row, col)
    return f"{get_column_letter(anchor_col)}{anchor_row}"


def _ooxml_comments_by_sheet(
    workbook_path: str | Path,
    issues: list[QcIssue],
) -> dict[str, list[tuple[str, str, str]]]:
    by_sheet: dict[str, list[tuple[str, str, str]]] = {}
    wb = openpyxl.load_workbook(workbook_path)
    try:
        for issue in issues:
            sheet = (issue.source_sheet or "").strip()
            if not sheet or sheet not in wb.sheetnames:
                continue
            cell = _annotation_cell_ref(wb[sheet], issue.source_row, _issue_col(issue.source_col))
            if not cell:
                continue
            by_sheet.setdefault(sheet, []).append((cell, _issue_comment_text(issue), _AUTHOR))
    finally:
        wb.close()
    return by_sheet


def export_annotated_workbook(
    report: QcReport,
    input_path: str | Path,
    output_path: str | Path | None = None,
) -> Path:
    """
    生成带标注底稿副本：

    - Sheet1 ``Comments【归档前删除】``：其他程序 findings 逐条 + FA list 共性问题合并行
    - Sheet2 ``Comments【FA list】``：FA list findings 逐条明细
    - 业务 sheet：单元格批注（有 source_row 时；**含外部链接时不 save**，避免 A3 #REF）

    Comments 表通过 ZIP/OXML 注入，不调用 ``openpyxl.save`` 整本保存。
    """
    input_path = Path(input_path)
    if input_path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError("仅支持 Excel 底稿（.xlsx / .xlsm）生成标注副本")

    out = Path(output_path) if output_path else annotated_workbook_path(input_path)
    shutil.copy2(input_path, out)

    issues = _finding_issues(report)
    fa_issues, other_issues = split_fa_list_issues(issues)
    overall = worst_severity([i.severity for i in issues]) if issues else Severity.PASS
    footer = (
        f" | 整体 {overall.value}"
        f" | 主表: 其他 {len(other_issues)} 条 + FA 共性 {len(_aggregate_fa_list_issues(fa_issues))} 行"
        f" | FA 明细 {len(fa_issues)} 条"
    )
    has_external = workbook_has_external_links(input_path)
    if has_external:
        footer += " | 业务表批注使用 OOXML 注入（保留 A3 等外部链接）"

    main_rows = build_main_comments_rows(other_issues, fa_issues)
    main_links = build_main_comments_hyperlinks(other_issues, fa_issues)
    fa_rows = build_fa_list_detail_rows(fa_issues)
    fa_links = build_fa_list_detail_hyperlinks(fa_issues)
    locator_rows = build_locator_rows(issues)
    locator_links = build_locator_hyperlinks(issues)
    llm_ingest_rows = build_llm_ingest_review_rows(report)
    llm_ingest_links = build_llm_ingest_review_hyperlinks(report)
    inject_worksheets_at_front(
        out,
        [
            (
                COMMENTS_SHEET_NAME,
                build_worksheet_xml(
                    _COMMENT_HEADERS,
                    main_rows,
                    footer=f"源文件: {input_path.name}{footer}",
                    hyperlinks=main_links,
                ),
            ),
            (
                FA_LIST_COMMENTS_SHEET_NAME,
                build_worksheet_xml(
                    _COMMENT_HEADERS,
                    fa_rows,
                    footer=f"源文件: {input_path.name} | FA list 专项明细",
                    hyperlinks=fa_links,
                ),
            ),
            (
                LOCATOR_SHEET_NAME,
                build_worksheet_xml(
                    _LOCATOR_HEADERS,
                    locator_rows,
                    footer=(
                        f"源文件: {input_path.name} | 定位表用于快速检索 Tab/Cell，"
                        "业务表批注通过 OOXML 写入（兼容外部链接底稿）"
                    ),
                    hyperlinks=locator_links,
                ),
            ),
            (
                LLM_INGEST_REVIEW_SHEET_NAME,
                build_worksheet_xml(
                    _LLM_INGEST_REVIEW_HEADERS,
                    llm_ingest_rows,
                    footer=(
                        f"源文件: {input_path.name} | 读取层复核提示仅用于判断 Agent 是否读对底稿，"
                        "不等同于业务规则 finding"
                    ),
                    hyperlinks=llm_ingest_links,
                ),
            ),
        ],
        remove_sheet_names=(
            COMMENTS_SHEET_NAME,
            FA_LIST_COMMENTS_SHEET_NAME,
            LOCATOR_SHEET_NAME,
            LLM_INGEST_REVIEW_SHEET_NAME,
        ),
    )

    annotation_result = inject_cell_comments(out, _ooxml_comments_by_sheet(out, issues))
    skipped = annotation_result.get("skipped_sheets") or []
    if skipped and not has_external:
        wb = openpyxl.load_workbook(out)
        _apply_cell_annotations(wb, issues)
        wb.save(out)
        wb.close()

    return out


def comments_summary_stats(
    report: QcReport,
    *,
    source_path: str | Path | None = None,
) -> dict[str, int | str | bool]:
    issues = _finding_issues(report)
    fa_issues, other_issues = split_fa_list_issues(issues)
    overall = worst_severity([i.severity for i in issues]) if issues else Severity.PASS
    has_external = (
        workbook_has_external_links(source_path) if source_path is not None else False
    )
    return {
        "overall_severity": overall.value,
        "finding_count": len(issues),
        "fail_count": sum(1 for i in issues if i.severity == Severity.FAIL),
        "warn_count": sum(1 for i in issues if i.severity == Severity.WARN),
        "need_review_count": sum(1 for i in issues if i.severity == Severity.NEED_REVIEW),
        "comments_sheet": COMMENTS_SHEET_NAME,
        "fa_list_comments_sheet": FA_LIST_COMMENTS_SHEET_NAME,
        "locator_sheet": LOCATOR_SHEET_NAME,
        "llm_ingest_review_sheet": LLM_INGEST_REVIEW_SHEET_NAME,
        "llm_ingest_review_count": len(build_llm_ingest_review_rows(report)),
        "other_finding_count": len(other_issues),
        "fa_list_finding_count": len(fa_issues),
        "fa_list_summary_row_count": len(_aggregate_fa_list_issues(fa_issues)),
        "has_external_links": has_external,
        "cell_annotations_applied": True,
    }
