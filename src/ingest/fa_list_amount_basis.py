from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from itertools import product
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formula import Tokenizer

from ingest.field_mapping import match_standard_field
from ingest.header_detection import scan_rows_for_headers
from ingest.models import (
    AmountCurrencyRole,
    AmountPeriodRole,
    FaListAmountBasis,
    FaListAmountBasisSource,
    FaListAmountBasisStatus,
    SheetKind,
)

_MEASURES = ("original_value", "accumulated_depreciation", "impairment_provision", "net_value")
_REQUIRED = {"original_value", "accumulated_depreciation", "net_value"}
_REF_RE = re.compile(
    r"^(?:(?:'((?:[^']|'')+)'|([^!]+))!)?\$?([A-Z]{1,3})(?:\$?(\d+))?:\$?([A-Z]{1,3})(?:\$?(\d+))?$"
)


@dataclass(frozen=True)
class _FormulaCriterion:
    range_ref: tuple[str | None, int, int | None, int | None]
    expression: str


def resolve_fa_list_amount_basis(
    path: str | Path,
    *,
    fa_sheet: str,
    k01_sheet: str | None,
    data_rows: list[tuple[Any, ...]],
    k01_route_reason: str | None = None,
) -> FaListAmountBasis:
    header_row, header_cells, _ = scan_rows_for_headers(data_rows, sheet_kind=SheetKind.FA_LIST)
    header_basis = resolve_unique_header_basis(header_cells, header_row=header_row)
    formula_wb = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        fa_ws = formula_wb[fa_sheet]
        if k01_route_reason:
            k01_basis = FaListAmountBasis(
                status=FaListAmountBasisStatus.AMBIGUOUS,
                conflicts=[k01_route_reason],
            )
        else:
            k01_basis = (
                _resolve_k01_formula_basis(formula_wb[k01_sheet], fa_sheet, header_cells)
                if k01_sheet and k01_sheet in formula_wb.sheetnames
                else FaListAmountBasis(status=FaListAmountBasisStatus.NOT_FOUND)
            )
        summary_basis = _resolve_fa_summary_basis(
            fa_ws,
            header_row,
            detail_header_cells=header_cells,
            expected_basis=(
                k01_basis
                if _can_supply_k01_bindings(k01_basis)
                else None
            ),
        )
        k01_basis = _validate_basis_semantics(fa_ws, header_cells, k01_basis)
        summary_basis = _validate_basis_semantics(fa_ws, header_cells, summary_basis)
        if (
            k01_basis.status == FaListAmountBasisStatus.INCOMPLETE
            and _can_supply_k01_bindings(k01_basis)
            and summary_basis.status == FaListAmountBasisStatus.CONFIRMED
        ):
            conflicts = _basis_conflicts(k01_basis, summary_basis)
            if conflicts:
                k01_basis = FaListAmountBasis(
                    status=FaListAmountBasisStatus.AMBIGUOUS,
                    evidence=[*k01_basis.evidence, *summary_basis.evidence],
                    conflicts=conflicts,
                )
            else:
                k01_basis = _complete_k01_basis(k01_basis, summary_basis)
                k01_basis = _validate_basis_semantics(fa_ws, header_cells, k01_basis)
        if header_basis.status == FaListAmountBasisStatus.CONFIRMED:
            header_basis.data_end_row = len(data_rows)
        header_basis = _validate_basis_semantics(fa_ws, header_cells, header_basis)
    finally:
        formula_wb.close()

    if k01_basis.status == FaListAmountBasisStatus.AMBIGUOUS:
        return k01_basis
    if k01_basis.status == FaListAmountBasisStatus.INCOMPLETE:
        return summary_basis if summary_basis.status == FaListAmountBasisStatus.AMBIGUOUS else k01_basis
    if k01_basis.status == FaListAmountBasisStatus.CONFIRMED:
        if summary_basis.status == FaListAmountBasisStatus.AMBIGUOUS:
            return summary_basis
        if summary_basis.status == FaListAmountBasisStatus.CONFIRMED:
            conflicts = _basis_conflicts(k01_basis, summary_basis)
            if conflicts:
                return FaListAmountBasis(
                    status=FaListAmountBasisStatus.AMBIGUOUS,
                    evidence=[*k01_basis.evidence, *summary_basis.evidence],
                    conflicts=conflicts,
                )
            if k01_basis.data_start_row is None:
                k01_basis.data_start_row = summary_basis.data_start_row
            if k01_basis.data_end_row is None:
                k01_basis.data_end_row = summary_basis.data_end_row
            k01_basis.evidence.extend(summary_basis.evidence)
        return k01_basis
    if summary_basis.status in {
        FaListAmountBasisStatus.CONFIRMED,
        FaListAmountBasisStatus.AMBIGUOUS,
        FaListAmountBasisStatus.INCOMPLETE,
    }:
        return summary_basis
    return header_basis


def resolve_unique_header_basis(
    header_cells: list[tuple[int, str]], *, header_row: int | None
) -> FaListAmountBasis:
    candidates: dict[str, list[int]] = {measure: [] for measure in _MEASURES}
    category_column: int | None = None
    for column, text in header_cells:
        field = match_standard_field(text, SheetKind.FA_LIST)
        if field in candidates:
            candidates[field].append(column)
        elif field == "asset_category":
            category_column = column
    duplicated = [field for field, columns in candidates.items() if len(columns) > 1]
    if duplicated:
        return FaListAmountBasis(
            status=FaListAmountBasisStatus.AMBIGUOUS,
            evidence=[f"header candidates: {field}={candidates[field]}" for field in duplicated],
            conflicts=["multiple amount header candidates"],
        )
    bindings = {field: columns[0] for field, columns in candidates.items() if len(columns) == 1}
    if not _REQUIRED.issubset(bindings):
        return FaListAmountBasis(
            status=FaListAmountBasisStatus.INCOMPLETE,
            bindings=bindings,
            category_column=category_column,
            conflicts=["required amount headers are incomplete"],
        )
    return FaListAmountBasis(
        status=FaListAmountBasisStatus.CONFIRMED,
        source=FaListAmountBasisSource.UNIQUE_HEADERS,
        bindings=bindings,
        category_column=category_column,
        data_start_row=(header_row + 1) if header_row else None,
        evidence=["unique FA list amount headers"],
    )


def _resolve_k01_formula_basis(
    ws,
    fa_sheet: str,
    source_header_cells: list[tuple[int, str]],
) -> FaListAmountBasis:
    anchors: list[tuple[int, int]] = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), max_col=min(ws.max_column, 60)):
        for cell in row:
            if _norm(cell.value) == "表2":
                anchors.append((cell.row, cell.column))
    if not anchors:
        return FaListAmountBasis(status=FaListAmountBasisStatus.NOT_FOUND)

    results: list[FaListAmountBasis] = []
    for anchor_row, anchor_col in anchors:
        right_boundary = min(ws.max_column, anchor_col + 12)
        for row in ws.iter_rows(
            min_row=max(1, anchor_row - 1),
            max_row=min(ws.max_row, anchor_row + 4),
            min_col=anchor_col + 1,
            max_col=min(ws.max_column, anchor_col + 20),
        ):
            table3_cols = [cell.column for cell in row if _norm(cell.value) == "表3"]
            if table3_cols:
                right_boundary = min(right_boundary, min(table3_cols) - 1)
        for header_row in range(anchor_row, min(ws.max_row, anchor_row + 8) + 1):
            duplicate_fields = _duplicate_measure_fields(
                ws, header_row, anchor_col, right_boundary
            )
            if duplicate_fields:
                results.append(
                    FaListAmountBasis(
                        status=FaListAmountBasisStatus.AMBIGUOUS,
                        conflicts=[f"duplicate K.01 amount headers: {duplicate_fields}"],
                    )
                )
                continue
            columns = _measure_columns(ws, header_row, anchor_col, right_boundary)
            if not _REQUIRED.issubset(columns):
                continue
            results.append(
                _formula_table_basis(
                    ws,
                    header_row,
                    columns,
                    fa_sheet,
                    source_header_cells,
                )
            )
    ambiguous = [item for item in results if item.status == FaListAmountBasisStatus.AMBIGUOUS]
    if ambiguous:
        return ambiguous[0]
    candidates = [
        item
        for item in results
        if item.status in {
            FaListAmountBasisStatus.CONFIRMED,
            FaListAmountBasisStatus.INCOMPLETE,
        }
    ]
    unique = {
        (
            tuple(sorted(item.bindings.items())),
            item.category_column,
            item.data_start_row,
            item.data_end_row,
            item.criteria_columns,
        ): item
        for item in candidates
    }
    if len(unique) != 1:
        return FaListAmountBasis(
            status=FaListAmountBasisStatus.AMBIGUOUS if unique else FaListAmountBasisStatus.NOT_FOUND,
            conflicts=["multiple K.01 table 2 candidates"] if len(unique) > 1 else [],
        )
    return next(iter(unique.values()))


def _formula_table_basis(
    ws,
    header_row: int,
    columns: dict[str, int],
    fa_sheet: str,
    source_header_cells: list[tuple[int, str]],
) -> FaListAmountBasis:
    measure_refs: dict[str, set[int]] = {measure: set() for measure in columns}
    category_refs: set[int] = set()
    data_ranges: set[tuple[int, int]] = set()
    criteria_signatures: set[tuple[tuple[str, int, int | None, int | None], ...]] = set()
    criteria_value_signatures: dict[int, set[tuple]] = defaultdict(set)
    currency_criteria_values: set[str] = set()
    source_fields = {
        column: match_standard_field(text, SheetKind.FA_LIST)
        for column, text in source_header_cells
    }
    formula_cells: list[str] = []
    incomplete_rows: list[int] = []
    invalid_rows: list[int] = []
    first_detail_row: int | None = None
    for row_no in range(header_row + 1, min(ws.max_row, header_row + 40) + 1):
        if (
            first_detail_row is not None
            and _is_k01_table_total_row(
                ws,
                row_no=row_no,
                columns=columns,
                first_detail_row=first_detail_row,
            )
        ):
            break
        row_measures: set[str] = set()
        for measure, column in columns.items():
            value = ws.cell(row_no, column).value
            if not isinstance(value, str) or not value.startswith("="):
                continue
            parsed = _summary_formula_refs(value, current_sheet=ws.title)
            if parsed is None:
                invalid_rows.append(row_no)
                continue
            amount_ref, criteria = parsed
            criteria_refs = tuple(item.range_ref for item in criteria)
            if not criteria_refs or not all(
                _same_sheet(ref[0], fa_sheet) for ref in (amount_ref, *criteria_refs)
            ):
                invalid_rows.append(row_no)
                continue
            if any(
                (amount_ref[2], amount_ref[3]) != (ref[2], ref[3])
                for ref in criteria_refs
            ):
                invalid_rows.append(row_no)
                continue
            row_measures.add(measure)
            measure_refs[measure].add(amount_ref[1])
            if amount_ref[2] is not None and amount_ref[3] is not None:
                data_ranges.add((amount_ref[2], amount_ref[3]))
            category_candidates = [
                item.range_ref[1]
                for item in criteria
                if source_fields.get(item.range_ref[1]) == "asset_category"
            ]
            if len(category_candidates) != 1:
                invalid_rows.append(row_no)
                continue
            category_refs.add(category_candidates[0])
            criteria_signatures.add(
                tuple(
                    (_norm(ref[0]), ref[1], ref[2], ref[3])
                    for ref in criteria_refs
                )
            )
            criteria_value_signatures[row_no].add(
                tuple(
                    (
                        _norm(item.range_ref[0]),
                        item.range_ref[1],
                        item.range_ref[2],
                        item.range_ref[3],
                        item.expression,
                    )
                    for item in criteria
                )
            )
            currency_criteria_values.update(
                item.expression
                for item in criteria
                if source_fields.get(item.range_ref[1]) == "currency"
            )
            formula_cells.append(ws.cell(row_no, column).coordinate)
        if row_measures and not _REQUIRED.issubset(row_measures):
            incomplete_rows.append(row_no)
        elif _REQUIRED.issubset(row_measures) and first_detail_row is None:
            first_detail_row = row_no
        if formula_cells and not row_measures:
            break
    conflicts = [measure for measure, refs in measure_refs.items() if len(refs) != 1]
    if (
        conflicts
        or len(category_refs) > 1
        or len(data_ranges) > 1
        or len(criteria_signatures) > 1
        or any(len(signatures) > 1 for signatures in criteria_value_signatures.values())
        or len(currency_criteria_values) > 1
        or incomplete_rows
        or invalid_rows
    ):
        return FaListAmountBasis(
            status=FaListAmountBasisStatus.AMBIGUOUS,
            evidence=[f"K.01 table2 formulas: {formula_cells[:12]}"],
            conflicts=(
                [f"inconsistent formula references: {conflicts}"]
                + (["inconsistent category column"] if len(category_refs) > 1 else [])
                + (["inconsistent formula row ranges"] if len(data_ranges) > 1 else [])
                + (["inconsistent SUMIFS criteria dimensions"] if len(criteria_signatures) > 1 else [])
                + (["amount formulas use different criteria values in the same K.01 row"] if any(len(signatures) > 1 for signatures in criteria_value_signatures.values()) else [])
                + (["amount formulas use inconsistent currency criteria"] if len(currency_criteria_values) > 1 else [])
                + ([f"incomplete category formula rows: {sorted(set(incomplete_rows))}"] if incomplete_rows else [])
                + ([f"unsupported or cross-sheet formula rows: {sorted(set(invalid_rows))}"] if invalid_rows else [])
            ),
        )
    bindings = {measure: next(iter(refs)) for measure, refs in measure_refs.items() if refs}
    criteria_columns = tuple(
        ref[1]
        for ref in next(iter(criteria_signatures), ())
    )
    if not _REQUIRED.issubset(bindings):
        return FaListAmountBasis(
            status=FaListAmountBasisStatus.INCOMPLETE,
            source=FaListAmountBasisSource.K01_FORMULA,
            bindings=bindings,
            category_column=next(iter(category_refs)) if category_refs else None,
            criteria_columns=criteria_columns,
            period_role=AmountPeriodRole.ENDING,
            currency_values=tuple(sorted(currency_criteria_values)),
            evidence=[f"K.01 table2 formulas: {formula_cells[:12]}"],
            conflicts=["required K.01 amount formula bindings are incomplete"],
        )
    if not data_ranges:
        return FaListAmountBasis(
            status=FaListAmountBasisStatus.INCOMPLETE,
            source=FaListAmountBasisSource.K01_FORMULA,
            bindings=bindings,
            category_column=next(iter(category_refs)) if category_refs else None,
            criteria_columns=criteria_columns,
            period_role=AmountPeriodRole.ENDING,
            currency_values=tuple(sorted(currency_criteria_values)),
            evidence=[f"K.01 table2 formulas: {formula_cells[:12]}"],
            conflicts=["K.01 formulas do not provide a bounded FA detail range"],
        )
    return FaListAmountBasis(
        status=FaListAmountBasisStatus.CONFIRMED,
        source=FaListAmountBasisSource.K01_FORMULA,
        bindings=bindings,
        category_column=next(iter(category_refs)) if category_refs else None,
        data_start_row=next(iter(data_ranges))[0] if data_ranges else None,
        data_end_row=next(iter(data_ranges))[1] if data_ranges else None,
        criteria_columns=criteria_columns,
        currency_values=tuple(sorted(currency_criteria_values)),
        period_role=AmountPeriodRole.ENDING,
        evidence=[f"K.01 table2 formulas: {formula_cells[:12]}"],
    )


@dataclass(frozen=True)
class _SummaryColumnEvidence:
    measure: str
    output_column: int
    amount_column: int
    category_column: int
    start_row: int | None
    end_row: int | None
    criteria_columns: tuple[int, ...]
    criteria_signature: tuple[tuple[int, str, str], ...]


def _resolve_fa_summary_basis(
    ws,
    detail_header_row: int | None,
    *,
    detail_header_cells: list[tuple[int, str]],
    expected_basis: FaListAmountBasis | None,
) -> FaListAmountBasis:
    if not detail_header_row or detail_header_row <= 1:
        return FaListAmountBasis(status=FaListAmountBasisStatus.NOT_FOUND)
    candidates: list[_SummaryColumnEvidence] = []
    saw_summary_formula = False
    saw_invalid_summary_formula = False
    source_fields = {
        column: match_standard_field(text, SheetKind.FA_LIST)
        for column, text in detail_header_cells
    }
    for header_row in range(1, detail_header_row):
        for column in range(1, min(ws.max_column, 60) + 1):
            measure = match_standard_field(
                str(ws.cell(header_row, column).value or ""), SheetKind.FA_LIST
            )
            if measure not in _MEASURES:
                continue
            signatures: set[
                tuple[
                    int,
                    tuple[int, ...],
                    int | None,
                    int | None,
                    tuple[tuple[int, str, str], ...],
                ]
            ] = set()
            invalid = False
            for row_no in range(header_row + 1, detail_header_row):
                value = ws.cell(row_no, column).value
                if not (
                    isinstance(value, str)
                    and value.startswith("=")
                    and "SUMIF" in value.upper()
                ):
                    continue
                saw_summary_formula = True
                parsed = _summary_formula_refs(value, current_sheet=ws.title)
                if parsed is None:
                    invalid = True
                    continue
                amount_ref, criteria = parsed
                criteria_refs = tuple(item.range_ref for item in criteria)
                if not criteria_refs:
                    invalid = True
                    continue
                category_ref = criteria_refs[0]
                category_candidates = [
                    item.range_ref
                    for item in criteria
                    if source_fields.get(item.range_ref[1]) == "asset_category"
                ]
                if len(category_candidates) != 1:
                    invalid = True
                    continue
                category_ref = category_candidates[0]
                if not (
                    all(_same_sheet(ref[0], ws.title) for ref in (amount_ref, *criteria_refs))
                    and all(
                        (amount_ref[2], amount_ref[3]) == (ref[2], ref[3])
                        for ref in criteria_refs
                    )
                ):
                    invalid = True
                    continue
                signatures.add((
                    amount_ref[1],
                    tuple(ref[1] for ref in criteria_refs),
                    amount_ref[2],
                    amount_ref[3],
                    tuple(
                        (
                            item.range_ref[1],
                            source_fields.get(item.range_ref[1]) or "unknown",
                            (
                                item.expression
                                if source_fields.get(item.range_ref[1]) == "currency"
                                else "{row-dimension}"
                            ),
                        )
                        for item in criteria
                    ),
                ))
            if invalid or len(signatures) != 1:
                if invalid:
                    saw_invalid_summary_formula = True
                continue
            amount_column, criteria_columns, start_row, end_row, criteria_signature = next(iter(signatures))
            candidates.append(
                _SummaryColumnEvidence(
                    measure=measure,
                    output_column=column,
                    amount_column=amount_column,
                    category_column=category_ref[1],
                    start_row=start_row,
                    end_row=end_row,
                    criteria_columns=criteria_columns,
                    criteria_signature=criteria_signature,
                )
            )

    results = _summary_basis_candidates(candidates, expected_basis=expected_basis)
    if len(results) == 1:
        return results[0]
    if len(results) > 1:
        return FaListAmountBasis(
            status=FaListAmountBasisStatus.AMBIGUOUS,
            conflicts=["multiple FA summary formula groups"],
        )
    if expected_basis is not None and candidates:
        return FaListAmountBasis(
            status=FaListAmountBasisStatus.AMBIGUOUS,
            conflicts=["K.01 and FA summary disagree: amount bindings or category column"],
        )
    if saw_summary_formula:
        return FaListAmountBasis(
            status=FaListAmountBasisStatus.AMBIGUOUS,
            conflicts=[
                (
                    "FA list summary formula uses cross-sheet or inconsistent ranges"
                    if saw_invalid_summary_formula
                    else "FA summary formulas do not form one complete matching amount group"
                )
            ],
        )
    return FaListAmountBasis(status=FaListAmountBasisStatus.NOT_FOUND)


def _summary_basis_candidates(
    candidates: list[_SummaryColumnEvidence],
    *,
    expected_basis: FaListAmountBasis | None,
) -> list[FaListAmountBasis]:
    by_measure = {
        measure: [candidate for candidate in candidates if candidate.measure == measure]
        for measure in _MEASURES
    }
    required_measures = list(_REQUIRED)
    if expected_basis and "impairment_provision" in expected_basis.bindings:
        required_measures.append("impairment_provision")
    elif by_measure["impairment_provision"]:
        required_measures.append("impairment_provision")
    if any(not by_measure[measure] for measure in required_measures):
        return []

    unique: dict[tuple, FaListAmountBasis] = {}
    for group in product(*(by_measure[measure] for measure in required_measures)):
        if len({item.output_column for item in group}) != len(group):
            continue
        category_columns = {item.category_column for item in group}
        ranges = {(item.start_row, item.end_row) for item in group}
        criteria_groups = {item.criteria_columns for item in group}
        criteria_signatures = {item.criteria_signature for item in group}
        if (
            len(category_columns) != 1
            or len(ranges) != 1
            or len(criteria_groups) != 1
            or len(criteria_signatures) != 1
        ):
            continue
        bindings = {item.measure: item.amount_column for item in group}
        category_column = next(iter(category_columns))
        if expected_basis is not None:
            if any(
                bindings.get(measure) != column
                for measure, column in expected_basis.bindings.items()
            ):
                continue
            if (
                expected_basis.category_column is not None
                and category_column != expected_basis.category_column
            ):
                continue
        start_row, end_row = next(iter(ranges))
        if start_row is None or end_row is None:
            continue
        signature = (
            tuple(sorted(bindings.items())),
            category_column,
            start_row,
            end_row,
            next(iter(criteria_groups)),
        )
        unique[signature] = FaListAmountBasis(
            status=FaListAmountBasisStatus.CONFIRMED,
            source=FaListAmountBasisSource.FA_SUMMARY_FORMULA,
            bindings=bindings,
            category_column=category_column,
            data_start_row=start_row,
            data_end_row=end_row,
            criteria_columns=next(iter(criteria_groups)),
            currency_values=tuple(
                sorted(
                    value
                    for _, field, value in next(iter(criteria_signatures))
                    if field == "currency"
                )
            ),
            evidence=["FA list summary formula group matched by source columns"],
        )
    return list(unique.values())


def _measure_columns(ws, row_no: int, left: int, right: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for column in range(left, right + 1):
        field = match_standard_field(str(ws.cell(row_no, column).value or ""), SheetKind.FA_LIST)
        if field not in _MEASURES:
            continue
        result.setdefault(field, column)
    return result


def _duplicate_measure_fields(ws, row_no: int, left: int, right: int) -> list[str]:
    counts: dict[str, int] = {}
    for column in range(left, right + 1):
        field = match_standard_field(str(ws.cell(row_no, column).value or ""), SheetKind.FA_LIST)
        if field in _MEASURES:
            counts[field] = counts.get(field, 0) + 1
    return sorted(field for field, count in counts.items() if count > 1)


def _is_k01_table_total_row(
    ws,
    *,
    row_no: int,
    columns: dict[str, int],
    first_detail_row: int,
) -> bool:
    if row_no <= first_detail_row:
        return False
    for column in columns.values():
        formula = ws.cell(row_no, column).value
        if not isinstance(formula, str) or not formula.startswith("="):
            return False
        refs = _formula_refs(formula, current_sheet=ws.title)
        if len(refs) != 1:
            return False
        sheet, ref_column, start_row, end_row = refs[0]
        if not _same_sheet(sheet, ws.title):
            return False
        if ref_column != column or start_row != first_detail_row or end_row != row_no - 1:
            return False
    return True


def _formula_refs(formula: str, *, current_sheet: str) -> list[tuple[str | None, int, int | None, int | None]]:
    if "INDIRECT" in formula.upper() or "[" in formula:
        return []
    refs: list[tuple[str | None, int, int | None, int | None]] = []
    try:
        items = Tokenizer(formula).items
    except Exception:
        return []
    for item in items:
        if item.type != "OPERAND" or item.subtype != "RANGE":
            continue
        match = _REF_RE.match(item.value)
        if not match:
            continue
        quoted_sheet, plain_sheet, start_col, start_row, end_col, end_row = match.groups()
        if start_col != end_col:
            continue
        sheet = (quoted_sheet or plain_sheet or current_sheet).replace("''", "'")
        refs.append((sheet, _column_number(start_col), int(start_row) if start_row else None, int(end_row) if end_row else None))
    return refs


def _summary_formula_refs(
    formula: str, *, current_sheet: str
) -> tuple[
    tuple[str | None, int, int | None, int | None],
    tuple[_FormulaCriterion, ...],
] | None:
    function_name, arguments = _summary_function_arguments(formula)
    if function_name == "SUMIFS" and len(arguments) >= 3 and len(arguments) % 2 == 1:
        amount_ref = _single_range_ref(arguments[0], current_sheet)
        pairs = zip(arguments[1::2], arguments[2::2])
    elif function_name == "SUMIF" and len(arguments) == 3:
        amount_ref = _single_range_ref(arguments[2], current_sheet)
        pairs = [(arguments[0], arguments[1])]
    else:
        return None
    if amount_ref is None:
        return None
    criteria: list[_FormulaCriterion] = []
    for range_expression, criterion_expression in pairs:
        range_ref = _single_range_ref(range_expression, current_sheet)
        if range_ref is None:
            return None
        criteria.append(
            _FormulaCriterion(
                range_ref=range_ref,
                expression=_normalize_criterion_expression(criterion_expression),
            )
        )
    if criteria:
        return amount_ref, tuple(criteria)
    return None


def _summary_function_arguments(formula: str) -> tuple[str | None, list[str]]:
    try:
        items = Tokenizer(formula).items
    except Exception:
        return None, []
    active = False
    depth = 0
    function_name: str | None = None
    arguments: list[str] = []
    current: list[str] = []
    completed = False
    for item in items:
        if completed:
            if item.type == "WSPACE":
                continue
            return None, []
        if item.type == "FUNC" and item.subtype == "OPEN":
            name = item.value[:-1].upper()
            if not active and name in {"SUMIF", "SUMIFS"}:
                active = True
                depth = 1
                function_name = name
                continue
            if active:
                depth += 1
        if not active:
            continue
        if item.type == "FUNC" and item.subtype == "CLOSE":
            depth -= 1
            if depth == 0:
                arguments.append("".join(current))
                active = False
                completed = True
                continue
        if item.type == "SEP" and item.subtype == "ARG" and depth == 1:
            arguments.append("".join(current))
            current = []
            continue
        current.append(item.value)
    return (function_name, arguments) if completed else (None, [])


def _single_range_ref(
    expression: str,
    current_sheet: str,
) -> tuple[str | None, int, int | None, int | None] | None:
    refs = _formula_refs("=" + expression, current_sheet=current_sheet)
    return refs[0] if len(refs) == 1 else None


def _normalize_criterion_expression(expression: str) -> str:
    return re.sub(r"\s+", "", str(expression or "")).casefold()


def _basis_conflicts(left: FaListAmountBasis, right: FaListAmountBasis) -> list[str]:
    conflicts = [
        measure for measure in _REQUIRED | {"impairment_provision"}
        if measure in left.bindings and measure in right.bindings and left.bindings[measure] != right.bindings[measure]
    ]
    if left.category_column and right.category_column and left.category_column != right.category_column:
        conflicts.append("asset_category")
    if (
        left.data_start_row is not None
        and right.data_start_row is not None
        and left.data_start_row != right.data_start_row
    ) or (
        left.data_end_row is not None
        and right.data_end_row is not None
        and left.data_end_row != right.data_end_row
    ):
        conflicts.append("data_range")
    if left.criteria_columns and right.criteria_columns and left.criteria_columns != right.criteria_columns:
        conflicts.append("criteria_dimensions")
    if (
        left.currency_values
        and right.currency_values
        and left.currency_values != right.currency_values
    ):
        conflicts.append("currency_criteria_values")
    return [f"K.01 and FA summary disagree: {', '.join(sorted(conflicts))}"] if conflicts else []


def _can_supply_k01_bindings(basis: FaListAmountBasis) -> bool:
    return bool(
        basis.source == FaListAmountBasisSource.K01_FORMULA
        and basis.status in {
            FaListAmountBasisStatus.CONFIRMED,
            FaListAmountBasisStatus.INCOMPLETE,
        }
        and _REQUIRED.issubset(basis.bindings)
        and basis.category_column is not None
    )


def _complete_k01_basis(
    k01_basis: FaListAmountBasis,
    summary_basis: FaListAmountBasis,
) -> FaListAmountBasis:
    return FaListAmountBasis(
        status=FaListAmountBasisStatus.CONFIRMED,
        source=FaListAmountBasisSource.K01_FORMULA,
        bindings=dict(k01_basis.bindings),
        category_column=k01_basis.category_column,
        data_start_row=summary_basis.data_start_row,
        data_end_row=summary_basis.data_end_row,
        evidence=[
            *k01_basis.evidence,
            *summary_basis.evidence,
            "K.01 formula bindings completed with matching bounded FA summary range",
        ],
        period_role=k01_basis.period_role,
        currency_role=k01_basis.currency_role,
        criteria_columns=k01_basis.criteria_columns,
        currency_values=k01_basis.currency_values,
    )


def _validate_basis_semantics(ws, header_cells, basis: FaListAmountBasis) -> FaListAmountBasis:
    if basis.status != FaListAmountBasisStatus.CONFIRMED:
        return basis
    if basis.data_start_row is None or basis.data_end_row is None:
        basis.status = FaListAmountBasisStatus.INCOMPLETE
        basis.conflicts.append("amount basis does not provide a bounded detail range")
        return basis

    header_by_column = {column: str(text or "") for column, text in header_cells}
    period_by_column = [
        _header_period_role(header_by_column.get(column, ""))
        for column in basis.bindings.values()
    ]
    currency_by_column = [
        _header_currency_role(header_by_column.get(column, ""))
        for column in basis.bindings.values()
    ]
    period_roles = {role for role in period_by_column if role != AmountPeriodRole.UNKNOWN}
    currency_roles = {role for role in currency_by_column if role != AmountCurrencyRole.UNKNOWN}
    if len(period_roles) > 1 or len(currency_roles) > 1:
        basis.status = FaListAmountBasisStatus.AMBIGUOUS
        basis.conflicts.append("selected amount columns mix period or currency semantics")
        return basis
    if period_roles and any(role == AmountPeriodRole.UNKNOWN for role in period_by_column):
        basis.status = FaListAmountBasisStatus.INCOMPLETE
        basis.conflicts.append("period semantics are explicit for only part of the amount group")
        return basis
    if currency_roles and any(role == AmountCurrencyRole.UNKNOWN for role in currency_by_column):
        basis.status = FaListAmountBasisStatus.INCOMPLETE
        basis.conflicts.append("currency semantics are explicit for only part of the amount group")
        return basis
    if period_roles:
        basis.period_role = next(iter(period_roles))
    if currency_roles:
        basis.currency_role = next(iter(currency_roles))

    currency_columns = [
        column
        for column, text in header_cells
        if match_standard_field(text, SheetKind.FA_LIST) == "currency"
    ]
    if len(currency_columns) == 1:
        currency_column = currency_columns[0]
        values = {
            _norm(ws.cell(row, currency_column).value)
            for row in range(basis.data_start_row, basis.data_end_row + 1)
            if _norm(ws.cell(row, currency_column).value)
        }
        basis.currency_values = tuple(sorted(values))
        if (
            len(values) > 1
            and basis.currency_role != AmountCurrencyRole.REPORTING
            and currency_column not in basis.criteria_columns
        ):
            basis.status = FaListAmountBasisStatus.INCOMPLETE
            basis.conflicts.append("multiple currencies exist but K.01 formulas do not constrain currency")
        elif len(values) > 1 and basis.currency_role == AmountCurrencyRole.UNKNOWN:
            basis.currency_role = AmountCurrencyRole.ORIGINAL
        elif (
            len(values) <= 1
            and basis.currency_role == AmountCurrencyRole.UNKNOWN
            and basis.source == FaListAmountBasisSource.K01_FORMULA
        ):
            basis.currency_role = AmountCurrencyRole.REPORTING
        elif len(values) == 1 and basis.currency_role == AmountCurrencyRole.UNKNOWN:
            basis.currency_role = AmountCurrencyRole.ORIGINAL
            basis.evidence.append("one row-level currency value establishes a single-currency amount group")
    elif len(currency_columns) > 1 and basis.currency_role == AmountCurrencyRole.UNKNOWN:
        basis.status = FaListAmountBasisStatus.AMBIGUOUS
        basis.conflicts.append("multiple currency columns exist without an explicit amount currency role")
    elif basis.currency_role == AmountCurrencyRole.UNKNOWN and basis.source == FaListAmountBasisSource.K01_FORMULA:
        basis.currency_role = AmountCurrencyRole.REPORTING
        basis.evidence.append("K.01 table 2 establishes one reporting-currency balance group")
    return basis


def _header_period_role(header: str) -> AmountPeriodRole:
    text = _norm(header)
    if any(token in text for token in ("期初", "年初", "opening")):
        return AmountPeriodRole.OPENING
    if any(token in text for token in ("期末", "年末", "ending", "closing")):
        return AmountPeriodRole.ENDING
    return AmountPeriodRole.UNKNOWN


def _header_currency_role(header: str) -> AmountCurrencyRole:
    text = _norm(header)
    if any(token in text for token in ("cny", "rmb", "人民币", "本位币", "本币")):
        return AmountCurrencyRole.REPORTING
    if any(token in text for token in ("原币", "外币", "交易币")):
        return AmountCurrencyRole.ORIGINAL
    return AmountCurrencyRole.UNKNOWN


def _same_sheet(value: str | None, expected: str) -> bool:
    return _norm(value) == _norm(expected)


def _column_number(value: str) -> int:
    result = 0
    for char in value:
        result = result * 26 + ord(char.upper()) - 64
    return result


def _norm(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()
