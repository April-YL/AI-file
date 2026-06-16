from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from ingest.addition_test_sheet import ModuleAssessment
from ingest.models import SheetKind
from ingest.records import FaListDataset
from ingest.sheet_loader import find_sheets_by_kind
from ingest.summary_sheet import PspProgramRow, SummarySheetDataset
from ingest.workbook_reader import read_worksheet_rows


@dataclass
class DisposalAmountItem:
    label: str
    amount: str | None
    source_row: int
    source_column: int | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "label": self.label,
            "amount": self.amount,
            "source_row": self.source_row,
            "source_column": self.source_column,
        }


@dataclass
class DisposalReconciliationCell:
    value: str | None
    formula: str | None
    source_row: int
    source_column: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "value": self.value,
            "formula": self.formula,
            "source_row": self.source_row,
            "source_column": self.source_column,
            "cell_ref": f"{get_column_letter(self.source_column)}{self.source_row}",
        }


@dataclass
class DisposalReconciliationRow:
    row_key: str
    label: str
    source_row: int
    measures: dict[str, DisposalReconciliationCell] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "row_key": self.row_key,
            "label": self.label,
            "source_row": self.source_row,
            "measures": {key: cell.to_dict() for key, cell in self.measures.items()},
        }


@dataclass
class DisposalReconciliationMatrix:
    header_row: int | None = None
    measure_columns: dict[str, int] = field(default_factory=dict)
    rows: dict[str, DisposalReconciliationRow] = field(default_factory=dict)
    recognition_confidence: float = 0.0
    recognition_evidence: list[str] = field(default_factory=list)
    missing_components: list[str] = field(default_factory=list)
    ambiguous_candidates: list[str] = field(default_factory=list)
    usable_for_rules: bool = False

    def to_dict(self) -> dict[str, Any]:
        return {
            "header_row": self.header_row,
            "measure_columns": self.measure_columns,
            "rows": {key: row.to_dict() for key, row in self.rows.items()},
            "recognition_confidence": self.recognition_confidence,
            "recognition_evidence": self.recognition_evidence,
            "missing_components": self.missing_components,
            "ambiguous_candidates": self.ambiguous_candidates,
            "usable_for_rules": self.usable_for_rules,
        }


@dataclass
class DisposalParameterItem:
    label: str
    value: str | None
    source_row: int
    source_column: int | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "label": self.label,
            "value": self.value,
            "source_row": self.source_row,
            "source_column": self.source_column,
        }


@dataclass
class DisposalTestedSampleRow:
    source_row: int
    sample_type: str | None = None
    asset_category: str | None = None
    asset_id: str | None = None
    asset_name: str | None = None
    sale_price: str | None = None
    disposal_voucher_no: str | None = None
    disposal_gain_loss: str | None = None
    support_sale_price: str | None = None
    sale_price_difference: str | None = None
    original_value: str | None = None
    accumulated_depreciation: str | None = None
    impairment_provision: str | None = None
    net_value: str | None = None
    disposal_date: str | None = None
    disposal_method: str | None = None
    evidence_amount: str | None = None
    evidence_description: str | None = None
    amount_difference: str | None = None
    attribute_results: list[str | None] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_row": self.source_row,
            "sample_type": self.sample_type,
            "asset_category": self.asset_category,
            "asset_id": self.asset_id,
            "asset_name": self.asset_name,
            "sale_price": self.sale_price,
            "disposal_voucher_no": self.disposal_voucher_no,
            "disposal_gain_loss": self.disposal_gain_loss,
            "support_sale_price": self.support_sale_price,
            "sale_price_difference": self.sale_price_difference,
            "original_value": self.original_value,
            "accumulated_depreciation": self.accumulated_depreciation,
            "impairment_provision": self.impairment_provision,
            "net_value": self.net_value,
            "disposal_date": self.disposal_date,
            "disposal_method": self.disposal_method,
            "evidence_amount": self.evidence_amount,
            "evidence_description": self.evidence_description,
            "amount_difference": self.amount_difference,
            "attribute_results": self.attribute_results,
        }


@dataclass
class DisposalSampleRow:
    source_row: int
    sample_type: str | None = None
    sample_source_no: str | None = None
    sampling_id: str | None = None
    asset_category: str | None = None
    asset_id: str | None = None
    asset_name: str | None = None
    original_value: str | None = None
    accumulated_depreciation: str | None = None
    impairment_provision: str | None = None
    net_value: str | None = None
    disposal_date: str | None = None
    disposal_method: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_row": self.source_row,
            "sample_type": self.sample_type,
            "sample_source_no": self.sample_source_no,
            "sampling_id": self.sampling_id,
            "asset_category": self.asset_category,
            "asset_id": self.asset_id,
            "asset_name": self.asset_name,
            "original_value": self.original_value,
            "accumulated_depreciation": self.accumulated_depreciation,
            "impairment_provision": self.impairment_provision,
            "net_value": self.net_value,
            "disposal_date": self.disposal_date,
            "disposal_method": self.disposal_method,
        }


@dataclass
class DisposalTestSheetDataset:
    source_file: str
    source_sheet: str
    waiver_note_text: str | None = None
    waiver_note_rows: list[int] = field(default_factory=list)
    amounts: dict[str, DisposalAmountItem] = field(default_factory=dict)
    reconciliation_matrix: DisposalReconciliationMatrix | None = None
    tested_samples: list[DisposalTestedSampleRow] = field(default_factory=list)
    module_assessments: list[ModuleAssessment] = field(default_factory=list)
    recognition_confidence: float = 0.0
    usable_for_rules: bool = False
    notes: list[str] = field(default_factory=list)


@dataclass
class DisposalSampleOutputDataset:
    source_file: str
    source_sheet: str
    parameters: dict[str, DisposalParameterItem] = field(default_factory=dict)
    amounts: dict[str, DisposalAmountItem] = field(default_factory=dict)
    selected_samples: list[DisposalSampleRow] = field(default_factory=list)
    module_assessments: list[ModuleAssessment] = field(default_factory=list)
    recognition_confidence: float = 0.0
    usable_for_rules: bool = False
    notes: list[str] = field(default_factory=list)


@dataclass
class DisposalExecutionPathDataset:
    path_kind: str
    recognition_confidence: float
    summary_status: str | None = None
    summary_waiver_reason: str | None = None
    summary_source_row: int | None = None
    disposal_list_sheet: str | None = None
    disposal_test_sheet: str | None = None
    disposal_sample_output_sheet: str | None = None
    test_sheet_waiver_note: str | None = None
    test_sheet_waiver_rows: list[int] = field(default_factory=list)
    missing_components: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "path_kind": self.path_kind,
            "recognition_confidence": self.recognition_confidence,
            "summary_status": self.summary_status,
            "summary_waiver_reason": self.summary_waiver_reason,
            "summary_source_row": self.summary_source_row,
            "disposal_list_sheet": self.disposal_list_sheet,
            "disposal_test_sheet": self.disposal_test_sheet,
            "disposal_sample_output_sheet": self.disposal_sample_output_sheet,
            "test_sheet_waiver_note": self.test_sheet_waiver_note,
            "test_sheet_waiver_rows": self.test_sheet_waiver_rows,
            "missing_components": self.missing_components,
            "notes": self.notes,
        }


_AMOUNT_CHARS = re.compile(r"[¥$€,\s]")
_PAREN_NEGATIVE = re.compile(r"^\((.+)\)$")
_WAIVER_TERMS = (
    "无处置",
    "没有处置",
    "无需执行",
    "无需测试",
    "无需抽样",
    "不执行",
    "不再执行",
    "未执行",
    "未抽样",
    "小于te",
    "低于te",
    "小于tt",
    "低于tt",
    "小于sad",
    "低于sad",
    "净值小于",
    "处置资产净值小于",
    "无性质异常",
)
_GUIDANCE_TERMS = ("基础操作指引", "进阶实操提示", "易错点", "canvas form", "sop", "审计抽样指南")

_TEST_AMOUNT_ANCHORS: dict[str, tuple[str, ...]] = {
    "sale_scrap_net_value": (
        "出售报废净值",
        "出售+报废净值",
        "出售和报废净值",
        "处置/报废总金额",
        "处置报废总金额",
        "处置总金额",
        "处置测试总体",
        "样本总体金额",
    ),
    "sale_net_value": ("出售净值", "出售资产净值"),
    "scrap_net_value": ("报废净值", "报废资产净值"),
    "other_disposal_net_value": ("其他减少净值", "其他处置净值"),
    "rollforward_disposal_net_value": (
        "后推处置净值",
        "k01处置净值",
        "bkd处置净值",
        "breakdown中处置净值",
        "breakdown中处置/报废金额",
        "breakdown中处置报废金额",
        "breakdown中处置金额",
    ),
    "difference_amount": ("差异",),
    "key_item_amount": ("关键项目金额", "定量关键项目金额"),
    "remaining_population_amount": ("代表性抽样的剩余总体", "剩余总体", "代表性总体"),
}
_SAMPLE_AMOUNT_ANCHORS: dict[str, tuple[str, ...]] = {
    "uploaded_data_amount": ("已上传数据", "上传数据"),
    "necessary_exclusion_amount": ("必要的数据排除项", "剔除项金额"),
    "sample_pool_amount": ("样本池总体金额", "样本池总金额"),
    "representative_population_amount": ("代表性总体价值", "代表性总体金额"),
    "total_amount": ("总金额",),
    "accounting_record_amount": ("会计记录的重大账户余额或活动", "会计记录金额"),
    "difference_amount": ("差额", "差异"),
    "key_item_count": ("关键项数量",),
    "key_item_amount": ("定量关键项金额", "关键项金额"),
    "representative_sample_size": ("代表性样本量",),
    "total_sample_size": ("代表性样本与关键项数量合计", "样本合计"),
    "sample_method": ("样本选择方法", "抽样方法"),
}
_SAMPLE_PARAMETER_ANCHORS: dict[str, tuple[str, ...]] = {
    "te": ("可容忍误差", "TE", "Tolerable Error"),
    "covered_assertions": ("测试涵盖的认定", "涵盖的认定", "assertions covered"),
    "fraud_or_special_risk": ("舞弊或特别风险", "特别风险"),
    "cra": ("综合风险评估", "CRA", "Combined Risk Assessment"),
}
_RECONCILIATION_MEASURE_TERMS: dict[str, tuple[str, ...]] = {
    "net_value": ("账面净值", "净值", "账面价值"),
    "original_value": ("原值", "资产原值"),
    "accumulated_depreciation": ("累计折旧", "累折"),
    "impairment_provision": ("减值准备", "减值"),
}
_RECONCILIATION_ROW_TERMS: dict[str, tuple[str, ...]] = {
    "disposal_list": ("处置/报废总金额", "处置报废总金额", "处置总金额"),
    "rollforward": (
        "breakdown中处置/报废金额",
        "breakdown中处置报废金额",
        "breakdown中处置金额",
        "后推处置金额",
    ),
    "difference": ("差异",),
    "investigation": ("差异是否需要进一步调查", "是否需要进一步调查"),
}


def _extract_reconciliation_matrix(
    rows: list[tuple[Any, ...]],
    formula_rows: list[tuple[Any, ...]],
) -> DisposalReconciliationMatrix:
    header_candidates: list[tuple[int, dict[str, int], int]] = []
    for r_idx, row in enumerate(rows, 1):
        mapping: dict[str, int] = {}
        for c_idx, cell in enumerate(row[:18], 1):
            text = _norm(cell)
            if not text:
                continue
            for measure, terms in _RECONCILIATION_MEASURE_TERMS.items():
                if measure not in mapping and any(_norm(term) == text for term in terms):
                    mapping[measure] = c_idx
                    break
        if len(mapping) >= 3 and all(
            key in mapping
            for key in ("original_value", "accumulated_depreciation", "impairment_provision")
        ):
            first_measure_col = min(mapping.values())
            context_hits = _reconciliation_context_hits(
                rows,
                header_row=r_idx,
                max_label_col=max(1, first_measure_col - 1),
            )
            header_candidates.append((r_idx, mapping, context_hits))

    matrix = DisposalReconciliationMatrix()
    if not header_candidates:
        matrix.missing_components.append("measure_header")
        return matrix

    contextual_candidates = [item for item in header_candidates if item[2] >= 2]
    if contextual_candidates:
        header_candidates = contextual_candidates
    header_candidates.sort(key=lambda item: (item[2], len(item[1]), -item[0]), reverse=True)
    matrix.header_row, matrix.measure_columns, context_hits = header_candidates[0]
    matrix.recognition_evidence.append(f"context_row_hits={context_hits}")
    if len(header_candidates) > 1 and header_candidates[1][2] >= 2:
        matrix.ambiguous_candidates.extend(
            f"measure_header_row:{row_no}"
            for row_no, _mapping, candidate_hits in header_candidates[1:]
            if candidate_hits >= 2
        )

    row_candidates: dict[str, list[DisposalReconciliationRow]] = {
        key: [] for key in _RECONCILIATION_ROW_TERMS
    }
    start = max(1, matrix.header_row - 2)
    end = min(len(rows), matrix.header_row + 14)
    for r_idx in range(start, end + 1):
        row = rows[r_idx - 1]
        label = _first_business_label(
            row,
            max_col=max(1, min(matrix.measure_columns.values()) - 1),
        )
        if not label:
            continue
        normalized = _norm(label)
        matched_row_keys: list[str] = []
        if "调查" in normalized and any(
            _norm(term) in normalized for term in _RECONCILIATION_ROW_TERMS["investigation"]
        ):
            matched_row_keys = ["investigation"]
        elif any(_norm(term) in normalized for term in _RECONCILIATION_ROW_TERMS["difference"]):
            matched_row_keys = ["difference"]
        else:
            matched_row_keys = [
                row_key
                for row_key in ("disposal_list", "rollforward")
                if any(
                    _norm(term) in normalized
                    for term in _RECONCILIATION_ROW_TERMS[row_key]
                )
            ]
        for row_key in matched_row_keys:
            terms = _RECONCILIATION_ROW_TERMS[row_key]
            if not any(_norm(term) in normalized for term in terms):
                continue
            measures: dict[str, DisposalReconciliationCell] = {}
            for measure, col in matrix.measure_columns.items():
                raw_value = row[col - 1] if col <= len(row) else None
                formula_row = formula_rows[r_idx - 1] if r_idx <= len(formula_rows) else ()
                raw_formula = formula_row[col - 1] if col <= len(formula_row) else None
                formula = str(raw_formula).strip() if _is_formula(raw_formula) else None
                value = _value_at(row, col)
                if value is not None or formula is not None:
                    measures[measure] = DisposalReconciliationCell(
                        value=value,
                        formula=formula,
                        source_row=r_idx,
                        source_column=col,
                    )
            row_candidates[row_key].append(
                DisposalReconciliationRow(
                    row_key=row_key,
                    label=label,
                    source_row=r_idx,
                    measures=measures,
                )
            )

    for row_key, candidates in row_candidates.items():
        if not candidates:
            matrix.missing_components.append(row_key)
            continue
        candidates.sort(key=lambda item: (len(item.measures), -item.source_row), reverse=True)
        matrix.rows[row_key] = candidates[0]
        if len(candidates) > 1:
            matrix.ambiguous_candidates.extend(
                f"{row_key}_row:{candidate.source_row}" for candidate in candidates[1:]
            )

    base_measures = {"original_value", "accumulated_depreciation", "impairment_provision"}
    matrix.recognition_evidence.append(f"header_row={matrix.header_row}")
    matrix.recognition_evidence.append(
        "measure_columns="
        + ",".join(f"{key}:{col}" for key, col in matrix.measure_columns.items())
    )
    matrix.recognition_evidence.append("rows=" + ",".join(matrix.rows))

    score = 0.25
    if base_measures.issubset(matrix.measure_columns):
        score += 0.25
    if "net_value" in matrix.measure_columns:
        score += 0.1
    if "disposal_list" in matrix.rows:
        score += 0.15
    if "rollforward" in matrix.rows:
        score += 0.15
    if "difference" in matrix.rows:
        score += 0.05
    if "investigation" in matrix.rows:
        score += 0.05
    disposal_formulas = _row_formulas(matrix.rows.get("disposal_list"))
    rollforward_formulas = _row_formulas(matrix.rows.get("rollforward"))
    disposal_source_recognized = any("处置清单" in f for f in disposal_formulas)
    rollforward_source_recognized = any(
        "k.01" in f.lower() or "k01" in f.lower() for f in rollforward_formulas
    )
    matrix.recognition_evidence.append(
        "disposal_list_formula_source="
        + ("recognized" if disposal_source_recognized else "unconfirmed")
    )
    matrix.recognition_evidence.append(
        "rollforward_formula_source="
        + ("recognized" if rollforward_source_recognized else "unconfirmed")
    )

    net_formula_checks = [
        _net_formula_uses_base_measures(row, matrix.measure_columns)
        for row in (
            matrix.rows.get("disposal_list"),
            matrix.rows.get("rollforward"),
        )
        if row is not None
    ]
    if net_formula_checks:
        matrix.recognition_evidence.append(
            "net_formula_relationship="
            + ("recognized" if all(net_formula_checks) else "partial")
        )

    if matrix.ambiguous_candidates:
        score -= 0.15
    if not disposal_source_recognized:
        score -= 0.05
    if not rollforward_source_recognized:
        score -= 0.05
    matrix.recognition_confidence = max(0.0, min(1.0, round(score, 3)))
    matrix.usable_for_rules = (
        matrix.recognition_confidence >= 0.75
        and not matrix.ambiguous_candidates
        and base_measures.issubset(matrix.measure_columns)
        and all(key in matrix.rows for key in ("disposal_list", "rollforward"))
        and disposal_source_recognized
        and rollforward_source_recognized
    )
    return matrix


def _build_test_module_assessments(
    rows: list[tuple[Any, ...]],
    *,
    waiver_text: str | None,
    reconciliation_matrix: DisposalReconciliationMatrix,
    amounts: dict[str, DisposalAmountItem],
    tested_samples: list[DisposalTestedSampleRow],
) -> list[ModuleAssessment]:
    return [
        _text_module_assessment(
            rows,
            module_key="population_definition",
            module_name="处置测试总体定义",
            terms=("详细测试所涵盖的总体", "样本总体", "处置/报废减少", "处置测试总体"),
        ),
        ModuleAssessment(
            module_key="amount_reconciliation",
            module_name="处置总体金额勾稽",
            status=(
                "recognized"
                if reconciliation_matrix.usable_for_rules
                else "partial"
                if reconciliation_matrix.rows or amounts
                else "missing"
            ),
            confidence=reconciliation_matrix.recognition_confidence,
            evidence=list(reconciliation_matrix.recognition_evidence),
            notes=(
                list(reconciliation_matrix.missing_components)
                + list(reconciliation_matrix.ambiguous_candidates)
            ),
        ),
        _text_module_assessment(
            rows,
            module_key="key_item_representation",
            module_name="关键项目与代表性抽样",
            terms=("测试的关键项目", "代表性抽样的剩余总体", "选择关键项目的理由"),
        ),
        _text_module_assessment(
            rows,
            module_key="test_attributes",
            module_name="处置测试属性",
            terms=("测试属性", "正确移除", "处置损益", "处置收入金额"),
        ),
        ModuleAssessment(
            module_key="sample_table",
            module_name="处置测试样本表",
            status="recognized" if tested_samples else "missing",
            confidence=0.88 if tested_samples else 0.25,
            evidence=[f"tested_samples={len(tested_samples)}"],
            notes=[],
        ),
        _text_module_assessment(
            rows,
            module_key="exception_summary",
            module_name="异常说明与结论",
            terms=("无异常情况", "异常情况", "如何解决", "执行的其他实质性程序"),
            waiver_text=waiver_text,
        ),
    ]


def _build_sample_output_module_assessments(
    rows: list[tuple[Any, ...]],
    *,
    parameters: dict[str, DisposalParameterItem],
    amounts: dict[str, DisposalAmountItem],
    selected_samples: list[DisposalSampleRow],
) -> list[ModuleAssessment]:
    return [
        _mapping_module_assessment(
            "sampling_prerequisites",
            "抽样参数与测试认定",
            parameters,
            required=("te", "covered_assertions", "cra"),
        ),
        _mapping_module_assessment(
            "source_data_summary",
            "源数据与样本池摘要",
            amounts,
            required=("uploaded_data_amount", "sample_pool_amount"),
        ),
        _text_module_assessment(
            rows,
            module_key="sampling_strategy",
            module_name="抽样策略与样本量",
            terms=("抽样策略", "关键项数量", "代表性样本量", "样本选择方法"),
        ),
        _mapping_module_assessment(
            "accounting_reconciliation",
            "总体与会计记录核对",
            amounts,
            required=("total_amount", "accounting_record_amount", "difference_amount"),
        ),
        ModuleAssessment(
            module_key="selected_samples",
            module_name="已选取样本明细",
            status="recognized" if selected_samples else "missing",
            confidence=0.9 if selected_samples else 0.25,
            evidence=[f"selected_samples={len(selected_samples)}"],
            notes=[],
        ),
    ]


def _text_module_assessment(
    rows: list[tuple[Any, ...]],
    *,
    module_key: str,
    module_name: str,
    terms: tuple[str, ...],
    waiver_text: str | None = None,
) -> ModuleAssessment:
    hits = _collect_text_hits(rows, terms)
    evidence = hits[:4]
    if waiver_text:
        evidence.append(f"waiver_text={waiver_text}")
    status = "recognized" if len(hits) >= 2 or waiver_text else "partial" if hits else "missing"
    return ModuleAssessment(
        module_key=module_key,
        module_name=module_name,
        status=status,
        confidence=0.78 if status == "recognized" else 0.52 if status == "partial" else 0.25,
        evidence=evidence,
        notes=[],
    )


def _mapping_module_assessment(
    module_key: str,
    module_name: str,
    mapping: dict[str, Any],
    *,
    required: tuple[str, ...],
) -> ModuleAssessment:
    present = [key for key in required if key in mapping]
    status = "recognized" if len(present) == len(required) else "partial" if present else "missing"
    return ModuleAssessment(
        module_key=module_key,
        module_name=module_name,
        status=status,
        confidence=0.86 if status == "recognized" else 0.56 if status == "partial" else 0.25,
        evidence=[f"present={','.join(present)}"] if present else [],
        notes=[],
    )


def _collect_text_hits(
    rows: list[tuple[Any, ...]],
    terms: tuple[str, ...],
    *,
    limit: int = 6,
) -> list[str]:
    hits: list[str] = []
    for row in rows:
        joined = " ".join(str(v).strip() for v in row[:24] if v not in (None, ""))
        if not joined:
            continue
        normalized = _norm(joined)
        if any(term in normalized for term in map(_norm, _GUIDANCE_TERMS)):
            continue
        if any(_norm(term) in normalized for term in terms):
            hits.append(_truncate(joined, 200))
            if len(hits) >= limit:
                break
    return hits


def _first_business_label(row: tuple[Any, ...], *, max_col: int = 18) -> str | None:
    for value in row[:max_col]:
        text = _clean(value)
        if not text or len(text) > 120:
            continue
        normalized = _norm(text)
        if any(term in normalized for term in map(_norm, _GUIDANCE_TERMS)):
            return None
        if any(
            _norm(term) in normalized
            for terms in _RECONCILIATION_ROW_TERMS.values()
            for term in terms
        ):
            return text
    return None


def _reconciliation_context_hits(
    rows: list[tuple[Any, ...]],
    *,
    header_row: int,
    max_label_col: int,
) -> int:
    hit_keys: set[str] = set()
    for r_idx in range(max(1, header_row - 2), min(len(rows), header_row + 14) + 1):
        label = _first_business_label(rows[r_idx - 1], max_col=max_label_col)
        if not label:
            continue
        normalized = _norm(label)
        for row_key in ("disposal_list", "rollforward", "difference", "investigation"):
            if any(
                _norm(term) in normalized
                for term in _RECONCILIATION_ROW_TERMS[row_key]
            ):
                hit_keys.add(row_key)
    return len(hit_keys)


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _row_formulas(row: DisposalReconciliationRow | None) -> list[str]:
    if row is None:
        return []
    return [cell.formula for cell in row.measures.values() if cell.formula]


def _net_formula_uses_base_measures(
    row: DisposalReconciliationRow,
    measure_columns: dict[str, int],
) -> bool:
    net_cell = row.measures.get("net_value")
    if net_cell is None or not net_cell.formula:
        return False
    formula = net_cell.formula.upper()
    expected_refs = [
        f"{get_column_letter(measure_columns[key])}{row.source_row}"
        for key in ("original_value", "accumulated_depreciation", "impairment_provision")
        if key in measure_columns
    ]
    return len(expected_refs) == 3 and all(ref.upper() in formula for ref in expected_refs)


def load_disposal_test_from_workbook(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    max_rows: int | None = 150,
) -> DisposalTestSheetDataset | None:
    path = Path(path)
    candidate = _choose_candidate(path, SheetKind.DISPOSAL_TEST, sheet_name, max_rows=max_rows)
    if candidate is None:
        return None
    rows = candidate["rows"]
    formula_rows = candidate["formula_rows"]
    waiver_text, waiver_rows = _scan_waiver_notes(rows)
    amounts = _extract_amount_items(rows, _TEST_AMOUNT_ANCHORS)
    reconciliation_matrix = _extract_reconciliation_matrix(rows, formula_rows)
    tested_samples = _extract_tested_samples(rows)
    module_assessments = _build_test_module_assessments(
        rows,
        waiver_text=waiver_text,
        reconciliation_matrix=reconciliation_matrix,
        amounts=amounts,
        tested_samples=tested_samples,
    )
    usable_for_rules = (
        reconciliation_matrix.usable_for_rules
        and float(candidate["confidence"]) >= 0.7
    )
    notes = [f"disposal_test_sheet_detected:{candidate['sheet_name']}"]
    if waiver_text:
        notes.append("disposal_test_waiver_note_detected")
    if amounts:
        notes.append(f"disposal_test_amounts_detected:{len(amounts)}")
    if tested_samples:
        notes.append(f"disposal_test_samples_detected:{len(tested_samples)}")
    if module_assessments:
        notes.append(f"disposal_test_modules_detected:{len(module_assessments)}")
    if not usable_for_rules:
        notes.append("disposal_test_not_usable_for_deterministic_rules")
    return DisposalTestSheetDataset(
        source_file=str(path),
        source_sheet=candidate["sheet_name"],
        waiver_note_text=waiver_text,
        waiver_note_rows=waiver_rows,
        amounts=amounts,
        reconciliation_matrix=reconciliation_matrix,
        tested_samples=tested_samples,
        module_assessments=module_assessments,
        recognition_confidence=float(candidate["confidence"]),
        usable_for_rules=usable_for_rules,
        notes=notes,
    )


def load_disposal_sample_output_from_workbook(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    max_rows: int | None = 150,
) -> DisposalSampleOutputDataset | None:
    path = Path(path)
    candidate = _choose_candidate(path, SheetKind.DISPOSAL_SAMPLE_OUTPUT, sheet_name, max_rows=max_rows)
    if candidate is None:
        return None
    rows = candidate["rows"]
    parameters = _extract_parameter_items(rows, _SAMPLE_PARAMETER_ANCHORS)
    amounts = _extract_amount_items(rows, _SAMPLE_AMOUNT_ANCHORS)
    selected_samples = _extract_selected_samples(rows)
    module_assessments = _build_sample_output_module_assessments(
        rows,
        parameters=parameters,
        amounts=amounts,
        selected_samples=selected_samples,
    )
    usable_for_rules = (
        float(candidate["confidence"]) >= 0.7
        and any(m.module_key == "selected_samples" and m.status == "recognized" for m in module_assessments)
    )
    notes = [f"disposal_sample_output_sheet_detected:{candidate['sheet_name']}"]
    if parameters:
        notes.append(f"disposal_sample_output_parameters_detected:{len(parameters)}")
    if amounts:
        notes.append(f"disposal_sample_output_amounts_detected:{len(amounts)}")
    if selected_samples:
        notes.append(f"disposal_sample_output_rows_detected:{len(selected_samples)}")
    if module_assessments:
        notes.append(f"disposal_sample_output_modules_detected:{len(module_assessments)}")
    if not usable_for_rules:
        notes.append("disposal_sample_output_not_usable_for_deterministic_rules")
    return DisposalSampleOutputDataset(
        source_file=str(path),
        source_sheet=candidate["sheet_name"],
        parameters=parameters,
        amounts=amounts,
        selected_samples=selected_samples,
        module_assessments=module_assessments,
        recognition_confidence=float(candidate["confidence"]),
        usable_for_rules=usable_for_rules,
        notes=notes,
    )


def build_disposal_execution_path(
    *,
    summary: SummarySheetDataset | None,
    disposal_list: FaListDataset | None,
    disposal_test: DisposalTestSheetDataset | None,
    disposal_sample_output: DisposalSampleOutputDataset | None,
) -> DisposalExecutionPathDataset:
    row = _find_summary_disposal_row(summary)
    summary_status = _normalize_status(row.execution_status if row else None)
    summary_reason = _clean(row.waiver_reason if row else None)
    summary_source_row = row.source_row if row else None

    disposal_list_sheet = _sheet_name(disposal_list)
    disposal_test_sheet = disposal_test.source_sheet if disposal_test else None
    sample_output_sheet = disposal_sample_output.source_sheet if disposal_sample_output else None

    missing: list[str] = []
    notes: list[str] = []
    if row is None:
        notes.append("summary_disposal_row_not_detected")
    if summary_status:
        notes.append(f"summary_status:{summary_status}")

    waiver_note = disposal_test.waiver_note_text if disposal_test else None
    waiver_rows = disposal_test.waiver_note_rows if disposal_test else []

    if summary_status not in {"no"} and not waiver_note:
        if not disposal_list_sheet:
            missing.append("处置清单")
        if not disposal_test_sheet:
            missing.append("K.02.2 处置测试")
        if not sample_output_sheet:
            missing.append("K.02.2a 处置选样输出")
        if missing:
            notes.append("missing_components:" + ",".join(missing))

    if summary_status == "no":
        path_kind = "summary_waived"
        confidence = 0.82 if summary_reason else 0.68
    elif waiver_note:
        path_kind = "test_sheet_waiver_note"
        confidence = 0.72
    elif summary_status == "yes" and not missing:
        path_kind = "executed_package_complete"
        confidence = 0.86
    elif summary_status == "yes" and missing:
        path_kind = "executed_package_incomplete"
        confidence = 0.76
    else:
        path_kind = "unclear"
        confidence = 0.45 if (disposal_list_sheet or disposal_test_sheet or sample_output_sheet) else 0.2

    return DisposalExecutionPathDataset(
        path_kind=path_kind,
        recognition_confidence=confidence,
        summary_status=summary_status,
        summary_waiver_reason=summary_reason,
        summary_source_row=summary_source_row,
        disposal_list_sheet=disposal_list_sheet,
        disposal_test_sheet=disposal_test_sheet,
        disposal_sample_output_sheet=sample_output_sheet,
        test_sheet_waiver_note=waiver_note,
        test_sheet_waiver_rows=waiver_rows,
        missing_components=missing,
        notes=notes,
    )


def _choose_candidate(
    path: Path,
    kind: SheetKind,
    sheet_name: str | None,
    *,
    max_rows: int | None,
) -> dict[str, Any] | None:
    if sheet_name:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        formula_wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        try:
            ws = wb[sheet_name]
            formula_ws = formula_wb[sheet_name]
            rows = read_worksheet_rows(ws, max_rows=max_rows)
            formula_rows = read_worksheet_rows(formula_ws, max_rows=max_rows)
        finally:
            wb.close()
            formula_wb.close()
        return {
            "sheet_name": sheet_name,
            "confidence": 0.9,
            "rows": rows,
            "formula_rows": formula_rows,
        }
    candidates = find_sheets_by_kind(path, kind, max_rows=max_rows or 150)
    if not candidates:
        return None
    chosen = candidates[0]
    formula_wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        formula_rows = read_worksheet_rows(formula_wb[chosen.sheet_name], max_rows=max_rows)
    finally:
        formula_wb.close()
    return {
        "sheet_name": chosen.sheet_name,
        "confidence": chosen.confidence,
        "rows": chosen.rows,
        "formula_rows": formula_rows,
    }


def _scan_waiver_notes(rows: list[tuple[Any, ...]]) -> tuple[str | None, list[int]]:
    hits: list[str] = []
    hit_rows: list[int] = []
    for r_idx, row in enumerate(rows, 1):
        texts = [_clean(v) for v in row[:18]]
        joined = " ".join(t for t in texts if t)
        if not joined:
            continue
        low = _norm(joined)
        if any(term in low for term in map(_norm, _GUIDANCE_TERMS)):
            continue
        if any(term in low for term in map(_norm, _WAIVER_TERMS)):
            hits.append(_truncate(joined, 240))
            hit_rows.append(r_idx)
    if not hits:
        return None, []
    return "；".join(hits[:3]), hit_rows[:6]


def _extract_parameter_items(
    rows: list[tuple[Any, ...]],
    anchors: dict[str, tuple[str, ...]],
    *,
    max_anchor_col: int = 18,
) -> dict[str, DisposalParameterItem]:
    found: dict[str, DisposalParameterItem] = {}
    for r_idx, row in enumerate(rows, 1):
        for c_idx, cell in enumerate(row[:max_anchor_col], 1):
            label = _clean(cell)
            if not label or len(label) > 160:
                continue
            normalized = _norm(label)
            if any(term in normalized for term in map(_norm, _GUIDANCE_TERMS)):
                continue
            for key, terms in anchors.items():
                if key in found or not any(_norm(term) in normalized for term in terms):
                    continue
                value, value_col = _first_value_to_right(row, c_idx, require_numeric=key == "te")
                if value is None:
                    continue
                found[key] = DisposalParameterItem(label, _stringify_cell(value), r_idx, value_col)
    return found


def _extract_amount_items(
    rows: list[tuple[Any, ...]],
    anchors: dict[str, tuple[str, ...]],
    *,
    max_anchor_col: int = 18,
) -> dict[str, DisposalAmountItem]:
    found: dict[str, DisposalAmountItem] = {}
    for r_idx, row in enumerate(rows, 1):
        for c_idx, cell in enumerate(row[:max_anchor_col], 1):
            label = _clean(cell)
            if not label or len(label) > 120:
                continue
            normalized = _norm(label)
            if any(term in normalized for term in map(_norm, _GUIDANCE_TERMS)):
                continue
            for key, terms in anchors.items():
                if key in found or not any(_norm(term) in normalized for term in terms):
                    continue
                value, value_col = _first_value_to_right(
                    row,
                    c_idx,
                    require_numeric=key != "sample_method",
                    require_short_text=key == "sample_method",
                )
                if value is None:
                    continue
                found[key] = DisposalAmountItem(label, _stringify_cell(value), r_idx, value_col)
    return found


def _first_value_to_right(
    row: tuple[Any, ...],
    label_col: int,
    *,
    require_numeric: bool = False,
    require_short_text: bool = False,
) -> tuple[Any | None, int | None]:
    for c_idx in range(label_col + 1, min(len(row), label_col + 9) + 1):
        if c_idx - 1 >= len(row):
            break
        value = row[c_idx - 1]
        if value is None or str(value).strip() == "":
            continue
        if require_numeric and _parse_amount(value) is None:
            continue
        text = str(value).strip()
        if len(text) > 120:
            continue
        if require_short_text and "\n" in text:
            continue
        return value, c_idx
    return None, None


def _extract_selected_samples(rows: list[tuple[Any, ...]]) -> list[DisposalSampleRow]:
    header_row, mapping = _find_table_header(
        rows,
        _sample_field_terms(include_evidence=False),
        required=("sample_source_no", "sampling_id", "sample_type", "net_value"),
    )
    if header_row is None:
        return []
    out: list[DisposalSampleRow] = []
    for r_idx, row in _iter_table_rows(rows, header_row + 1, mapping):
        out.append(
            DisposalSampleRow(
                source_row=r_idx,
                sample_type=_value_at(row, mapping.get("sample_type")),
                sample_source_no=_value_at(row, mapping.get("sample_source_no")),
                sampling_id=_value_at(row, mapping.get("sampling_id")),
                asset_category=_value_at(row, mapping.get("asset_category")),
                asset_id=_value_at(row, mapping.get("asset_id")),
                asset_name=_value_at(row, mapping.get("asset_name")),
                original_value=_value_at(row, mapping.get("original_value")),
                accumulated_depreciation=_value_at(row, mapping.get("accumulated_depreciation")),
                impairment_provision=_value_at(row, mapping.get("impairment_provision")),
                net_value=_value_at(row, mapping.get("net_value")),
                disposal_date=_value_at(row, mapping.get("disposal_date")),
                disposal_method=_value_at(row, mapping.get("disposal_method")),
            )
        )
    return out


def _extract_tested_samples(rows: list[tuple[Any, ...]]) -> list[DisposalTestedSampleRow]:
    header_row, mapping = _find_table_header(
        rows,
        _sample_field_terms(include_evidence=True),
        required=("asset_id", "asset_name", "net_value"),
    )
    if header_row is None:
        return []
    attribute_cols = [col for field, col in mapping.items() if field.startswith("attribute_")]
    out: list[DisposalTestedSampleRow] = []
    for r_idx, row in _iter_table_rows(rows, header_row + 1, mapping):
        sample = DisposalTestedSampleRow(
            source_row=r_idx,
            sample_type=_value_at(row, mapping.get("sample_type")),
            asset_category=_value_at(row, mapping.get("asset_category")),
            asset_id=_value_at(row, mapping.get("asset_id")),
            asset_name=_value_at(row, mapping.get("asset_name")),
            sale_price=_value_at(row, mapping.get("sale_price")),
            disposal_voucher_no=_value_at(row, mapping.get("disposal_voucher_no")),
            disposal_gain_loss=_value_at(row, mapping.get("disposal_gain_loss")),
            support_sale_price=_value_at(row, mapping.get("support_sale_price")),
            sale_price_difference=_value_at(row, mapping.get("sale_price_difference")),
            original_value=_value_at(row, mapping.get("original_value")),
            accumulated_depreciation=_value_at(row, mapping.get("accumulated_depreciation")),
            impairment_provision=_value_at(row, mapping.get("impairment_provision")),
            net_value=_value_at(row, mapping.get("net_value")),
            disposal_date=_value_at(row, mapping.get("disposal_date")),
            disposal_method=_value_at(row, mapping.get("disposal_method")),
            evidence_amount=(
                _value_at(row, mapping.get("evidence_amount"))
                or _value_at(row, mapping.get("support_sale_price"))
            ),
            evidence_description=_value_at(row, mapping.get("evidence_description")),
            amount_difference=_value_at(row, mapping.get("amount_difference")),
            attribute_results=[_value_at(row, col) for col in attribute_cols],
        )
        if _is_valid_sample_row(sample):
            out.append(sample)
    return out


def _sample_field_terms(*, include_evidence: bool) -> dict[str, tuple[str, ...]]:
    terms = {
        "sample_source_no": ("源样本#", "源样本号", "样本#"),
        "sampling_id": ("抽样id", "抽样ID", "随机抽样ID"),
        "sample_type": ("样本类型",),
        "asset_category": ("固定资产类别", "资产类别"),
        "asset_id": ("固定资产编号", "资产编号", "卡片编号", "卡片编码"),
        "asset_name": ("固定资产名称", "资产名称"),
        "sale_price": ("出售价格", "出售价", "处置售价"),
        "disposal_voucher_no": ("处置交易凭证号", "交易凭证号", "凭证号"),
        "disposal_gain_loss": ("处置损益", "处置收益", "处置亏损"),
        "support_sale_price": ("出售价格（通过审计证据/支持性文件取得）", "支持性文件取得", "通过审计证据", "支持性文件金额"),
        "sale_price_difference": ("出售价格差异", "售价差异"),
        "evidence_amount": ("出售价格（通过审计证据/支持性文件取得）", "支持性文件取得", "通过审计证据", "支持性文件金额"),
        "original_value": ("资产原价", "原值", "固定资产原值", "处置原值"),
        "accumulated_depreciation": ("累计折旧", "处置累计折旧", "减少累计折旧"),
        "impairment_provision": ("减值准备", "减值"),
        "net_value": ("净值", "处置净值", "账面净值", "账面价值"),
        "disposal_date": ("处置日期", "处置/报废日", "减少日期", "报废日期", "业务日期"),
        "disposal_method": ("减少方式", "处置方式", "处置情况", "报废方式", "变动方式"),
    }
    if include_evidence:
        terms.update(
            {
                "evidence_description": ("获得的证据", "支持的描述", "证据描述"),
                "amount_difference": ("金额差异", "净值差异", "差异"),
            }
        )
    return terms


def _is_valid_sample_row(row: DisposalTestedSampleRow) -> bool:
    if row.asset_id and row.asset_name and row.net_value:
        return True
    if row.net_value and (row.asset_id or row.asset_name):
        return True
    if (
        row.sale_price
        or row.disposal_voucher_no
        or row.disposal_gain_loss
        or row.support_sale_price
        or row.sale_price_difference
        or row.evidence_amount
    ) and (row.asset_id or row.asset_name or row.net_value):
        return True
    if row.attribute_results and any(v is not None for v in row.attribute_results):
        return True
    return False


def _find_table_header(
    rows: list[tuple[Any, ...]],
    field_terms: dict[str, tuple[str, ...]],
    *,
    required: tuple[str, ...],
) -> tuple[int | None, dict[str, int]]:
    best_row: int | None = None
    best_mapping: dict[str, int] = {}
    best_score = 0
    for r_idx, row in enumerate(rows, 1):
        mapping: dict[str, int] = {}
        for c_idx, cell in enumerate(row, 1):
            text = _norm(cell)
            if not text:
                continue
            matched = False
            for field, terms in field_terms.items():
                if field in mapping:
                    continue
                if any(_norm(term) == text or _norm(term) in text for term in terms):
                    mapping[field] = c_idx
                    matched = True
                    break
            if not matched and text in {"1", "2", "3", "4"}:
                mapping[f"attribute_{text}"] = c_idx
        score = len(mapping)
        if score > best_score and all(field in mapping for field in required):
            best_score = score
            best_row = r_idx
            best_mapping = mapping
    return best_row, best_mapping


def _iter_table_rows(
    rows: list[tuple[Any, ...]],
    start_row: int,
    mapping: dict[str, int],
) -> list[tuple[int, tuple[Any, ...]]]:
    out: list[tuple[int, tuple[Any, ...]]] = []
    blank_streak = 0
    identity_cols = [
        col
        for key, col in mapping.items()
        if key in {"asset_id", "asset_name", "net_value", "sample_source_no", "sampling_id", "sample_type"}
    ]
    for r_idx in range(start_row, len(rows) + 1):
        row = rows[r_idx - 1]
        has_identity = any(_value_at(row, col) for col in identity_cols)
        if not has_identity:
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0
        joined = _norm(" ".join(str(v) for v in row if v is not None))
        if any(term in joined for term in map(_norm, _GUIDANCE_TERMS)):
            continue
        out.append((r_idx, row))
    return out


def _find_summary_disposal_row(summary: SummarySheetDataset | None) -> PspProgramRow | None:
    if summary is None:
        return None
    candidates = [row for row in summary.programs if _is_disposal_program_row(row)]
    if not candidates:
        return None
    candidates.sort(key=lambda r: (0 if r.execution_status else 1, r.source_row or 0))
    return candidates[0]


def _is_disposal_program_row(row: PspProgramRow) -> bool:
    text = _norm(f"{row.procedure_name} {row.sheet_ref or ''}")
    if "k022a" in text or "k022b" in text:
        return False
    if "k022" in text:
        return True
    return any(token in text for token in ("处置", "减少", "报废")) and any(
        token in text for token in ("测试", "细节", "tod")
    )


def _normalize_status(value: str | None) -> str | None:
    text = _norm(value)
    if not text:
        return None
    if text in {"是", "yes", "y", "执行", "已执行"} or ("执行" in text and "不执行" not in text):
        return "yes"
    if text in {"否", "no", "n", "不执行", "未执行"} or "不执行" in text or "未执行" in text:
        return "no"
    return "unknown"


def _sheet_name(dataset: FaListDataset | None) -> str | None:
    if dataset is None or not dataset.source_sheet:
        return None
    return dataset.source_sheet


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _stringify_cell(value: Any) -> str:
    if isinstance(value, Decimal):
        value = _normalize_decimal_amount(value)
        text = format(value, "f")
        return text.rstrip("0").rstrip(".") if "." in text else text
    if isinstance(value, (int, float)):
        dec = _normalize_decimal_amount(Decimal(str(value)))
        text = format(dec, "f")
        return text.rstrip("0").rstrip(".") if "." in text else text
    return str(value).strip()


def _normalize_decimal_amount(value: Decimal) -> Decimal:
    rounded = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if abs(value - rounded) <= Decimal("0.000001"):
        return rounded
    return value


def _parse_amount(value: Any) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in ("-", "—", "N/A", "n/a", "#N/A"):
        return None
    text = _AMOUNT_CHARS.sub("", text)
    paren = _PAREN_NEGATIVE.match(text)
    if paren:
        text = f"-{paren.group(1)}"
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _value_at(row: tuple[Any, ...], col: int | None) -> str | None:
    if col is None or col <= 0 or col > len(row):
        return None
    value = row[col - 1]
    if value is None or str(value).strip() == "":
        return None
    amount = _parse_amount(value)
    if amount is not None:
        return str(amount)
    return str(value).strip()


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip().lower())


def _truncate(value: str, max_len: int) -> str:
    return value if len(value) <= max_len else value[: max_len - 1] + "…"
