from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl

from ingest.field_mapping import map_headers
from ingest.fa_list_amount_basis import resolve_fa_list_amount_basis, resolve_unique_header_basis
from ingest.fa_list_field_semantics import (
    resolve_fa_list_identity_basis,
    resolve_fa_list_salvage_basis,
)
from ingest.fa_list_population import build_fa_list_population
from ingest.fa_list_routing import choose_fa_list_route
from ingest.header_detection import scan_rows_for_headers
from ingest.amount_field_groups import build_amount_field_groups, select_amount_field_group
from ingest.models import (
    AmountFieldGroup,
    AmountGroupStatus,
    AssetRecord,
    FaListAmountBasis,
    FaListAmountBasisStatus,
    FaListReviewProfile,
    FaListRoutingDecision,
    FaListRoutingStatus,
    FaListSalvageMode,
    FieldMapping,
    SheetKind,
)
from ingest.sheet_classifier import classify_sheet, score_by_name
from ingest.sheet_period_routing import choose_sheet_candidate, sort_sheet_candidates
from ingest.workbook_reader import read_worksheet_rows

_RECORD_FIELDS = (
    "asset_id",
    "asset_name",
    "asset_category",
    "start_date",
    "useful_life_months",
    "salvage_rate",
    "salvage_value",
    "entity_name",
    "currency",
    "original_value",
    "accumulated_depreciation",
    "impairment_provision",
    "net_value",
    "addition_method",
    "disposal_date",
    "disposal_method",
)


@dataclass
class FaListDataset:
    source_file: str
    source_sheet: str
    mapped_fields: list[FieldMapping]
    records: list[AssetRecord]
    amount_groups: list[AmountFieldGroup] = field(default_factory=list)
    selected_amount_group_id: str | None = None
    amount_basis: FaListAmountBasis | None = None
    fa_profile: FaListReviewProfile | None = None


@dataclass
class DisposalMethodBucket:
    bucket_key: str
    bucket_label: str
    record_count: int = 0
    original_value_total: str = "0"
    accumulated_depreciation_total: str = "0"
    impairment_provision_total: str = "0"
    net_value_total: str = "0"
    asset_ids: list[str] = field(default_factory=list)
    methods: list[str] = field(default_factory=list)
    source_rows: list[int] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "bucket_key": self.bucket_key,
            "bucket_label": self.bucket_label,
            "record_count": self.record_count,
            "original_value_total": self.original_value_total,
            "accumulated_depreciation_total": self.accumulated_depreciation_total,
            "impairment_provision_total": self.impairment_provision_total,
            "net_value_total": self.net_value_total,
            "asset_ids": self.asset_ids,
            "methods": self.methods,
            "source_rows": self.source_rows,
        }


@dataclass
class DisposalListSummary:
    source_file: str
    source_sheet: str
    record_count: int
    total_original_value: str = "0"
    total_accumulated_depreciation: str = "0"
    total_impairment_provision: str = "0"
    total_net_value: str = "0"
    sale_net_value: str = "0"
    scrap_net_value: str = "0"
    sale_scrap_net_value: str = "0"
    other_reduction_net_value: str = "0"
    unclassified_net_value: str = "0"
    buckets: list[DisposalMethodBucket] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    amount_group_status: AmountGroupStatus | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_file": self.source_file,
            "source_sheet": self.source_sheet,
            "record_count": self.record_count,
            "total_original_value": self.total_original_value,
            "total_accumulated_depreciation": self.total_accumulated_depreciation,
            "total_impairment_provision": self.total_impairment_provision,
            "total_net_value": self.total_net_value,
            "sale_net_value": self.sale_net_value,
            "scrap_net_value": self.scrap_net_value,
            "sale_scrap_net_value": self.sale_scrap_net_value,
            "other_reduction_net_value": self.other_reduction_net_value,
            "unclassified_net_value": self.unclassified_net_value,
            "buckets": [b.to_dict() for b in self.buckets],
            "notes": self.notes,
        }


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _decimal_or_zero(value: str | None) -> Decimal:
    if value is None:
        return Decimal("0")
    text = str(value).strip()
    if not text:
        return Decimal("0")
    normalized = text.replace(",", "").replace(" ", "")
    if normalized.startswith("(") and normalized.endswith(")"):
        normalized = f"-{normalized[1:-1]}"
    if normalized in {"-", "—"}:
        return Decimal("0")
    try:
        return Decimal(normalized)
    except InvalidOperation:
        return Decimal("0")


def _stringify_decimal(value: Decimal) -> str:
    rounded = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if abs(value - rounded) <= Decimal("0.000001"):
        value = rounded
    return format(value.normalize(), "f") if value == value.to_integral() else format(value, "f")


def _normalize_disposal_method(text: str | None) -> str:
    return re.sub(r"\s+", "", (text or "")).lower()


def _disposal_bucket_for_method(text: str | None) -> tuple[str, str]:
    normalized = _normalize_disposal_method(text)
    if not normalized:
        return "unknown", "未识别"
    sale_tokens = ("出售", "转让", "售出", "销售")
    scrap_tokens = ("报废", "核销")
    other_tokens = ("其他", "调拨", "重分类", "转出", "转入", "合并", "拆分")
    has_sale = any(token in normalized for token in sale_tokens)
    has_scrap = any(token in normalized for token in scrap_tokens)
    has_other = any(token in normalized for token in other_tokens)
    if "处置及报废" in normalized or "处置/报废" in normalized or "处置报废" in normalized:
        return "sale_scrap", "出售+报废"
    if has_other and not (has_sale or has_scrap):
        return "other", "其他减少"
    if has_sale and has_scrap:
        return "sale_scrap", "出售+报废"
    if has_sale:
        return "sale", "出售"
    if has_scrap:
        return "scrap", "报废"
    if "处置" in normalized:
        return "sale_scrap", "出售+报废"
    if has_other:
        return "other", "其他减少"
    return "unknown", "未识别"


def build_disposal_list_summary(dataset: FaListDataset | None) -> DisposalListSummary | None:
    if dataset is None or not dataset.source_sheet:
        return None

    buckets: dict[str, DisposalMethodBucket] = {}
    total_original = Decimal("0")
    total_accumulated = Decimal("0")
    total_impairment = Decimal("0")
    total_net = Decimal("0")

    for record in dataset.records:
        bucket_key, bucket_label = _disposal_bucket_for_method(record.disposal_method)
        bucket = buckets.setdefault(
            bucket_key,
            DisposalMethodBucket(bucket_key=bucket_key, bucket_label=bucket_label),
        )
        bucket.record_count += 1
        bucket.asset_ids.append((record.asset_id or "").strip())
        if record.disposal_method:
            bucket.methods.append(record.disposal_method)
        if record.source_row is not None:
            bucket.source_rows.append(record.source_row)

        original = _decimal_or_zero(record.original_value)
        accumulated = _decimal_or_zero(record.accumulated_depreciation)
        impairment = _decimal_or_zero(record.impairment_provision)
        net = _decimal_or_zero(record.net_value)

        total_original += original
        total_accumulated += accumulated
        total_impairment += impairment
        total_net += net

        bucket.original_value_total = _stringify_decimal(
            _decimal_or_zero(bucket.original_value_total) + original
        )
        bucket.accumulated_depreciation_total = _stringify_decimal(
            _decimal_or_zero(bucket.accumulated_depreciation_total) + accumulated
        )
        bucket.impairment_provision_total = _stringify_decimal(
            _decimal_or_zero(bucket.impairment_provision_total) + impairment
        )
        bucket.net_value_total = _stringify_decimal(_decimal_or_zero(bucket.net_value_total) + net)

    sale_net = _decimal_or_zero(buckets.get("sale").net_value_total if buckets.get("sale") else None)
    scrap_net = _decimal_or_zero(buckets.get("scrap").net_value_total if buckets.get("scrap") else None)
    sale_scrap_net = (
        _decimal_or_zero(buckets.get("sale_scrap").net_value_total if buckets.get("sale_scrap") else None)
        + sale_net
        + scrap_net
    )
    other_net = _decimal_or_zero(buckets.get("other").net_value_total if buckets.get("other") else None)
    unknown_net = _decimal_or_zero(buckets.get("unknown").net_value_total if buckets.get("unknown") else None)

    notes: list[str] = []
    if buckets.get("unknown"):
        notes.append("disposal_methods_unclassified")
    if buckets.get("other"):
        notes.append("disposal_other_reduction_detected")

    selected_amount_group = next(
        (group for group in dataset.amount_groups if group.group_id == dataset.selected_amount_group_id),
        None,
    )
    return DisposalListSummary(
        source_file=dataset.source_file,
        source_sheet=dataset.source_sheet,
        record_count=len(dataset.records),
        total_original_value=_stringify_decimal(total_original),
        total_accumulated_depreciation=_stringify_decimal(total_accumulated),
        total_impairment_provision=_stringify_decimal(total_impairment),
        total_net_value=_stringify_decimal(total_net),
        sale_net_value=_stringify_decimal(sale_net),
        scrap_net_value=_stringify_decimal(scrap_net),
        sale_scrap_net_value=_stringify_decimal(sale_scrap_net),
        other_reduction_net_value=_stringify_decimal(other_net),
        unclassified_net_value=_stringify_decimal(unknown_net),
        buckets=sorted(buckets.values(), key=lambda b: b.bucket_key),
        amount_group_status=(
            selected_amount_group.status
            if selected_amount_group is not None
            else AmountGroupStatus.NOT_FOUND
        ),
        notes=notes,
    )


def _build_record(
    row_values: dict[str, Any],
    col_by_field: dict[str, int],
    source_row: int,
) -> AssetRecord:
    data: dict[str, Any] = {"source_row": source_row}
    for field_name in _RECORD_FIELDS:
        col = col_by_field.get(field_name)
        if col is None:
            data[field_name] = None
        else:
            data[field_name] = _cell_str(row_values.get(col))
    return AssetRecord(**data)


def _is_non_asset_summary_row(
    record: AssetRecord,
    *,
    sheet_kind: SheetKind = SheetKind.FA_LIST,
) -> bool:
    """过滤 FA list / 清单尾部重分类/合计等非资产明细行。"""
    aid = (record.asset_id or "").strip()
    name = (record.asset_name or "").strip()
    summary_tokens = ("资产类别重分类", "重分类", "合计", "小计", "总计", "尾差")
    text_fields = (
        aid,
        name,
        (record.asset_category or "").strip(),
        (record.addition_method or "").strip(),
        (record.disposal_method or "").strip(),
    )
    if not any(text_fields):
        return True
    has_summary_marker = any(
        token in text for text in text_fields if text for token in summary_tokens
    )
    has_identity = bool(aid or name)

    if sheet_kind == SheetKind.FA_LIST and not has_identity:
        return True
    if has_summary_marker and not has_identity:
        return True
    if any(token in aid for token in summary_tokens) and not name:
        return True
    if any(token in (record.addition_method or "") for token in summary_tokens):
        return True
    if any(token in (record.disposal_method or "") for token in summary_tokens):
        return True
    if any(token in (record.asset_category or "") for token in summary_tokens) and not name:
        return True
    if any(token in name for token in summary_tokens) and not aid:
        return True
    if sheet_kind == SheetKind.DISPOSAL_LIST:
        method = (record.disposal_method or "").strip()
        if method.endswith("小计") or method.endswith("合计"):
            return True
        if not has_identity and not method:
            amount_fields = (
                record.original_value,
                record.accumulated_depreciation,
                record.net_value,
            )
            if not any(v and str(v).strip() not in {"", "-", "—"} for v in amount_fields):
                return True
    return aid in {"-", "—", "N/A", "NA"}


def parse_fa_list_rows(
    rows: list[tuple[Any, ...]],
    *,
    source_file: str = "",
    source_sheet: str = "FA list",
    sheet_kind: SheetKind = SheetKind.FA_LIST,
    amount_basis: FaListAmountBasis | None = None,
    routing: FaListRoutingDecision | None = None,
    number_formats: dict[int, list[str]] | None = None,
) -> FaListDataset:
    header_row, header_cells, _ = scan_rows_for_headers(rows, sheet_kind=sheet_kind)
    if not header_cells:
        profile = None
        if sheet_kind == SheetKind.FA_LIST:
            route = routing or FaListRoutingDecision(
                status=FaListRoutingStatus.CONFIRMED,
                selected_sheet=source_sheet,
                candidates=[source_sheet] if source_sheet else [],
                reason="direct FA list source",
            )
            basis = amount_basis or FaListAmountBasis(
                status=FaListAmountBasisStatus.NOT_FOUND,
                conflicts=["FA list header was not found"],
            )
            population = build_fa_list_population([], amount_basis=basis)
            profile = FaListReviewProfile(
                routing=route,
                amount_basis=basis,
                population=population,
                identity_basis=resolve_fa_list_identity_basis(population, [], route),
                salvage_basis=resolve_fa_list_salvage_basis(
                    header_cells=[],
                    rows=rows,
                    header_row=None,
                    mapped_fields=[],
                    amount_basis=basis,
                ),
            )
        return FaListDataset(
            source_file=source_file,
            source_sheet=source_sheet,
            mapped_fields=[],
            records=[],
            amount_basis=profile.amount_basis if profile else amount_basis,
            fa_profile=profile,
        )

    mapped_fields, _ = map_headers(header_cells, sheet_kind=sheet_kind)
    salvage_basis = None
    if sheet_kind == SheetKind.FA_LIST:
        amount_basis = amount_basis or resolve_unique_header_basis(
            header_cells, header_row=header_row
        )
        if amount_basis.status == FaListAmountBasisStatus.CONFIRMED:
            amount_names = {"original_value", "accumulated_depreciation", "impairment_provision", "net_value"}
            header_by_column = {column: text for column, text in header_cells}
            mapped_fields = [item for item in mapped_fields if item.standard_field not in amount_names]
            mapped_fields.extend(
                FieldMapping(
                    standard_field=measure,
                    source_header=header_by_column.get(column, f"column_{column}"),
                    column_index=column,
                )
                for measure, column in amount_basis.bindings.items()
            )
        else:
            # Unconfirmed candidates remain diagnostic evidence on the basis;
            # they must not become canonical rule inputs.
            amount_names = {"original_value", "accumulated_depreciation", "impairment_provision", "net_value"}
            mapped_fields = [item for item in mapped_fields if item.standard_field not in amount_names]
        salvage_basis = resolve_fa_list_salvage_basis(
            header_cells=header_cells,
            rows=rows,
            header_row=header_row,
            mapped_fields=mapped_fields,
            amount_basis=amount_basis,
            number_formats=number_formats,
        )
        if salvage_basis.mode in {
            FaListSalvageMode.EXPLICIT_RATE,
            FaListSalvageMode.DERIVED_FROM_VALUE,
            FaListSalvageMode.RATE_AND_VALUE,
        }:
            bindings = []
            if salvage_basis.rate_column is not None:
                bindings.append(("salvage_rate", salvage_basis.rate_column))
            if salvage_basis.value_column is not None:
                bindings.append(("salvage_value", salvage_basis.value_column))
            header_by_column = {col: text for col, text in header_cells}
            for field_name, column in bindings:
                if not any(item.standard_field == field_name for item in mapped_fields):
                    mapped_fields.append(
                        FieldMapping(
                            standard_field=field_name,
                            source_header=header_by_column.get(column, f"column_{column}"),
                            column_index=column,
                        )
                    )
    amount_groups = (
        build_amount_field_groups(header_cells, sheet_kind=sheet_kind)
        if sheet_kind == SheetKind.DISPOSAL_LIST
        else []
    )
    selected_amount_group = select_amount_field_group(amount_groups)
    if selected_amount_group is not None:
        amount_names = {"original_value", "accumulated_depreciation", "impairment_provision", "net_value"}
        mapped_fields = [item for item in mapped_fields if item.standard_field not in amount_names]
        mapped_fields.extend(
            FieldMapping(
                standard_field=measure,
                source_header=candidate.source_header,
                column_index=candidate.column_index,
            )
            for measure, candidate in selected_amount_group.members.items()
        )
    col_by_field = {m.standard_field: m.column_index for m in mapped_fields}

    records: list[AssetRecord] = []
    start_idx = (header_row or 1)
    for r_idx in range(start_idx, len(rows)):
        source_row = r_idx + 1
        row = rows[r_idx]
        if row is None:
            continue
        row_values = {i + 1: row[i] if i < len(row) else None for i in range(len(row))}
        if not any(v is not None and str(v).strip() for v in row_values.values()):
            continue
        record = _build_record(row_values, col_by_field, source_row=source_row)
        if sheet_kind != SheetKind.FA_LIST and _is_non_asset_summary_row(record, sheet_kind=sheet_kind):
            continue
        records.append(record)

    fa_profile = None
    if sheet_kind == SheetKind.FA_LIST:
        routing = routing or FaListRoutingDecision(
            status=FaListRoutingStatus.CONFIRMED,
            selected_sheet=source_sheet,
            candidates=[source_sheet] if source_sheet else [],
            reason="direct FA list parse",
        )
        population = build_fa_list_population(records, amount_basis=amount_basis)
        identity_basis = resolve_fa_list_identity_basis(population, mapped_fields, routing)
        assert amount_basis is not None
        assert salvage_basis is not None
        fa_profile = FaListReviewProfile(
            routing=routing,
            amount_basis=amount_basis,
            population=population,
            identity_basis=identity_basis,
            salvage_basis=salvage_basis,
        )
        records = population.asset_records

    return FaListDataset(
        source_file=source_file,
        source_sheet=source_sheet,
        mapped_fields=mapped_fields,
        records=records,
        amount_groups=amount_groups,
        selected_amount_group_id=selected_amount_group.group_id if selected_amount_group else None,
        amount_basis=amount_basis,
        fa_profile=fa_profile,
    )


def load_fa_list_csv(path: str | Path, *, source_sheet: str = "FA list") -> FaListDataset:
    path = Path(path)
    rows: list[tuple[Any, ...]] = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(tuple(row))
    dataset = parse_fa_list_rows(
        rows,
        source_file=str(path),
        source_sheet=source_sheet,
    )
    return dataset


@dataclass
class FaListSheetCandidate:
    sheet_name: str
    confidence: float
    rows: list[tuple[Any, ...]]
    number_formats: dict[int, list[str]]


def find_fa_list_sheets(
    path: str | Path,
    *,
    max_rows: int | None = None,
) -> list[FaListSheetCandidate]:
    """扫描工作簿，返回识别为 FA list 的工作表（按 confidence 降序）。"""
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    candidates: list[FaListSheetCandidate] = []
    try:
        for ws in wb.worksheets:
            name_kind, *_ = score_by_name(ws.title)
            if name_kind == SheetKind.SKIP:
                continue
            rows = read_worksheet_rows(ws, max_rows=max_rows)
            kind, confidence, *_ = classify_sheet(ws.title, rows)
            if kind == SheetKind.FA_LIST:
                candidates.append(
                    FaListSheetCandidate(
                        sheet_name=ws.title,
                        confidence=confidence,
                        rows=rows,
                        number_formats=_number_format_samples(ws, max_rows=max_rows),
                    )
                )
    finally:
        wb.close()
    return sort_sheet_candidates(
        candidates,
        name=lambda c: c.sheet_name,
        confidence=lambda c: c.confidence,
        source_path=path,
    )


def load_fa_list_from_workbook(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    max_rows: int | None = None,
    k01_sheet_name: str | None = None,
    routing: FaListRoutingDecision | None = None,
    k01_route_reason: str | None = None,
) -> FaListDataset:
    """从 Excel 底稿读取 FA list 工作表并解析为 AssetRecord 列表。"""
    path = Path(path)
    if routing is not None and routing.status != FaListRoutingStatus.CONFIRMED:
        return _routing_unresolved_dataset(path, routing)
    if sheet_name:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            route = routing or choose_fa_list_route(wb.sheetnames, explicit_sheet=sheet_name)
            if route.status != FaListRoutingStatus.CONFIRMED:
                return _routing_unresolved_dataset(path, route)
            ws = wb[sheet_name]
            rows = read_worksheet_rows(ws, max_rows=max_rows)
            number_formats = _number_format_samples(ws, max_rows=max_rows)
        finally:
            wb.close()
        amount_basis = resolve_fa_list_amount_basis(
            path,
            fa_sheet=sheet_name,
            k01_sheet=k01_sheet_name,
            data_rows=rows,
            k01_route_reason=k01_route_reason,
        )
        return parse_fa_list_rows(
            rows,
            source_file=str(path),
            source_sheet=sheet_name,
            amount_basis=amount_basis,
            routing=route,
            number_formats=number_formats,
        )

    candidates = find_fa_list_sheets(path, max_rows=max_rows)

    if candidates:
        routing = choose_fa_list_route([candidate.sheet_name for candidate in candidates])
        if routing.status != FaListRoutingStatus.CONFIRMED or not routing.selected_sheet:
            return _routing_unresolved_dataset(path, routing)
        chosen = next(candidate for candidate in candidates if candidate.sheet_name == routing.selected_sheet)
    else:
        return _routing_unresolved_dataset(
            path,
            choose_fa_list_route([]),
        )

    amount_basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet=chosen.sheet_name,
        k01_sheet=k01_sheet_name,
        data_rows=chosen.rows,
        k01_route_reason=k01_route_reason,
    )
    return parse_fa_list_rows(
        chosen.rows,
        source_file=str(path),
        source_sheet=chosen.sheet_name,
        amount_basis=amount_basis,
        routing=routing,
        number_formats=chosen.number_formats,
    )


def _number_format_samples(ws, *, max_rows: int | None) -> dict[int, list[str]]:
    last_row = min(ws.max_row, max_rows) if max_rows else ws.max_row
    samples: dict[int, list[str]] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(last_row, 250)):
        for cell in row:
            if cell.value is None or not cell.number_format:
                continue
            values = samples.setdefault(cell.column, [])
            if cell.number_format not in values and len(values) < 8:
                values.append(cell.number_format)
    return samples


def _routing_unresolved_dataset(
    path: Path,
    routing: FaListRoutingDecision,
) -> FaListDataset:
    basis = FaListAmountBasis(
        status=FaListAmountBasisStatus.NOT_FOUND,
        conflicts=[routing.reason],
    )
    population = build_fa_list_population([], amount_basis=basis)
    profile = FaListReviewProfile(
        routing=routing,
        amount_basis=basis,
        population=population,
        identity_basis=resolve_fa_list_identity_basis(population, [], routing),
        salvage_basis=resolve_fa_list_salvage_basis(
            header_cells=[],
            rows=[],
            header_row=None,
            mapped_fields=[],
            amount_basis=basis,
        ),
    )
    return FaListDataset(
        source_file=str(path),
        source_sheet="",
        mapped_fields=[],
        records=[],
        amount_basis=basis,
        fa_profile=profile,
    )
