from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from ingest.field_mapping import map_headers
from ingest.header_detection import scan_rows_for_headers
from ingest.models import FieldMapping, SheetKind
from ingest.sheet_classifier import classify_sheet
from ingest.workbook_reader import read_worksheet_rows

K03_BRANCH_DEPRECIATION_TEST = "depreciation_test"
K03_BRANCH_POLICY_REVIEW = "depreciation_policy_review"

EXECUTION_PATH_SAP_MEDIUM = "sap_medium_precision"
EXECUTION_PATH_SAP_HIGH = "sap_high_precision"
EXECUTION_PATH_TOD_SAMPLING = "tod_sampling"
EXECUTION_PATH_TOD_BY_ITEM = "tod_by_item"
EXECUTION_PATH_SAP_PLUS_TOD_SAMPLING = "sap_plus_tod_sampling"
EXECUTION_PATH_POLICY_REVIEW = "policy_review"
EXECUTION_PATH_UNKNOWN = "unknown"

COMPONENT_STATE_EXECUTED = "EXECUTED"
COMPONENT_STATE_TEMPLATE_ONLY = "TEMPLATE_ONLY"
COMPONENT_STATE_INCOMPLETE = "INCOMPLETE"
COMPONENT_STATE_AMBIGUOUS = "AMBIGUOUS"

INGEST_DEPTH_DETAILED = "detailed"
INGEST_DEPTH_LIGHTWEIGHT = "lightweight"
INGEST_DEPTH_TEMPLATE_DETECTION = "template_detection"

RULE_STATUS_READY_FOR_LATER_RULES = "ready_for_later_rules"
RULE_STATUS_LATER_PHASE = "later_phase"

_K03_KINDS = {
    SheetKind.DEPRECIATION_TOD,
    SheetKind.DEPRECIATION_TOD_SAMPLE,
    SheetKind.SAP,
    SheetKind.DEPRECIATION_POLICY,
}

_AMOUNT_FIELDS = {
    "original_value",
    "accumulated_depreciation",
    "impairment_provision",
    "net_value",
    "current_depreciation",
    "management_depreciation",
    "audit_recalculated_depreciation",
    "depreciation_difference",
}
_DATE_FIELDS = {"start_date", "depreciation_start_date", "disposal_date"}
_BY_ITEM_CORE_FIELDS = {
    "asset_id",
    "asset_name",
    "original_value",
    "useful_life_months",
    "salvage_rate",
}


@dataclass
class K03Area:
    start_row: int | None = None
    end_row: int | None = None
    start_col: int | None = None
    end_col: int | None = None
    text: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "start_row": self.start_row,
            "end_row": self.end_row,
            "start_col": self.start_col,
            "end_col": self.end_col,
            "text": self.text,
        }


@dataclass
class K03Column:
    source_header: str
    column_index: int
    column_letter: str
    standard_field: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_header": self.source_header,
            "column_index": self.column_index,
            "column_letter": self.column_letter,
            "standard_field": self.standard_field,
        }


@dataclass
class K03DetailRow:
    source_row: int
    raw_values: dict[str, Any] = field(default_factory=dict)
    normalized_values: dict[str, Any] = field(default_factory=dict)
    cell_refs: dict[str, str] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_row": self.source_row,
            "raw_values": self.raw_values,
            "normalized_values": self.normalized_values,
            "cell_refs": self.cell_refs,
        }


@dataclass
class K03DetailTableRef:
    source_file: str
    sheet_name: str
    start_row: int | None = None
    end_row: int | None = None
    start_col: int | None = None
    end_col: int | None = None
    header_row: int | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_file": self.source_file,
            "sheet_name": self.sheet_name,
            "start_row": self.start_row,
            "end_row": self.end_row,
            "start_col": self.start_col,
            "end_col": self.end_col,
            "header_row": self.header_row,
        }


@dataclass
class K03DetailTable:
    detail_rows: list[K03DetailRow] = field(default_factory=list)
    total_rows: list[K03DetailRow] = field(default_factory=list)


@dataclass
class K03PolicyRow:
    source_row: int
    asset_category: Any = None
    current_method: Any = None
    current_useful_life: Any = None
    current_salvage_rate: Any = None
    current_annual_rate: Any = None
    prior_method: Any = None
    prior_useful_life: Any = None
    prior_salvage_rate: Any = None
    prior_annual_rate: Any = None
    useful_life_same_marker: Any = None
    salvage_rate_same_marker: Any = None
    difference_explanation: Any = None
    cell_refs: dict[str, str] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_row": self.source_row,
            "asset_category": self.asset_category,
            "current_method": self.current_method,
            "current_useful_life": self.current_useful_life,
            "current_salvage_rate": self.current_salvage_rate,
            "current_annual_rate": self.current_annual_rate,
            "prior_method": self.prior_method,
            "prior_useful_life": self.prior_useful_life,
            "prior_salvage_rate": self.prior_salvage_rate,
            "prior_annual_rate": self.prior_annual_rate,
            "useful_life_same_marker": self.useful_life_same_marker,
            "salvage_rate_same_marker": self.salvage_rate_same_marker,
            "difference_explanation": self.difference_explanation,
            "cell_refs": self.cell_refs,
        }


@dataclass
class K03PolicyTable:
    range: K03Area | None = None
    header_row: int | None = None
    column_map: dict[str, K03Column] = field(default_factory=dict)
    rows: list[K03PolicyRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "range": self.range.to_dict() if self.range else None,
            "header_row": self.header_row,
            "column_map": {k: v.to_dict() for k, v in self.column_map.items()},
            "rows": [r.to_dict() for r in self.rows],
            "warnings": self.warnings,
        }


@dataclass
class K03SheetDataset:
    workbook_name: str
    source_file: str
    sheet_name: str
    k03_branch: str
    execution_path: str = EXECUTION_PATH_UNKNOWN
    template_type: str = "unknown"
    ingest_depth: str = INGEST_DEPTH_LIGHTWEIGHT
    rule_status: str = RULE_STATUS_LATER_PHASE
    detected_sections: list[str] = field(default_factory=list)
    header_rows: list[int] = field(default_factory=list)
    detail_table_ref: K03DetailTableRef | None = None
    detail_table_range: K03Area | None = None
    total_rows: list[int] = field(default_factory=list)
    conclusion_area: K03Area | None = None
    note_area: K03Area | None = None
    instruction_area: K03Area | None = None
    policy_table: K03PolicyTable | None = None
    raw_columns: list[K03Column] = field(default_factory=list)
    normalized_column_map: dict[str, K03Column] = field(default_factory=dict)
    unmapped_columns: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    row_count: int = 0
    column_count: int = 0
    amount_columns: list[str] = field(default_factory=list)
    date_columns: list[str] = field(default_factory=list)
    unsupported_or_later_phase: bool = False
    summary: dict[str, Any] = field(default_factory=dict)
    preview_rows: list[dict[str, Any]] = field(default_factory=list)
    llm_candidate_context: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "workbook_name": self.workbook_name,
            "source_file": self.source_file,
            "sheet_name": self.sheet_name,
            "k03_branch": self.k03_branch,
            "execution_path": self.execution_path,
            "template_type": self.template_type,
            "ingest_depth": self.ingest_depth,
            "rule_status": self.rule_status,
            "detected_sections": self.detected_sections,
            "header_rows": self.header_rows,
            "detail_table_ref": (
                self.detail_table_ref.to_dict() if self.detail_table_ref else None
            ),
            "detail_table_range": (
                self.detail_table_range.to_dict() if self.detail_table_range else None
            ),
            "total_rows": self.total_rows,
            "conclusion_area": (
                self.conclusion_area.to_dict() if self.conclusion_area else None
            ),
            "note_area": self.note_area.to_dict() if self.note_area else None,
            "instruction_area": (
                self.instruction_area.to_dict() if self.instruction_area else None
            ),
            "policy_table": self.policy_table.to_dict() if self.policy_table else None,
            "raw_columns": [c.to_dict() for c in self.raw_columns],
            "normalized_column_map": {
                k: v.to_dict() for k, v in self.normalized_column_map.items()
            },
            "unmapped_columns": self.unmapped_columns,
            "warnings": self.warnings,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "amount_columns": self.amount_columns,
            "date_columns": self.date_columns,
            "unsupported_or_later_phase": self.unsupported_or_later_phase,
            "summary": self.summary,
            "preview_rows": self.preview_rows,
            "llm_candidate_context": self.llm_candidate_context,
        }


@dataclass
class K03SamplingRow:
    source_row: int
    sample_type: str | None = None
    asset_id: str | None = None
    values: dict[str, Any] = field(default_factory=dict)
    cell_refs: dict[str, str] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_row": self.source_row,
            "sample_type": self.sample_type,
            "asset_id": self.asset_id,
            "values": self.values,
            "cell_refs": self.cell_refs,
        }


@dataclass
class K03RoleEvidence:
    role: str
    status: str
    required_hits: list[str] = field(default_factory=list)
    supporting_hits: list[str] = field(default_factory=list)
    missing_required: list[str] = field(default_factory=list)
    conflicts: list[str] = field(default_factory=list)
    matched_header_fields: list[str] = field(default_factory=list)
    matched_rows: list[int] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "role": self.role,
            "status": self.status,
            "required_hits": self.required_hits,
            "supporting_hits": self.supporting_hits,
            "missing_required": self.missing_required,
            "conflicts": self.conflicts,
            "matched_header_fields": self.matched_header_fields,
            "matched_rows": self.matched_rows,
        }


@dataclass
class K03ComponentSheet:
    role: str
    sheet_name: str
    execution_path: str
    template_type: str
    execution_state: str = COMPONENT_STATE_AMBIGUOUS
    evidence: dict[str, Any] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "role": self.role,
            "sheet_name": self.sheet_name,
            "execution_path": self.execution_path,
            "template_type": self.template_type,
            "execution_state": self.execution_state,
            "evidence": self.evidence,
            "warnings": self.warnings,
        }


@dataclass
class K03LeadLinkage:
    assertion: str | None = None
    cra: str | None = None
    tt: str | None = None
    tt_overall: str | None = None
    source_row: int | None = None
    cra_cell: str | None = None
    tt_cell: str | None = None
    tt_overall_cell: str | None = None
    source: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "assertion": self.assertion,
            "cra": self.cra,
            "tt": self.tt,
            "tt_overall": self.tt_overall,
            "source_row": self.source_row,
            "cra_cell": self.cra_cell,
            "tt_cell": self.tt_cell,
            "tt_overall_cell": self.tt_overall_cell,
            "source": self.source,
        }


@dataclass
class K03ExecutionProfile:
    primary_depreciation_path: str = EXECUTION_PATH_UNKNOWN
    executed_depreciation_paths: list[str] = field(default_factory=list)
    component_sheets: dict[str, list[K03ComponentSheet]] = field(default_factory=dict)
    evidence_completeness: dict[str, Any] = field(default_factory=dict)
    lead_linkage: K03LeadLinkage | None = None
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "primary_depreciation_path": self.primary_depreciation_path,
            "executed_depreciation_paths": self.executed_depreciation_paths,
            "component_sheets": {
                role: [item.to_dict() for item in items]
                for role, items in self.component_sheets.items()
            },
            "evidence_completeness": self.evidence_completeness,
            "lead_linkage": self.lead_linkage.to_dict() if self.lead_linkage else None,
            "warnings": self.warnings,
        }


def load_k03_sheets_from_workbook(
    path: str | Path,
    *,
    max_rows: int | None = None,
) -> list[K03SheetDataset]:
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    datasets: list[K03SheetDataset] = []
    try:
        for ws in wb.worksheets:
            preview_rows = read_worksheet_rows(ws, max_rows=max_rows or 200)
            kind, confidence, *_ = classify_sheet(ws.title, preview_rows)
            dynamic_tod = _looks_like_dynamic_tod(preview_rows)
            auxiliary_depreciation = (
                any(token in _norm(ws.title) for token in ("本期计提", "currentdepreciation"))
                and not _looks_like_k03_sheet(ws.title)
            )
            if auxiliary_depreciation:
                continue
            if kind not in _K03_KINDS and not _looks_like_k03_sheet(ws.title) and not dynamic_tod:
                continue
            if dynamic_tod and kind not in _K03_KINDS:
                kind = SheetKind.DEPRECIATION_TOD_SAMPLE
                confidence = 0.8
            rows = read_worksheet_rows(ws, max_rows=None)
            datasets.append(
                _parse_k03_sheet(
                    path=path,
                    sheet_name=ws.title,
                    rows=rows,
                    classified_kind=kind,
                    classification_confidence=confidence,
                )
            )
    finally:
        wb.close()
    return datasets


def _looks_like_dynamic_tod(rows: list[tuple[Any, ...]]) -> bool:
    evidences = [
        evaluate_tod_sampling_main(rows, ""),
        evaluate_tod_sampling_output(rows, ""),
        evaluate_tod_by_item(rows, ""),
    ]
    if resolve_tod_role(rows, "").status == "FOUND":
        return True
    # Keep sufficiently evidenced incomplete/ambiguous K.03 candidates so the
    # parser can record DATA_INSUFFICIENT instead of silently dropping them.
    # Requiring three independent semantic requirements prevents generic asset
    # tables with only a sample-like header from entering the K.03 pipeline.
    return max((len(item.required_hits) for item in evidences), default=0) >= 3


def evaluate_tod_sampling_main(rows: list[tuple[Any, ...]], sheet_name: str = "") -> K03RoleEvidence:
    header_row, detail_rows, table_state = _detect_sampling_detail_table(rows)
    fields = set((detail_rows[0].values or {}).keys()) if detail_rows else set()
    text = _norm(" ".join(_text(value) for row in rows for value in row if _text(value)))
    population, population_row, _, population_state = _find_unique_parameter(
        rows, ("折旧费用总体", "depreciation population", "population amount"), end_row=header_row
    )
    requirements = {
        "sampling_detail_table": table_state == "FOUND",
        "sample_type_and_asset_id": {"sample_type", "asset_id"}.issubset(fields),
        "depreciation_calculation_fields": {
            "current_depreciation", "recalculated_depreciation", "depreciation_difference"
        }.issubset(fields),
        "sampling_population_section": population_state == "FOUND" and _looks_numeric_parameter(population),
    }
    hits = [name for name, matched in requirements.items() if matched]
    missing = [name for name, matched in requirements.items() if not matched]
    status = "FOUND" if not missing else ("INCOMPLETE" if len(hits) >= 2 else "UNKNOWN")
    supporting = []
    normalized_name = _norm(sheet_name)
    if "抽样" in normalized_name or "sampling" in normalized_name:
        supporting.append("sheet_name_sampling_hint")
    if "关键项目" in text or "keyitem" in text:
        supporting.append("key_item_section")
    return K03RoleEvidence(
        role="tod_sampling",
        status=status,
        required_hits=hits,
        supporting_hits=supporting,
        missing_required=missing,
        matched_header_fields=sorted(fields),
        matched_rows=[row for row in (population_row, header_row) if row],
    )


def evaluate_tod_sampling_output(rows: list[tuple[Any, ...]], sheet_name: str = "") -> K03RoleEvidence:
    header_row, detail_rows, table_state = _detect_sampling_detail_table(rows)
    fields = set((detail_rows[0].values or {}).keys()) if detail_rows else set()
    parameter_hits = _sampling_parameter_hits(rows, header_row)
    parameter_conflicts = _sampling_parameter_conflicts(rows, header_row)
    key_count, key_count_row, _, key_count_state = _find_unique_parameter(
        rows, ("关键项目数量", "key item quantity"), end_row=header_row
    )
    representative_count, representative_row, _, representative_state = _find_unique_parameter(
        rows, ("代表性样本量", "representative sample quantity"), end_row=header_row
    )
    requirements = {
        "selected_sample_table": table_state == "FOUND",
        "sample_type_and_asset_id": {"sample_type", "asset_id"}.issubset(fields),
        "sampling_parameter_cluster": len(parameter_hits) >= 3,
        "sample_summary": _has_parameter_label(
            rows,
            ("关键项目数量", "代表性样本量", "key item quantity", "representative sample quantity"),
            end_row=header_row,
        ),
    }
    hits = [name for name, matched in requirements.items() if matched]
    missing = [name for name, matched in requirements.items() if not matched]
    status = "FOUND" if not missing and not parameter_conflicts else ("INCOMPLETE" if len(hits) >= 2 else "UNKNOWN")
    supporting = []
    normalized_name = _norm(sheet_name)
    if "选样输出" in normalized_name or "sampleoutput" in normalized_name:
        supporting.append("sheet_name_output_hint")
    return K03RoleEvidence(
        role="tod_sampling_output",
        status=status,
        required_hits=hits,
        supporting_hits=supporting,
        missing_required=missing,
        conflicts=parameter_conflicts,
        matched_header_fields=sorted(fields),
        matched_rows=[row for row in (key_count_row, representative_row, header_row) if row],
    )


def _sampling_parameter_hits(rows: list[tuple[Any, ...]], header_row: int | None) -> list[str]:
    labels = {
        "te": ("可容忍误差", "TE"),
        "sampling_currency": ("抽样单位货币列", "抽样货币单元", "sampling currency"),
        "sampling_method": ("抽样方法", "sampling method"),
        "sample_pool": ("样本池总体金额", "样本池", "sample pool"),
    }
    hits: list[str] = []
    validators = {
        "te": _looks_numeric_parameter,
        "sampling_currency": lambda value: "折旧" in _norm(_text(value)) or "depreciation" in _norm(_text(value)),
        "sampling_method": lambda value: bool(_text(value)) and not _looks_like_parameter_label(value),
        "sample_pool": _looks_numeric_parameter,
    }
    for name, aliases in labels.items():
        value, _, _, state = _find_unique_parameter(rows, aliases, end_row=header_row)
        if state == "FOUND" and validators[name](value):
            hits.append(name)
    return hits


def _sampling_parameter_conflicts(rows: list[tuple[Any, ...]], header_row: int | None) -> list[str]:
    conflicts: list[str] = []
    for name, aliases in (
        ("te", ("可容忍误差", "TE")),
        ("sampling_currency", ("抽样单位货币列", "抽样货币单元", "sampling currency")),
    ):
        _, _, _, state = _find_unique_parameter(rows, aliases, end_row=header_row)
        if state == "AMBIGUOUS":
            conflicts.append(f"ambiguous_{name}")
    return conflicts


def _looks_numeric_parameter(value: Any) -> bool:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    text = _text(value).strip().replace(",", "").replace("，", "")
    if text.endswith("%"):
        text = text[:-1]
    return bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text))


def _looks_like_parameter_label(value: Any) -> bool:
    text = _norm(_text(value))
    labels = (
        "可容忍误差", "抽样单位货币列", "抽样货币单元", "抽样方法",
        "样本池", "关键项目数量", "代表性样本量",
    )
    return any(text == _norm(label) or text.startswith(_norm(label)) for label in labels)


def evaluate_tod_by_item(rows: list[tuple[Any, ...]], sheet_name: str = "") -> K03RoleEvidence:
    header_row, header_cells, _ = scan_rows_for_headers(rows, sheet_kind=SheetKind.DEPRECIATION_TOD)
    mapped_fields, _ = _map_k03_headers(header_cells)
    normalized = {item.standard_field: item for item in mapped_fields}
    score = _by_item_score(normalized)
    sampling = evaluate_tod_sampling_main(rows, sheet_name)
    requirements = {
        "detailed_depreciation_fields": score >= 6,
        "no_sampling_structure": sampling.status not in {"FOUND", "INCOMPLETE"},
    }
    hits = [name for name, matched in requirements.items() if matched]
    missing = [name for name, matched in requirements.items() if not matched]
    status = "FOUND" if not missing else ("INCOMPLETE" if score >= 4 and requirements["no_sampling_structure"] else "UNKNOWN")
    supporting = []
    normalized_name = _norm(sheet_name)
    if "byitem" in normalized_name or "逐项" in normalized_name:
        supporting.append("sheet_name_by_item_hint")
    return K03RoleEvidence(
        role="tod_by_item",
        status=status,
        required_hits=hits,
        supporting_hits=supporting,
        missing_required=missing,
        matched_header_fields=sorted(normalized),
        matched_rows=[header_row] if header_row else [],
    )


def resolve_tod_role(rows: list[tuple[Any, ...]], sheet_name: str) -> K03RoleEvidence:
    candidates = [
        evaluate_tod_sampling_main(rows, sheet_name),
        evaluate_tod_sampling_output(rows, sheet_name),
        evaluate_tod_by_item(rows, sheet_name),
    ]
    found = [item for item in candidates if item.status == "FOUND"]
    if len(found) == 1:
        return found[0]
    if len(found) > 1:
        return K03RoleEvidence(
            role="unknown",
            status="AMBIGUOUS",
            conflicts=[item.role for item in found],
            supporting_hits=[hint for item in found for hint in item.supporting_hits],
            matched_header_fields=sorted({field for item in found for field in item.matched_header_fields}),
            matched_rows=sorted({row for item in found for row in item.matched_rows}),
        )
    incomplete = [item for item in candidates if item.status == "INCOMPLETE"]
    if len(incomplete) == 1:
        return incomplete[0]
    if len(incomplete) > 1:
        return K03RoleEvidence(
            role="unknown",
            status="AMBIGUOUS",
            conflicts=[item.role for item in incomplete],
        )
    return K03RoleEvidence(role="unknown", status="UNKNOWN")


def build_k03_execution_profile(
    k03_sheets: list[K03SheetDataset] | None,
    *,
    lead: Any | None = None,
    workbook_sheet_names: list[str] | None = None,
) -> K03ExecutionProfile:
    """Build workbook-level K.03 path recognition without applying QC rules."""
    datasets = k03_sheets or []
    components: dict[str, list[K03ComponentSheet]] = {}
    for dataset in datasets:
        role = _component_role(dataset)
        if role is None:
            continue
        evidence = _component_evidence(dataset)
        components.setdefault(role, []).append(
            K03ComponentSheet(
                role=role,
                sheet_name=dataset.sheet_name,
                execution_path=dataset.execution_path,
                template_type=dataset.template_type,
                execution_state=_component_execution_state(dataset, evidence),
                evidence=evidence,
                warnings=list(dataset.warnings),
            )
        )

    auxiliary_names = _auxiliary_current_depreciation_sheets(workbook_sheet_names or [])
    if auxiliary_names:
        components["auxiliary_current_depreciation"] = [
            K03ComponentSheet(
                role="auxiliary_current_depreciation",
                sheet_name=name,
                execution_path=EXECUTION_PATH_UNKNOWN,
                template_type="auxiliary_current_depreciation",
                evidence={"is_required_procedure_page": False},
            )
            for name in auxiliary_names
        ]

    lead_linkage = _build_k03_lead_linkage(lead)
    executed_paths = _executed_depreciation_paths(components)
    primary = _select_primary_depreciation_path(components, lead_linkage)
    evidence = _profile_evidence_completeness(components)
    warnings = _profile_warnings(components, lead_linkage, primary)
    return K03ExecutionProfile(
        primary_depreciation_path=primary,
        executed_depreciation_paths=executed_paths,
        component_sheets=components,
        evidence_completeness=evidence,
        lead_linkage=lead_linkage,
        warnings=warnings,
    )


def _component_role(dataset: K03SheetDataset) -> str | None:
    if dataset.execution_path == EXECUTION_PATH_SAP_MEDIUM:
        return "sap_medium"
    if dataset.execution_path == EXECUTION_PATH_SAP_HIGH:
        return "sap_high"
    if dataset.execution_path == EXECUTION_PATH_TOD_BY_ITEM:
        return "tod_by_item"
    if dataset.execution_path == EXECUTION_PATH_TOD_SAMPLING:
        if dataset.template_type == "tod_sampling_output":
            return "tod_sampling_output"
        return "tod_sampling"
    if dataset.execution_path == EXECUTION_PATH_POLICY_REVIEW:
        return "policy_review"
    if dataset.k03_branch == K03_BRANCH_DEPRECIATION_TEST:
        return "unknown_depreciation_test"
    if dataset.k03_branch == K03_BRANCH_POLICY_REVIEW:
        return "policy_review"
    return None


def _component_evidence(dataset: K03SheetDataset) -> dict[str, Any]:
    summary = dataset.summary or {}
    if dataset.execution_path in {EXECUTION_PATH_SAP_MEDIUM, EXECUTION_PATH_SAP_HIGH}:
        return {
            "has_parameter_area": bool(
                summary.get("sap_entity_type")
                or summary.get("sap_te")
                or summary.get("sap_cra")
            ),
            "has_calculation_area": bool(summary.get("sap_deviation_rows")),
            "has_deviation_threshold_area": bool(summary.get("sap_deviation_rows")),
            "has_conclusion_area": bool(summary.get("sap_conclusion_text")),
            "has_note_area": bool(summary.get("sap_note_text")),
            "deviation_row_count": len(summary.get("sap_deviation_rows") or []),
            "over_threshold_count": summary.get("sap_deviation_over_threshold_count"),
        }
    if dataset.template_type == "tod_sampling_output":
        return {
            "has_parameter_area": bool(
                summary.get("sample_output_te")
                or summary.get("sample_output_sampling_currency")
            ),
            "has_data_area": bool(summary.get("sample_output_summary_text")),
            "has_sampling_currency": bool(summary.get("sample_output_sampling_currency")),
            "is_required_for_tod_sampling_path": True,
            "selected_sample_row_count": summary.get("sample_output_selected_rows_count", 0),
        }
    if dataset.execution_path == EXECUTION_PATH_TOD_SAMPLING:
        return {
            "has_parameter_area": bool(
                summary.get("tod_population_amount")
                or summary.get("tod_breakdown_depreciation_amount")
            ),
            "has_data_area": bool(dataset.detail_table_range),
            "has_conclusion_area": bool(summary.get("has_conclusion_area")),
            "has_note_area": bool(summary.get("has_note_area")),
            "sample_rows_count": summary.get("tod_sample_rows_count"),
        }
    if dataset.execution_path == EXECUTION_PATH_TOD_BY_ITEM:
        return {
            "has_header_area": bool(dataset.header_rows),
            "has_data_area": bool(dataset.detail_table_range),
            "has_conclusion_area": bool(summary.get("has_conclusion_area")),
            "has_note_area": bool(summary.get("has_note_area")),
            "detail_row_count": dataset.row_count,
            "mapped_fields": summary.get("mapped_fields", []),
        }
    if dataset.execution_path == EXECUTION_PATH_POLICY_REVIEW:
        return {
            "has_policy_table": bool(summary.get("has_policy_table")),
            "has_note_area": bool(summary.get("has_note_area")),
            "policy_row_count": summary.get("policy_row_count"),
            "is_independent_required_procedure": True,
        }
    return {
        "has_header_area": bool(dataset.header_rows),
        "has_data_area": bool(dataset.detail_table_range),
        "has_conclusion_area": bool(summary.get("has_conclusion_area")),
        "has_note_area": bool(summary.get("has_note_area")),
    }


def _component_execution_state(
    dataset: K03SheetDataset,
    evidence: dict[str, Any],
) -> str:
    if dataset.execution_path in {EXECUTION_PATH_SAP_MEDIUM, EXECUTION_PATH_SAP_HIGH}:
        if evidence.get("has_parameter_area") and evidence.get("has_calculation_area"):
            return COMPONENT_STATE_EXECUTED
        if evidence.get("has_parameter_area") or evidence.get("has_calculation_area"):
            return COMPONENT_STATE_INCOMPLETE
        return COMPONENT_STATE_TEMPLATE_ONLY
    if dataset.template_type == "tod_sampling_output":
        if evidence.get("selected_sample_row_count") or (
            evidence.get("has_parameter_area") and evidence.get("has_data_area")
        ):
            return COMPONENT_STATE_EXECUTED
        if evidence.get("has_parameter_area") or evidence.get("has_data_area"):
            return COMPONENT_STATE_INCOMPLETE
        return COMPONENT_STATE_TEMPLATE_ONLY
    if dataset.execution_path == EXECUTION_PATH_TOD_SAMPLING:
        if evidence.get("has_parameter_area") and evidence.get("sample_rows_count"):
            return COMPONENT_STATE_EXECUTED
        if evidence.get("has_parameter_area") or evidence.get("sample_rows_count"):
            return COMPONENT_STATE_INCOMPLETE
        return COMPONENT_STATE_TEMPLATE_ONLY
    if dataset.execution_path == EXECUTION_PATH_TOD_BY_ITEM:
        if evidence.get("has_data_area") and evidence.get("detail_row_count"):
            return COMPONENT_STATE_EXECUTED
        if evidence.get("has_data_area") or evidence.get("detail_row_count"):
            return COMPONENT_STATE_INCOMPLETE
        return COMPONENT_STATE_TEMPLATE_ONLY
    if dataset.execution_path == EXECUTION_PATH_POLICY_REVIEW:
        if evidence.get("has_policy_table") and evidence.get("policy_row_count"):
            return COMPONENT_STATE_EXECUTED
        if evidence.get("has_policy_table") or evidence.get("has_note_area"):
            return COMPONENT_STATE_INCOMPLETE
        return COMPONENT_STATE_TEMPLATE_ONLY
    return COMPONENT_STATE_AMBIGUOUS


def _executed_depreciation_paths(
    components: dict[str, list[K03ComponentSheet]],
) -> list[str]:
    role_paths = (
        ("sap_medium", EXECUTION_PATH_SAP_MEDIUM),
        ("sap_high", EXECUTION_PATH_SAP_HIGH),
        ("tod_by_item", EXECUTION_PATH_TOD_BY_ITEM),
        ("tod_sampling", EXECUTION_PATH_TOD_SAMPLING),
    )
    return [
        path
        for role, path in role_paths
        if any(
            item.execution_state == COMPONENT_STATE_EXECUTED
            for item in components.get(role, [])
        )
    ]


def _auxiliary_current_depreciation_sheets(sheet_names: list[str]) -> list[str]:
    result: list[str] = []
    for name in sheet_names:
        text = _norm(name)
        if "本期计提" in text or "currentdepreciation" in text:
            result.append(name)
    return result


def _build_k03_lead_linkage(lead: Any | None) -> K03LeadLinkage | None:
    if lead is None or not getattr(lead, "cra_rows", None):
        return None
    row = _find_depreciation_assertion_row(getattr(lead, "cra_rows", []))
    if row is None:
        return None
    sheet = getattr(lead, "source_sheet", "") or ""

    def cell(col: int | None) -> str | None:
        if not getattr(row, "source_row", None) or not col:
            return None
        return f"{sheet}!{get_column_letter(col)}{row.source_row}"

    return K03LeadLinkage(
        assertion=getattr(row, "assertion", None),
        cra=getattr(row, "cra", None),
        tt=getattr(row, "tt", None),
        tt_overall=getattr(row, "tt_overall", None),
        source_row=getattr(row, "source_row", None),
        cra_cell=cell(getattr(row, "source_col_cra", None)),
        tt_cell=cell(getattr(row, "source_col_tt", None)),
        tt_overall_cell=cell(getattr(row, "source_col_tt_overall", None)),
        source="lead_cra_rows",
    )


def _find_depreciation_assertion_row(cra_rows: list[Any]) -> Any | None:
    for row in cra_rows:
        assertion = _norm(getattr(row, "assertion", ""))
        if (
            "计价" in assertion
            or "计量" in assertion
            or "v/m" in assertion
            or "valuation" in assertion
            or "measurement" in assertion
        ):
            return row
    return None


def _select_primary_depreciation_path(
    components: dict[str, list[K03ComponentSheet]],
    lead_linkage: K03LeadLinkage | None,
) -> str:
    def active(role: str) -> bool:
        items = components.get(role, [])
        executed = [item for item in items if item.execution_state == COMPONENT_STATE_EXECUTED]
        return bool(executed or items)

    has_medium = active("sap_medium")
    has_high = active("sap_high")
    has_by_item = active("tod_by_item")
    has_sampling = active("tod_sampling")
    has_sampling_output = active("tod_sampling_output")
    lead_cra = lead_linkage.cra if lead_linkage else None

    if has_medium and (has_sampling or has_sampling_output) and not has_high:
        if not _is_minimal_cra_value(lead_cra):
            return EXECUTION_PATH_SAP_PLUS_TOD_SAMPLING
    if has_high and not _is_minimal_cra_value(lead_cra):
        return EXECUTION_PATH_SAP_HIGH
    if has_medium and _is_minimal_cra_value(lead_cra):
        return EXECUTION_PATH_SAP_MEDIUM
    if has_high and not has_medium:
        return EXECUTION_PATH_SAP_HIGH
    if has_medium and not has_high:
        return EXECUTION_PATH_SAP_MEDIUM
    if has_by_item and not (has_sampling or has_sampling_output):
        return EXECUTION_PATH_TOD_BY_ITEM
    if has_sampling or has_sampling_output:
        return EXECUTION_PATH_TOD_SAMPLING
    if has_by_item:
        return EXECUTION_PATH_TOD_BY_ITEM
    return EXECUTION_PATH_UNKNOWN


def _profile_evidence_completeness(
    components: dict[str, list[K03ComponentSheet]],
) -> dict[str, Any]:
    sap_roles = ("sap_medium", "sap_high")
    sap_items = [item for role in sap_roles for item in components.get(role, [])]
    tod_sampling = components.get("tod_sampling", [])
    tod_sampling_output = components.get("tod_sampling_output", [])
    return {
        "sap": [
            {
                "sheet_name": item.sheet_name,
                "template_type": item.template_type,
                "execution_state": item.execution_state,
                **item.evidence,
            }
            for item in sap_items
        ],
        "tod_sampling": {
            "has_test_sheet": bool(tod_sampling),
            "has_sampling_output": bool(tod_sampling_output),
        },
        "tod_by_item": {
            "has_test_sheet": bool(components.get("tod_by_item")),
        },
        "policy_review": {
            "exists": bool(components.get("policy_review")),
            "is_independent_required_procedure": True,
        },
        "auxiliary_current_depreciation": {
            "exists": bool(components.get("auxiliary_current_depreciation")),
            "is_required_procedure_page": False,
        },
    }


def _profile_warnings(
    components: dict[str, list[K03ComponentSheet]],
    lead_linkage: K03LeadLinkage | None,
    primary: str,
) -> list[str]:
    warnings: list[str] = []
    has_medium = bool(components.get("sap_medium"))
    has_high = bool(components.get("sap_high"))
    has_sampling = bool(components.get("tod_sampling"))
    has_sampling_output = bool(components.get("tod_sampling_output"))
    has_by_item = bool(components.get("tod_by_item"))
    has_policy = bool(components.get("policy_review"))
    lead_cra = lead_linkage.cra if lead_linkage else None

    if not lead_linkage:
        warnings.append("k03_lead_depreciation_cra_tt_not_identified")
    if has_medium and has_high and lead_cra is None:
        warnings.append("k03_multiple_sap_paths_without_lead_cra")
    if has_medium and not has_high and not _is_minimal_cra_value(lead_cra) and not (has_sampling or has_sampling_output):
        warnings.append("k03_sap_medium_without_high_or_tod_for_non_minimal_cra")
    if has_sampling and not has_sampling_output:
        warnings.append("k03_tod_sampling_missing_sampling_output")
    if has_sampling_output and not has_sampling:
        warnings.append("k03_tod_sampling_output_without_test_sheet")
    if not has_policy:
        warnings.append("k03_policy_review_missing")
    if primary == EXECUTION_PATH_UNKNOWN and components.get("auxiliary_current_depreciation"):
        warnings.append("k03_only_auxiliary_current_depreciation_without_procedure_page")
    if sum(bool(v) for v in (has_medium or has_high, has_by_item, has_sampling or has_sampling_output)) > 1:
        warnings.append("k03_multiple_depreciation_test_components_detected")
    return warnings


def _is_minimal_cra_value(value: Any) -> bool:
    text = _norm(_text(value))
    return text in {"minimal", "最低", "最小", "极低"} or "minimal" in text


def load_k03_detail_table(dataset: K03SheetDataset) -> K03DetailTable:
    """Read full K.03 TOD-by item details only when deterministic rules need them."""
    ref = dataset.detail_table_ref
    if ref is None or not ref.source_file or not ref.sheet_name:
        return K03DetailTable()
    if not ref.header_row or not ref.start_row or not ref.end_row:
        return K03DetailTable()

    wb = openpyxl.load_workbook(ref.source_file, read_only=True, data_only=True)
    try:
        if ref.sheet_name not in wb.sheetnames:
            return K03DetailTable()
        ws = wb[ref.sheet_name]
        header_by_col = {
            col.column_index: col.source_header
            for col in dataset.raw_columns
            if col.column_index is not None
        }
        field_by_col = {
            col.column_index: field
            for field, col in dataset.normalized_column_map.items()
            if col.column_index is not None
        }
        if not header_by_col:
            return K03DetailTable()

        detail_rows: list[K03DetailRow] = []
        total_rows: list[K03DetailRow] = []
        read_start = (ref.header_row or ref.start_row) + 1
        min_col = min(header_by_col)
        max_col = max(header_by_col)
        row_iter = ws.iter_rows(
            min_row=read_start,
            max_row=ref.end_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
        for row_number, row_values in zip(range(read_start, ref.end_row + 1), row_iter):
            raw_values: dict[str, Any] = {}
            normalized_values: dict[str, Any] = {}
            cell_refs: dict[str, str] = {}
            for col in sorted(header_by_col):
                header = header_by_col[col]
                value = row_values[col - min_col]
                raw_values[header] = value
                field = field_by_col.get(col)
                if field:
                    normalized_values[field] = value
                    cell_refs[field] = f"{get_column_letter(col)}{row_number}"
            if _is_blank(raw_values.values()):
                continue
            row = K03DetailRow(
                source_row=row_number,
                raw_values=raw_values,
                normalized_values=normalized_values,
                cell_refs=cell_refs,
            )
            if row_number in dataset.total_rows or _row_has_token(
                raw_values.values(),
                ("合计", "总计", "小计", "total"),
            ):
                total_rows.append(row)
            else:
                detail_rows.append(row)
        return K03DetailTable(detail_rows=detail_rows, total_rows=total_rows)
    finally:
        wb.close()


def _parse_k03_sheet(
    *,
    path: Path,
    sheet_name: str,
    rows: list[tuple[Any, ...]],
    classified_kind: SheetKind,
    classification_confidence: float,
) -> K03SheetDataset:
    if _is_policy_review(sheet_name, rows, classified_kind):
        return _parse_policy_review(path, sheet_name, rows, classification_confidence)
    if _is_sap_sheet(sheet_name, classified_kind):
        return _parse_sap_sheet(path, sheet_name, rows)
    return _parse_tod_sheet(path, sheet_name, rows, classified_kind)


def _parse_policy_review(
    path: Path,
    sheet_name: str,
    rows: list[tuple[Any, ...]],
    confidence: float,
) -> K03SheetDataset:
    text_rows = _rows_containing(rows, ("政策", "折旧", "结论", "说明", "复核"))
    conclusion = _area_for_rows(rows, _rows_containing(rows, ("结论",)))
    policy_table = _extract_policy_table(rows)
    note = _detect_policy_note_area(rows, policy_table.range if policy_table else None)
    detected_sections = ["policy_review"]
    if policy_table and policy_table.rows:
        detected_sections.append("policy_table")
    if note:
        detected_sections.append("note_area")
    warnings = [] if confidence >= 0.45 else ["k03_policy_review_low_confidence"]
    if policy_table is None or not policy_table.rows:
        warnings.append("k03_policy_table_not_identified")
    elif policy_table.warnings:
        warnings.extend(policy_table.warnings)
    ds = K03SheetDataset(
        workbook_name=path.name,
        source_file=str(path),
        sheet_name=sheet_name,
        k03_branch=K03_BRANCH_POLICY_REVIEW,
        execution_path=EXECUTION_PATH_POLICY_REVIEW,
        template_type="policy_review",
        ingest_depth=INGEST_DEPTH_LIGHTWEIGHT,
        rule_status=RULE_STATUS_LATER_PHASE,
        detected_sections=detected_sections,
        conclusion_area=conclusion,
        note_area=note,
        policy_table=policy_table,
        warnings=warnings,
        summary={
            "text_row_count": len(text_rows),
            "policy_row_count": len(policy_table.rows) if policy_table else 0,
            "has_policy_table": bool(policy_table and policy_table.rows),
            "has_note_area": note is not None,
        },
        unsupported_or_later_phase=False,
        llm_candidate_context={
            "text_rows": text_rows[:20],
            "warnings": warnings,
            "candidate_for": "depreciation_policy_semantic_review",
            "policy_table_summary": {
                "row_count": len(policy_table.rows) if policy_table else 0,
                "mapped_fields": sorted(policy_table.column_map) if policy_table else [],
            },
        },
    )
    return ds


def _extract_policy_table(rows: list[tuple[Any, ...]]) -> K03PolicyTable | None:
    header_row, column_map, warnings = _detect_policy_header(rows)
    if not header_row or not column_map:
        return None
    required = {"asset_category", "current_useful_life", "prior_useful_life"}
    if not (required & set(column_map)):
        warnings.append("k03_policy_table_core_columns_missing")

    row_items: list[K03PolicyRow] = []
    end_row = header_row
    max_col = max(c.column_index for c in column_map.values())
    min_col = min(c.column_index for c in column_map.values())
    for idx in range(header_row, len(rows)):
        row_number = idx + 1
        row = rows[idx]
        values = {
            field: _cell_value(row, col.column_index)
            for field, col in column_map.items()
        }
        if _is_blank(values.values()):
            if row_items:
                break
            continue
        if _row_has_token(row, ("Notes", "注：", "注:", "说明", "结论")) and row_items:
            break
        category = values.get("asset_category")
        if not _text(category):
            if row_items:
                break
            continue
        if _looks_like_policy_non_data_row(category):
            if row_items:
                break
            continue

        cell_refs = {
            field: f"{get_column_letter(col.column_index)}{row_number}"
            for field, col in column_map.items()
        }
        row_items.append(
            K03PolicyRow(
                source_row=row_number,
                asset_category=category,
                current_method=values.get("current_method"),
                current_useful_life=values.get("current_useful_life"),
                current_salvage_rate=values.get("current_salvage_rate"),
                current_annual_rate=values.get("current_annual_rate"),
                prior_method=values.get("prior_method"),
                prior_useful_life=values.get("prior_useful_life"),
                prior_salvage_rate=values.get("prior_salvage_rate"),
                prior_annual_rate=values.get("prior_annual_rate"),
                useful_life_same_marker=values.get("useful_life_same_marker"),
                salvage_rate_same_marker=values.get("salvage_rate_same_marker"),
                difference_explanation=values.get("difference_explanation"),
                cell_refs=cell_refs,
            )
        )
        end_row = row_number

    return K03PolicyTable(
        range=K03Area(
            start_row=header_row,
            end_row=end_row,
            start_col=min_col,
            end_col=max_col,
        ),
        header_row=header_row,
        column_map=column_map,
        rows=row_items,
        warnings=warnings,
    )


def _detect_policy_header(
    rows: list[tuple[Any, ...]],
) -> tuple[int | None, dict[str, K03Column], list[str]]:
    best_row: int | None = None
    best_map: dict[str, K03Column] = {}
    best_score = 0
    warnings: list[str] = []
    for idx, row in enumerate(rows[:80], start=1):
        mapping = _map_policy_header_row(rows, idx)
        fields = set(mapping)
        score = len(fields & {
            "asset_category",
            "current_method",
            "current_useful_life",
            "current_salvage_rate",
            "prior_method",
            "prior_useful_life",
            "prior_salvage_rate",
            "useful_life_same_marker",
            "salvage_rate_same_marker",
            "difference_explanation",
        })
        if "asset_category" in fields:
            score += 2
        if {"current_useful_life", "prior_useful_life"} <= fields:
            score += 2
        if {"current_salvage_rate", "prior_salvage_rate"} <= fields:
            score += 2
        if score > best_score:
            best_row = idx
            best_map = mapping
            best_score = score
    if best_score < 4:
        return None, {}, ["k03_policy_table_header_not_identified"]
    return best_row, best_map, warnings


def _map_policy_header_row(
    rows: list[tuple[Any, ...]],
    row_number: int,
) -> dict[str, K03Column]:
    row = rows[row_number - 1]
    prev = rows[row_number - 2] if row_number >= 2 else ()
    prev_filled: list[str] = []
    carried = ""
    for value in prev:
        if _text(value):
            carried = _text(value)
        prev_filled.append(carried)
    mapping: dict[str, K03Column] = {}
    used_fields: set[str] = set()
    metric_occurrences: dict[str, int] = {}
    repeated_policy_metrics = sum(
        1 for value in row if _norm(_text(value)) in {"使用寿命", "使用年限"}
    ) >= 2
    for col_idx, value in enumerate(row, start=1):
        header = _text(value)
        normalized_header = _norm(header)
        prev_header = prev_filled[col_idx - 1] if col_idx - 1 < len(prev_filled) else ""
        combined = " ".join(v for v in (prev_header, header) if v)
        metric_occurrences[normalized_header] = metric_occurrences.get(normalized_header, 0) + 1
        occurrence = metric_occurrences[normalized_header]
        if repeated_policy_metrics and normalized_header == "折旧政策":
            field = "asset_category"
        elif normalized_header in {"使用寿命", "使用年限"} and occurrence >= 3:
            field = "useful_life_same_marker"
        elif normalized_header in {"残值率", "净残值率"} and occurrence >= 3:
            field = "salvage_rate_same_marker"
        else:
            field = _policy_field_for_header(combined, header, prev_header, used_fields)
        if not field:
            continue
        used_fields.add(field)
        mapping[field] = K03Column(
            source_header=combined or header,
            column_index=col_idx,
            column_letter=get_column_letter(col_idx),
            standard_field=field,
        )
    return mapping


def _policy_field_for_header(
    combined: str,
    header: str,
    prev_header: str,
    used_fields: set[str],
) -> str | None:
    n = _norm(combined)
    h = _norm(header)
    p = _norm(prev_header)
    if not n:
        return None
    if any(token in n for token in ("固定资产类别", "资产类别", "类别")) and "asset_category" not in used_fields:
        return "asset_category"
    is_prior = any(token in n for token in ("上期", "上年", "以前", "prior", "previous", "2024"))
    is_current = any(token in n for token in ("本期", "本年", "当前", "current", "2025")) or (
        not is_prior and any(token in p for token in ("本期", "本年", "current"))
    )
    is_difference = any(token in n for token in ("差异", "是否一致", "变化", "变动", "difference"))

    if any(token in n for token in ("差异说明", "说明", "备注", "原因", "解释", "note", "comment")):
        return "difference_explanation"
    if is_difference and any(token in n for token in ("寿命", "年限", "使用年限")):
        return "useful_life_same_marker"
    if is_difference and any(token in n for token in ("残值", "残值率")):
        return "salvage_rate_same_marker"
    if any(token in n for token in ("折旧方法", "折旧政策", "折旧方式")):
        return "prior_method" if is_prior else "current_method" if is_current or "current_method" not in used_fields else "prior_method"
    if any(token in n for token in ("使用寿命", "使用年限", "折旧年限", "寿命")):
        return "prior_useful_life" if is_prior else "current_useful_life" if is_current or "current_useful_life" not in used_fields else "prior_useful_life"
    if any(token in n for token in ("残值率", "净残值率", "残值")):
        return "prior_salvage_rate" if is_prior else "current_salvage_rate" if is_current or "current_salvage_rate" not in used_fields else "prior_salvage_rate"
    if any(token in n for token in ("年折旧率", "折旧率")):
        return "prior_annual_rate" if is_prior else "current_annual_rate" if is_current or "current_annual_rate" not in used_fields else "prior_annual_rate"
    # Two-level headers sometimes put only the metric in the leaf row.
    if h in {"折旧方法", "使用寿命", "使用年限", "残值率", "年折旧率"}:
        if "上期" in p or "prior" in p:
            prefix = "prior"
        else:
            prefix = "current"
        suffix = {
            "折旧方法": "method",
            "使用寿命": "useful_life",
            "使用年限": "useful_life",
            "残值率": "salvage_rate",
            "年折旧率": "annual_rate",
        }[h]
        return f"{prefix}_{suffix}"
    return None


def _detect_policy_note_area(
    rows: list[tuple[Any, ...]],
    policy_range: K03Area | None,
) -> K03Area | None:
    start = policy_range.end_row if policy_range and policy_range.end_row else 0
    matches: list[int] = []
    for idx, row in enumerate(rows, start=1):
        if idx <= start:
            continue
        if _row_has_token(row, ("Notes", "注：", "注:", "说明", "差异说明", "结论")):
            matches.append(idx)
    return _area_for_rows(rows, matches)


def _cell_value(row: tuple[Any, ...], col_index: int) -> Any:
    return row[col_index - 1] if col_index - 1 < len(row) else None


def _looks_like_policy_non_data_row(value: Any) -> bool:
    text = _text(value)
    if not text:
        return True
    lower = text.lower()
    return any(token in lower for token in ("合计", "小计", "总计", "notes", "说明", "结论", "表1"))


def _parse_sap_sheet(
    path: Path,
    sheet_name: str,
    rows: list[tuple[Any, ...]],
) -> K03SheetDataset:
    text = _combined_text(sheet_name, rows)
    if "高精确度" in text or "高精度" in text:
        execution_path = EXECUTION_PATH_SAP_HIGH
        template_type = "sap_high_precision"
    elif "中精确度" in text or "中精度" in text:
        execution_path = EXECUTION_PATH_SAP_MEDIUM
        template_type = "sap_medium_precision"
    else:
        execution_path = EXECUTION_PATH_UNKNOWN
        template_type = "sap"

    warnings = []
    if execution_path == EXECUTION_PATH_UNKNOWN:
        warnings.append("k03_sap_precision_not_identified")
    return K03SheetDataset(
        workbook_name=path.name,
        source_file=str(path),
        sheet_name=sheet_name,
        k03_branch=K03_BRANCH_DEPRECIATION_TEST,
        execution_path=execution_path,
        template_type=template_type,
        ingest_depth=INGEST_DEPTH_TEMPLATE_DETECTION,
        rule_status=RULE_STATUS_LATER_PHASE,
        detected_sections=["sap_template"],
        warnings=warnings,
        summary={"template_detection_only": True},
        unsupported_or_later_phase=True,
    )


def _parse_tod_sheet(
    path: Path,
    sheet_name: str,
    rows: list[tuple[Any, ...]],
    classified_kind: SheetKind,
) -> K03SheetDataset:
    header_row, header_cells, _ = scan_rows_for_headers(
        rows,
        sheet_kind=SheetKind.DEPRECIATION_TOD,
    )
    mapped_fields, unmapped = _map_k03_headers(header_cells)
    normalized = {m.standard_field: m for m in mapped_fields}
    raw_columns = [
        K03Column(
            source_header=text,
            column_index=col,
            column_letter=get_column_letter(col),
            standard_field=next(
                (m.standard_field for m in mapped_fields if m.column_index == col),
                None,
            ),
        )
        for col, text in header_cells
    ]
    by_item_score = _by_item_score(normalized)
    sample_score = _sample_score(sheet_name, rows, header_cells)
    warnings: list[str] = []

    if by_item_score >= 6 and sample_score < 3:
        execution_path = EXECUTION_PATH_TOD_BY_ITEM
        ingest_depth = INGEST_DEPTH_DETAILED
        template_type = "tod_by_item"
    elif by_item_score >= 7:
        execution_path = EXECUTION_PATH_TOD_BY_ITEM
        ingest_depth = INGEST_DEPTH_DETAILED
        template_type = "tod_by_item"
    elif sample_score >= 2 or (
        classified_kind == SheetKind.DEPRECIATION_TOD_SAMPLE and bool(header_cells)
    ):
        execution_path = EXECUTION_PATH_TOD_SAMPLING
        ingest_depth = INGEST_DEPTH_LIGHTWEIGHT
        template_type = "tod_sampling"
    else:
        execution_path = EXECUTION_PATH_UNKNOWN
        ingest_depth = INGEST_DEPTH_LIGHTWEIGHT
        template_type = "tod_unknown"
        warnings.append("k03_tod_execution_path_not_identified")

    detail_rows, detail_range, total_rows = _extract_detail_rows(
        rows,
        header_row=header_row,
        header_cells=header_cells,
        normalized=normalized,
    )
    conclusion = _detect_conclusion_area(rows, header_row, detail_range)
    note = _area_for_rows(rows, _rows_containing(rows, ("说明", "注：", "注:")))
    instruction = _area_for_rows(rows, _rows_containing(rows, ("获取", "编制", "按照", "根据")))
    detected_sections = []
    if header_row:
        detected_sections.append("header")
    if detail_range:
        detected_sections.append("detail_table")
    if total_rows:
        detected_sections.append("total_rows")
    if conclusion:
        detected_sections.append("conclusion_area")
    if note:
        detected_sections.append("note_area")
    if instruction:
        detected_sections.append("instruction_area")

    if unmapped:
        warnings.append("k03_unmapped_columns_present")
    missing_noncritical = sorted(_BY_ITEM_CORE_FIELDS - set(normalized))
    if execution_path == EXECUTION_PATH_TOD_BY_ITEM and missing_noncritical:
        warnings.append("k03_tod_by_item_missing_noncritical_fields:" + ",".join(missing_noncritical))

    preview_rows = [row.to_dict() for row in detail_rows[:5]]
    field_summary = {
        "raw_column_count": len(raw_columns),
        "mapped_field_count": len(normalized),
        "mapped_fields": sorted(normalized),
        "unmapped_column_count": len(unmapped),
        "amount_columns": sorted(set(normalized) & _AMOUNT_FIELDS),
        "date_columns": sorted(set(normalized) & _DATE_FIELDS),
    }
    table_ref = (
        K03DetailTableRef(
            source_file=str(path),
            sheet_name=sheet_name,
            start_row=detail_range.start_row,
            end_row=detail_range.end_row,
            start_col=detail_range.start_col,
            end_col=detail_range.end_col,
            header_row=header_row,
        )
        if detail_range
        else None
    )

    return K03SheetDataset(
        workbook_name=path.name,
        source_file=str(path),
        sheet_name=sheet_name,
        k03_branch=K03_BRANCH_DEPRECIATION_TEST,
        execution_path=execution_path,
        template_type=template_type,
        ingest_depth=ingest_depth,
        rule_status=RULE_STATUS_LATER_PHASE,
        detected_sections=detected_sections,
        header_rows=[header_row] if header_row else [],
        detail_table_ref=table_ref,
        detail_table_range=detail_range,
        total_rows=total_rows,
        conclusion_area=conclusion,
        note_area=note,
        instruction_area=instruction,
        raw_columns=raw_columns,
        normalized_column_map={
            field: K03Column(
                source_header=m.source_header,
                column_index=m.column_index,
                column_letter=get_column_letter(m.column_index),
                standard_field=m.standard_field,
            )
            for field, m in normalized.items()
        },
        unmapped_columns=unmapped,
        warnings=warnings,
        row_count=len(detail_rows),
        column_count=len(raw_columns),
        amount_columns=sorted(set(normalized) & _AMOUNT_FIELDS),
        date_columns=sorted(set(normalized) & _DATE_FIELDS),
        unsupported_or_later_phase=False,
        summary={
            **field_summary,
            "total_row_count": len(total_rows),
            "has_conclusion_area": conclusion is not None,
            "has_note_area": note is not None,
            "has_instruction_area": instruction is not None,
        },
        preview_rows=preview_rows,
        llm_candidate_context={
            "instruction_area": instruction.to_dict() if instruction else None,
            "note_area": note.to_dict() if note else None,
            "conclusion_area": conclusion.to_dict() if conclusion else None,
            "warnings": warnings,
            "field_summary": field_summary,
            "anomaly_row_refs": [],
            "preview_rows": preview_rows[:3],
        },
    )


def _map_k03_headers(header_cells: list[tuple[int, str]]) -> tuple[list[FieldMapping], list[str]]:
    mapped, unmapped = map_headers(header_cells, SheetKind.DEPRECIATION_TOD)
    used_cols = {m.column_index for m in mapped}
    extra_mapped: list[FieldMapping] = []
    extra_unmapped: list[str] = []
    for col, text in header_cells:
        if col in used_cols:
            continue
        field = _match_k03_extra_field(text)
        if field:
            extra_mapped.append(FieldMapping(field, text.strip(), col))
            used_cols.add(col)
        elif text.strip():
            extra_unmapped.append(text.strip())
    return mapped + extra_mapped, list(dict.fromkeys(unmapped + extra_unmapped))


def _match_k03_extra_field(text: str) -> str | None:
    n = _norm(text)
    if not n:
        return None
    checks = (
        ("sample_type", ("样本类型", "sample type")),
        ("evidence_description", ("获得的证据", "支持的描述", "evidence description")),
        ("test_attribute_1", ("测试属性1", "属性1", "attribute 1")),
        ("test_attribute_2", ("测试属性2", "属性2", "attribute 2")),
        ("management_depreciation", ("管理层计算折旧", "管理层测算折旧", "客户计算折旧", "账面折旧")),
        (
            "audit_recalculated_depreciation",
            ("审计重新计算折旧", "审计重算折旧", "审计测算折旧", "重新计算折旧"),
        ),
        ("depreciation_difference", ("差异", "差额", "diff", "difference")),
        ("depreciation_start_date", ("折旧起始日期", "开始折旧日期", "折旧开始日期")),
        ("conclusion", ("结论", "测试结论", "复核结论")),
        ("current_depreciation", ("本期折旧", "本年折旧", "本期计提折旧")),
    )
    for field, synonyms in checks:
        for synonym in synonyms:
            ns = _norm(synonym)
            if n == ns or ns in n or n in ns:
                return field
    return None


def _extract_detail_rows(
    rows: list[tuple[Any, ...]],
    *,
    header_row: int | None,
    header_cells: list[tuple[int, str]],
    normalized: dict[str, FieldMapping],
) -> tuple[list[K03DetailRow], K03Area | None, list[int]]:
    if not header_row or not header_cells:
        return [], None, []

    header_by_col = {col: text for col, text in header_cells}
    field_by_col = {m.column_index: m.standard_field for m in normalized.values()}
    min_col = min(header_by_col)
    max_col = max(header_by_col)
    detail_rows: list[K03DetailRow] = []
    total_rows: list[int] = []
    end_row = header_row

    for idx in range(header_row, len(rows)):
        row_number = idx + 1
        row = rows[idx]
        values = {col: row[col - 1] if col - 1 < len(row) else None for col in header_by_col}
        if _is_blank(values.values()):
            if detail_rows:
                break
            continue
        if _row_has_token(values.values(), ("结论", "说明", "注：", "注:")) and not detail_rows:
            continue
        if _row_has_token(values.values(), ("合计", "总计", "小计", "total")):
            total_rows.append(row_number)
            end_row = row_number
            continue
        if _row_has_token(values.values(), ("结论",)) and detail_rows:
            break

        raw_values: dict[str, Any] = {}
        normalized_values: dict[str, Any] = {}
        cell_refs: dict[str, str] = {}
        for col, header in header_by_col.items():
            value = values.get(col)
            raw_values[header] = value
            field = field_by_col.get(col)
            if field:
                normalized_values[field] = value
                cell_refs[field] = f"{get_column_letter(col)}{row_number}"
        if not _is_blank(raw_values.values()):
            detail_rows.append(
                K03DetailRow(
                    source_row=row_number,
                    raw_values=raw_values,
                    normalized_values=normalized_values,
                    cell_refs=cell_refs,
                )
            )
            end_row = row_number

    if not detail_rows and not total_rows:
        return [], None, []
    start_row = detail_rows[0].source_row if detail_rows else total_rows[0]
    return (
        detail_rows,
        K03Area(start_row=start_row, end_row=end_row, start_col=min_col, end_col=max_col),
        total_rows,
    )


def _detect_conclusion_area(
    rows: list[tuple[Any, ...]],
    header_row: int | None,
    detail_range: K03Area | None,
) -> K03Area | None:
    start = detail_range.end_row if detail_range and detail_range.end_row else header_row or 0
    matches = []
    for idx, row in enumerate(rows, start=1):
        if idx <= start:
            continue
        if _row_has_token(row, ("结论", "未见异常", "可以接受", "无需调整")):
            matches.append(idx)
    return _area_for_rows(rows, matches)


def _area_for_rows(rows: list[tuple[Any, ...]], row_numbers: list[int]) -> K03Area | None:
    if not row_numbers:
        return None
    texts = []
    start_col = None
    end_col = None
    for row_number in row_numbers:
        row = rows[row_number - 1]
        for col, value in enumerate(row, start=1):
            text = _text(value)
            if not text:
                continue
            texts.append(text)
            start_col = col if start_col is None else min(start_col, col)
            end_col = col if end_col is None else max(end_col, col)
    return K03Area(
        start_row=min(row_numbers),
        end_row=max(row_numbers),
        start_col=start_col,
        end_col=end_col,
        text=" ".join(texts)[:500] if texts else None,
    )


def _rows_containing(rows: list[tuple[Any, ...]], tokens: tuple[str, ...]) -> list[int]:
    result: list[int] = []
    for idx, row in enumerate(rows, start=1):
        if _row_has_token(row, tokens):
            result.append(idx)
    return result


def _by_item_score(normalized: dict[str, FieldMapping]) -> int:
    fields = set(normalized)
    score = len(fields & _BY_ITEM_CORE_FIELDS)
    if "current_depreciation" in fields:
        score += 1
    if "audit_recalculated_depreciation" in fields:
        score += 1
    if "management_depreciation" in fields:
        score += 1
    if "depreciation_difference" in fields:
        score += 1
    if "conclusion" in fields:
        score += 1
    return score


def _sample_score(
    sheet_name: str,
    rows: list[tuple[Any, ...]],
    header_cells: list[tuple[int, str]],
) -> int:
    text = _combined_text(sheet_name, rows[:30])
    headers = " ".join(text for _, text in header_cells)
    score = 0
    for token in ("抽样", "样本", "选样", "sample", "凭证", "检查程序"):
        if token in text.lower() or token in headers.lower():
            score += 1
    return score


def _is_policy_review(
    sheet_name: str,
    rows: list[tuple[Any, ...]],
    kind: SheetKind,
) -> bool:
    if kind == SheetKind.DEPRECIATION_POLICY:
        return True
    text = _combined_text(sheet_name, rows[:20])
    return "K.03.3" in sheet_name or ("折旧政策" in text and "复核" in text)


def _is_sap_sheet(sheet_name: str, kind: SheetKind) -> bool:
    return kind == SheetKind.SAP or "K.03.1" in sheet_name or "SAP" in sheet_name.upper()


def _looks_like_k03_sheet(sheet_name: str) -> bool:
    normalized = sheet_name.replace(" ", "").upper()
    return "K.03" in normalized or "K03" in normalized


def _combined_text(sheet_name: str, rows: list[tuple[Any, ...]]) -> str:
    cells = [sheet_name]
    for row in rows:
        for value in row:
            text = _text(value)
            if text:
                cells.append(text)
    return " ".join(cells)


def _row_has_token(values: Any, tokens: tuple[str, ...]) -> bool:
    text = " ".join(_text(value) for value in values if _text(value))
    lower = text.lower()
    return any(token.lower() in lower for token in tokens)


def _is_blank(values: Any) -> bool:
    return not any(_text(value) for value in values)


def _text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def _norm(text: str) -> str:
    return re.sub(r"\s+", "", str(text).strip().lower())


def _parse_sap_sheet(
    path: Path,
    sheet_name: str,
    rows: list[tuple[Any, ...]],
) -> K03SheetDataset:
    name_text = _norm(sheet_name)
    head_text = _norm(_combined_text(sheet_name, rows[:15]))
    if "高精确度" in name_text or "高精度" in name_text or "highprecision" in head_text:
        execution_path = EXECUTION_PATH_SAP_HIGH
        template_type = "sap_high_precision"
    elif "中精确度" in name_text or "中精度" in name_text or "mediumprecision" in head_text:
        execution_path = EXECUTION_PATH_SAP_MEDIUM
        template_type = "sap_medium_precision"
    else:
        execution_path = EXECUTION_PATH_UNKNOWN
        template_type = "sap"

    warnings: list[str] = []
    if execution_path == EXECUTION_PATH_UNKNOWN:
        warnings.append("k03_sap_precision_not_identified")

    summary = _extract_sap_summary(rows)
    return K03SheetDataset(
        workbook_name=path.name,
        source_file=str(path),
        sheet_name=sheet_name,
        k03_branch=K03_BRANCH_DEPRECIATION_TEST,
        execution_path=execution_path,
        template_type=template_type,
        ingest_depth=INGEST_DEPTH_LIGHTWEIGHT,
        rule_status=RULE_STATUS_READY_FOR_LATER_RULES,
        detected_sections=_sap_detected_sections(summary),
        warnings=warnings,
        row_count=sum(1 for item in summary.get("sap_deviation_rows", []) if item),
        column_count=0,
        unsupported_or_later_phase=False,
        summary=summary,
        llm_candidate_context={
            "sap_expectation_text": summary.get("sap_expectation_text"),
            "sap_precision_reason_text": summary.get("sap_precision_reason_text"),
            "sap_conclusion_text": summary.get("sap_conclusion_text"),
            "warnings": warnings,
        },
    )


def _parse_tod_sheet(
    path: Path,
    sheet_name: str,
    rows: list[tuple[Any, ...]],
    classified_kind: SheetKind,
) -> K03SheetDataset:
    header_row, header_cells, _ = scan_rows_for_headers(
        rows,
        sheet_kind=SheetKind.DEPRECIATION_TOD,
    )
    mapped_fields, unmapped = _map_k03_headers(header_cells)
    normalized = {m.standard_field: m for m in mapped_fields}
    raw_columns = [
        K03Column(
            source_header=text,
            column_index=col,
            column_letter=get_column_letter(col),
            standard_field=next(
                (m.standard_field for m in mapped_fields if m.column_index == col),
                None,
            ),
        )
        for col, text in header_cells
    ]
    role_evidence = resolve_tod_role(rows, sheet_name)
    warnings: list[str] = []

    if role_evidence.status == "FOUND" and role_evidence.role == "tod_sampling_output":
        execution_path = EXECUTION_PATH_TOD_SAMPLING
        ingest_depth = INGEST_DEPTH_LIGHTWEIGHT
        template_type = "tod_sampling_output"
    elif role_evidence.status == "FOUND" and role_evidence.role == "tod_sampling":
        execution_path = EXECUTION_PATH_TOD_SAMPLING
        ingest_depth = INGEST_DEPTH_LIGHTWEIGHT
        template_type = "tod_sampling"
    elif role_evidence.status == "FOUND" and role_evidence.role == "tod_by_item":
        execution_path = EXECUTION_PATH_TOD_BY_ITEM
        ingest_depth = INGEST_DEPTH_DETAILED
        template_type = "tod_by_item"
    else:
        execution_path = EXECUTION_PATH_UNKNOWN
        ingest_depth = INGEST_DEPTH_LIGHTWEIGHT
        template_type = "tod_ambiguous" if role_evidence.status == "AMBIGUOUS" else "tod_incomplete"
        warnings.append("k03_tod_role_ambiguous" if role_evidence.status == "AMBIGUOUS" else "k03_tod_execution_path_not_identified")

    detail_rows, detail_range, total_rows = _extract_detail_rows(
        rows,
        header_row=header_row,
        header_cells=header_cells,
        normalized=normalized,
    )
    conclusion = _detect_conclusion_area(rows, header_row, detail_range)
    note = _area_for_rows(rows, _rows_containing(rows, ("说明", "注：", "注", "Notes", "note")))
    instruction = _area_for_rows(rows, _rows_containing(rows, ("获取", "编制", "按照", "根据")))
    detected_sections: list[str] = []
    if header_row:
        detected_sections.append("header")
    if detail_range:
        detected_sections.append("detail_table")
    if total_rows:
        detected_sections.append("total_rows")
    if conclusion:
        detected_sections.append("conclusion_area")
    if note:
        detected_sections.append("note_area")
    if instruction:
        detected_sections.append("instruction_area")

    if unmapped:
        warnings.append("k03_unmapped_columns_present")
    missing_noncritical = sorted(_BY_ITEM_CORE_FIELDS - set(normalized))
    if execution_path == EXECUTION_PATH_TOD_BY_ITEM and missing_noncritical:
        warnings.append("k03_tod_by_item_missing_noncritical_fields:" + ",".join(missing_noncritical))

    preview_rows = [row.to_dict() for row in detail_rows[:5]]
    field_summary = {
        "raw_column_count": len(raw_columns),
        "mapped_field_count": len(normalized),
        "mapped_fields": sorted(normalized),
        "unmapped_column_count": len(unmapped),
        "amount_columns": sorted(set(normalized) & _AMOUNT_FIELDS),
        "date_columns": sorted(set(normalized) & _DATE_FIELDS),
    }
    sampling_summary = _extract_tod_sampling_summary(rows, template_type, detail_rows)
    table_ref = (
        K03DetailTableRef(
            source_file=str(path),
            sheet_name=sheet_name,
            start_row=detail_range.start_row,
            end_row=detail_range.end_row,
            start_col=detail_range.start_col,
            end_col=detail_range.end_col,
            header_row=header_row,
        )
        if detail_range
        else None
    )

    return K03SheetDataset(
        workbook_name=path.name,
        source_file=str(path),
        sheet_name=sheet_name,
        k03_branch=K03_BRANCH_DEPRECIATION_TEST,
        execution_path=execution_path,
        template_type=template_type,
        ingest_depth=ingest_depth,
        rule_status=RULE_STATUS_READY_FOR_LATER_RULES,
        detected_sections=detected_sections,
        header_rows=[header_row] if header_row else [],
        detail_table_ref=table_ref,
        detail_table_range=detail_range,
        total_rows=total_rows,
        conclusion_area=conclusion,
        note_area=note,
        instruction_area=instruction,
        raw_columns=raw_columns,
        normalized_column_map={
            field: K03Column(
                source_header=m.source_header,
                column_index=m.column_index,
                column_letter=get_column_letter(m.column_index),
                standard_field=m.standard_field,
            )
            for field, m in normalized.items()
        },
        unmapped_columns=unmapped,
        warnings=warnings,
        row_count=len(detail_rows),
        column_count=len(raw_columns),
        amount_columns=sorted(set(normalized) & _AMOUNT_FIELDS),
        date_columns=sorted(set(normalized) & _DATE_FIELDS),
        unsupported_or_later_phase=False,
        summary={
            **field_summary,
            **sampling_summary,
            "role_evidence": role_evidence.to_dict(),
            "total_row_count": len(total_rows),
            "has_conclusion_area": conclusion is not None,
            "has_note_area": note is not None,
            "has_instruction_area": instruction is not None,
        },
        preview_rows=preview_rows,
        llm_candidate_context={
            "instruction_area": instruction.to_dict() if instruction else None,
            "note_area": note.to_dict() if note else None,
            "conclusion_area": conclusion.to_dict() if conclusion else None,
            "warnings": warnings,
        },
    )


def _extract_sap_summary(rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    labels = {
        "sap_entity_type": ("实体类型", "entity type"),
        "sap_te": ("可容忍误差", "te"),
        "sap_cra": ("cra",),
        "sap_evidence_level": ("计划的保证水平", "planned assurance"),
        "sap_threshold": ("偏差阈值", "deviation threshold"),
    }
    summary: dict[str, Any] = {}
    for key, candidates in labels.items():
        value, row_no, col_no = _find_label_value(rows, candidates)
        if value is not None:
            summary[key] = value
            summary[f"{key}_row"] = row_no
            summary[f"{key}_col"] = col_no
    summary["sap_expectation_text"] = _collect_rows_after_tokens(rows, ("预期", "expectation"), max_rows=8)
    summary["sap_precision_reason_text"] = _collect_rows_after_tokens(rows, ("精确", "precision"), max_rows=8)
    summary["sap_decomposition_text"] = _collect_rows_after_tokens(rows, ("细分", "分解", "decomposition"), max_rows=8)
    summary["sap_note_text"] = _collect_rows_after_tokens(rows, ("说明", "note", "备注"), max_rows=8)
    summary["sap_conclusion_text"] = _collect_rows_after_tokens(rows, ("结论", "conclusion"), max_rows=8)
    deviation_rows = _extract_sap_deviation_rows(rows)
    summary["sap_deviation_rows"] = deviation_rows
    summary["sap_deviation_over_threshold_count"] = sum(
        1 for item in deviation_rows if _looks_yes(item.get("over_threshold"))
    )
    return summary


def _extract_tod_sampling_summary(
    rows: list[tuple[Any, ...]],
    template_type: str,
    detail_rows: list[K03DetailRow],
) -> dict[str, Any]:
    if template_type == "tod_sampling_output":
        return _extract_tod_sampling_output_summary(rows)
    if template_type != "tod_sampling":
        return {}
    summary: dict[str, Any] = {
        "tod_population_amount": _find_label_value(rows, ("折旧费用总体", "population"))[0],
        "tod_breakdown_depreciation_amount": _find_label_value(rows, ("breakdown", "折旧计提金额"))[0],
        "tod_key_item_amount": _find_label_value(rows, ("测试的关键项目", "key item"))[0],
        "tod_remaining_population": _find_label_value(rows, ("剩余总体", "remaining population"))[0],
        "tod_key_item_reason_text": _collect_rows_after_tokens(rows, ("关键项目", "key item"), max_rows=6),
        "tod_conclusion_text": _collect_rows_after_tokens(rows, ("结论", "conclusion"), max_rows=8),
        "tod_sample_rows_count": len(detail_rows),
    }
    _, sampling_rows, sampling_state = _detect_sampling_detail_table(rows)
    summary["tod_sampling_detail_state"] = sampling_state
    summary["tod_sampling_rows"] = [item.to_dict() for item in sampling_rows]
    if detail_rows:
        summary["tod_sample_preview"] = [row.to_dict() for row in detail_rows[:5]]
    return summary


def _extract_tod_sampling_output_summary(rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    header_row, sampling_rows, table_state = _detect_sampling_detail_table(rows)
    sample_type_counts = {"key": 0, "representative": 0, "replacement": 0}
    for item in sampling_rows:
        if item.sample_type in sample_type_counts:
            sample_type_counts[item.sample_type] += 1
    te, te_row, te_col, te_state = _find_unique_parameter(
        rows, ("可容忍误差", "TE"), end_row=header_row
    )
    currency, currency_row, currency_col, currency_state = _find_unique_parameter(
        rows, ("抽样单位货币列", "抽样货币单元", "sampling currency"), end_row=header_row
    )
    return {
        "sample_output_te": te,
        "sample_output_te_row": te_row,
        "sample_output_te_col": te_col,
        "sample_output_te_state": te_state,
        "sample_output_sampling_currency": currency,
        "sample_output_sampling_currency_row": currency_row,
        "sample_output_sampling_currency_col": currency_col,
        "sample_output_sampling_currency_state": currency_state,
        "sample_output_population_amount": _find_label_value(rows, ("总体金额", "population amount"), max_offset=12)[0],
        "sample_output_key_item_count": _find_label_value(rows, ("关键项目数量", "key item quantity"))[0],
        "sample_output_representative_count": _find_label_value(rows, ("代表性样本量", "representative sample quantity"))[0],
        "sample_output_selected_count": _find_label_value(rows, ("代表性样本与关键项数量合计", "selected sample total"))[0],
        "sample_output_key_item_amount": _find_label_value(rows, ("关键项目金额", "key item amount"))[0],
        "sample_output_dual_purpose": _find_label_value(rows, ("双重目的", "dual purpose"))[0],
        "sample_output_overstatement": _find_label_value(rows, ("高估", "overstatement"))[0],
        "sample_output_assurance_level": _find_label_value(rows, ("保证水平", "assurance"))[0],
        "sample_output_sample_pool_amount": _find_label_value(rows, ("样本池", "sample pool"))[0],
        "sample_output_expected_misstatement": _find_label_value(rows, ("预期错报", "expected misstatement"))[0],
        "sample_output_sampling_method": _find_label_value(rows, ("抽样方法", "sampling method"))[0],
        "sample_output_key_item_rows_count": sample_type_counts["key"],
        "sample_output_representative_rows_count": sample_type_counts["representative"],
        "sample_output_replacement_rows_count": sample_type_counts["replacement"],
        "sample_output_selected_rows_count": sample_type_counts["key"] + sample_type_counts["representative"],
        "sample_output_detail_header_row": header_row,
        "sample_output_detail_state": table_state,
        "sample_output_rows": [item.to_dict() for item in sampling_rows],
        "sample_output_summary_text": _collect_rows_after_tokens(rows, ("样本", "sample"), max_rows=10),
    }


def _sampling_output_parameter_score(rows: list[tuple[Any, ...]]) -> int:
    tokens = ("可容忍误差", "抽样货币单元", "抽样单位货币列", "抽样方法", "样本池")
    text = _norm(" ".join(_text(value) for row in rows for value in row if _text(value)))
    return sum(1 for token in tokens if _norm(token) in text)


def _sampling_header_field(value: Any) -> str | None:
    text = _norm(_text(value))
    aliases = {
        "sample_type": ("样本类型", "sampletype"),
        "asset_id": ("固定资产编号", "资产编号", "assetid", "fano"),
        "asset_name": ("固定资产名称", "资产名称", "assetname"),
        "current_depreciation": ("本期计提折旧", "账面计提折旧费用", "bookdepreciation"),
        "recalculated_depreciation": ("重新计算折旧费用", "重新计算折旧", "重算折旧", "recalculateddepreciation"),
        "depreciation_difference": ("差异", "折旧差异", "difference"),
        "evidence_description": ("获得的证据/支持的描述", "获得的证据支持的描述", "获得的证据", "支持的描述", "evidencedescription"),
        "note": ("notes", "note", "差异说明", "异常说明", "跟进说明"),
    }
    for field, names in aliases.items():
        if any(
            text == _norm(name) or (len(_norm(name)) >= 4 and _norm(name) in text)
            for name in names
        ):
            return field
    return None


def _normalize_sample_type(value: Any) -> str | None:
    text = _norm(_text(value))
    if "替换样本" in text or "replacementsample" in text:
        return "replacement"
    if "代表性样本" in text or "representativesample" in text:
        return "representative"
    if "关键项" in text or "keyitem" in text:
        return "key"
    return None


def _normalize_asset_id(value: Any) -> str | None:
    text = _text(value).strip()
    if not text:
        return None
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text.replace(" ", "").upper()


def _detect_sampling_detail_table(
    rows: list[tuple[Any, ...]],
) -> tuple[int | None, list[K03SamplingRow], str]:
    candidates: list[tuple[int, dict[int, str]]] = []
    for row_no, row in enumerate(rows, start=1):
        current_fields = {
            col: field for col, value in enumerate(row, start=1)
            if (field := _sampling_header_field(value))
        }
        if len(current_fields) < 2 and _header_fragment_count(row) < 2:
            continue
        fields = dict(current_fields)
        combined_values = _combined_header_values(rows, row_no)
        for col, value in enumerate(combined_values, start=1):
            if col not in fields and (field := _sampling_header_field(value)):
                fields[col] = field
        if {"current_depreciation", "recalculated_depreciation", "depreciation_difference"}.issubset(fields.values()):
            numeric_cols = [col for col, value in enumerate(row, start=1) if _norm(_text(value)) in {"1", "2"}]
            if numeric_cols:
                fields[numeric_cols[0]] = "test_attribute_1"
            if len(numeric_cols) > 1:
                fields[numeric_cols[1]] = "test_attribute_2"
        if {"sample_type", "asset_id"}.issubset(fields.values()) and len(fields) >= 3:
            candidates.append((row_no, fields))
    if not candidates:
        return None, [], "MISSING"
    if len(candidates) > 1:
        return None, [], "AMBIGUOUS"
    header_row, fields = candidates[0]
    result: list[K03SamplingRow] = []
    blank_streak = 0
    for row_no in range(header_row + 1, len(rows) + 1):
        row = rows[row_no - 1]
        values = {field: _cell_value(row, col) for col, field in fields.items()}
        sample_type = _normalize_sample_type(values.get("sample_type"))
        asset_id = _normalize_asset_id(values.get("asset_id"))
        if not sample_type and not asset_id:
            blank_streak += 1
            if result and blank_streak >= 2:
                break
            continue
        blank_streak = 0
        result.append(K03SamplingRow(
            source_row=row_no,
            sample_type=sample_type,
            asset_id=asset_id,
            values=values,
            cell_refs={field: f"{get_column_letter(col)}{row_no}" for col, field in fields.items()},
        ))
    return header_row, result, "FOUND" if result else "INCOMPLETE"


def _combined_header_values(rows: list[tuple[Any, ...]], row_no: int) -> list[str]:
    current = rows[row_no - 1]
    width = len(current)
    result: list[str] = []
    for col in range(1, width + 1):
        parts: list[str] = []
        for candidate_row in range(max(1, row_no - 2), row_no + 1):
            source = rows[candidate_row - 1]
            if candidate_row < row_no and _header_fragment_count(source) < 2:
                continue
            value = _cell_value(source, col)
            if not _text(value) and candidate_row < row_no:
                for prior_col in range(col - 1, max(0, col - 4), -1):
                    carried = _cell_value(source, prior_col)
                    if _text(carried):
                        value = carried
                        break
            if _text(value):
                parts.append(_text(value))
        result.append(" ".join(dict.fromkeys(parts)))
    return result


def _header_fragment_count(row: tuple[Any, ...]) -> int:
    fragments = ("样本", "类型", "固定资产", "资产", "编号", "名称", "折旧", "差异", "证据", "属性",
                 "sample", "type", "asset", "id", "name", "depreciation", "difference", "evidence")
    return sum(
        1 for value in row
        if any(_norm(fragment) == _norm(_text(value)) for fragment in fragments)
    )


def _find_unique_parameter(
    rows: list[tuple[Any, ...]], labels: tuple[str, ...], *, end_row: int | None = None
) -> tuple[Any, int | None, int | None, str]:
    normalized_labels = tuple(_norm(label) for label in labels)
    matches: list[tuple[Any, int, int]] = []
    for row_no, row in enumerate(rows, start=1):
        if end_row is not None and row_no >= end_row:
            break
        for col_no, value in enumerate(row, start=1):
            text = _norm(_text(value))
            if not text or not any(
                text == label or (len(label) >= 4 and label in text)
                for label in normalized_labels
            ):
                continue
            for offset in range(1, min(12, len(row) - col_no) + 1):
                candidate = _cell_value(row, col_no + offset)
                if _text(candidate):
                    matches.append((candidate, row_no, col_no + offset))
                    break
    unique: dict[str, tuple[Any, int, int]] = {}
    for value, row, col in matches:
        unique.setdefault(_norm(str(value)), (value, row, col))
    if not unique:
        return None, None, None, "MISSING"
    if len(unique) > 1:
        return None, None, None, "AMBIGUOUS"
    value, row, col = next(iter(unique.values()))
    return value, row, col, "FOUND"


def _has_parameter_label(
    rows: list[tuple[Any, ...]], labels: tuple[str, ...], *, end_row: int | None = None
) -> bool:
    normalized_labels = tuple(_norm(label) for label in labels)
    for row_no, row in enumerate(rows, start=1):
        if end_row is not None and row_no >= end_row:
            break
        for value in row:
            text = _norm(_text(value))
            if text and any(
                text == label or (len(label) >= 4 and label in text)
                for label in normalized_labels
            ):
                return True
    return False


def _find_label_value(
    rows: list[tuple[Any, ...]],
    labels: tuple[str, ...],
    *,
    max_offset: int = 4,
) -> tuple[Any, int | None, int | None]:
    compact_labels = tuple(_norm(label) for label in labels)
    for row_no, row in enumerate(rows, start=1):
        for col_no, value in enumerate(row, start=1):
            text = _norm(_text(value))
            if not text or not any(label in text for label in compact_labels):
                continue
            for offset in range(1, max_offset + 1):
                candidate = _cell_value(row, col_no + offset)
                if _text(candidate):
                    return candidate, row_no, col_no + offset
    return None, None, None


def _collect_rows_after_tokens(
    rows: list[tuple[Any, ...]],
    tokens: tuple[str, ...],
    *,
    max_rows: int,
) -> str | None:
    compact_tokens = tuple(_norm(token) for token in tokens)
    collected: list[str] = []
    started = False
    for row in rows:
        row_text = " ".join(_text(value) for value in row if _text(value))
        compact = _norm(row_text)
        if not started:
            if not any(token in compact for token in compact_tokens):
                continue
            started = True
        if started and row_text:
            collected.append(row_text)
            if len(collected) >= max_rows:
                break
    return " ".join(collected) if collected else None


def _extract_sap_deviation_rows(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    marker_row = None
    vertical_columns: dict[str, int] = {}
    for row_no, row in enumerate(rows, start=1):
        normalized_cells = [_norm(_text(value)) for value in row]
        row_text = " ".join(normalized_cells)
        if "是否超过已分配偏差阈值" in row_text or "差异是否超过已分配偏差阈值" in row_text:
            for col_no, text in enumerate(normalized_cells, start=1):
                if "账面计提折旧" in text or "实际折旧" in text:
                    vertical_columns["actual_depreciation"] = col_no
                elif text == "差异" or "差异=" in text:
                    vertical_columns["deviation_amount"] = col_no
                elif "已分配偏差阈值" in text and "是否" not in text:
                    vertical_columns["threshold"] = col_no
                elif "是否超过" in text:
                    vertical_columns["over_threshold"] = col_no
            marker_row = row_no
            break
        if "偏差是否超过阈值" in row_text:
            marker_row = row_no
            break
    if marker_row is None:
        return []
    if vertical_columns.get("over_threshold"):
        result: list[dict[str, Any]] = []
        for row_no in range(marker_row + 1, min(len(rows), marker_row + 100) + 1):
            row = rows[row_no - 1]
            values = {
                key: _cell_value(row, col_no)
                for key, col_no in vertical_columns.items()
            }
            if not any(_text(value) for value in values.values()):
                continue
            first_text = _norm(" ".join(_text(value) for value in row if _text(value)))
            if any(token in first_text for token in ("绝对值偏差", "偏差阈值", "是否超过阈值")):
                break
            if not _text(values.get("deviation_amount")):
                continue
            result.append(
                {
                    "row": row_no,
                    "column": vertical_columns["over_threshold"],
                    **values,
                }
            )
        return result[:100]
    result: list[dict[str, Any]] = []
    marker = rows[marker_row - 1]
    deviation = rows[marker_row - 3] if marker_row >= 3 else ()
    threshold = rows[marker_row - 4] if marker_row >= 4 else ()
    actual = rows[marker_row - 2] if marker_row >= 2 else ()
    for col_no in range(1, max(len(marker), len(deviation), len(threshold), len(actual)) + 1):
        over_threshold = _cell_value(marker, col_no)
        if not _text(over_threshold):
            continue
        result.append(
            {
                "row": marker_row,
                "column": col_no,
                "actual_depreciation": _cell_value(actual, col_no),
                "deviation_amount": _cell_value(deviation, col_no),
                "threshold": _cell_value(threshold, col_no),
                "over_threshold": over_threshold,
            }
        )
    return result[:20]


def _looks_yes(value: Any) -> bool:
    text = _norm(_text(value))
    return text in {"是", "yes", "y", "true"} or text.startswith("是")


def _sap_detected_sections(summary: dict[str, Any]) -> list[str]:
    sections = ["sap_template"]
    for key, name in (
        ("sap_expectation_text", "expectation_area"),
        ("sap_precision_reason_text", "precision_area"),
        ("sap_deviation_rows", "deviation_table"),
        ("sap_conclusion_text", "conclusion_area"),
    ):
        if summary.get(key):
            sections.append(name)
    return sections


def _parse_k03_sheet(
    *,
    path: Path,
    sheet_name: str,
    rows: list[tuple[Any, ...]],
    classified_kind: SheetKind,
    classification_confidence: float,
) -> K03SheetDataset:
    sheet_text = _norm(sheet_name)
    if "选样输出" in sheet_text or "抽样" in sheet_text or "sampling" in sheet_text:
        return _parse_tod_sheet(path, sheet_name, rows, classified_kind)
    if _is_policy_review(sheet_name, rows, classified_kind):
        return _parse_policy_review(path, sheet_name, rows, classification_confidence)
    if _is_sap_sheet(sheet_name, classified_kind):
        return _parse_sap_sheet(path, sheet_name, rows)
    return _parse_tod_sheet(path, sheet_name, rows, classified_kind)
