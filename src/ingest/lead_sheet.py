from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Literal

import openpyxl
from openpyxl.utils import get_column_letter

from ingest.lead_sheet_blocks import (
    LeadBlock,
    LeadBlockKind,
    block_for_kind,
    detect_lead_blocks,
    slice_rows_for_block,
)
from ingest.models import SheetKind
from ingest.sheet_classifier import classify_sheet, score_by_name
from ingest.sheet_period_routing import choose_sheet_candidate, sort_sheet_candidates
from ingest.workbook_reader import read_worksheet_rows

_DEFAULT_MAX_ROWS = 200


@dataclass
class MaterialityCapture:
    """PM/TE/SAD 等重要性参数摘录。"""

    field_key: str
    label: str
    workpaper_value: str | None = None
    canvas_value: str | None = None
    source_row: int | None = None
    source_col_workpaper: int | None = None
    source_col_canvas: int | None = None

    def cell_ref_workpaper(self, sheet_name: str) -> str | None:
        if self.source_row and self.source_col_workpaper:
            return f"{sheet_name}!{get_column_letter(self.source_col_workpaper)}{self.source_row}"
        return None

    def cell_ref_canvas(self, sheet_name: str) -> str | None:
        if self.source_row and self.source_col_canvas:
            return f"{sheet_name}!{get_column_letter(self.source_col_canvas)}{self.source_row}"
        return None

    def to_dict(self, sheet_name: str) -> dict[str, Any]:
        return {
            "field_key": self.field_key,
            "label": self.label,
            "workpaper_value": self.workpaper_value,
            "canvas_or_external_value": self.canvas_value,
            "workpaper_cell": self.cell_ref_workpaper(sheet_name),
            "canvas_cell": self.cell_ref_canvas(sheet_name),
            "compare_status": "pending_manual",
        }


@dataclass
class LeadBasicInfoField:
    field_key: str
    label: str
    value: str | None = None
    source_row: int | None = None
    source_col: int | None = None

    def to_dict(self, sheet_name: str) -> dict[str, Any]:
        cell = None
        if self.source_row and self.source_col:
            cell = f"{sheet_name}!{get_column_letter(self.source_col)}{self.source_row}"
        return {
            "field_key": self.field_key,
            "label": self.label,
            "value": self.value,
            "source_cell": cell,
        }


@dataclass
class CraAssertionRow:
    """认定级 CRA / TT 摘录。"""

    assertion: str
    cra: str | None = None
    tt: str | None = None
    tt_overall: str | None = None
    source_row: int | None = None
    source_col_assertion: int | None = None
    source_col_cra: int | None = None
    source_col_tt: int | None = None
    source_col_tt_overall: int | None = None

    def to_dict(self, sheet_name: str) -> dict[str, Any]:
        def ref(col: int | None) -> str | None:
            if self.source_row and col:
                return f"{sheet_name}!{get_column_letter(col)}{self.source_row}"
            return None

        return {
            "assertion": self.assertion,
            "cra": self.cra,
            "tt": self.tt,
            "tt_overall": self.tt_overall,
            "assertion_cell": ref(self.source_col_assertion),
            "cra_cell": ref(self.source_col_cra),
            "tt_cell": ref(self.source_col_tt),
            "tt_overall_cell": ref(self.source_col_tt_overall),
            "compare_status": "pending_manual",
        }


@dataclass
class ExpectationRow:
    account_change: str
    expectation: str | None
    source_row: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "account_change": self.account_change,
            "expectation": self.expectation,
            "source_row": self.source_row,
        }


@dataclass
class VolatilityThreshold:
    amount: str | None = None
    percent: str | None = None
    source_row_amount: int | None = None
    source_row_percent: int | None = None
    """波动幅度金额口径：简版 Lead 无 CRA 区时可能直接 link TE，而非认定 TT。"""
    amount_source: Literal["te", "tt"] | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "amount": self.amount,
            "percent": self.percent,
            "source_row_amount": self.source_row_amount,
            "source_row_percent": self.source_row_percent,
            "amount_source": self.amount_source,
        }


@dataclass
class LeadMovementColumnBinding:
    role: str
    source_header: str
    column_index: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "role": self.role,
            "source_header": self.source_header,
            "column_index": self.column_index,
        }


@dataclass
class LeadMovementRow:
    account_label: str
    sheet_ref: str | None
    values: dict[str, str | None]
    source_row: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "account_label": self.account_label,
            "sheet_ref": self.sheet_ref,
            "values": self.values,
            "source_row": self.source_row,
        }


@dataclass
class CheckWithA3AccountLine:
    """引导表账户行与 Check with A3 / Diff 按金额列对齐的摘录。"""

    account_label: str
    amount_role: str | None
    movement_value: str | None = None
    a3_value: str | None = None
    diff_value: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "account_label": self.account_label,
            "amount_role": self.amount_role,
            "movement_value": self.movement_value,
            "a3_value": self.a3_value,
            "diff_value": self.diff_value,
        }


@dataclass
class LeadCheckWithA3:
    """K.00 引导主表末尾 Check with A3、Diff 与 Notes 区。"""

    check_source_row: int | None = None
    diff_source_row: int | None = None
    notes_source_row: int | None = None
    notes_text: str | None = None
    lines: list[CheckWithA3AccountLine] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "check_source_row": self.check_source_row,
            "diff_source_row": self.diff_source_row,
            "notes_source_row": self.notes_source_row,
            "notes_text": self.notes_text,
            "lines": [ln.to_dict() for ln in self.lines],
        }


@dataclass
class AdjustmentSummaryRow:
    adjustment_type: str | None
    source_row: int
    raw_cells: list[str | None] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "adjustment_type": self.adjustment_type,
            "source_row": self.source_row,
            "raw_cells": self.raw_cells,
        }


@dataclass
class LeadSheetDataset:
    source_file: str
    source_sheet: str
    blocks: list[LeadBlock] = field(default_factory=list)
    basic_info_fields: list[LeadBasicInfoField] = field(default_factory=list)
    materiality: list[MaterialityCapture] = field(default_factory=list)
    cra_rows: list[CraAssertionRow] = field(default_factory=list)
    expectations: list[ExpectationRow] = field(default_factory=list)
    volatility: VolatilityThreshold | None = None
    movement_bindings: list[LeadMovementColumnBinding] = field(default_factory=list)
    movement_rows: list[LeadMovementRow] = field(default_factory=list)
    check_with_a3: LeadCheckWithA3 | None = None
    fluctuation_notes: str | None = None
    adjustment_rows: list[AdjustmentSummaryRow] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    """Lead 版式：``no_cra_te_volatility`` = 无 CRA/TT 区且波动幅度金额取自 TE（案例 A）。"""
    layout_variant: Literal["standard", "no_cra_te_volatility"] | None = None
    usable_for_rules: bool = True

    def has_materiality_data(self) -> bool:
        return any(c.workpaper_value or c.canvas_value for c in self.materiality)

    def has_cra_data(self) -> bool:
        return len(self.cra_rows) > 0

    def block(self, kind: LeadBlockKind) -> LeadBlock | None:
        return block_for_kind(self.blocks, kind)


_BASIC_INFO_LABELS: dict[str, tuple[str, ...]] = {
    "client_name": ("客户名称", "clientname"),
    "period_end": ("期末", "periodend", "资产负债表日"),
    "analysis_date": ("分析日期", "analysisdate"),
    "te": ("可容忍误差", "te", "tolerableerror", "可容忍错报"),
    "sad": ("名义金额", "sad", "明显微小错报", "summaryauditdifference"),
    "pm": ("计划重要性", "pm", "planningmateriality", "重要性pm"),
    "gaap": ("适用会计准则", "gaap", "公认会计准则", "会计准则"),
    "currency": ("记账本位币", "currency", "本位币"),
}

_MATERIALITY_KEYS = frozenset({"pm", "te", "sad"})

_CANVAS_HEADER_HINTS = ("canvas", "a3", "外出取数", "系统", "最终", "canvasa3")

_CRA_HEADER_HINTS = ("cra", "combinedrisk", "风险", "风险评估")
_TT_HEADER_HINTS = ("tt", "测试阈值", "threshold", "各项认定")
_TT_OVERALL_HINTS = ("所有相关认定", "allrelevant")
_ASSERTION_HEADER_HINTS = ("认定", "相关认定", "assertion")

_MOVEMENT_ROLE_HEADERS: dict[str, tuple[str, ...]] = {
    "gl_code": ("总账科目编码", "科目编码", "accountcode"),
    "account_label": ("科目名称", "账户", "accountname"),
    "sheet_ref": ("索引号", "程序页", "sheet", "底稿索引"),
    "book_balance": ("期末账面数", "账面数", "bookbalance"),
    "book_adjustment": ("账表调整数", "账表调整", "bookadjustment"),
    "unaudited": ("期末未审数", "未审数", "unaudited"),
    "audit_adjustment": ("审计调整数", "auditadjustment"),
    "audited_ending": ("期末审定数", "审定数", "audited"),
    "py_audited": ("上期末审定数", "上期审定", "prioraudit"),
    "movement_amount": ("变动金额", "变动额", "changeamount"),
    "movement_pct": ("变动%", "变动%", "changepercent", "变动比例"),
    "notes": ("notes", "备注", "说明"),
    "investigate_quantitative": ("基于波动幅度判断", "波动幅度判断"),
    "investigate_qualitative": ("基于定性考虑判断", "定性考虑", "定性判断"),
}

_EXPECTATION_ROW_LABELS = (
    "新增",
    "减少",
    "在建工程",
    "转让",
    "外汇",
    "折旧方法",
    "折旧费用",
    "使用寿命",
)

_ACCOUNT_MOVEMENT_LABELS = (
    "原值",
    "累计折旧",
    "减值准备",
    "净值",
)

_AMOUNT_ROLES_FOR_A3 = ("audited_ending", "book_balance", "unaudited")

_CHECK_WITH_A3_LABEL_HINTS = ("checkwitha3", "checkwitha3", "与a3核对", "核对a3")
_DIFF_ROW_LABEL_HINTS = ("diff", "diff.")


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f").rstrip("0").rstrip(".") or "0"
    text = str(value).strip()
    return text if text else None


def _norm(text: str) -> str:
    return re.sub(r"[\s_\-（）()]", "", text.lower())


def _get_cell(rows: list[tuple[Any, ...]], row_idx: int, col_idx: int) -> str | None:
    if row_idx < 0 or row_idx >= len(rows):
        return None
    row = rows[row_idx]
    if col_idx < 0 or col_idx >= len(row):
        return None
    return _cell_str(row[col_idx])


def _label_matches(cell_norm: str, patterns: tuple[str, ...]) -> bool:
    if not cell_norm or len(cell_norm) > 50:
        return False
    for p in patterns:
        pn = _norm(p)
        if len(pn) <= 4 and pn.isascii():
            if cell_norm == pn:
                return True
            continue
        if cell_norm == pn or pn in cell_norm or cell_norm in pn:
            return True
    return False


def _is_exact_field_label(text: str) -> bool:
    """判断单元格文本是否像字段标签（整格匹配，不用子串，避免「企业会计准则」误判）。"""
    n = _norm(text)
    if not n:
        return False
    for patterns in _BASIC_INFO_LABELS.values():
        for p in patterns:
            pn = _norm(p)
            if len(pn) <= 4 and pn.isascii():
                if n == pn:
                    return True
                continue
            if n == pn:
                return True
    return False


def _is_probable_value(text: str) -> bool:
    n = _norm(text)
    if _is_exact_field_label(text):
        return False
    if n in ("认定", "cra", "tt", "canvas", "底稿值", "参考"):
        return False
    return True


def _find_canvas_column(rows: list[tuple[Any, ...]], scan_rows: int = 35) -> int | None:
    for r in range(min(scan_rows, len(rows))):
        row = rows[r]
        for c, val in enumerate(row):
            text = _cell_str(val)
            if not text:
                continue
            n = _norm(text)
            if any(h in n for h in _CANVAS_HEADER_HINTS):
                return c
    return None


def _header_col_index(header_cells: list[str | None], hints: tuple[str, ...]) -> int | None:
    for c, text in enumerate(header_cells):
        if not text:
            continue
        n = _norm(text)
        for h in hints:
            hn = _norm(h)
            if len(hn) <= 4 and hn.isascii():
                if n == hn:
                    return c
                continue
            if hn in n or n in hn:
                return c
    return None


def _match_movement_role(header: str) -> str | None:
    n = _norm(header)
    if not n:
        return None
    if "是否进一步调查" in n:
        if "波动幅度" in n:
            return "investigate_quantitative"
        if "定性" in n:
            return "investigate_qualitative"
    # 「上期末审定数」含子串「期末审定数」：PY 列仅用 hint⊆header，避免「期末审定数」反向命中 PY
    for h in _MOVEMENT_ROLE_HEADERS["py_audited"]:
        hn = _norm(h)
        if hn in n:
            return "py_audited"
    for role, hints in _MOVEMENT_ROLE_HEADERS.items():
        if role == "py_audited":
            continue
        for h in hints:
            hn = _norm(h)
            if hn in n or n in hn:
                return role
    return None


def _extract_basic_info(
    rows: list[tuple[Any, ...]],
    block: LeadBlock | None,
) -> list[LeadBasicInfoField]:
    scope = slice_rows_for_block(rows, block)
    row_offset = (block.start_row - 1) if block else 0
    canvas_col = _find_canvas_column(scope)
    found: dict[str, LeadBasicInfoField] = {}

    for r, row in enumerate(scope):
        for c, val in enumerate(row[:8]):
            text = _cell_str(val)
            if not text:
                continue
            n = _norm(text)
            for field_key, patterns in _BASIC_INFO_LABELS.items():
                if not _label_matches(n, patterns):
                    continue
                item = found.get(field_key)
                if item is None:
                    item = LeadBasicInfoField(field_key=field_key, label=text)
                    found[field_key] = item
                item.source_row = row_offset + r + 1
                for dc in (1, 2, 3):
                    v = _get_cell(scope, r, c + dc)
                    if v and _is_probable_value(v):
                        if canvas_col is not None and c + dc == canvas_col:
                            continue
                        item.value = v
                        item.source_col = c + dc + 1
                        break
                break

    display = {
        "client_name": "客户名称",
        "period_end": "期末",
        "analysis_date": "分析日期",
        "te": "可容忍误差 (TE)",
        "sad": "名义金额 (SAD)",
        "pm": "计划重要性 (PM)",
        "gaap": "适用会计准则",
        "currency": "记账本位币",
    }
    for key, item in found.items():
        item.label = display.get(key, item.label)
    return list(found.values())


def _basic_info_to_materiality(
    fields: list[LeadBasicInfoField],
    rows: list[tuple[Any, ...]],
) -> list[MaterialityCapture]:
    canvas_col = _find_canvas_column(rows)
    caps: list[MaterialityCapture] = []
    by_key = {f.field_key: f for f in fields}
    for key in ("pm", "te", "sad"):
        f = by_key.get(key)
        if f is None:
            continue
        cap = MaterialityCapture(
            field_key=key,
            label=f.label,
            workpaper_value=f.value,
            source_row=f.source_row,
            source_col_workpaper=f.source_col,
        )
        if canvas_col is not None and f.source_row:
            cv = _get_cell(rows, f.source_row - 1, canvas_col)
            if cv and _is_probable_value(cv):
                cap.canvas_value = cv
                cap.source_col_canvas = canvas_col + 1
        caps.append(cap)
    return caps


def _extract_cra_table(rows: list[tuple[Any, ...]], block: LeadBlock | None) -> list[CraAssertionRow]:
    scope = slice_rows_for_block(rows, block)
    row_offset = (block.start_row - 1) if block else 0
    header_row_idx: int | None = None
    col_assertion = col_cra = col_tt = col_tt_overall = None

    for r, row in enumerate(scope[:50]):
        cells = [_cell_str(v) for v in row[:18]]
        if not any(cells):
            continue
        ca = _header_col_index(cells, _ASSERTION_HEADER_HINTS)
        cc = _header_col_index(cells, _CRA_HEADER_HINTS)
        ct = None
        ct_all = None
        for c, text in enumerate(cells):
            if not text:
                continue
            n = _norm(text)
            if ca is not None and c == ca:
                continue
            if _label_matches(n, _TT_OVERALL_HINTS):
                ct_all = c
            elif _label_matches(n, _TT_HEADER_HINTS) and n != "认定":
                if ct is None:
                    ct = c
        hits = sum(x is not None for x in (ca, cc, ct, ct_all))
        if hits >= 2 and ca is not None:
            header_row_idx = r
            col_assertion, col_cra, col_tt, col_tt_overall = ca, cc, ct, ct_all
            break

    if header_row_idx is None:
        return []

    results: list[CraAssertionRow] = []
    for r in range(header_row_idx + 1, min(header_row_idx + 40, len(scope))):
        assertion = _get_cell(scope, r, col_assertion) if col_assertion is not None else None
        if not assertion or _norm(assertion) in ("合计", "总计", "认定"):
            if results:
                break
            continue
        if len(assertion) > 80:
            continue
        cra = _get_cell(scope, r, col_cra) if col_cra is not None else None
        tt = _get_cell(scope, r, col_tt) if col_tt is not None else None
        tt_overall = _get_cell(scope, r, col_tt_overall) if col_tt_overall is not None else None
        if not cra and not tt and not tt_overall:
            if results:
                break
            continue
        results.append(
            CraAssertionRow(
                assertion=assertion,
                cra=cra,
                tt=tt,
                tt_overall=tt_overall,
                source_row=row_offset + r + 1,
                source_col_assertion=col_assertion + 1 if col_assertion is not None else None,
                source_col_cra=col_cra + 1 if col_cra is not None else None,
                source_col_tt=col_tt + 1 if col_tt is not None else None,
                source_col_tt_overall=col_tt_overall + 1 if col_tt_overall is not None else None,
            )
        )
    return results


def _extract_expectations(
    rows: list[tuple[Any, ...]],
    block: LeadBlock | None,
) -> tuple[list[ExpectationRow], VolatilityThreshold | None]:
    scope = slice_rows_for_block(rows, block)
    row_offset = (block.start_row - 1) if block else 0
    header_r: int | None = None
    col_change = col_expect = None

    for r, row in enumerate(scope[:40]):
        cells = [_cell_str(v) for v in row[:10]]
        change_col = next(
            (i for i, t in enumerate(cells) if t and _label_matches(_norm(t), ("账户变更",))),
            None,
        )
        expect_col = next(
            (i for i, t in enumerate(cells) if t and _label_matches(_norm(t), ("预期及额外考虑", "预期"))),
            None,
        )
        if change_col is not None:
            header_r = r
            col_change = change_col
            col_expect = expect_col if expect_col is not None else change_col + 1
            break

    expectations: list[ExpectationRow] = []
    volatility = VolatilityThreshold()

    if header_r is not None:
        empty_streak = 0
        for r in range(header_r + 1, len(scope)):
            change = _get_cell(scope, r, col_change) if col_change is not None else None
            if change and _label_matches(_norm(change), ("波动范围",)):
                break
            if change and _label_matches(_norm(change), ("波动幅度",)):
                amt = _get_cell(scope, r, (col_expect or col_change or 0) + 0)
                if _norm(change).endswith("()") or "()" in change:
                    volatility.source_row_amount = row_offset + r + 1
                    volatility.amount = _get_cell(scope, r, (col_change or 0) + 1)
                continue
            if change and "波动幅度" in change and "%" in change:
                volatility.source_row_percent = row_offset + r + 1
                volatility.percent = _get_cell(scope, r, (col_change or 0) + 1)
                continue
            if not change:
                empty_streak += 1
                if empty_streak >= 3 and expectations:
                    break
                continue
            empty_streak = 0
            if len(change) > 60:
                continue
            exp = _get_cell(scope, r, col_expect) if col_expect is not None else None
            expectations.append(
                ExpectationRow(
                    account_change=change,
                    expectation=exp,
                    source_row=row_offset + r + 1,
                )
            )

    # 兜底：按已知账户变更标签扫描块内
    if not expectations:
        for r, row in enumerate(scope):
            change = _cell_str(row[1]) if len(row) > 1 else None
            if not change:
                continue
            n = _norm(change)
            if not any(_norm(lbl) in n or n.startswith(_norm(lbl)) for lbl in _EXPECTATION_ROW_LABELS):
                continue
            exp = _cell_str(row[2]) if len(row) > 2 else None
            expectations.append(
                ExpectationRow(account_change=change, expectation=exp, source_row=row_offset + r + 1)
            )

    # 波动阈值（可在预期块内任意位置）
    for r, row in enumerate(scope):
        label = _cell_str(row[1]) if len(row) > 1 else None
        if not label or "波动幅度" not in label:
            continue
        val = _cell_str(row[2]) if len(row) > 2 else None
        if "%" in label:
            volatility.source_row_percent = row_offset + r + 1
            volatility.percent = val
        else:
            volatility.source_row_amount = row_offset + r + 1
            volatility.amount = val

    vol_out = volatility if (volatility.amount or volatility.percent) else None
    return expectations, vol_out


def _is_check_with_a3_label(text: str) -> bool:
    n = _norm(text)
    if not n:
        return False
    if n in _DIFF_ROW_LABEL_HINTS:
        return False
    return "checkwitha3" in n or "checkwitha3" in n.replace(" ", "")


def _is_diff_row_label(text: str) -> bool:
    n = _norm(text)
    return n in _DIFF_ROW_LABEL_HINTS


def _is_notes_row_label(text: str) -> bool:
    n = _norm(text)
    return n.startswith("notes") or n in ("说明", "备注")


def _row_amount_values_by_role(
    scope: list[tuple[Any, ...]],
    row_idx: int,
    role_cols: dict[str, int],
) -> dict[str, str | None]:
    values: dict[str, str | None] = {}
    for role, col in role_cols.items():
        if role in ("account_label", "gl_code", "sheet_ref"):
            continue
        values[role] = _get_cell(scope, row_idx, col)
    return values


def _primary_amount_role(values: dict[str, str | None]) -> tuple[str | None, str | None]:
    for role in _AMOUNT_ROLES_FOR_A3:
        val = values.get(role)
        if val is not None and str(val).strip():
            return role, val
    return None, None


def _collect_notes_after_diff(
    scope: list[tuple[Any, ...]],
    start_r: int,
    label_col: int,
) -> tuple[int | None, str | None]:
    parts: list[str] = []
    notes_row: int | None = None
    for r in range(start_r, min(start_r + 8, len(scope))):
        row = scope[r]
        label = _get_cell(scope, r, label_col)
        notes_label_col = next(
            (
                c
                for c, value in enumerate(row)
                if value is not None and _is_notes_row_label(str(value))
            ),
            None,
        )
        if label and (_is_check_with_a3_label(label) or _is_diff_row_label(label)):
            continue
        if notes_label_col is not None:
            notes_row = r
            for c in range(len(row)):
                if c == notes_label_col:
                    continue
                text = _cell_str(row[c]) if c < len(row) else None
                if text and len(text) > 2:
                    parts.append(text)
            continue
        if notes_row is not None or parts:
            for c in range(len(row)):
                text = _cell_str(row[c]) if c < len(row) else None
                if not text:
                    continue
                n = _norm(text)
                if _is_notes_row_label(text) or n in ("波动说明",):
                    break
                if len(text) >= 4:
                    parts.append(text)
            if parts and r > (notes_row or start_r):
                break
    text = "\n".join(parts).strip() if parts else None
    return notes_row, text or None


def _extract_check_with_a3(
    scope: list[tuple[Any, ...]],
    *,
    role_cols: dict[str, int],
    movement_rows: list[LeadMovementRow],
    row_offset: int,
    start_r: int,
) -> LeadCheckWithA3 | None:
    """从引导主表末尾摘录 Check with A3、Diff 及 Notes（与四行按金额列对齐）。"""
    label_col = role_cols.get("account_label", 1)
    check_r: int | None = None
    diff_r: int | None = None
    check_by_role: dict[str, str | None] = {}
    diff_by_role: dict[str, str | None] = {}

    for r in range(start_r, min(start_r + 12, len(scope))):
        label = _get_cell(scope, r, label_col)
        if not label:
            continue
        if _is_check_with_a3_label(label):
            check_r = r
            check_by_role = _row_amount_values_by_role(scope, r, role_cols)
            continue
        if _is_diff_row_label(label):
            diff_r = r
            diff_by_role = _row_amount_values_by_role(scope, r, role_cols)
            break

    if check_r is None and diff_r is None:
        return None

    notes_row, notes_text = (None, None)
    if diff_r is not None:
        notes_row, notes_text = _collect_notes_after_diff(scope, diff_r + 1, label_col)
    elif check_r is not None:
        notes_row, notes_text = _collect_notes_after_diff(scope, check_r + 1, label_col)

    lines: list[CheckWithA3AccountLine] = []
    for mov in movement_rows:
        role, mov_val = _primary_amount_role(mov.values)
        lines.append(
            CheckWithA3AccountLine(
                account_label=mov.account_label,
                amount_role=role,
                movement_value=mov_val,
                a3_value=check_by_role.get(role) if role else None,
                diff_value=diff_by_role.get(role) if role else None,
            )
        )

    return LeadCheckWithA3(
        check_source_row=row_offset + check_r + 1 if check_r is not None else None,
        diff_source_row=row_offset + diff_r + 1 if diff_r is not None else None,
        notes_source_row=row_offset + notes_row + 1 if notes_row is not None else None,
        notes_text=notes_text,
        lines=lines,
    )


def _extract_movement_table(
    rows: list[tuple[Any, ...]],
    block: LeadBlock | None,
) -> tuple[
    list[LeadMovementColumnBinding],
    list[LeadMovementRow],
    LeadCheckWithA3 | None,
]:
    scope = slice_rows_for_block(rows, block)
    row_offset = (block.start_row - 1) if block else 0
    header_r: int | None = None

    for r, row in enumerate(scope[:15]):
        cells = [_cell_str(v) for v in row[:20]]
        if not any(cells):
            continue
        has_book = any(
            t
            and _label_matches(
                _norm(t),
                ("期末账面数", "账面数", "期末审定数", "审定数", "期末未审数"),
            )
            for t in cells
        )
        has_name = any(t and _label_matches(_norm(t), ("科目名称", "总账科目编码")) for t in cells)
        if has_book and has_name:
            header_r = r
            break

    if header_r is None:
        return [], [], None

    header_cells = [_cell_str(v) for v in scope[header_r][:22]]
    bindings: list[LeadMovementColumnBinding] = []
    role_cols: dict[str, int] = {}
    for c, text in enumerate(header_cells):
        if not text:
            continue
        role = _match_movement_role(text)
        if role and role not in role_cols:
            role_cols[role] = c
            bindings.append(
                LeadMovementColumnBinding(role=role, source_header=text, column_index=c + 1)
            )

    audited_cols = [
        c for c, text in enumerate(header_cells) if text and _norm(text) == _norm("审定数")
    ]
    if len(audited_cols) >= 2 and "py_audited" not in role_cols:
        prior_col = audited_cols[-1]
        role_cols["py_audited"] = prior_col
        bindings.append(
            LeadMovementColumnBinding(
                role="py_audited",
                source_header=header_cells[prior_col] or "审定数",
                column_index=prior_col + 1,
            )
        )

    investigate_cols = [
        c
        for c, text in enumerate(header_cells)
        if text and "进一步调查" in text
    ]
    for role, col in zip(
        ("investigate_quantitative", "investigate_qualitative"),
        investigate_cols[:2],
    ):
        if role in role_cols:
            continue
        role_cols[role] = col
        bindings.append(
            LeadMovementColumnBinding(
                role=role,
                source_header=header_cells[col] or "进一步调查",
                column_index=col + 1,
            )
        )

    for c, text in enumerate(header_cells):
        if text and "\u79d1\u76ee\u540d\u79f0" in text:
            role_cols["account_label"] = c
            bindings = [b for b in bindings if b.role != "account_label"]
            bindings.append(
                LeadMovementColumnBinding(
                    role="account_label", source_header=text, column_index=c + 1
                )
            )
            break

    data_rows: list[LeadMovementRow] = []
    check_with_a3: LeadCheckWithA3 | None = None
    empty_streak = 0
    for r in range(header_r + 1, len(scope)):
        label_col = role_cols.get("account_label", 2)
        label = _get_cell(scope, r, label_col)
        if label and (
            _is_check_with_a3_label(label) or _is_diff_row_label(label)
        ):
            check_with_a3 = _extract_check_with_a3(
                scope,
                role_cols=role_cols,
                movement_rows=data_rows,
                row_offset=row_offset,
                start_r=r,
            )
            break
        if not label:
            empty_streak += 1
            if empty_streak >= 3 and data_rows:
                break
            continue
        n = _norm(label)
        if not any(lbl in n for lbl in map(_norm, _ACCOUNT_MOVEMENT_LABELS)):
            if data_rows:
                break
            continue
        empty_streak = 0
        sheet_ref_val = (
            _get_cell(scope, r, role_cols.get("sheet_ref", -1))
            if "sheet_ref" in role_cols
            else None
        )
        values: dict[str, str | None] = {}
        for role, c in role_cols.items():
            if role in ("account_label", "gl_code", "sheet_ref"):
                continue
            values[role] = _get_cell(scope, r, c)
        if sheet_ref_val is not None:
            values["sheet_ref"] = sheet_ref_val
        data_rows.append(
            LeadMovementRow(
                account_label=label,
                sheet_ref=sheet_ref_val,
                values=values,
                source_row=row_offset + r + 1,
            )
        )
    return bindings, data_rows, check_with_a3


def _extract_fluctuation_notes(
    rows: list[tuple[Any, ...]],
    block: LeadBlock | None,
) -> str | None:
    scope = slice_rows_for_block(rows, block)
    parts: list[str] = []
    passed_header = False
    for r, row in enumerate(scope[:25]):
        for c in range(min(8, len(row))):
            text = _cell_str(row[c])
            if not text:
                continue
            if text == "波动说明":
                passed_header = True
                continue
            if not passed_header:
                continue
            if text.startswith("[") and text.endswith("]") and len(text) <= 4:
                continue
            if len(text) >= 4:
                parts.append(text)
                break
    return "\n".join(parts) if parts else None


def _extract_adjustment_summary(
    rows: list[tuple[Any, ...]],
    block: LeadBlock | None,
) -> list[AdjustmentSummaryRow]:
    scope = slice_rows_for_block(rows, block)
    row_offset = (block.start_row - 1) if block else 0
    header_r: int | None = None
    for r, row in enumerate(scope[:10]):
        cells = [_cell_str(v) for v in row[:12]]
        if any(t and _label_matches(_norm(t), ("调整类型",)) for t in cells):
            header_r = r
            break
    if header_r is None:
        return []

    results: list[AdjustmentSummaryRow] = []
    for r in range(header_r + 1, min(header_r + 30, len(scope))):
        cells = [_cell_str(v) for v in scope[r][:12]]
        if not any(cells):
            if results:
                break
            continue
        adj_type = cells[1] or cells[0]
        if _is_no_adjustment_conclusion(cells) or _is_non_adjustment_note(cells):
            continue
        results.append(
            AdjustmentSummaryRow(
                adjustment_type=adj_type,
                source_row=row_offset + r + 1,
                raw_cells=cells,
            )
        )
    return results


def _is_no_adjustment_conclusion(cells: list[str | None]) -> bool:
    text = " ".join(c for c in cells if c)
    compact = _norm(text)
    if not compact:
        return False
    no_adjustment_markers = (
        "本年度不涉及审计调整",
        "本年不涉及审计调整",
        "本期不涉及审计调整",
        "不涉及审计调整",
        "无审计调整",
        "无调整事项",
        "不涉及调整事项",
    )
    return any(_norm(marker) in compact for marker in no_adjustment_markers)


def _is_non_adjustment_note(cells: list[str | None]) -> bool:
    text_cells = [c for c in cells if c]
    if len(text_cells) != 1:
        return False
    text = text_cells[0]
    compact = _norm(text)
    if compact.startswith("nb") and ("te" in compact or "sad" in compact):
        return True
    if "执行阶段" in text and "审定阶段" in text and ("TE" in text or "SAD" in text):
        return True
    return False


def _normalize_amount_key(value: str | None) -> str | None:
    if value is None:
        return None
    t = re.sub(r"[\s,，]", "", str(value).strip())
    if t.endswith(".0"):
        t = t[:-2]
    return t or None


def _annotate_volatility_and_layout(
    *,
    basic_info_fields: list[LeadBasicInfoField],
    cra_rows: list[CraAssertionRow],
    cra_block: LeadBlock | None,
    volatility: VolatilityThreshold | None,
) -> tuple[VolatilityThreshold | None, Literal["standard", "no_cra_te_volatility"] | None, list[str]]:
    extra_notes: list[str] = []
    layout: Literal["standard", "no_cra_te_volatility"] | None = None
    if volatility is None:
        return None, layout, extra_notes

    te_field = next((f for f in basic_info_fields if f.field_key == "te"), None)
    te_val = _normalize_amount_key(te_field.value if te_field else None)
    vol_amt = _normalize_amount_key(volatility.amount)

    if cra_block is None and not cra_rows and te_val and vol_amt and te_val == vol_amt:
        volatility.amount_source = "te"
        layout = "no_cra_te_volatility"
        extra_notes.append(
            "简版 Lead（无 CRA/TT 认定表）：波动幅度金额与 TE 一致，项目口径为取 TE 而非认定 TT。"
        )
    elif cra_rows and vol_amt:
        volatility.amount_source = "tt"

    return volatility, layout, extra_notes


def parse_lead_sheet_rows(
    rows: list[tuple[Any, ...]],
    *,
    source_file: str = "",
    source_sheet: str = "K.00 Lead Sheet",
) -> LeadSheetDataset:
    blocks = detect_lead_blocks(rows)
    notes: list[str] = []

    basic_block = block_for_kind(blocks, LeadBlockKind.BASIC_INFO)
    cra_block = block_for_kind(blocks, LeadBlockKind.CRA_THRESHOLD)
    exp_block = block_for_kind(blocks, LeadBlockKind.EXPECTATION)
    mov_block = block_for_kind(blocks, LeadBlockKind.MOVEMENT_TABLE)
    fluc_block = block_for_kind(blocks, LeadBlockKind.FLUCTUATION_NOTES)
    adj_block = block_for_kind(blocks, LeadBlockKind.ADJUSTMENT_SUMMARY)

    if not blocks:
        notes.append("未识别 Lead 分块锚点，已回退全表扫描。")

    basic_info_fields = _extract_basic_info(rows, basic_block)
    materiality = _basic_info_to_materiality(basic_info_fields, rows)
    cra_rows = _extract_cra_table(rows, cra_block)
    expectations, volatility = _extract_expectations(rows, exp_block)
    movement_bindings, movement_rows, check_with_a3 = _extract_movement_table(
        rows, mov_block
    )
    fluctuation_notes = _extract_fluctuation_notes(rows, fluc_block)
    if not fluctuation_notes and check_with_a3 and check_with_a3.notes_text:
        fluctuation_notes = check_with_a3.notes_text
        notes.append("Lead fluctuation notes recovered from the movement-table Notes row.")
    adjustment_rows = _extract_adjustment_summary(rows, adj_block)

    if not basic_info_fields:
        notes.append("未识别 Lead 基础信息块（客户名称/期末/TE 等）。")
    if not materiality:
        notes.append("未摘录 PM/TE/SAD，请人工打开 K.00 核对。")
    if not cra_rows:
        if cra_block is None:
            notes.append(
                "未识别 CRA/TT 认定表；若为基础信息下直接列 tickmark、"
                "且波动幅度 link TE，属简版 Lead（见 layout_variant）。"
            )
        else:
            notes.append("未识别 CRA/TT 认定表，请人工核对各认定 CRA 与 TT。")
    if not expectations:
        notes.append("未识别预期分析（ARP）账户变更表。")
    if not movement_rows:
        notes.append("未识别两期固定资产引导主表。")
    elif check_with_a3 is None:
        notes.append("引导主表未识别 Check with A3 / Diff 行，无法自动核对 A3 差异。")
    if not adjustment_rows and adj_block:
        notes.append("已定位调整汇总表锚点，但尚无数据行。")

    for b in blocks:
        if b.confidence < 0.7:
            notes.append(f"块 {b.kind.value} 置信度较低 ({b.confidence:.0%})，锚点={b.anchor_text!r}。")

    volatility, layout_variant, vol_notes = _annotate_volatility_and_layout(
        basic_info_fields=basic_info_fields,
        cra_rows=cra_rows,
        cra_block=cra_block,
        volatility=volatility,
    )
    notes.extend(vol_notes)

    movement_labels = {_norm(row.account_label) for row in movement_rows}
    expected_movement_labels = {_norm(label) for label in _ACCOUNT_MOVEMENT_LABELS}
    usable_for_rules = bool(
        len(movement_labels & expected_movement_labels) >= 3
        and not any(len(row.account_label) > 80 for row in movement_rows)
    )
    if mov_block is not None and not usable_for_rules:
        notes.append(
            "Lead movement table ingest is unreliable; dependent deterministic rules were paused."
        )

    return LeadSheetDataset(
        source_file=source_file,
        source_sheet=source_sheet,
        blocks=blocks,
        basic_info_fields=basic_info_fields,
        materiality=materiality,
        cra_rows=cra_rows,
        expectations=expectations,
        volatility=volatility,
        movement_bindings=movement_bindings,
        movement_rows=movement_rows,
        check_with_a3=check_with_a3,
        fluctuation_notes=fluctuation_notes,
        adjustment_rows=adjustment_rows,
        notes=notes,
        layout_variant=layout_variant,
        usable_for_rules=usable_for_rules,
    )


def find_lead_sheets(
    path: str | Path,
    *,
    max_rows: int | None = _DEFAULT_MAX_ROWS,
) -> list[tuple[str, float, list[tuple[Any, ...]]]]:
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    found: list[tuple[str, float, list[tuple[Any, ...]]]] = []
    try:
        for ws in wb.worksheets:
            rows = read_worksheet_rows(ws, max_rows=max_rows)
            kind, confidence, *_ = classify_sheet(ws.title, rows)
            name_kind, name_score, _ = score_by_name(ws.title)
            if kind == SheetKind.LEAD:
                found.append((ws.title, confidence, rows))
            elif name_kind == SheetKind.LEAD and name_score >= 0.75:
                found.append((ws.title, name_score, rows))
    finally:
        wb.close()
    ordered = sort_sheet_candidates(
        found,
        name=lambda c: c[0],
        confidence=lambda c: c[1],
        source_path=path,
    )
    return ordered


def load_lead_from_workbook(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    max_rows: int | None = _DEFAULT_MAX_ROWS,
) -> LeadSheetDataset:
    path = Path(path)

    if sheet_name:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            rows = read_worksheet_rows(ws, max_rows=max_rows)
        finally:
            wb.close()
        return parse_lead_sheet_rows(rows, source_file=str(path), source_sheet=sheet_name)

    candidates = find_lead_sheets(path, max_rows=max_rows)

    if candidates:
        chosen = choose_sheet_candidate(
            candidates,
            name=lambda c: c[0],
            confidence=lambda c: c[1],
            source_path=path,
        )
        assert chosen is not None
        name, _, rows = chosen
        return parse_lead_sheet_rows(rows, source_file=str(path), source_sheet=name)

    return LeadSheetDataset(
        source_file=str(path),
        source_sheet="",
        notes=["未识别 K.00 Lead Sheet 工作表。"],
    )
