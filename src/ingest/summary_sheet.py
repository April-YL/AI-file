from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

import openpyxl

from ingest.models import SheetKind
from ingest.sheet_classifier import classify_sheet, score_by_name
from ingest.sheet_period_routing import choose_sheet_candidate, sort_sheet_candidates
from ingest.workbook_reader import read_worksheet_rows

SummaryLayout = Literal["swp", "classic"]


@dataclass
class PspProgramRow:
    """汇总页单行程序（PSP）记录。"""

    procedure_name: str
    sheet_ref: str | None
    execution_status: str | None
    waiver_reason: str | None
    notes: str | None
    source_row: int
    is_psp: bool = False


@dataclass
class SummaryColumnBinding:
    """汇总页唯一主表上的列角色（1-based 列号与 `FieldMapping` 一致）。"""

    role: str
    source_header: str
    column_index: int


@dataclass
class SummarySheetDataset:
    """汇总页（通常仅一张主程序表）解析结果。

    ``column_bindings``：主表列与标准角色；``last_data_row``：最后一行程序数据（1-based）；
    ``notes``：表头定位与结束条件等。
    ``layout``：``swp`` 为标准 K1 SWP（B/C 程序、F 程序页、G/H/I）；``classic`` 为四列简版。
    """

    source_file: str
    source_sheet: str
    header_row: int | None
    programs: list[PspProgramRow]
    column_bindings: list[SummaryColumnBinding] = field(default_factory=list)
    last_data_row: int | None = None
    notes: list[str] = field(default_factory=list)
    layout: SummaryLayout | None = None


# ---- Classic（四列简版）：程序 | 工作表 | 是否执行 | 不执行原因 ---------------------------------
_PROCEDURE_HEADERS = (
    "程序",
    "程序名称",
    "审计程序",
    "底稿程序",
    "审计程序名称",
    "程序描述",
)
_SHEET_HEADERS = ("工作表", "sheet", "底稿索引", "索引", "底稿名称", "工作表名称")
_NOTES_HEADERS = ("注意事项", "备注", "说明")

# ---- SWP 标准版式：F 程序页、G 执行、H 不执行的原因、I 注意事项；B/C 为程序编号与说明 -----------
def _match_program_page_col(headers: list[str]) -> int | None:
    """「程序页」列：不得与仅含「程序」二字的列混淆。"""
    for idx, h in enumerate(headers):
        if not str(h).strip():
            continue
        raw = str(h).strip()
        nh = _norm_header(h)
        if nh in ("程序页", "返回页", "底稿页", "索引页"):
            return idx
        if "程序页" in raw or raw.startswith("返回页"):
            return idx
    return None


_EMPTY_ROWS_END_TABLE = 3
_PSP_MARKERS = ("psp", "ps p", "显著风险", "specific performance")


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _norm_header(text: str) -> str:
    return re.sub(r"\s+", "", text.strip().lower())


def _match_column(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    """宽松匹配（用于工作表名、程序页等）；勿用于「执行」列。"""
    for idx, h in enumerate(headers):
        if not str(h).strip():
            continue
        nh = _norm_header(h)
        for c in candidates:
            nc = _norm_header(c)
            if nh == nc or nc in nh or nh in nc:
                return idx
    return None


def _match_exec_header_col(headers: list[str]) -> int | None:
    """「执行」列表头：匹配 是否执行 / 执行，且绝不会匹配「不执行的原因」。"""
    exact_order = ("是否执行", "是否需执行", "执行与否", "执行")
    for want in exact_order:
        wn = _norm_header(want)
        for idx, h in enumerate(headers):
            if not str(h).strip():
                continue
            nh = _norm_header(h)
            if "不执行" in nh or nh.startswith("未执行原因"):
                continue
            if nh == wn:
                return idx
    for idx, h in enumerate(headers):
        if not str(h).strip():
            continue
        nh = _norm_header(h)
        if "不执行" in nh or nh.startswith("未执行原因"):
            continue
        if nh == "执行":
            return idx
        if "是否" in nh and "执行" in nh and "不" not in nh[:1]:
            return idx
    return None


def _match_waiver_header_col(headers: list[str]) -> int | None:
    """「不执行原因」类表头。"""
    for idx, h in enumerate(headers):
        if not str(h).strip():
            continue
        raw = str(h).strip()
        nh = _norm_header(h)
        if nh == "执行" or nh == "是否执行":
            continue
        if ("不执行" in raw or "未执行" in nh) and ("原因" in raw or "理由" in raw):
            return idx
        if nh in ("拒绝理由", "拒绝执行理由"):
            return idx
    for idx, h in enumerate(headers):
        if not str(h).strip():
            continue
        nh = _norm_header(h)
        if "原因" in nh and "不执行" in nh:
            return idx
    return None


def _swp_header_score(headers: list[str]) -> int:
    """K1 SWP：G 列起为「执行」、H 为不执行原因；或含「程序页」；表头 B 常为 K.xx。"""
    ix_exec = _match_exec_header_col(headers)
    ix_waiver = _match_waiver_header_col(headers)
    if ix_exec is None or ix_waiver is None:
        return -1
    if ix_exec == ix_waiver:
        return -1
    has_page = _match_program_page_col(headers) is not None
    b1 = str(headers[1]).strip() if len(headers) > 1 else ""
    b_k = bool(re.match(r"^[kK]\.\s*\d+", b1))
    # 四列简版「是否执行」在 C 列 (index 2)，不得按 SWP 解析
    if not has_page and not b_k and ix_exec < 5:
        return -1
    score = 6
    if has_page:
        score += 2
    if _match_column(headers, _NOTES_HEADERS) is not None:
        score += 1
    if b_k:
        score += 1
    return score


def _match_classic_procedure_col(headers: list[str]) -> int | None:
    """简版「程序」列：不包含「程序页/返回页」整词，避免 F 列误命中。"""
    for idx, h in enumerate(headers):
        if not str(h).strip():
            continue
        nh = _norm_header(h)
        if "程序页" in nh or nh in ("返回页", "程序页", "索引页"):
            continue
        for c in _PROCEDURE_HEADERS:
            nc = _norm_header(c)
            if nh == nc or nc in nh or nh in nc:
                return idx
    return None


def _classic_header_score(headers: list[str]) -> int:
    col_proc = _match_classic_procedure_col(headers)
    col_exec = _match_exec_header_col(headers)
    col_waiver = _match_waiver_header_col(headers)
    if col_proc is None or (col_exec is None and col_waiver is None):
        return -1
    score = 2
    if _match_column(headers, _SHEET_HEADERS) is not None:
        score += 1
    if col_exec is not None:
        score += 1
    if col_waiver is not None:
        score += 1
    if _match_column(headers, _NOTES_HEADERS) is not None:
        score += 1
    return score


def _detect_header_row(
    rows: list[tuple[Any, ...]],
    max_scan: int = 50,
) -> tuple[int | None, list[str], SummaryLayout | None]:
    """返回 (表头行 1-based, 表头单元格文本列表, layout)。"""
    best_r: int | None = None
    best_headers: list[str] = []
    best_score = -1
    best_layout: SummaryLayout | None = None
    limit = min(len(rows), max_scan)
    for r_idx in range(limit):
        row = rows[r_idx]
        if not row:
            continue
        headers = [_cell_str(c) or "" for c in row]
        if sum(1 for h in headers if h.strip()) < 2:
            continue

        swp_s = _swp_header_score(headers)
        cl_s = _classic_header_score(headers)
        # 同分时优先简版，避免仅四列的假阳性命中 SWP
        use_swp = swp_s > cl_s
        score = swp_s if use_swp else cl_s
        layout: SummaryLayout | None = "swp" if use_swp and swp_s > 0 else ("classic" if cl_s > 0 else None)
        if score < 0 or layout is None:
            continue

        if score > best_score or (
            score == best_score
            and (best_r is None or r_idx + 1 < best_r)
        ):
            best_score = score
            best_r = r_idx + 1
            best_headers = headers
            best_layout = layout

    return best_r, best_headers, best_layout


def _build_column_bindings(
    headers: list[str],
    layout: SummaryLayout,
) -> list[SummaryColumnBinding]:
    bindings: list[SummaryColumnBinding] = []
    seen: set[str] = set()

    def add(role: str, col_idx: int | None, label: str) -> None:
        if col_idx is None or role in seen:
            return
        seen.add(role)
        h = headers[col_idx] if col_idx < len(headers) else ""
        bindings.append(
            SummaryColumnBinding(
                role=role,
                source_header=str(h).strip(),
                column_index=col_idx + 1,
            )
        )

    if layout == "swp":
        add("procedure_code", 1 if len(headers) > 1 else None, "")
        add("procedure_text", 2 if len(headers) > 2 else None, "")
        ix_f = _match_program_page_col(headers)
        add("sheet_ref", ix_f, "")
        add("execution_status", _match_exec_header_col(headers), "")
        add("waiver_reason", _match_waiver_header_col(headers), "")
        add("notes", _match_column(headers, _NOTES_HEADERS), "")
    else:
        ix_p = _match_classic_procedure_col(headers)
        add("procedure", ix_p, "")
        add("sheet_ref", _match_column(headers, _SHEET_HEADERS), "")
        add("execution_status", _match_exec_header_col(headers), "")
        add("waiver_reason", _match_waiver_header_col(headers), "")
        add("notes", _match_column(headers, _NOTES_HEADERS), "")
    return bindings


def _is_psp_row(procedure_name: str, notes: str | None) -> bool:
    blob = f"{procedure_name} {notes or ''}".lower()
    return any(m in blob for m in _PSP_MARKERS) or "psp" in blob.replace(" ", "")


def _is_swp_subheader_repeat(cells: list[Any], col_g: int, col_h: int) -> bool:
    """行间重复的「执行 | 不执行的原因」小表头。"""
    if col_g >= len(cells) or col_h >= len(cells):
        return False
    g = _cell_str(cells[col_g]) or ""
    h = _cell_str(cells[col_h]) or ""
    if _norm_header(g) == "执行" and ("原因" in h or "理由" in h):
        return True
    return False


def _is_classic_header_text(proc: str) -> bool:
    return _norm_header(proc) in ("程序", "程序名称", "审计程序")


def parse_summary_rows(
    rows: list[tuple[Any, ...]],
    *,
    source_file: str = "",
    source_sheet: str = "汇总",
) -> SummarySheetDataset:
    notes_out: list[str] = []
    header_row, headers, layout = _detect_header_row(rows)
    if not header_row or not headers or not layout:
        notes_out.append(
            "未识别到汇总主表表头（K1 SWP：G「执行」+ H「不执行原因」；"
            "或简版：「程序」列 + 执行/不执行原因列）。"
        )
        return SummarySheetDataset(
            source_file=source_file,
            source_sheet=source_sheet,
            header_row=None,
            programs=[],
            notes=notes_out,
            layout=None,
        )

    column_bindings = _build_column_bindings(headers, layout)
    notes_out.append(f"main_table_header_row={header_row}")
    notes_out.append(f"summary_layout={layout}")

    col_exec = _match_exec_header_col(headers)
    col_waiver = _match_waiver_header_col(headers)
    col_notes = _match_column(headers, _NOTES_HEADERS)

    programs: list[PspProgramRow] = []
    start_idx = header_row
    last_data_row: int | None = None
    empty_run = 0

    if layout == "swp":
        col_b, col_c = 1, 2
        col_sheet = _match_program_page_col(headers)
        for r_idx in range(start_idx, len(rows)):
            row = rows[r_idx]
            if row is None or not any(_cell_str(c) for c in row):
                empty_run += 1
                if programs and empty_run >= _EMPTY_ROWS_END_TABLE:
                    notes_out.append(f"main_table_end_after_blank_rows_{_EMPTY_ROWS_END_TABLE}")
                    break
                continue
            empty_run = 0
            cells = list(row)

            def _get(col: int | None) -> str | None:
                if col is None or col >= len(cells):
                    return None
                return _cell_str(cells[col])

            if col_exec is not None and col_waiver is not None:
                if _is_swp_subheader_repeat(cells, col_exec, col_waiver):
                    continue

            b = _get(col_b)
            c = _get(col_c)
            sheet_ref = _get(col_sheet)
            proc_name = f"{b or ''} {c or ''}".strip() or (b or c or "")
            if not proc_name and sheet_ref:
                # SWP 常见续行：B/C 为空，仅 F 列给出具体程序页（如 K.02.1a / K.02.2a）。
                proc_name = sheet_ref
            if not proc_name:
                continue
            low = proc_name.lower()
            if low in ("程序页", "返回页") and not sheet_ref:
                continue

            exec_status = _get(col_exec)
            waiver = _get(col_waiver)
            note_val = _get(col_notes)
            programs.append(
                PspProgramRow(
                    procedure_name=proc_name,
                    sheet_ref=sheet_ref,
                    execution_status=exec_status,
                    waiver_reason=waiver,
                    notes=note_val,
                    source_row=r_idx + 1,
                    is_psp=_is_psp_row(proc_name, note_val),
                )
            )
            last_data_row = r_idx + 1
    else:
        col_proc = _match_classic_procedure_col(headers)
        col_sheet = _match_column(headers, _SHEET_HEADERS)
        for r_idx in range(start_idx, len(rows)):
            row = rows[r_idx]
            if row is None or not any(_cell_str(c) for c in row):
                empty_run += 1
                if programs and empty_run >= _EMPTY_ROWS_END_TABLE:
                    notes_out.append(f"main_table_end_after_blank_rows_{_EMPTY_ROWS_END_TABLE}")
                    break
                continue
            empty_run = 0
            cells = list(row)

            def _get(col: int | None) -> str | None:
                if col is None or col >= len(cells):
                    return None
                return _cell_str(cells[col])

            proc = _get(col_proc)
            if not proc:
                continue
            if _is_classic_header_text(proc):
                continue

            sheet_ref = _get(col_sheet)
            exec_status = _get(col_exec)
            waiver = _get(col_waiver)
            note_val = _get(col_notes)
            programs.append(
                PspProgramRow(
                    procedure_name=proc,
                    sheet_ref=sheet_ref,
                    execution_status=exec_status,
                    waiver_reason=waiver,
                    notes=note_val,
                    source_row=r_idx + 1,
                    is_psp=_is_psp_row(proc, note_val),
                )
            )
            last_data_row = r_idx + 1

    if not programs:
        notes_out.append("主表已识别表头，但未解析到程序数据行。")

    return SummarySheetDataset(
        source_file=source_file,
        source_sheet=source_sheet,
        header_row=header_row,
        programs=programs,
        column_bindings=column_bindings,
        last_data_row=last_data_row,
        notes=notes_out,
        layout=layout,
    )


def find_summary_sheets(
    path: str | Path,
    *,
    max_rows: int | None = 200,
) -> list[tuple[str, float, list[tuple[Any, ...]]]]:
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    found: list[tuple[str, float, list[tuple[Any, ...]]]] = []
    try:
        for ws in wb.worksheets:
            rows = read_worksheet_rows(ws, max_rows=max_rows)
            name_kind, name_score, _ = score_by_name(ws.title)
            if name_kind == SheetKind.SUMMARY and name_score >= 0.75:
                found.append((ws.title, min(0.98, name_score + 0.1), rows))
                continue
            kind, confidence, *_ = classify_sheet(ws.title, rows)
            if kind == SheetKind.SUMMARY:
                found.append((ws.title, confidence, rows))
    finally:
        wb.close()
    return sort_sheet_candidates(
        found,
        name=lambda c: c[0],
        confidence=lambda c: c[1],
        source_path=path,
    )


def load_summary_from_workbook(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    max_rows: int | None = 200,
) -> SummarySheetDataset:
    path = Path(path)
    candidates = find_summary_sheets(path, max_rows=max_rows)

    if sheet_name:
        wanted = _norm_sheet_name(sheet_name)
        match = next((c for c in candidates if _norm_sheet_name(c[0]) == wanted), None)
        if match is None:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                real_name = _resolve_sheet_name(wb.sheetnames, sheet_name)
                ws = wb[real_name]
                rows = read_worksheet_rows(ws, max_rows=max_rows)
            finally:
                wb.close()
            return parse_summary_rows(
                rows, source_file=str(path), source_sheet=real_name
            )
        name, _, rows = match
        return parse_summary_rows(rows, source_file=str(path), source_sheet=name)

    if candidates:
        chosen = choose_sheet_candidate(
            candidates,
            name=lambda c: c[0],
            confidence=lambda c: c[1],
            source_path=path,
        )
        assert chosen is not None
        name, _, rows = chosen
        return parse_summary_rows(rows, source_file=str(path), source_sheet=name)

    return SummarySheetDataset(
        source_file=str(path),
        source_sheet="",
        header_row=None,
        programs=[],
        notes=["未识别到名称或内容像「汇总」的工作表。"],
        layout=None,
    )


def _norm_sheet_name(name: str) -> str:
    return re.sub(r"\s+", "", str(name).strip().lower())


def _resolve_sheet_name(sheet_names: list[str], requested: str) -> str:
    if requested in sheet_names:
        return requested
    wanted = _norm_sheet_name(requested)
    for name in sheet_names:
        if _norm_sheet_name(name) == wanted:
            return name
    raise KeyError(f"Worksheet {requested} does not exist.")
