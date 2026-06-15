from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING, Literal

import openpyxl

from rules.models import QcIssue, Severity

if TYPE_CHECKING:
    from ingest.workbook_context import WorkbookQcContext

DeliveryStage = Literal["first", "final"]

FIRST_DELIVERY_RULE_ID = "first_delivery_standard"
FINAL_DELIVERY_RULE_ID = "final_delivery_standard"

_PROCEDURE_RULE_IDS = {
    "psp_completion",
    "addition_test_package_complete",
    "disposal_test_package_complete",
}
_ADJUSTMENT_RULE_HINTS = ("adjustment", "调整")


@dataclass(frozen=True)
class DeliveryCompletionContext:
    """Delivery stage selected by the reviewer."""

    stage: DeliveryStage


@dataclass(frozen=True)
class SampleEvidenceSummary:
    total_samples: int | None
    unfinished_samples: int | None
    completed_samples: int
    sources: tuple[str, ...] = ()


@dataclass(frozen=True)
class WorkpaperCleanupEvidence:
    readable: bool
    residue_count: int = 0
    examples: tuple[str, ...] = ()


def check_delivery_completion(
    ctx: DeliveryCompletionContext | None,
    *,
    prior_issues: list[QcIssue] | None = None,
    workbook_context: WorkbookQcContext | None = None,
    workbook_path: str | Path | None = None,
    workbook_sheet_titles: list[str] | None = None,
) -> list[QcIssue]:
    if ctx is None:
        return []
    prior = prior_issues or []
    if ctx.stage == "first":
        return _check_first_delivery(
            prior,
            workbook_context=workbook_context,
            workbook_path=workbook_path,
            workbook_sheet_titles=workbook_sheet_titles,
        )
    if ctx.stage == "final":
        return _check_final_delivery(
            prior,
            workbook_context=workbook_context,
            workbook_path=workbook_path,
        )
    return []


def _check_first_delivery(
    prior_issues: list[QcIssue],
    *,
    workbook_context: WorkbookQcContext | None,
    workbook_path: str | Path | None,
    workbook_sheet_titles: list[str] | None,
) -> list[QcIssue]:
    issues: list[QcIssue] = []
    if _has_open_procedure_issues(prior_issues):
        issues.append(
            _issue(
                FIRST_DELIVERY_RULE_ID,
                "psp_completion",
                Severity.FAIL,
                "首次交付未达到交付完成度标准：仍存在 PSP 执行完整性问题。",
                "先处理汇总页 PSP 执行状态、程序页勾稽或拒绝执行理由，再作为首次交付提交。",
            )
        )

    evidence = _sample_evidence_summary(workbook_context)
    has_tod_evidence = _has_depreciation_tod_evidence(workbook_path, workbook_sheet_titles)
    if evidence.total_samples is None and not has_tod_evidence:
        issues.append(
            _issue(
                FIRST_DELIVERY_RULE_ID,
                "tod_cutoff_evidence",
                Severity.NEED_REVIEW,
                "首次交付完成度需要复核：底稿中未识别到可结构化的 TOD/cutoff 支持性证据样本。",
                "请确认详细测试或截止性测试是否已开始获取支持性证据；如已获取，请补充到测试底稿的证据列。",
            )
        )
    elif (
        evidence.total_samples
        and evidence.total_samples > 0
        and evidence.completed_samples == 0
        and not has_tod_evidence
    ):
        issues.append(
            _issue(
                FIRST_DELIVERY_RULE_ID,
                "tod_cutoff_evidence",
                Severity.FAIL,
                "首次交付未达到交付完成度标准：已识别样本，但未见任何样本取得支持性证据。",
                "至少获取并记录部分详细测试或截止性测试支持性证据后，再作为首次交付提交。",
            )
        )

    if _has_open_misstatement_or_adjustment_issues(prior_issues):
        issues.append(
            _issue(
                FIRST_DELIVERY_RULE_ID,
                "risk_response_status",
                Severity.FAIL,
                "首次交付未达到交付完成度标准：仍存在重大错报风险或调整事项相关未闭环 finding。",
                "先补充风险应对、调整事项处理记录或相应说明，再作为首次交付提交。",
            )
        )

    return issues


def _check_final_delivery(
    prior_issues: list[QcIssue],
    *,
    workbook_context: WorkbookQcContext | None,
    workbook_path: str | Path | None,
) -> list[QcIssue]:
    issues: list[QcIssue] = []
    if _has_open_procedure_issues(prior_issues):
        issues.append(
            _issue(
                FINAL_DELIVERY_RULE_ID,
                "procedures_complete",
                Severity.FAIL,
                "整体交付未达到交付完成度标准：仍存在程序执行或程序包完整性问题。",
                "先处理汇总页 PSP、K.02 新增/处置程序包等未闭环事项，再作为整体交付提交。",
            )
        )

    cleanup = _scan_workpaper_cleanup(workbook_path)
    if not cleanup.readable:
        issues.append(
            _issue(
                FINAL_DELIVERY_RULE_ID,
                "comments_cleared",
                Severity.NEED_REVIEW,
                "整体交付完成度需要复核：无法读取工作簿以检查 Comments 是否清理。",
                "请确认底稿中 Review Notes、Open Items、Comments 和未解决标记已清理。",
            )
        )
    elif cleanup.residue_count:
        examples = "；".join(cleanup.examples[:3])
        issues.append(
            _issue(
                FINAL_DELIVERY_RULE_ID,
                "comments_cleared",
                Severity.FAIL,
                f"整体交付未达到交付完成度标准：底稿中仍识别到 {cleanup.residue_count} 处 Comments/未解决标记。",
                f"清理全部 review comments / open items 后再作为整体交付提交。示例：{examples}",
            )
        )

    sample_issue = _sample_completion_issue(_sample_evidence_summary(workbook_context))
    if sample_issue:
        issues.append(sample_issue)

    if _has_open_adjustment_issues(prior_issues):
        issues.append(
            _issue(
                FINAL_DELIVERY_RULE_ID,
                "adjustments_status",
                Severity.FAIL,
                "整体交付未达到交付完成度标准：仍存在审计调整相关未闭环 finding。",
                "确认审计调整均已确定；如仍未确定，应在底稿中保留已提示项目组 EIC 的记录。",
            )
        )

    return issues


def _has_open_procedure_issues(prior_issues: list[QcIssue]) -> bool:
    return any(
        issue.rule_id in _PROCEDURE_RULE_IDS and issue.severity != Severity.PASS
        for issue in prior_issues
    )


def _has_open_adjustment_issues(prior_issues: list[QcIssue]) -> bool:
    return any(
        issue.severity != Severity.PASS
        and any(h in f"{issue.rule_id} {issue.message}" for h in _ADJUSTMENT_RULE_HINTS)
        for issue in prior_issues
    )


def _has_open_misstatement_or_adjustment_issues(prior_issues: list[QcIssue]) -> bool:
    return any(
        issue.severity in {Severity.FAIL, Severity.NEED_REVIEW}
        and (
            (issue.problem_category or "") == "错报"
            or "错报风险" in (issue.qc_checkpoint or "")
            or any(h in f"{issue.rule_id} {issue.message}" for h in _ADJUSTMENT_RULE_HINTS)
        )
        for issue in prior_issues
    )


def _sample_evidence_summary(
    workbook_context: WorkbookQcContext | None,
) -> SampleEvidenceSummary:
    if workbook_context is None:
        return SampleEvidenceSummary(None, None, 0)

    total: int | None = None
    completed = 0
    sources: list[str] = []

    addition_test = workbook_context.addition_test
    addition_sample_output = workbook_context.addition_sample_output

    selected = list(addition_sample_output.selected_samples) if addition_sample_output else []
    tested = list(addition_test.tested_samples) if addition_test else []

    if selected:
        total = len(selected)
        tested_by_id = {
            _sample_key(row.asset_id, row.asset_name, row.original_value): row
            for row in tested
        }
        for sample in selected:
            tested_row = tested_by_id.get(
                _sample_key(sample.asset_id, sample.asset_name, sample.original_value)
            )
            if tested_row and _tested_sample_has_evidence(tested_row):
                completed += 1
        sources.append(f"新增选样输出:{len(selected)}")
    elif tested:
        total = len(tested)
        completed = sum(1 for row in tested if _tested_sample_has_evidence(row))
        sources.append(f"新增测试:{len(tested)}")

    if total is None:
        return SampleEvidenceSummary(None, None, completed, tuple(sources))
    unfinished = max(total - completed, 0)
    return SampleEvidenceSummary(total, unfinished, completed, tuple(sources))


def _sample_key(asset_id: object, asset_name: object, amount: object) -> tuple[str, str, str]:
    return (_norm(asset_id), _norm(asset_name), _norm(amount))


def _tested_sample_has_evidence(row: object) -> bool:
    values = [
        getattr(row, "evidence_amount", None),
        getattr(row, "evidence_description", None),
        getattr(row, "amount_difference", None),
        *(getattr(row, "attribute_results", None) or []),
    ]
    return any(str(v).strip() for v in values if v is not None)


def _sample_completion_issue(summary: SampleEvidenceSummary) -> QcIssue | None:
    if summary.total_samples is None or summary.unfinished_samples is None:
        return _issue(
            FINAL_DELIVERY_RULE_ID,
            "supporting_evidence_samples",
            Severity.NEED_REVIEW,
            "整体交付完成度需要复核：底稿中未识别到可结构化的支持性证据样本完成情况。",
            "请确认 K.02/K.03 测试底稿是否已列示样本和支持性证据检查结果。",
        )
    if summary.total_samples == 0:
        return None
    ratio = summary.unfinished_samples / summary.total_samples
    if ratio > 0.5:
        source = f"（来源：{'；'.join(summary.sources)}）" if summary.sources else ""
        return _issue(
            FINAL_DELIVERY_RULE_ID,
            "supporting_evidence_samples",
            Severity.FAIL,
            (
                "整体交付未达到交付完成度标准：未获取并检查支持性证据的样本数"
                f"为 {summary.unfinished_samples}/{summary.total_samples}，超过 50%。{source}"
            ),
            "继续获取并检查支持性证据，直至未完成样本不超过总样本数量的 50%。",
        )
    return None


def _scan_workpaper_cleanup(workbook_path: str | Path | None) -> WorkpaperCleanupEvidence:
    if workbook_path is None:
        return WorkpaperCleanupEvidence(False)
    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
    except Exception:
        return WorkpaperCleanupEvidence(False)

    examples: list[str] = []
    residue_count = 0
    try:
        for ws in wb.worksheets:
            if _is_comment_sheet(ws.title):
                count, sheet_examples = _scan_comment_sheet(ws)
                residue_count += count
                examples.extend(sheet_examples)
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if cell.comment is not None:
                        residue_count += 1
                        examples.append(f"{ws.title}!{cell.coordinate}: 单元格批注")
                    text = str(cell.value).strip() if cell.value is not None else ""
                    if text and _looks_like_unresolved_note(text):
                        residue_count += 1
                        examples.append(f"{ws.title}!{cell.coordinate}: {text[:40]}")
                    if len(examples) >= 5:
                        return WorkpaperCleanupEvidence(True, residue_count, tuple(examples))
    finally:
        wb.close()
    return WorkpaperCleanupEvidence(True, residue_count, tuple(examples))


def _is_comment_sheet(title: str) -> bool:
    t = _norm(title)
    return "comment" in t or "reviewnote" in t or "openitem" in t or "批注" in title


def _scan_comment_sheet(ws) -> tuple[int, list[str]]:
    count = 0
    examples: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = [str(v).strip() if v is not None else "" for v in row]
        if not any(cells):
            continue
        closed_text = " ".join(cells[-3:]).lower()
        if any(x in closed_text for x in ("closed", "yes", "已关闭", "已清理", "done")):
            continue
        count += 1
        if len(examples) < 5:
            examples.append(f"{ws.title}: {' | '.join(c for c in cells[:4] if c)[:80]}")
    return count, examples


def _looks_like_unresolved_note(text: str) -> bool:
    t = _norm(text)
    markers = (
        "reviewnote",
        "openitem",
        "unresolved",
        "notcleared",
        "待清理",
        "未清理",
        "未解决",
        "待解决",
        "os事项",
    )
    return any(m in t for m in markers)


def _has_depreciation_tod_evidence(
    workbook_path: str | Path | None,
    workbook_sheet_titles: list[str] | None,
) -> bool:
    if workbook_path is None or not workbook_sheet_titles:
        return False
    candidates = [
        title
        for title in workbook_sheet_titles
        if ("k032" in _norm(title) or "tod" in _norm(title) or "byitem" in _norm(title))
        and ("折旧" in title or "depreciation" in _norm(title))
    ]
    if not candidates:
        return False
    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        for title in candidates:
            if title in wb.sheetnames and _sheet_has_substantive_content(wb[title]):
                return True
    finally:
        wb.close()
    return False


def _sheet_has_substantive_content(ws) -> bool:
    non_empty = 0
    numeric = 0
    terms = set()
    for row in ws.iter_rows(min_row=1, max_row=80, max_col=30, values_only=True):
        for value in row:
            if value is None or str(value).strip() == "":
                continue
            text = str(value).strip()
            non_empty += 1
            if isinstance(value, (int, float)) or text.replace(".", "", 1).isdigit():
                numeric += 1
            normalized = _norm(text)
            for term in ("折旧", "原值", "差异", "使用寿命", "depreciation", "useful"):
                if _norm(term) in normalized:
                    terms.add(term)
            if non_empty >= 8 and numeric >= 2 and len(terms) >= 2:
                return True
    return False


def _norm(value: object) -> str:
    if value is None:
        return ""
    return "".join(ch.lower() for ch in str(value).strip() if ch.isalnum())


def _issue(
    rule_id: str,
    field: str,
    severity: Severity,
    message: str,
    suggestion: str,
) -> QcIssue:
    return QcIssue(
        asset_id=None,
        rule_id=rule_id,
        field=field,
        severity=severity,
        message=message,
        suggestion=suggestion,
        procedure_code="GLOBAL",
        source_sheet="workbook",
    )
