from __future__ import annotations

import re
from collections.abc import Sequence
from dataclasses import dataclass
from typing import Callable, Literal

import openpyxl

from ingest.summary_sheet import PspProgramRow, SummarySheetDataset
from rules.models import QcIssue, Severity
from rules.psp_sheet_matcher import count_non_empty_cells, find_matching_sheet

RULE_ID = "psp_completion"

_MIN_WAIVER_LEN = 8
_MIN_SHEET_MATCH_SCORE = 0.48
_STRONG_SHEET_MATCH_SCORE = 0.72
_MIN_SUBSTANCE_CELLS = 8
_MAX_SIGNAL_SCAN_ROWS = 80
_MAX_SIGNAL_SCAN_COLS = 20
_MIN_TOD_CONTENT_TERMS = 2
_MIN_TOD_NUMERIC_CELLS = 2

_K_SEGMENT = re.compile(r"[kK]\.\d+(?:\.\d+)*[a-zA-Z]?")

_SKIP_PROCEDURE_NAMES = frozenset(
    {
        "程序",
        "程序名称",
        "审计程序",
        "返回汇总页",
        "返回汇总",
    }
)

_EXEC_YES = re.compile(
    r"^(是|已执行|执行|完成|y|yes|true|√|✓|1|done)$",
    re.IGNORECASE,
)
_EXEC_NO = re.compile(
    r"^(否|不执行|未执行|拒绝|n|no|false|0|不适用|n/a|na)$",
    re.IGNORECASE,
)
_EXEC_PARTIAL = re.compile(r"部分|partial", re.IGNORECASE)


@dataclass(frozen=True)
class WaiverSemanticReview:
    adequacy: Literal["sufficient", "insufficient", "unclear"]
    rationale: str = ""
    suggested_action: str = ""


WaiverReasonReviewer = Callable[[PspProgramRow], WaiverSemanticReview | None]


def _rule_review_waiver_reason(row: PspProgramRow) -> WaiverSemanticReview | None:
    waiver = _norm_token(row.waiver_reason)
    if not waiver:
        return None
    proc = _norm_token(f"{row.procedure_name or ''} {row.sheet_ref or ''}")
    empty_conclusions = ("无需执行", "不执行", "不适用", "na", "n/a", "金额小", "不重大")
    if waiver in empty_conclusions or any(waiver == _norm_token(x) for x in empty_conclusions):
        return WaiverSemanticReview(
            adequacy="insufficient",
            rationale="理由仅为结论，未说明业务原因、金额依据、性质风险或替代程序。",
            suggested_action="补充可复核的不执行依据，例如金额与 TE/TT/SAD 的关系、性质风险判断或替代程序。",
        )
    if "减值" in proc and "无减值迹象" in waiver:
        has_process = any(x in waiver for x in ("识别", "评估", "判断", "支持", "canvas", "swp", "检查"))
        if not has_process:
            return WaiverSemanticReview(
                adequacy="insufficient",
                rationale="减值测试理由仅说明“无减值迹象”，未说明减值迹象识别程序、判断原因或支持文件。",
                suggested_action="补充减值迹象识别过程、支持文件或 Canvas/SWP 减值评估索引。",
            )
    if ("处置" in proc or "新增" in proc) and ("小于te" in waiver or "低于te" in waiver):
        has_tt_nature = ("tt" in waiver and any(x in waiver for x in ("性质", "异常", "单项"))) or "sad" in waiver
        if not has_tt_nature:
            return WaiverSemanticReview(
                adequacy="insufficient",
                rationale="理由仅说明金额小于 TE，未同时说明单项 TT/性质风险判断，也未说明金额小于 SAD。",
                suggested_action="补充总体金额、单项 TT、性质异常项和/或 SAD 口径，说明为何可不执行该测试。",
            )
    return None


def _should_skip_row(row: PspProgramRow) -> bool:
    name = (row.procedure_name or "").strip()
    sheet_ref = (row.sheet_ref or "").strip()
    if not name or name in _SKIP_PROCEDURE_NAMES:
        return True
    if name.startswith("返回"):
        return True
    if sheet_ref == "\u7a0b\u5e8f\u9875":
        return True
    if not (row.execution_status or "").strip() and re.search(
        r"[kK]\.02\.[12]a\b", sheet_ref or name
    ):
        return True
    return False


def normalize_execution_status(raw: str | None) -> str:
    if raw is None:
        return "empty"
    text = str(raw).strip()
    if not text:
        return "empty"
    if _EXEC_PARTIAL.search(text):
        return "partial"
    if _EXEC_YES.search(text):
        return "yes"
    if _EXEC_NO.search(text):
        return "no"
    if "不执行" in text or "未执行" in text:
        return "no"
    if "部分" in text:
        return "partial"
    if "执行" in text and "不" not in text:
        return "yes"
    return "ambiguous"


def _fallback_sheet_ref_from_procedure(name: str | None) -> str | None:
    if not name:
        return None
    m = _K_SEGMENT.search(name)
    return m.group(0) if m else None


def _ref_for_sheet_match(row: PspProgramRow) -> str | None:
    ref = (row.sheet_ref or "").strip()
    if ref:
        return ref
    return _fallback_sheet_ref_from_procedure(row.procedure_name)


def _check_yes_program_sheet(
    row: PspProgramRow,
    source_sheet: str,
    *,
    workbook_sheet_titles: Sequence[str] | None,
    workbook_path: str | None,
) -> list[QcIssue]:
    """已执行程序与工作表名称及表内非空内容的形式勾稽（可选，需传入 sheet 列表）。"""
    issues: list[QcIssue] = []
    if workbook_sheet_titles is None:
        return issues

    label = row.procedure_name
    if row.sheet_ref:
        label = f"{label} ({row.sheet_ref})"

    ref = _ref_for_sheet_match(row)
    if ref and "\u9879\u76ee\u7ec4\u81ea\u884c\u586b\u5199\u5e95\u7a3f\u7d22\u5f15" in ref:
        return issues
    if not ref:
        issues.append(
            QcIssue(
                asset_id=None,
                rule_id=RULE_ID,
                field="sheet_ref",
                severity=Severity.NEED_REVIEW,
                message=f"程序「{label}」标为已执行，但程序页/工作表引用为空，无法与底稿工作表勾稽",
                suggestion="补充程序页列或与底稿一致的工作表引用（含 K.xx 程序页编号）",
                procedure_code="SUMMARY",
                source_sheet=source_sheet,
                source_row=row.source_row,
            )
        )
        return issues

    titles_list = list(workbook_sheet_titles)
    if not titles_list:
        return issues

    matched, score, _reason = find_matching_sheet(ref, titles_list)
    if matched is None or score < _MIN_SHEET_MATCH_SCORE:
        issues.append(
            QcIssue(
                asset_id=None,
                rule_id=RULE_ID,
                field="sheet_ref",
                severity=Severity.FAIL,
                message=(
                    f"程序「{label}」标为已执行，但工作簿中未找到与程序页引用「{ref}」"
                    "相匹配的工作表（已做名称规范化与模糊匹配）"
                ),
                suggestion="核对程序页名称与底稿工作表名称是否一致，或修正汇总页超链接目标",
                procedure_code="SUMMARY",
                source_sheet=source_sheet,
                source_row=row.source_row,
            )
        )
        return issues

    if score < _STRONG_SHEET_MATCH_SCORE:
        issues.append(
            QcIssue(
                asset_id=None,
                rule_id=RULE_ID,
                field="sheet_ref",
                severity=Severity.NEED_REVIEW,
                message=(
                    f"程序「{label}」对应的底稿页可能为「{matched}」"
                    f"（名称匹配置信度约 {score:.0%}），请人工确认是否为正确工作表"
                ),
                suggestion="确认汇总页程序页与打开底稿后的实际表名一致",
                procedure_code="SUMMARY",
                source_sheet=source_sheet,
                source_row=row.source_row,
            )
        )

    if workbook_path and matched:
        ncells = count_non_empty_cells(workbook_path, matched, max_rows=40)
        if 0 <= ncells < _MIN_SUBSTANCE_CELLS:
            issues.append(
                QcIssue(
                    asset_id=None,
                    rule_id=RULE_ID,
                    field="sheet_substance",
                    severity=Severity.WARN,
                    message=(
                        f"程序「{label}」对应工作表「{matched}」前若干行有效内容较少"
                        f"（非空单元格约 {ncells} 个），请确认是否已实质执行并完成底稿"
                    ),
                    suggestion="检查该工作表是否为空模板或未保存结果",
                    procedure_code="SUMMARY",
                    source_sheet=source_sheet,
                    source_row=row.source_row,
                )
            )

    return issues


def _check_program_row(
    row: PspProgramRow,
    source_sheet: str,
    *,
    execution_status_override: str | None = None,
    skip_waiver_reason_check: bool = False,
    workbook_sheet_titles: Sequence[str] | None = None,
    workbook_path: str | None = None,
    waiver_reason_reviewer: WaiverReasonReviewer | None = None,
) -> list[QcIssue]:
    issues: list[QcIssue] = []
    status = normalize_execution_status(
        execution_status_override
        if execution_status_override is not None
        else row.execution_status
    )
    label = row.procedure_name
    if row.sheet_ref:
        label = f"{label} ({row.sheet_ref})"

    if status == "empty":
        issues.append(
            QcIssue(
                asset_id=None,
                rule_id=RULE_ID,
                field="execution_status",
                severity=Severity.NEED_REVIEW,
                message=f"程序「{label}」执行状态为空，无法自动判断",
                suggestion="在汇总页补充是否执行，或由 reviewer 人工确认",
                procedure_code="SUMMARY",
                source_sheet=source_sheet,
                source_row=row.source_row,
            )
        )
        return issues

    if status == "partial":
        issues.append(
            QcIssue(
                asset_id=None,
                rule_id=RULE_ID,
                field="execution_status",
                severity=Severity.NEED_REVIEW,
                message=f"程序「{label}」为部分执行，需人工判断证据是否充分",
                suggestion="在底稿中说明部分执行范围及替代程序",
                procedure_code="SUMMARY",
                source_sheet=source_sheet,
                source_row=row.source_row,
            )
        )
        return issues

    if status == "no":
        if skip_waiver_reason_check:
            return issues
        waiver = (row.waiver_reason or "").strip()
        if not waiver:
            issues.append(
                QcIssue(
                    asset_id=None,
                    rule_id=RULE_ID,
                    field="waiver_reason",
                    severity=Severity.FAIL,
                    message=f"程序「{label}」标记为不执行，但未填写不执行/拒绝理由",
                    suggestion="在汇总页补充恰当的不执行原因",
                    procedure_code="SUMMARY",
                    source_sheet=source_sheet,
                    source_row=row.source_row,
                )
            )
        elif waiver_reason_reviewer is not None:
            reviewed = waiver_reason_reviewer(row)
            if reviewed is None:
                issues.append(
                    QcIssue(
                        asset_id=None,
                        rule_id=RULE_ID,
                        field="waiver_reason",
                        severity=Severity.NEED_REVIEW,
                        message=(
                            f"程序「{label}」已填写不执行理由，但语义复核未返回有效结果，"
                            "请人工判断理由是否充分合理"
                        ),
                        suggestion="检查 LLM 配置与连通性，或由 reviewer 人工复核该理由",
                        procedure_code="SUMMARY",
                        source_sheet=source_sheet,
                        source_row=row.source_row,
                    )
                )
            elif reviewed.adequacy != "sufficient":
                severity = (
                    Severity.WARN
                    if reviewed.adequacy == "insufficient"
                    else Severity.NEED_REVIEW
                )
                suggestion = reviewed.suggested_action or (
                    "补充与审计准则、项目风险和替代程序相匹配的不执行说明"
                )
                msg_suffix = f"；模型提示：{reviewed.rationale}" if reviewed.rationale else ""
                issues.append(
                    QcIssue(
                        asset_id=None,
                        rule_id=RULE_ID,
                        field="waiver_reason",
                        severity=severity,
                        message=(
                            f"程序「{label}」不执行理由语义上"
                            + ("不足" if reviewed.adequacy == "insufficient" else "不明确")
                            + msg_suffix
                        ),
                        suggestion=suggestion,
                        procedure_code="SUMMARY",
                        source_sheet=source_sheet,
                        source_row=row.source_row,
                        review_source="规则+LLM",
                        llm_review_type="PSP不执行理由充分性",
                    )
                )
        else:
            rule_reviewed = _rule_review_waiver_reason(row)
            if rule_reviewed is not None:
                issues.append(
                    QcIssue(
                        asset_id=None,
                        rule_id=RULE_ID,
                        field="waiver_reason",
                        severity=Severity.WARN,
                        message=(
                            f"程序「{label}」不执行理由语义上不足；"
                            f"规则提示：{rule_reviewed.rationale}"
                        ),
                        suggestion=rule_reviewed.suggested_action
                        or "补充与审计准则、项目风险和替代程序相匹配的不执行说明",
                        procedure_code="SUMMARY",
                        source_sheet=source_sheet,
                        source_row=row.source_row,
                    )
                )
            elif len(waiver) < _MIN_WAIVER_LEN:
                issues.append(
                    QcIssue(
                        asset_id=None,
                        rule_id=RULE_ID,
                        field="waiver_reason",
                        severity=Severity.WARN,
                        message=f"程序「{label}」不执行理由过短，建议补充充分说明",
                        suggestion="补充与审计准则、项目风险相匹配的拒绝理由",
                        procedure_code="SUMMARY",
                        source_sheet=source_sheet,
                        source_row=row.source_row,
                    )
                )
            else:
                issues.append(
                    QcIssue(
                        asset_id=None,
                        rule_id=RULE_ID,
                        field="waiver_reason",
                        severity=Severity.NEED_REVIEW,
                        message=(
                            f"程序「{label}」标记为不执行，但未启用语义复核，"
                            "需人工判断拒绝执行理由是否充分合理"
                        ),
                        suggestion="启用 LLM 语义复核（--llm）或由 reviewer 人工复核理由充分性",
                        procedure_code="SUMMARY",
                        source_sheet=source_sheet,
                        source_row=row.source_row,
                    )
                )
        return issues

    if status == "ambiguous":
        issues.append(
            QcIssue(
                asset_id=None,
                rule_id=RULE_ID,
                field="execution_status",
                severity=Severity.NEED_REVIEW,
                message=f"程序「{label}」执行状态「{row.execution_status}」无法自动解析",
                suggestion="请使用「是/否」或明确的不执行标记，并填写理由",
                procedure_code="SUMMARY",
                source_sheet=source_sheet,
                source_row=row.source_row,
            )
        )

    if status == "yes":
        issues.extend(
            _check_yes_program_sheet(
                row,
                source_sheet,
                workbook_sheet_titles=workbook_sheet_titles,
                workbook_path=workbook_path,
            )
        )

    return issues


def check_psp_completion(
    dataset: SummarySheetDataset,
    *,
    workbook_sheet_titles: Sequence[str] | None = None,
    workbook_path: str | None = None,
    waiver_reason_reviewer: WaiverReasonReviewer | None = None,
    enforce_template_completeness: bool = False,
) -> list[QcIssue]:
    issues: list[QcIssue] = []
    if not dataset.programs:
        issues.append(
            QcIssue(
                asset_id=None,
                rule_id=RULE_ID,
                field=None,
                severity=Severity.NEED_REVIEW,
                message="未在底稿中识别到汇总页程序表或表头不匹配",
                suggestion="确认存在「汇总」工作表，且含程序名称与是否执行/不执行原因列",
                procedure_code="SUMMARY",
                source_sheet=dataset.source_sheet or "汇总",
            )
        )
        return issues

    programs = [p for p in dataset.programs if not _should_skip_row(p)]
    # 行级执行与拒绝理由检查应覆盖汇总页全部有效程序行；
    # is_psp 仅用于标识高关注项，不应用于缩小检查范围。
    targets = programs
    inherited_status = _infer_merged_execution_status(targets)
    inherited_status_sources = _infer_merged_execution_status_sources(targets)
    issues.extend(
        _dep_alt_selection_consistency_issues(
            targets,
            source_sheet=dataset.source_sheet or "汇总",
            inherited_status=inherited_status,
            workbook_sheet_titles=workbook_sheet_titles,
            workbook_path=workbook_path,
        )
    )
    skip_waiver_row_ids = _rows_skip_waiver_by_dep_alt(
        targets,
        inherited_status=inherited_status,
        workbook_sheet_titles=workbook_sheet_titles,
        workbook_path=workbook_path,
    )

    for idx, row in enumerate(targets):
        merged_inherited_no = (
            idx in inherited_status_sources
            and normalize_execution_status(inherited_status.get(idx, row.execution_status)) == "no"
        )
        issues.extend(
            _check_program_row(
                row,
                dataset.source_sheet or "汇总",
                execution_status_override=inherited_status.get(idx),
                skip_waiver_reason_check=(idx in skip_waiver_row_ids) or merged_inherited_no,
                workbook_sheet_titles=workbook_sheet_titles,
                workbook_path=workbook_path,
                waiver_reason_reviewer=waiver_reason_reviewer,
            )
        )

    if enforce_template_completeness and dataset.layout == "swp":
        issues.extend(
            _check_template_program_completeness(
                dataset.programs,
                dataset.source_sheet or "汇总",
                workbook_sheet_titles=workbook_sheet_titles,
            )
        )

    return issues


_EXPECTED_TEMPLATE_PROGRAMS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("K.00", ("K.00 Lead Sheet",)),
    ("K.01", ("K.01 Agree SL to GL",)),
    ("K.02.1", ("K.02.1 新增测试",)),
    ("K.02.1a", ("K.02.1a 新增选样输出",)),
    ("K.02_addition_list", ("新增清单",)),
    ("K.02.2", ("K.02.2 处置测试",)),
    ("K.02.2a", ("K.02.2a 处置选样输出",)),
    ("K.02_disposal_list", ("处置清单",)),
    # 折旧详细测试：SAP 与 TOD 二选一执行即可。
    ("K.03_dep_test_alt", ("K.03.1 SAP", "K.03.2 折旧测试TOD")),
    ("K.03.3", ("K.03.3 折旧政策复核",)),
    ("K.04", ("K.04 固定资产减值",)),
)


def _norm_text(s: str | None) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", "", str(s)).lower()


def _row_matches_expected(row: PspProgramRow, expected_ref: str) -> bool:
    expected = _norm_text(expected_ref)
    proc = _norm_text(row.procedure_name)
    sref = _norm_text(row.sheet_ref)
    if expected in proc or expected in sref:
        return True
    fallback = _fallback_sheet_ref_from_procedure(row.procedure_name)
    return bool(fallback and _norm_text(fallback) in expected)


def _check_template_program_completeness(
    programs: list[PspProgramRow],
    source_sheet: str,
    *,
    workbook_sheet_titles: Sequence[str] | None,
) -> list[QcIssue]:
    issues: list[QcIssue] = []
    # Group headers and sample-output reference rows are not executable
    # program rows, but they still evidence that the template module exists.
    present_rows = programs
    for _, expected_refs in _EXPECTED_TEMPLATE_PROGRAMS:
        has_row = any(
            _row_matches_expected(row, expected_ref)
            for row in present_rows
            for expected_ref in expected_refs
        )
        if has_row:
            continue
        found_sheet = False
        expected_label = " / ".join(expected_refs)
        if workbook_sheet_titles:
            for expected_ref in expected_refs:
                matched, score, _ = find_matching_sheet(expected_ref, list(workbook_sheet_titles))
                if matched and score >= _STRONG_SHEET_MATCH_SCORE:
                    found_sheet = True
                    break
        sev = Severity.FAIL
        issues.append(
            QcIssue(
                asset_id=None,
                rule_id=RULE_ID,
                field="program_completeness",
                severity=sev,
                message=(
                    f"汇总页未识别到模板关键程序「{expected_label}」；"
                    "请确认程序是否缺失、被隐藏或命名变体未被识别"
                    + ("（工作簿已发现相关程序页）" if found_sheet else "（工作簿中也未发现相关程序页）")
                ),
                suggestion=(
                    "补充汇总页对应程序行，或调整程序页引用/命名使其可被识别"
                ),
                procedure_code="SUMMARY",
                source_sheet=source_sheet,
            )
        )
    return issues


def _norm_token(s: str | None) -> str:
    if not s:
        return ""
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", str(s).lower())


def _k02_exec_group_key(row: PspProgramRow) -> str | None:
    ref = _ref_for_sheet_match(row) or row.procedure_name or ""
    m = _K_SEGMENT.search(ref)
    if not m:
        return None
    code = _norm_token(m.group(0))
    if code in {"k021", "k021a"}:
        return "k021"
    if code in {"k022", "k022a"}:
        return "k022"
    return None


def _infer_merged_execution_status(programs: list[PspProgramRow]) -> dict[int, str]:
    """汇总页执行列合并单元格时，将同组程序的执行状态前向继承。"""
    groups: dict[str, list[tuple[int, PspProgramRow]]] = {}
    for idx, row in enumerate(programs):
        key = _k02_exec_group_key(row)
        if key is None:
            continue
        groups.setdefault(key, []).append((idx, row))

    inherited: dict[int, str] = {}
    for pairs in groups.values():
        pairs.sort(key=lambda x: x[1].source_row or 0)
        fallback_status: str | None = None
        for _, row in pairs:
            text = (row.execution_status or "").strip()
            if text:
                fallback_status = text
                break
        if not fallback_status:
            continue
        for idx, row in pairs:
            if not (row.execution_status or "").strip():
                inherited[idx] = fallback_status
    return inherited


def _infer_merged_execution_status_sources(programs: list[PspProgramRow]) -> dict[int, int]:
    """返回通过合并单元格逻辑继承状态的行 -> 主行索引。"""
    groups: dict[str, list[tuple[int, PspProgramRow]]] = {}
    for idx, row in enumerate(programs):
        key = _k02_exec_group_key(row)
        if key is None:
            continue
        groups.setdefault(key, []).append((idx, row))

    inherited_from: dict[int, int] = {}
    for pairs in groups.values():
        pairs.sort(key=lambda x: x[1].source_row or 0)
        source_idx: int | None = None
        for idx, row in pairs:
            if (row.execution_status or "").strip():
                source_idx = idx
                break
        if source_idx is None:
            continue
        for idx, row in pairs:
            if idx == source_idx:
                continue
            if not (row.execution_status or "").strip():
                inherited_from[idx] = source_idx
    return inherited_from


def _dep_alt_kind(row: PspProgramRow) -> Literal["sap", "tod"] | None:
    blob = _norm_token(f"{row.procedure_name} {row.sheet_ref or ''}")
    if "k031" in blob or "sap" in blob:
        return "sap"
    if "k032" in blob or "tod" in blob:
        return "tod"
    return None


def _has_tod_evidence_in_workbook(
    workbook_sheet_titles: Sequence[str] | None,
    workbook_path: str | None,
) -> bool:
    if not workbook_sheet_titles:
        return False
    candidates: list[str] = []
    for title in workbook_sheet_titles:
        if _has_tod_number_signal(title) and _has_tod_semantic_signal(title):
            candidates.append(title)
    if not candidates:
        return False
    if not workbook_path:
        return False
    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        wb = None
    if wb is not None:
        try:
            for title in candidates:
                if title not in wb.sheetnames:
                    continue
                if _has_tod_content_signal(wb[title]):
                    return True
        finally:
            wb.close()
        return False
    for title in candidates:
        # 回退：仅在无法打开工作簿时，使用“有实质内容”作为弱证据。
        if count_non_empty_cells(workbook_path, title, max_rows=40) >= _MIN_SUBSTANCE_CELLS:
            return True
    return False


def _has_tod_number_signal(text: str | None) -> bool:
    t = _norm_token(text)
    return "k032" in t or "k032" in t.replace("o", "0")


def _has_tod_semantic_signal(text: str | None) -> bool:
    t = _norm_token(text)
    dep_terms = ("折旧", "depreciation", "累计折旧")
    tod_terms = ("tod", "byitem", "逐项", "重算", "重新计算", "详细测试")
    return any(x in t for x in dep_terms) and any(x in t for x in tod_terms)


def _has_tod_content_signal(ws: openpyxl.worksheet.worksheet.Worksheet) -> bool:
    non_empty = 0
    numeric_cells = 0
    term_hits: set[str] = set()
    content_terms = (
        "资产编号",
        "原值",
        "累计折旧",
        "账面折旧",
        "重算折旧",
        "差异",
        "使用寿命",
        "残值率",
        "折旧",
        "depreciation",
        "useful",
        "salvage",
        "byitem",
        "重算",
        "逐项",
    )
    for row in ws.iter_rows(
        min_row=1,
        max_row=_MAX_SIGNAL_SCAN_ROWS,
        min_col=1,
        max_col=_MAX_SIGNAL_SCAN_COLS,
        values_only=True,
    ):
        for v in row:
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            non_empty += 1
            t = _norm_token(s)
            if isinstance(v, (int, float)) or re.fullmatch(r"-?\d+(?:\.\d+)?", s):
                numeric_cells += 1
            for kw in content_terms:
                if _norm_token(kw) in t:
                    term_hits.add(kw)
            if (
                len(term_hits) >= _MIN_TOD_CONTENT_TERMS
                and numeric_cells >= _MIN_TOD_NUMERIC_CELLS
                and non_empty >= _MIN_SUBSTANCE_CELLS
            ):
                return True
    return False


def _rows_skip_waiver_by_dep_alt(
    programs: list[PspProgramRow],
    *,
    inherited_status: dict[int, str],
    workbook_sheet_titles: Sequence[str] | None,
    workbook_path: str | None,
) -> set[int]:
    """K.03 折旧测试 SAP/TOD 二选一；TOD by item 视为 TOD 已执行证据。"""
    sap_idx: int | None = None
    tod_idx: int | None = None
    sap_yes = False
    tod_yes = False
    for idx, row in enumerate(programs):
        kind = _dep_alt_kind(row)
        if not kind:
            continue
        status_text = inherited_status.get(idx, row.execution_status)
        status = normalize_execution_status(status_text)
        if kind == "sap":
            sap_idx = idx if sap_idx is None else sap_idx
            sap_yes = sap_yes or status == "yes"
        else:
            tod_idx = idx if tod_idx is None else tod_idx
            tod_yes = tod_yes or status == "yes"

    if not tod_yes:
        tod_yes = _has_tod_evidence_in_workbook(workbook_sheet_titles, workbook_path)
    if not (sap_yes or tod_yes):
        return set()

    skip: set[int] = set()
    for idx in (sap_idx, tod_idx):
        if idx is None:
            continue
        status_text = inherited_status.get(idx, programs[idx].execution_status)
        if normalize_execution_status(status_text) == "no":
            skip.add(idx)
    return skip


def _dep_alt_selection_consistency_issues(
    programs: list[PspProgramRow],
    *,
    source_sheet: str,
    inherited_status: dict[int, str],
    workbook_sheet_titles: Sequence[str] | None,
    workbook_path: str | None,
) -> list[QcIssue]:
    """当底稿证据显示折旧测试已执行，但汇总页勾选为否时，提示勾选口径不一致。"""
    sap_idx: int | None = None
    tod_idx: int | None = None
    sap_yes = False
    tod_yes = False
    for idx, row in enumerate(programs):
        kind = _dep_alt_kind(row)
        if not kind:
            continue
        status_text = inherited_status.get(idx, row.execution_status)
        status = normalize_execution_status(status_text)
        if kind == "sap":
            sap_idx = idx if sap_idx is None else sap_idx
            sap_yes = sap_yes or status == "yes"
        else:
            tod_idx = idx if tod_idx is None else tod_idx
            tod_yes = tod_yes or status == "yes"

    tod_evidence = _has_tod_evidence_in_workbook(workbook_sheet_titles, workbook_path)
    issues: list[QcIssue] = []
    if tod_evidence and tod_idx is not None:
        tod_row = programs[tod_idx]
        tod_status = normalize_execution_status(inherited_status.get(tod_idx, tod_row.execution_status))
        if tod_status == "no":
            issues.append(
                QcIssue(
                    asset_id=None,
                    rule_id=RULE_ID,
                    field="execution_status_consistency",
                    severity=Severity.NEED_REVIEW,
                    message=(
                        f"程序「{tod_row.procedure_name}」在汇总页标记为不执行，"
                        "但工作簿存在「K.03.2 TOD/by item」且有实质测试内容，"
                        "请核对是否为勾选口径错误"
                    ),
                    suggestion=(
                        "确认 K.03.2 TOD（含 by item）是否已执行；若已执行请更新汇总页执行勾选，"
                        "若确未执行请补充与现有底稿证据一致的说明"
                    ),
                    procedure_code="SUMMARY",
                    source_sheet=source_sheet,
                    source_row=tod_row.source_row,
                )
            )

    if not issues and not (sap_yes or tod_yes or tod_evidence):
        return issues
    if sap_idx is None and tod_idx is None:
        return issues
    return issues
