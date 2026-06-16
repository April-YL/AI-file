from pathlib import Path

import openpyxl
import pytest

from ingest.addition_test_sheet import AdditionExecutionPathDataset, AdditionTestSheetDataset
from ingest.disposal_test_sheet import (
    DisposalExecutionPathDataset,
    DisposalSampleOutputDataset,
    DisposalSampleRow,
    DisposalTestSheetDataset,
    DisposalTestedSampleRow,
    DisposalAmountItem,
)
from ingest.lead_sheet import LeadMovementRow, LeadSheetDataset
from ingest.models import AssetRecord, FieldMapping, RollforwardLayoutProfile
from ingest.records import FaListDataset
from ingest.rollforward_sheet import (
    K01SectionRegion,
    MovementTransactionAmount,
    RollforwardSheetDataset,
)
from ingest.workbook_context import WorkbookQcContext
from report.pipeline import run_workbook_qc, run_workbook_qc_from_path
from rules.delivery_completion import DeliveryCompletionContext
from rules.models import QcIssue, Severity

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures"
CASE_B = (
    Path(__file__).resolve().parents[2]
    / "固定资产质检agent"
    / "案例库"
    / "K1 SWP 固定资产 20251231 B医疗公司.xlsx"
)


@pytest.fixture
def workbook_demo(tmp_path: Path) -> Path:
    path = tmp_path / "workbook_demo.xlsx"
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "汇总"
    ws_sum.append(["程序", "工作表", "是否执行", "不执行原因"])
    ws_sum.append(["K.01 后推", "K.01", "是", ""])
    ws_sum.append(["PSP-折旧测试", "K.03.1", "否", ""])
    ws_sum.append(["PSP-新增", "K.02.1", "否", "已提供合同约定的购置清单"])
    ws_fa = wb.create_sheet("FA list")
    with (FIXTURES / "fa_list_mixed.csv").open(encoding="utf-8-sig") as f:
        for line in f:
            ws_fa.append(line.strip().split(","))
    wb.save(path)
    wb.close()
    return path


def test_workbook_qc_includes_psp_and_fa_list(workbook_demo: Path):
    report = run_workbook_qc_from_path(str(workbook_demo), llm=False)
    rule_ids = {i.rule_id for i in report.issues}
    assert "psp_completion" in rule_ids or any(
        i.dict_rule_code == "AE-003" for i in report.issues
    )
    assert "fa_list_required_fields" in rule_ids
    assert report.procedure_code == "WORKBOOK"
    severities = {i.severity for i in report.issues}
    assert Severity.FAIL in severities


def test_workbook_qc_delivery_stage_selects_first_rule(workbook_demo: Path):
    report = run_workbook_qc_from_path(
        str(workbook_demo),
        llm=False,
        delivery_context=DeliveryCompletionContext(stage="first"),
    )

    delivery_issues = [
        i for i in report.issues if i.rule_id in {"first_delivery_standard", "final_delivery_standard"}
    ]
    assert {i.rule_id for i in delivery_issues} == {"first_delivery_standard"}
    assert "first_delivery_standard" in report.rule_ids
    assert "final_delivery_standard" not in report.rule_ids


def test_workbook_qc_delivery_stage_selects_final_rule(workbook_demo: Path):
    report = run_workbook_qc_from_path(
        str(workbook_demo),
        llm=False,
        delivery_context=DeliveryCompletionContext(stage="final"),
    )

    delivery_issues = [
        i for i in report.issues if i.rule_id in {"first_delivery_standard", "final_delivery_standard"}
    ]
    assert {i.rule_id for i in delivery_issues} == {"final_delivery_standard"}
    assert "final_delivery_standard" in report.rule_ids
    assert "first_delivery_standard" not in report.rule_ids


def test_k03_ingest_dataset_does_not_emit_findings(tmp_path: Path):
    path = tmp_path / "k03_only.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.2 折旧测试"
    ws.append(["说明", "TOD-by item 全量折旧测试"])
    ws.append(
        [
            "资产编号",
            "资产名称",
            "原值",
            "残值率",
            "折旧年限",
            "管理层计算折旧",
            "审计重新计算折旧",
            "差异",
            "结论",
        ]
    )
    ws.append(["FA-TEST-001", "设备A", 1200, "5%", 60, 228, 228, 0, "通过"])
    ws.append(["合计", "", 1200, "", "", 228, 228, 0, ""])
    ws.append(["结论", "未见重大差异"])
    wb.save(path)
    wb.close()

    report = run_workbook_qc_from_path(str(path), llm=False)

    assert not [
        issue
        for issue in report.issues
        if (issue.procedure_code or "").startswith("K.03")
        or "depreciation" in issue.rule_id
    ]


@pytest.mark.skipif(not CASE_B.exists(), reason="B company case workbook not available")
def test_workbook_qc_b_company_includes_addition_sheet_section():
    report = run_workbook_qc_from_path(str(CASE_B), llm=False)
    data = report.to_dict()
    section = data["addition_sheet_section"]
    preview = section["consistency_preview"]

    assert section["addition_test"]["source_sheet"] == "K.02.1 新增测试 "
    assert section["addition_sample_output"]["source_sheet"] == "K.02.1a 新增选样输出"
    assert section["addition_sample_output"]["parameters"]["te"]["value"] == "241,890.00"
    assert section["addition_sample_output"]["parameters"]["cra"]["value"] == "最低"
    assert preview["selected_count"] == 1
    assert preview["tested_count"] == 1
    assert preview["matched_count"] == 1
    assert preview["key_item_selected_amount"] == "380000"
    assert preview["key_item_tested_amount"] == "380000"
    assert "addition_sample_match" in report.rule_ids
    assert not [issue for issue in report.issues if issue.rule_id == "addition_sample_match"]


def test_workbook_qc_includes_addition_llm_issue(monkeypatch):
    monkeypatch.setenv("FA_QC_LLM_API_KEY", "sk-test")
    monkeypatch.setenv("FA_QC_LLM_ENABLED", "true")

    from rules.models import QcIssue

    ctx = WorkbookQcContext(
        source_file="addition_llm_demo.xlsx",
        fa_list=None,
        summary=None,
        lead=None,
        rollforward=None,
        addition_list=None,
        addition_test=AdditionTestSheetDataset(
            source_file="addition_llm_demo.xlsx",
            source_sheet="K.02.1 新增测试",
            waiver_note_text="No narrative support provided.",
            waiver_note_rows=[12],
        ),
        addition_sample_output=None,
        addition_execution_path=AdditionExecutionPathDataset(
            path_kind="test_sheet_waiver_note",
            recognition_confidence=0.9,
            summary_status="waived",
            summary_waiver_reason="below SAD",
            summary_source_row=8,
            addition_test_sheet="K.02.1 新增测试",
            test_sheet_waiver_note="No narrative support provided.",
            test_sheet_waiver_rows=[12],
        ),
        disposal_list=None,
        structure=None,
        reconciliations=[],
    )

    mock_issue = QcIssue(
        asset_id=None,
        rule_id="addition_semantic_review",
        field="waiver_reason",
        severity=Severity.WARN,
        message="mock addition semantic issue",
        suggestion="document waiver rationale",
        procedure_code="K.02.1",
        source_sheet="K.02.1 新增测试",
        source_row=12,
    )

    monkeypatch.setattr("llm.addition_review.build_addition_llm_issues", lambda *args, **kwargs: [mock_issue])
    monkeypatch.setattr(
        "report.pipeline.enrich_report_with_llm",
        lambda report, config, summary=None, workbook=None: report,
    )
    report = run_workbook_qc(ctx, llm=True)

    assert any(i.rule_id == "addition_semantic_review" for i in report.issues)
    assert "addition_semantic" in {d["key"] for d in report.runtime_timings["llm_details"]}


def test_workbook_qc_addition_purchase_population_excludes_default_cip_and_llm_does_not_escalate(monkeypatch):
    monkeypatch.setenv("FA_QC_LLM_API_KEY", "sk-test")
    monkeypatch.setenv("FA_QC_LLM_ENABLED", "true")

    addition_list = FaListDataset(
        source_file="stage0_addition_guard.xlsx",
        source_sheet="新增清单",
        mapped_fields=[
            FieldMapping("asset_id", "固定资产编号", 1),
            FieldMapping("asset_name", "固定资产名称", 2),
            FieldMapping("asset_category", "固定资产类别", 3),
            FieldMapping("start_date", "入账开始日期", 4),
            FieldMapping("original_value", "原值", 5),
            FieldMapping("addition_method", "新增方式", 6),
        ],
        records=[
            AssetRecord(
                source_row=2,
                asset_id="FA-TEST-001",
                asset_name="设备A",
                asset_category="机器设备",
                start_date="2025-01-01",
                original_value="100",
                addition_method="购置",
            ),
            AssetRecord(
                source_row=3,
                asset_id="FA-TEST-002",
                asset_name="设备B",
                asset_category="机器设备",
                start_date="2025-02-01",
                original_value="900",
                addition_method="在建工程转入",
            ),
        ],
    )
    rollforward = RollforwardSheetDataset(
        source_file="stage0_addition_guard.xlsx",
        source_sheet="K.01 Agree SL to GL",
        header_row=1,
        mapped_fields=[],
        movement_transactions=[
            MovementTransactionAmount(
                transaction_key="purchase",
                transaction_label="购置",
                measure="original_value",
                amount=100,
                source_row=12,
            )
        ],
    )
    ctx = WorkbookQcContext(
        source_file="stage0_addition_guard.xlsx",
        fa_list=None,
        summary=None,
        lead=None,
        rollforward=rollforward,
        addition_list=addition_list,
        structure=None,
        reconciliations=[],
    )
    mock_review = {
        "topics": [
            {
                "topic": "special_addition_source",
                "assessment": "insufficient",
                "rationale": "在建工程转入未说明测试安排",
                "missing_evidence": [],
                "suggested_action": "补充说明",
            }
        ]
    }

    monkeypatch.setattr("llm.addition_review.chat_completion_json", lambda *args, **kwargs: mock_review)
    monkeypatch.setattr("llm.ingest_review.run_workbook_ingest_reviews", lambda *args, **kwargs: [])
    monkeypatch.setattr(
        "report.pipeline.enrich_report_with_llm",
        lambda report, config, summary=None, workbook=None: report,
    )

    report = run_workbook_qc(ctx, llm=True)

    assert not [
        issue
        for issue in report.issues
        if issue.rule_id == "addition_rollforward_reconciliation"
    ]
    assert not [
        issue
        for issue in report.issues
        if issue.rule_id == "addition_semantic_review"
        and issue.field == "special_addition_source"
    ]


def test_workbook_qc_k01_without_opening_anchor_skips_opening_fail_and_merges_derived_net():
    lead = LeadSheetDataset(
        source_file="stage0_k01_guard.xlsx",
        source_sheet="K.00 Lead Sheet",
        movement_rows=[
            LeadMovementRow("原值", "K.01", {"py_audited": "1000", "audited_ending": "1100"}, 10),
            LeadMovementRow("累计折旧", "K.01", {"py_audited": "300", "audited_ending": "350"}, 11),
            LeadMovementRow("减值准备", "K.01", {"py_audited": "0", "audited_ending": "0"}, 12),
            LeadMovementRow("净值", None, {"py_audited": "700", "audited_ending": "750"}, 13),
        ],
    )
    rollforward = RollforwardSheetDataset(
        source_file="stage0_k01_guard.xlsx",
        source_sheet="K.01 Agree SL to GL",
        header_row=1,
        mapped_fields=[],
        ending_totals={
            "original_value": 1100,
            "accumulated_depreciation": 340,
            "impairment_provision": 0,
            "net_value": 760,
        },
    )
    ctx = WorkbookQcContext(
        source_file="stage0_k01_guard.xlsx",
        fa_list=None,
        summary=None,
        lead=lead,
        rollforward=rollforward,
        addition_list=None,
        structure=None,
        reconciliations=[],
    )

    report = run_workbook_qc(ctx, llm=False)
    k01_link_issues = [
        issue for issue in report.issues if issue.rule_id == "lead_rollforward_tb_reconciliation"
    ]

    assert len(k01_link_issues) == 1
    assert k01_link_issues[0].field == "期末|accumulated_depreciation"
    assert "并导致净值差异" in k01_link_issues[0].message
    assert not any((issue.field or "").startswith("期初|") for issue in k01_link_issues)


def test_workbook_qc_includes_disposal_sample_issue():
    ctx = WorkbookQcContext(
        source_file="disposal_demo.xlsx",
        fa_list=None,
        summary=None,
        lead=None,
        rollforward=None,
        addition_list=None,
        addition_test=None,
        addition_sample_output=None,
        addition_execution_path=None,
        disposal_list=None,
        disposal_sample_output=DisposalSampleOutputDataset(
            source_file="disposal_demo.xlsx",
            source_sheet="K.02.2a 处置选样输出",
            amounts={
                "key_item_count": DisposalAmountItem(
                    label="关键项数量",
                    amount="0",
                    source_row=50,
                    source_column=6,
                )
            },
            selected_samples=[
                DisposalSampleRow(
                    source_row=102,
                    sample_type="代表性样本",
                    asset_id="FA-D-001",
                    asset_name="旧设备A",
                    net_value="300",
                )
            ],
        ),
        disposal_test=DisposalTestSheetDataset(
            source_file="disposal_demo.xlsx",
            source_sheet="K.02.2 处置测试",
            tested_samples=[
                DisposalTestedSampleRow(
                    source_row=69,
                    sample_type="关键项（key item）",
                    asset_id="FA-D-001",
                    asset_name="旧设备A",
                    net_value="300",
                )
            ],
        ),
        disposal_execution_path=DisposalExecutionPathDataset(
            path_kind="executed_package_complete",
            recognition_confidence=0.9,
            summary_status="yes",
            disposal_list_sheet="处置清单",
            disposal_test_sheet="K.02.2 处置测试",
            disposal_sample_output_sheet="K.02.2a 处置选样输出",
        ),
        structure=None,
        reconciliations=[],
    )

    report = run_workbook_qc(ctx, llm=False)
    assert any(issue.rule_id == "disposal_sample_match" for issue in report.issues)


def test_workbook_qc_includes_disposal_llm_issue(monkeypatch):
    ctx = WorkbookQcContext(
        source_file="disposal_llm_demo.xlsx",
        fa_list=None,
        summary=None,
        lead=None,
        disposal_test=DisposalTestSheetDataset(
            source_file="disposal_llm_demo.xlsx",
            source_sheet="K.02.2 处置测试",
        ),
    )
    mock_issue = QcIssue(
        asset_id=None,
        rule_id="disposal_semantic_review",
        field="evidence_description",
        severity=Severity.NEED_REVIEW,
        message="处置证据描述需复核",
        suggestion="补充具体证据索引",
        procedure_code="K.02.2",
        source_sheet="K.02.2 处置测试",
        review_source="LLM辅助判断",
    )
    monkeypatch.setattr("report.pipeline.load_llm_config", lambda cli_enabled=None: type("Config", (), {"enabled": True})())
    monkeypatch.setattr("llm.disposal_review.build_disposal_llm_issues", lambda *args, **kwargs: [mock_issue])
    monkeypatch.setattr("llm.ingest_review.run_workbook_ingest_reviews", lambda *args, **kwargs: [])
    monkeypatch.setattr("report.pipeline.enrich_report_with_llm", lambda report, config, **kwargs: report)

    report = run_workbook_qc(ctx)

    assert any(issue.rule_id == "disposal_semantic_review" for issue in report.issues)
    assert "disposal_semantic" in {d["key"] for d in report.runtime_timings["llm_details"]}


def test_workbook_qc_includes_k01_ingest_review_section(monkeypatch):
    monkeypatch.setenv("FA_QC_LLM_API_KEY", "sk-test")
    monkeypatch.setenv("FA_QC_LLM_ENABLED", "true")

    ctx = WorkbookQcContext(
        source_file="k01_ingest_review_demo.xlsx",
        fa_list=None,
        summary=None,
        lead=None,
        rollforward=RollforwardSheetDataset(
            source_file="k01_ingest_review_demo.xlsx",
            source_sheet="K.01 Agree SL to GL",
            header_row=None,
            mapped_fields=[],
            layout_profile=RollforwardLayoutProfile.HYBRID,
            section_presence={
                "b1_bkd_main_table": True,
                "b2_movement_tb_reconciliation": True,
                "b3_table2_fa_summary": False,
                "b4_table3_check_with_table1": False,
                "b5_table4_depreciation_pl": True,
                "b6_notes_investigation_routing": True,
            },
            section_regions={
                "b1_bkd_main_table": K01SectionRegion(
                    section_id="b1_bkd_main_table",
                    anchor_row=12,
                    evidence=["表1", "固定资产类别"],
                ),
                "b5_table4_depreciation_pl": K01SectionRegion(
                    section_id="b5_table4_depreciation_pl",
                    anchor_row=82,
                    evidence=["表4", "折旧费用与利润表"],
                ),
            },
            section_conflicts=["duplicate_anchor:b4_table3_check_with_table1"],
            recognition_confidence=0.58,
        ),
        addition_list=None,
        disposal_list=None,
        structure=None,
        reconciliations=[],
    )

    mock_result = {
        "assessment": "suspicious",
        "risk_level": "high",
        "risk_area": "missing_module",
        "suspected_object": "b4_table3_check_with_table1",
        "candidate_sheet": "K.01 Agree SL to GL",
        "candidate_rows": [82],
        "evidence_anchors": ["表4", "折旧费用与利润表"],
        "rationale": "mock k01 ingest risk",
        "suggested_action": "review K.01 section boundaries",
        "should_retry_deterministic_ingest": True,
        "manual_review_focus": "人工核对 K.01 表3/表4边界。",
    }

    monkeypatch.setattr("llm.ingest_review.chat_completion_json", lambda *args, **kwargs: mock_result)
    monkeypatch.setattr(
        "report.pipeline.enrich_report_with_llm",
        lambda report, config, summary=None, workbook=None: report,
    )

    report = run_workbook_qc(ctx, llm=True)
    data = report.to_dict()

    assert "ingest_review_section" in data
    reviews = data["ingest_review_section"]["reviews"]
    assert reviews[0]["assessment"] == "suspicious"
    assert reviews[0]["procedure_code"] == "K.01"
    assert reviews[0]["source_sheet"] == "K.01 Agree SL to GL"
    assert not [issue for issue in report.issues if issue.rule_id == "llm_ingest_review"]
    assert "ingest_review" in {d["key"] for d in report.runtime_timings["llm_details"]}


def test_workbook_qc_reviews_missing_k01_candidate(monkeypatch, tmp_path: Path):
    monkeypatch.setenv("FA_QC_LLM_API_KEY", "sk-test")
    monkeypatch.setenv("FA_QC_LLM_ENABLED", "true")

    path = tmp_path / "missing_k01_candidate.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K01 SL-GL"
    ws.append(["表1", "固定资产类别", "年初余额", "年末余额", "审定数"])
    ws.append(["表2 check with 表1", "差异", "Notes"])
    wb.save(path)
    wb.close()

    ctx = WorkbookQcContext(
        source_file=str(path),
        fa_list=None,
        summary=None,
        lead=None,
        rollforward=None,
        addition_list=None,
        disposal_list=None,
        structure=None,
        reconciliations=[],
    )

    mock_result = {
        "assessment": "suspicious",
        "risk_level": "high",
        "risk_area": "missing_sheet",
        "suspected_object": "K.01 Agree SL to GL",
        "candidate_sheet": "K01 SL-GL",
        "candidate_rows": [1],
        "evidence_anchors": ["表1", "固定资产类别"],
        "rationale": "mock missing K.01 ingest risk",
        "suggested_action": "review candidate K.01 sheet",
        "should_retry_deterministic_ingest": True,
        "manual_review_focus": "人工核对 K01 SL-GL 是否为 K.01。",
    }
    monkeypatch.setattr("llm.ingest_review.chat_completion_json", lambda *args, **kwargs: mock_result)
    monkeypatch.setattr(
        "report.pipeline.enrich_report_with_llm",
        lambda report, config, summary=None, workbook=None: report,
    )

    report = run_workbook_qc(ctx, llm=True)
    data = report.to_dict()

    reviews = data["ingest_review_section"]["reviews"]
    assert reviews[0]["assessment"] == "suspicious"
    assert reviews[0]["risk_area"] == "missing_sheet"
    assert reviews[0]["source_sheet"] == "K01 SL-GL"
    assert not [issue for issue in report.issues if issue.rule_id == "llm_ingest_review"]
