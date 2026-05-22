import json
from pathlib import Path

import openpyxl
import pytest

from report.export_review_html import export_review_html
from report.manual_review import build_manual_review_sections
from report.pipeline import run_workbook_qc_from_path

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures"


@pytest.fixture
def full_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "full.xlsx"
    wb = openpyxl.Workbook()
    ws_lead = wb.active
    ws_lead.title = "K.00 Lead Sheet"
    ws_lead.append(["项目", "底稿值", "Canvas"])
    ws_lead.append(["可容忍误差", "100000", "100000"])
    ws_lead.append(["名义金额", "50000", "50000"])
    ws_lead.append(["计划重要性", "200000", "200000"])
    ws_lead.append([])
    ws_lead.append(["认定", "CRA", "TT"])
    ws_lead.append(["固定资产", "High", "120000"])
    ws_sum = wb.create_sheet("汇总")
    ws_sum.append(["程序", "工作表", "是否执行", "不执行原因"])
    ws_sum.append(["K.01", "K.01", "是", ""])
    ws_fa = wb.create_sheet("FA list")
    with (FIXTURES / "fa_list_valid.csv").open(encoding="utf-8-sig") as f:
        for line in f:
            ws_fa.append(line.strip().split(","))
    wb.save(path)
    wb.close()
    return path


def test_report_contains_manual_review_sections(full_workbook: Path, tmp_path: Path):
    report = run_workbook_qc_from_path(str(full_workbook), llm=False)
    assert len(report.manual_review_sections) == 2
    data = report.to_dict()
    assert "manual_review_sections" in data
    ae1 = next(s for s in data["manual_review_sections"] if s["dict_rule_code"] == "AE-001")
    assert "PM/TE/SAD" in ae1["checklist_prompt"]
    assert len(ae1["items"]) >= 2
    codes = {i["dict_rule_code"] for i in data["issues"]}
    assert "AE-001" in codes
    assert "AE-002" in codes

    html_path = tmp_path / "review.html"
    export_review_html(report, html_path)
    text = html_path.read_text(encoding="utf-8")
    assert "AE-001" in text
    assert 'id="findings"' in text
    assert "100000" in json.dumps(data, ensure_ascii=False)
