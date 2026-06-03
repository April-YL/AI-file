"""OXML 注入 Comments 表且不破坏外部链接。"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from ingest.case_library import iter_case_workbooks
from report.export_annotated_workbook import (
    COMMENTS_SHEET_NAME,
    FA_LIST_COMMENTS_SHEET_NAME,
    export_annotated_workbook,
)
from report.ooxml_workbook import _insert_legacy_drawing, workbook_has_external_links
from report.pipeline import run_workbook_qc_from_path

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures"


def test_insert_legacy_drawing_does_not_duplicate_r_namespace():
    xml = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        b'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        b"<sheetData /></worksheet>"
    )

    out = _insert_legacy_drawing(xml, "rId9").decode("utf-8")

    root_tag = out[out.index("<worksheet") : out.index(">", out.index("<worksheet"))]
    assert root_tag.count("xmlns:r=") == 1
    assert "<legacyDrawing" in out
    assert 'r:id="rId9"' in out


def _b_medical_path() -> Path | None:
    for ref in iter_case_workbooks():
        if "B医疗" in ref.path.name and not ref.skipped:
            return ref.path
    return None


@pytest.mark.parametrize("path", [FIXTURES / "workbook_with_lead.xlsx"])
def test_inject_comment_sheets_without_external_links(path: Path, tmp_path: Path):
    if not path.is_file():
        pytest.skip("fixture missing")
    assert not workbook_has_external_links(path)
    report = run_workbook_qc_from_path(str(path), llm=False)
    out = tmp_path / "out.xlsx"
    export_annotated_workbook(report, path, out)
    wb = openpyxl.load_workbook(out, read_only=True)
    assert wb.sheetnames[0] == COMMENTS_SHEET_NAME
    assert wb.sheetnames[1] == FA_LIST_COMMENTS_SHEET_NAME
    wb.close()


def test_external_links_preserved_on_b_medical(tmp_path: Path):
    src = _b_medical_path()
    if src is None:
        pytest.skip("B医疗 case workbook not in case library")
    assert workbook_has_external_links(src)

    orig_wb = openpyxl.load_workbook(src, data_only=True)
    orig_ws = orig_wb["K.00 Lead Sheet"]
    orig_g = orig_ws.cell(55, 7).value
    orig_i = orig_ws.cell(55, 9).value
    orig_j = orig_ws.cell(55, 10).value
    orig_wb.close()

    report = run_workbook_qc_from_path(str(src), llm=False)
    out = tmp_path / "ann.xlsx"
    export_annotated_workbook(report, src, out)

    ann_wb = openpyxl.load_workbook(out, data_only=True)
    ann_ws = ann_wb["K.00 Lead Sheet"]
    assert ann_ws.cell(55, 7).value == orig_g
    assert ann_ws.cell(55, 9).value == orig_i
    assert ann_ws.cell(55, 10).value == orig_j
    ann_wb.close()

    ann_formula = openpyxl.load_workbook(out, data_only=False)
    ws = ann_formula["K.00 Lead Sheet"]
    assert str(ws.cell(55, 7).value).startswith("=[")
    ann_formula.close()
