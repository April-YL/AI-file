from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

from ingest.lead_sheet import load_lead_from_workbook
from report.lead_sheet_report import build_lead_sheet_section
from rules.lead_required_fields import check_lead_required_fields
from rules.materiality_consistency import check_materiality_consistency
from rules.models import QcIssue, Severity
from rules.registry import attach_rule_metadata


def test_build_lead_sheet_section_includes_blocks_and_rules(tmp_path: Path):
    path = tmp_path / "lead.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["B2"] = "客户名称"
    ws["C2"] = "XYZ"
    ws["B3"] = "期末"
    ws["C3"] = datetime(2025, 12, 31)
    ws["B4"] = "分析日期"
    ws["C4"] = datetime(2026, 1, 15)
    ws["B5"] = "可容忍误差"
    ws["C5"] = 100
    ws["B6"] = "名义金额"
    ws["C6"] = 10
    ws["B7"] = "适用会计准则"
    ws["C7"] = "PRC GAAP"
    ws["B8"] = "记账本位币"
    ws["C8"] = "CNY"
    wb.save(path)
    wb.close()

    lead = load_lead_from_workbook(path)
    issues = attach_rule_metadata(check_lead_required_fields(lead))
    issues.extend(attach_rule_metadata(check_materiality_consistency(lead)))
    sec = build_lead_sheet_section(lead, issues)

    assert sec["ingested"] is True
    assert sec["source_sheet"] == "K.00 Lead Sheet"
    assert sec["blocks"]
    assert sec["lead_qc"]["overall_severity"] in ("PASS", "NEED_REVIEW", "WARN", "FAIL")
    assert "lead_required_fields" in sec["lead_qc"]["rules"]
    assert sec["lead_qc"]["rules"]["materiality_consistency"]["dict_rule_code"] == "AE-001"


def test_lead_section_reflects_failures():
    from ingest.lead_sheet import LeadSheetDataset

    ds = LeadSheetDataset(source_file="t.xlsx", source_sheet="K.00 Lead Sheet")
    issues = [
        QcIssue(
            asset_id=None,
            rule_id="lead_required_fields",
            field="client_name",
            severity=Severity.FAIL,
            message="missing",
            suggestion="fix",
            procedure_code="K.00",
            source_sheet="K.00 Lead Sheet",
        )
    ]
    sec = build_lead_sheet_section(ds, issues)
    assert sec["lead_qc"]["overall_severity"] == "FAIL"
    assert sec["lead_qc"]["rules"]["lead_required_fields"]["issue_count"] == 1
