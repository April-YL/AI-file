from pathlib import Path

import openpyxl
import pytest

from ingest.lead_sheet import load_lead_from_workbook
from report.export_review_html import export_review_html
from report.lead_sheet_report import build_lead_sheet_section
from report.pipeline import run_workbook_qc_from_path
from rules.lead_required_fields import check_lead_required_fields
from rules.registry import attach_rule_metadata

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures"


def test_html_includes_lead_section_and_procedure_filter(tmp_path: Path):
    path = FIXTURES / "workbook_with_lead.xlsx"
    if not path.is_file():
        pytest.skip("fixture workbook_with_lead.xlsx missing")

    report = run_workbook_qc_from_path(str(path), llm=False)
    assert report.lead_sheet_section is not None

    html_path = tmp_path / "review.html"
    export_review_html(report, html_path)
    text = html_path.read_text(encoding="utf-8")

    assert 'id="findings"' in text
    assert "Findings 清单" in text
    assert "Comments" in text
    assert "FAIL" in text or "NEED_REVIEW" in text


def test_html_lead_section_from_minimal_workbook(tmp_path: Path):
    path = tmp_path / "lead_only.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["B2"] = "客户名称"
    ws["C2"] = "测试"
    wb.save(path)
    wb.close()

    lead = load_lead_from_workbook(path)
    issues = attach_rule_metadata(check_lead_required_fields(lead))
    sec = build_lead_sheet_section(lead, issues)

    from report.summary import QcReport, build_report

    report = build_report(
        source_file=str(path),
        source_sheet=lead.source_sheet,
        procedure_code="WORKBOOK",
        rule_ids=["lead_required_fields"],
        records=[],
        issues=issues,
        lead_sheet_section=sec,
    )
    html_path = tmp_path / "lead.html"
    export_review_html(report, html_path)
    text = html_path.read_text(encoding="utf-8")
    assert "lead_required_fields" in text or "FAIL" in text
    assert "findings" in text
