from pathlib import Path
import zipfile

import openpyxl
import pytest

from report.export_annotated_workbook import (
    COMMENTS_SHEET_NAME,
    EXECUTION_TRACE_SHEET_NAME,
    FA_LIST_COMMENTS_SHEET_NAME,
    LLM_INGEST_REVIEW_SHEET_NAME,
    LOCATOR_SHEET_NAME,
    build_comments_rows,
    build_fa_list_detail_rows,
    build_llm_ingest_review_rows,
    build_locator_rows,
    build_main_comments_rows,
    export_annotated_workbook,
    split_fa_list_issues,
)
from report.pipeline import run_workbook_qc_from_path
from report.summary import ReportSummary, QcReport
from rules.models import QcIssue, Severity

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures"


def test_split_fa_list_issues():
    path = FIXTURES / "workbook_with_lead.xlsx"
    if not path.is_file():
        pytest.skip("fixture missing")
    report = run_workbook_qc_from_path(str(path), llm=False)
    issues = [i for i in report.issues if i.severity != Severity.PASS]
    fa, other = split_fa_list_issues(issues)
    assert all(i.procedure_code == "FA_LIST" for i in fa)
    assert all(i.procedure_code != "FA_LIST" for i in other)


def test_main_rows_fewer_than_fa_detail_when_many_fa_dupes():
    path = FIXTURES / "fa_list_mixed.xlsx"
    if not path.is_file():
        pytest.skip("fixture missing")
    from ingest.records import load_fa_list_from_workbook
    from rules.runner import run_fa_list_rules
    from rules.registry import attach_rule_metadata
    from rules.models import ColumnContext
    from report.summary import build_report

    ds = load_fa_list_from_workbook(path)
    ctx = ColumnContext(
        mapped_fields={m.standard_field for m in ds.mapped_fields},
        source_sheet=ds.source_sheet,
    )
    issues = attach_rule_metadata(run_fa_list_rules(ds.records, ctx))
    issues = [i for i in issues if i.severity != Severity.PASS]
    fa, other = split_fa_list_issues(issues)
    main_rows = build_main_comments_rows(other, fa)
    detail_rows = build_fa_list_detail_rows(fa)
    assert len(detail_rows) == len(fa)
    if len(fa) > 2:
        assert len(main_rows) < len(fa) + len(other)


def test_fa_list_net_value_summary_uses_difference_totals():
    issue = QcIssue(
        asset_id="FA-TEST-001",
        rule_id="fa_list_recalc",
        dict_rule_code="FA-RC-003",
        field="net_value",
        severity=Severity.FAIL,
        message="净值勾稽不一致：原值 100，累计折旧 40，减值 0，净值 50，差异 10。",
        suggestion="核对净值公式。",
        procedure_code="FA_LIST",
        source_sheet="FA list",
        source_row=10,
    )

    rows = build_main_comments_rows([], [issue])

    assert len(rows) == 1
    assert "共 1 条同类问题" in rows[0][3]
    assert "差异绝对值合计=10" in rows[0][3]
    assert "最大单项差异=10" in rows[0][3]
    assert "代表性问题" not in rows[0][3]
    assert "[FAIL]" not in rows[0][3]
    assert "FA-RC-003" not in rows[0][3]
    assert rows[0][8] == "FAIL"
    assert rows[0][9] == "FA-RC-003"
    assert rows[0][10] == "net_value"


def test_fa_list_net_value_summary_shows_total_and_max_difference_without_representative():
    issues = [
        QcIssue(
            asset_id="FA-TEST-001",
            rule_id="asset_value_consistency",
            dict_rule_code="FA-RC-003",
            field="net_value",
            severity=Severity.FAIL,
            message="净值与原值减累计折旧不一致：净值=37479.03，计算值=57401.55，差异=19922.52（允差=11.8704）",
            suggestion="核对净值公式。",
            procedure_code="FA_LIST",
            source_sheet="FA list",
            source_row=10,
        ),
        QcIssue(
            asset_id="FA-TEST-002",
            rule_id="asset_value_consistency",
            dict_rule_code="FA-RC-003",
            field="net_value",
            severity=Severity.FAIL,
            message="净值与原值减累计折旧不一致：净值=10，计算值=30，差异=20（允差=0.01）",
            suggestion="核对净值公式。",
            procedure_code="FA_LIST",
            source_sheet="FA list",
            source_row=11,
        ),
    ]

    rows = build_main_comments_rows([], issues)

    assert len(rows) == 1
    assert "共 2 条同类问题" in rows[0][3]
    assert "差异绝对值合计=19942.52" in rows[0][3]
    assert "最大单项差异=19922.52" in rows[0][3]
    assert f"详见《{FA_LIST_COMMENTS_SHEET_NAME}》" in rows[0][3]
    assert "代表性问题" not in rows[0][3]


def test_fa_list_detail_keeps_net_value_difference_message():
    issue = QcIssue(
        asset_id="FA-TEST-001",
        rule_id="asset_value_consistency",
        dict_rule_code="FA-RC-003",
        field="net_value",
        severity=Severity.FAIL,
        message="净值与原值减累计折旧不一致：净值=37479.03，计算值=57401.55，差异=19922.52（允差=11.8704）",
        suggestion="核对净值公式。",
        procedure_code="FA_LIST",
        source_sheet="FA list",
        source_row=10,
    )

    rows = build_fa_list_detail_rows([issue])

    assert "净值=37479.03" in rows[0][3]
    assert "计算值=57401.55" in rows[0][3]
    assert "差异=19922.52" in rows[0][3]
    assert "允差=11.8704" in rows[0][3]


def test_export_two_comment_sheets(tmp_path: Path):
    src = FIXTURES / "workbook_with_lead.xlsx"
    if not src.is_file():
        pytest.skip("fixture missing")
    report = run_workbook_qc_from_path(str(src), llm=False)
    out = tmp_path / "out.xlsx"
    export_annotated_workbook(report, src, out)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames[0] == COMMENTS_SHEET_NAME
    assert wb.sheetnames[1] == FA_LIST_COMMENTS_SHEET_NAME
    assert wb.sheetnames[2] == LOCATOR_SHEET_NAME
    assert wb.sheetnames[3] == EXECUTION_TRACE_SHEET_NAME
    assert wb.sheetnames[4] == LLM_INGEST_REVIEW_SHEET_NAME
    visible_sheetnames = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    assert LOCATOR_SHEET_NAME not in visible_sheetnames
    assert wb[LOCATOR_SHEET_NAME].sheet_state == "hidden"
    assert wb[COMMENTS_SHEET_NAME].cell(1, 1).value == "EY Ref."
    assert wb[COMMENTS_SHEET_NAME].cell(1, 4).value == "Question/Comment"
    assert wb[COMMENTS_SHEET_NAME].cell(1, 9).value == "Severity"
    assert wb[COMMENTS_SHEET_NAME].cell(1, 10).value == "Rule"
    assert wb[COMMENTS_SHEET_NAME].cell(1, 11).value == "Field"
    assert wb[COMMENTS_SHEET_NAME].column_dimensions["D"].width >= 75
    assert wb[COMMENTS_SHEET_NAME].column_dimensions["G"].width >= 50
    assert wb[FA_LIST_COMMENTS_SHEET_NAME].column_dimensions["D"].width >= 75
    assert wb[FA_LIST_COMMENTS_SHEET_NAME].column_dimensions["G"].width >= 50
    assert wb[LOCATOR_SHEET_NAME].cell(1, 1).value == "EY Ref."
    wb.close()


def test_export_llm_ingest_review_sheet(tmp_path: Path):
    src = tmp_path / "source.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K01 SL-GL"
    ws["A1"] = "表1"
    wb.save(src)
    wb.close()

    report = QcReport(
        source_file=str(src),
        source_sheet="workbook",
        procedure_code="WORKBOOK",
        rule_ids=[],
        issues=[],
        asset_results=[],
        summary=ReportSummary(
            total_records=0,
            pass_count=0,
            warn_count=0,
            fail_count=0,
            need_review_count=0,
            overall_severity=Severity.PASS,
        ),
        ingest_review_section={
            "description": "读取结果复核提示（LLM 辅助，不等同于业务规则 finding）。",
            "reviews": [
                {
                    "procedure_code": "K.01",
                    "review_type": "k01_ingest_review",
                    "assessment": "suspicious",
                    "risk_level": "high",
                    "risk_area": "missing_sheet",
                    "candidate_sheet": "K01 SL-GL",
                    "candidate_rows": [1],
                    "evidence_anchors": ["表1", "固定资产类别"],
                    "rationale": "候选 sheet 出现后推锚点。",
                    "suggested_action": "人工核对是否为 K.01。",
                    "manual_review_focus": "打开候选 sheet 第1行附近。",
                    "note": "读取结果复核提示，不等同于业务规则 finding。",
                }
            ],
        },
    )

    rows = build_llm_ingest_review_rows(report)
    assert rows[0][1] == "K.01"
    assert rows[0][3] == "suspicious"

    out = tmp_path / "out.xlsx"
    export_annotated_workbook(report, src, out)

    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2[LLM_INGEST_REVIEW_SHEET_NAME]
    assert ws2.cell(1, 1).value == "序号"
    assert ws2.cell(2, 4).value == "suspicious"
    assert ws2.cell(2, 7).value == "K01 SL-GL"
    assert ws2.cell(2, 8).hyperlink is not None
    assert ws2.cell(2, 8).hyperlink.location == "'K01 SL-GL'!B1"
    wb2.close()


def test_main_comments_cell_ref_has_internal_hyperlink(tmp_path: Path):
    src = tmp_path / "source.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["B7"] = "问题位置"
    wb.save(src)
    wb.close()

    issue = QcIssue(
        asset_id=None,
        rule_id="lead_required_fields",
        field="gaap",
        severity=Severity.FAIL,
        message="缺少适用会计准则",
        suggestion="补充",
        procedure_code="K.00",
        source_sheet="K.00 Lead Sheet",
        source_row=7,
        dict_rule_code="LEAD-001",
    )
    report = QcReport(
        source_file=str(src),
        source_sheet="K.00 Lead Sheet",
        procedure_code="K.00",
        rule_ids=["lead_required_fields"],
        issues=[issue],
        asset_results=[],
        summary=ReportSummary(
            total_records=0,
            pass_count=0,
            warn_count=0,
            fail_count=1,
            need_review_count=0,
            overall_severity=Severity.FAIL,
        ),
    )
    out = tmp_path / "out.xlsx"
    export_annotated_workbook(report, src, out)

    wb2 = openpyxl.load_workbook(out)
    cell = wb2[COMMENTS_SHEET_NAME]["C2"]
    assert cell.value == "$B$7"
    assert cell.hyperlink is not None
    assert cell.hyperlink.location == "'K.00 Lead Sheet'!B7"
    assert wb2["K.00 Lead Sheet"]["B7"].comment is not None
    assert "缺少适用会计准则" in wb2["K.00 Lead Sheet"]["B7"].comment.text
    wb2.close()


def test_main_comments_uses_source_column_for_anchor(tmp_path: Path):
    src = tmp_path / "source_col.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.02.1a 新增选样输出"
    ws["F41"] = 100
    wb.save(src)
    wb.close()

    issue = QcIssue(
        asset_id=None,
        rule_id="addition_sample_pool_purchase_amount_match",
        field="sample_pool_amount",
        severity=Severity.FAIL,
        message="样本池金额与后推金额不一致",
        suggestion="核对样本池金额",
        procedure_code="K.02.1",
        source_sheet="K.02.1a 新增选样输出",
        source_row=41,
        source_col=6,
    )
    report = QcReport(
        source_file=str(src),
        source_sheet="K.02.1a 新增选样输出",
        procedure_code="K.02.1",
        rule_ids=["addition_sample_pool_purchase_amount_match"],
        issues=[issue],
        asset_results=[],
        summary=ReportSummary(
            total_records=0,
            pass_count=0,
            warn_count=0,
            fail_count=1,
            need_review_count=0,
            overall_severity=Severity.FAIL,
        ),
    )
    out = tmp_path / "out.xlsx"
    export_annotated_workbook(report, src, out)

    wb2 = openpyxl.load_workbook(out)
    cell = wb2[COMMENTS_SHEET_NAME]["C2"]
    assert cell.value == "$F$41"
    assert cell.hyperlink.location == "'K.02.1a 新增选样输出'!F41"
    assert wb2["K.02.1a 新增选样输出"]["F41"].comment is not None
    assert wb2["K.02.1a 新增选样输出"]["B41"].comment is None
    wb2.close()


def test_external_link_workbook_gets_ooxml_cell_comment(tmp_path: Path):
    src = tmp_path / "source_external.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["B7"] = "问题位置"
    wb.save(src)
    wb.close()

    with zipfile.ZipFile(src, "a", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("xl/externalLinks/externalLink1.xml", "<externalLink/>")

    issue = QcIssue(
        asset_id=None,
        rule_id="lead_required_fields",
        field="gaap",
        severity=Severity.FAIL,
        message="缺少适用会计准则",
        suggestion="补充",
        procedure_code="K.00",
        source_sheet="K.00 Lead Sheet",
        source_row=7,
        dict_rule_code="LEAD-001",
    )
    report = QcReport(
        source_file=str(src),
        source_sheet="K.00 Lead Sheet",
        procedure_code="K.00",
        rule_ids=["lead_required_fields"],
        issues=[issue],
        asset_results=[],
        summary=ReportSummary(
            total_records=0,
            pass_count=0,
            warn_count=0,
            fail_count=1,
            need_review_count=0,
            overall_severity=Severity.FAIL,
        ),
    )
    out = tmp_path / "out.xlsx"
    export_annotated_workbook(report, src, out)

    with zipfile.ZipFile(out, "r") as zf:
        names = set(zf.namelist())
        assert "xl/externalLinks/externalLink1.xml" in names
        comment_parts = [n for n in names if n.startswith("xl/comments/comment")]
        assert comment_parts
        comment_xml = zf.read(comment_parts[0]).decode("utf-8")
        assert 'ref="B7"' in comment_xml
        assert "缺少适用会计准则" in comment_xml
        assert any(n.startswith("xl/drawings/vmlDrawing") for n in names)


def test_merged_cell_comment_anchors_to_top_left(tmp_path: Path):
    src = tmp_path / "merged_source.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.02.1 addition test"
    ws.merge_cells("B15:F15")
    ws["B15"] = "merged header"
    wb.save(src)
    wb.close()

    issue = QcIssue(
        asset_id=None,
        rule_id="addition_semantic_review",
        field="sample_selection",
        severity=Severity.WARN,
        message="样本选择依据不足",
        suggestion="补充样本选择依据",
        procedure_code="K.02.1",
        source_sheet="K.02.1 addition test",
        source_row=15,
    )
    report = QcReport(
        source_file=str(src),
        source_sheet="K.02.1 addition test",
        procedure_code="K.02.1",
        rule_ids=["addition_semantic_review"],
        issues=[issue],
        asset_results=[],
        summary=ReportSummary(
            total_records=0,
            pass_count=0,
            warn_count=1,
            fail_count=0,
            need_review_count=0,
            overall_severity=Severity.WARN,
        ),
    )
    out = tmp_path / "merged_out.xlsx"
    export_annotated_workbook(report, src, out)

    wb2 = openpyxl.load_workbook(out)
    assert wb2["K.02.1 addition test"]["B15"].comment is not None
    assert "样本选择依据不足" in wb2["K.02.1 addition test"]["B15"].comment.text
    wb2.close()


def test_answer_blank_agent_ref_in_last_column():
    issue = QcIssue(
        asset_id=None,
        rule_id="lead_required_fields",
        field="gaap",
        severity=Severity.FAIL,
        message="缺少适用会计准则",
        suggestion="在 Lead 表补充适用会计准则",
        procedure_code="K.00",
        source_sheet="K.00 Lead Sheet",
        source_row=7,
    )
    rows = build_main_comments_rows([issue], [])
    assert len(rows) == 1
    assert rows[0][4] is None
    assert "Agent 参考" not in (rows[0][3] or "")
    assert rows[0][6] == "在 Lead 表补充适用会计准则"
    assert rows[0][7] == "规则判断"
    assert rows[0][8] == "FAIL"
    assert rows[0][9] == "lead_required_fields"
    assert rows[0][10] == "gaap"


def test_comment_rows_include_llm_review_source():
    issue = QcIssue(
        asset_id=None,
        rule_id="psp_completion",
        field="waiver_reason",
        severity=Severity.WARN,
        message="程序不执行理由语义上不足；模型提示：仅说明小于TE。",
        suggestion="补充单项TT和性质异常判断",
        procedure_code="SUMMARY",
        source_sheet="汇总",
        source_row=18,
        review_source="规则+LLM",
        llm_review_type="PSP不执行理由充分性",
    )
    rows = build_main_comments_rows([issue], [])
    assert rows[0][7] == "规则+LLM（PSP不执行理由充分性）"
    data = issue.to_dict()
    assert data["review_source"] == "规则+LLM"
    assert data["llm_review_type"] == "PSP不执行理由充分性"


def test_build_comments_rows_compat():
    path = FIXTURES / "workbook_with_lead.xlsx"
    if not path.is_file():
        pytest.skip("fixture missing")
    report = run_workbook_qc_from_path(str(path), llm=False)
    issues = [i for i in report.issues if i.severity != Severity.PASS]
    rows = build_comments_rows(issues)
    assert rows and rows[0][0] == 1


def test_main_comment_rows_order_summary_then_lead_then_other():
    issues = [
        QcIssue(
            asset_id=None,
            rule_id="lead_required_fields",
            field="gaap",
            severity=Severity.FAIL,
            message="Lead 缺少会计准则",
            suggestion="补充",
            procedure_code="K.00",
            source_sheet="K.00 Lead Sheet",
            source_row=10,
        ),
        QcIssue(
            asset_id=None,
            rule_id="psp_completion",
            field="execution_status",
            severity=Severity.WARN,
            message="汇总页执行状态异常",
            suggestion="补充",
            procedure_code="SUMMARY",
            source_sheet="汇总",
            source_row=20,
        ),
        QcIssue(
            asset_id=None,
            rule_id="rollforward_exists",
            field=None,
            severity=Severity.NEED_REVIEW,
            message="后推页缺失",
            suggestion="补充",
            procedure_code="K.01",
            source_sheet="K.01 Agree SL to GL",
            source_row=5,
        ),
    ]
    rows = build_main_comments_rows(issues, [])
    tabs = [r[1] for r in rows]
    assert tabs == ["汇总", "K.00 Lead Sheet", "K.01 Agree SL to GL"]


def test_question_comment_is_not_truncated_with_ellipsis():
    issue = QcIssue(
        asset_id=None,
        rule_id="psp_completion",
        field="waiver_reason",
        severity=Severity.WARN,
        message=(
            "程序「K.03.2 折旧测试TOD」不执行理由语义上不足；模型提示："
            "未说明业务风险、阈值判断与替代程序，建议补充详细说明并给出证据来源。"
        ),
        suggestion="补充说明",
        procedure_code="SUMMARY",
        source_sheet="汇总",
        source_row=18,
    )
    rows = build_main_comments_rows([issue], [])
    question = rows[0][3]
    assert not question.startswith("[WARN]")
    assert "模型提示" not in question
    assert "..." not in question
    assert rows[0][8] == "WARN"
    assert rows[0][9] == "psp_completion"
    assert rows[0][10] == "waiver_reason"


def test_question_comment_uses_short_title_mapping():
    issue = QcIssue(
        asset_id=None,
        rule_id="psp_completion",
        field="execution_status_consistency",
        severity=Severity.NEED_REVIEW,
        message="很长的原始信息，不应直接出现在 Question/Comment 里。",
        suggestion="补充说明",
        procedure_code="SUMMARY",
        source_sheet="汇总",
        source_row=18,
        dict_rule_code="AE-003",
    )
    rows = build_main_comments_rows([issue], [])
    question = rows[0][3]
    assert question == "汇总勾选与底稿证据不一致（K.03.2/TOD）"
    assert rows[0][8] == "NEED_REVIEW"
    assert rows[0][9] == "AE-003"
    assert rows[0][10] == "execution_status_consistency"


def test_locator_rows_include_navigate_ref():
    issue = QcIssue(
        asset_id=None,
        rule_id="lead_required_fields",
        field="gaap",
        severity=Severity.FAIL,
        message="缺少适用会计准则",
        suggestion="补充",
        procedure_code="K.00",
        source_sheet="K.00 Lead Sheet",
        source_row=7,
        dict_rule_code="LEAD-001",
    )
    rows = build_locator_rows([issue])
    assert len(rows) == 1
    assert rows[0][2] == "LEAD-001"
    assert rows[0][4] == "$B$7"
    assert rows[0][7] == "K.00 Lead Sheet!B7"
    assert rows[0][8] == "规则判断"
