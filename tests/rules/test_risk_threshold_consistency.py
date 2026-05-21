from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from ingest.lead_sheet import load_lead_from_workbook
from rules.risk_threshold_consistency import check_risk_threshold_consistency


@pytest.fixture
def no_cra_lead_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "lead_no_cra.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["B2"] = "客户名称"
    ws["C2"] = "A公司"
    ws["B5"] = "可容忍误差（TE）"
    ws["C5"] = 1148000000
    ws["B26"] = "账户变更"
    ws["C26"] = "预期及额外考虑"
    ws["B37"] = "波动幅度 (CNY):"
    ws["C37"] = 1148000000
    wb.save(path)
    wb.close()
    return path


def test_no_cra_layout_skips_ae002(no_cra_lead_xlsx: Path):
    lead = load_lead_from_workbook(no_cra_lead_xlsx)
    assert lead.layout_variant == "no_cra_te_volatility"
    assert check_risk_threshold_consistency(lead) == []
