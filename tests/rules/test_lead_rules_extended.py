from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from ingest.lead_sheet import LeadMovementRow, LeadSheetDataset, load_lead_from_workbook
from ingest.models import RollforwardColumnBinding, RollforwardPeriodRole
from ingest.rollforward_sheet import RollforwardSheetDataset, parse_rollforward_rows
from rules.lead_expectation_analysis import check_lead_expectation_analysis
from rules.lead_movement_rows_complete import check_lead_movement_rows_complete
from rules.lead_rollforward_tb_reconciliation import (
    build_lead_rollforward_tb_reconciliation_observation,
    check_lead_rollforward_tb_reconciliation,
)
from rules.lead_runner import run_lead_rules
from rules.execution_recorder import RuleExecutionRecorder
from rules.lead_tt_gam_range import check_lead_tt_gam_range
from rules.lead_tt_overall_min import check_lead_tt_overall_min
from rules.lead_volatility_threshold_link import check_lead_volatility_threshold_link
from rules.models import Severity
from rules.registry import attach_rule_metadata
from rules.unexpected_movement_investigation import check_unexpected_movement_investigation


@pytest.fixture
def swp_lead_xlsx(tmp_path: Path) -> Path:
    """与 tests/ingest/test_lead_sheet.py 标准版式一致。"""
    path = tmp_path / "swp_lead.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["B2"] = "客户名称"
    ws["C2"] = "XYZ公司"
    ws["B3"] = "期末"
    ws["C3"] = datetime(2025, 12, 31)
    ws["B4"] = "分析日期"
    ws["C4"] = datetime(2026, 3, 1)
    ws["B5"] = "可容忍误差（TE）"
    ws["C5"] = 500000
    ws["B6"] = "名义金额（SAD）"
    ws["C6"] = 50000
    ws["B7"] = "适用会计准则（GAAP）"
    ws["C7"] = "CAS"
    ws["B8"] = "记账本位币"
    ws["C8"] = "CNY"
    ws["B14"] = "认定"
    ws["C14"] = "CRA"
    ws["D14"] = "各项认定1"
    ws["E14"] = "所有相关认定 2"
    ws["B15"] = "完整性（C）"
    ws["C15"] = "Minimal"
    ws["D15"] = 100000
    ws["E15"] = 100000
    ws["B16"] = "计价/计量（V/M）"
    ws["C16"] = "Low"
    ws["D16"] = 80000
    ws["B27"] = "账户变更"
    ws["C27"] = "预期及额外考虑"
    ws["B28"] = "新增"
    ws["C28"] = "无重大新增"
    ws["B37"] = "波动范围"
    ws["B38"] = "波动幅度 ():"
    ws["C38"] = 100000
    ws["B39"] = "波动幅度 (%):"
    ws["C39"] = 0.1
    ws["B48"] = "总账科目编码"
    ws["C48"] = "科目名称"
    ws["D48"] = "索引号"
    ws["E48"] = "期末账面数"
    ws["G48"] = "期末未审数"
    ws["I48"] = "期末审定数"
    ws["J48"] = "上期末审定数"
    ws["C49"] = "原值"
    ws["D49"] = "K.01"
    ws["E49"] = 1000
    ws["G49"] = 1000
    ws["I49"] = 1000
    ws["J49"] = 900
    ws["C50"] = "累计折旧"
    ws["D50"] = "K.01"
    ws["E50"] = 200
    ws["I50"] = 200
    ws["J50"] = 150
    ws["C51"] = "减值准备"
    ws["D51"] = "K.01"
    ws["E51"] = 0
    ws["I51"] = 0
    ws["J51"] = 0
    ws["C53"] = "净值"
    ws["I53"] = 800
    ws["J53"] = 750
    ws["B58"] = "波动说明"
    ws["B59"] = "无异常波动"
    wb.save(path)
    wb.close()
    return path


def test_tt_overall_min_fails_when_not_equal_min(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    issues = check_lead_tt_overall_min(lead)
    assert any(i.rule_id == "lead_tt_overall_min" and i.severity == Severity.FAIL for i in issues)


def test_tt_gam_range_warns_on_swp_fixture(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    issues = check_lead_tt_gam_range(lead)
    assert any(i.rule_id == "lead_tt_gam_range" and i.severity == Severity.WARN for i in issues)


def test_volatility_link_passes_when_amount_matches_overall_tt(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    issues = check_lead_volatility_threshold_link(lead)
    assert not issues


def test_movement_rows_complete_requires_impairment_row(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    issues = check_lead_movement_rows_complete(lead)
    assert not any(i.severity == Severity.FAIL for i in issues)
    assert not any(i.field == "py_audited" for i in issues)
    assert not any("sheet_ref" in (i.field or "") for i in issues)


def test_tt_gam_range_accepts_ratio_at_upper_bound():
    from ingest.lead_sheet import CraAssertionRow, LeadBasicInfoField, LeadSheetDataset
    from rules.lead_common import tt_ratio_within_gam_band

    assert tt_ratio_within_gam_band(
        Decimal("0.750000000000000035"),
        Decimal("0.50"),
        Decimal("0.75"),
    )
    lead = LeadSheetDataset(
        source_file="test.xlsx",
        source_sheet="K.00 Lead Sheet",
        basic_info_fields=[
            LeadBasicInfoField(field_key="te", label="TE", value="213730"),
        ],
        cra_rows=[
            CraAssertionRow(
                assertion="计价/计量（V/M）",
                cra="Low",
                tt="160297.5",
                source_row=16,
            )
        ],
    )
    issues = check_lead_tt_gam_range(lead)
    assert not any("计价" in i.message for i in issues)


def test_expectation_block_present(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    assert not check_lead_expectation_analysis(lead)


def test_ae004_trivial_note_when_threshold_exceeded(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    assert lead.volatility is not None
    lead.volatility.amount = "1"
    lead.fluctuation_notes = "无异常波动"
    orig = next(r for r in lead.movement_rows if r.account_label == "原值")
    orig.values["audited_ending"] = "500000"
    orig.values["py_audited"] = "0"
    issues = check_unexpected_movement_investigation(lead)
    assert any(i.rule_id == "unexpected_movement_investigation" for i in issues)


def test_ae004_threshold_requires_amount_and_percent(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    assert lead.volatility is not None
    lead.volatility.amount = "100000"
    lead.volatility.percent = "10%"
    lead.fluctuation_notes = "无异常波动"
    orig = next(r for r in lead.movement_rows if r.account_label == "原值")
    orig.values["movement_amount"] = "500000"
    orig.values["movement_pct"] = "5%"
    orig.values["investigate_quantitative"] = ""
    orig.values["investigate_qualitative"] = ""

    issues = check_unexpected_movement_investigation(lead)
    assert not any(i.rule_id == "unexpected_movement_investigation" for i in issues)


def test_ae004_threshold_triggers_when_amount_and_percent_exceed(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    assert lead.volatility is not None
    lead.volatility.amount = "100000"
    lead.volatility.percent = "10%"
    lead.fluctuation_notes = "无异常波动"
    orig = next(r for r in lead.movement_rows if r.account_label == "原值")
    orig.values["movement_amount"] = "500000"
    orig.values["movement_pct"] = "20%"
    orig.values["investigate_quantitative"] = ""
    orig.values["investigate_qualitative"] = ""

    issues = check_unexpected_movement_investigation(lead)
    assert any(i.rule_id == "unexpected_movement_investigation" for i in issues)


def test_rollforward_reconciliation_match(tmp_path: Path):
    path = tmp_path / "lead_rf.xlsx"
    wb = openpyxl.Workbook()
    ws_lead = wb.active
    ws_lead.title = "K.00 Lead Sheet"
    ws_lead["B5"] = "可容忍误差"
    ws_lead["C5"] = 100
    ws_lead["C40"] = "科目名称"
    ws_lead["E40"] = "期末账面数"
    ws_lead["C41"] = "原值"
    ws_lead["E41"] = 3000
    ws_lead["C42"] = "累计折旧"
    ws_lead["E42"] = 500
    ws_lead["C43"] = "减值准备"
    ws_lead["E43"] = 0
    ws_lead["C44"] = "净值"
    ws_lead["E44"] = 2500
    ws_rf = wb.create_sheet("K.01 Agree SL to GL")
    ws_rf.append(["类别", "原值", "累计折旧", "净值"])
    ws_rf.append(["合计", 3000, 500, 2500])
    wb.save(path)
    wb.close()

    lead = load_lead_from_workbook(path)
    import openpyxl as ox

    wb2 = ox.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb2["K.01 Agree SL to GL"].iter_rows(values_only=True))
    wb2.close()
    rf = parse_rollforward_rows(rows, source_sheet="K.01 Agree SL to GL")
    issues = check_lead_rollforward_tb_reconciliation(lead, rf)
    assert issues == []


def test_rollforward_reconciliation_mismatch(tmp_path: Path):
    path = tmp_path / "lead_rf_bad.xlsx"
    wb = openpyxl.Workbook()
    ws_lead = wb.active
    ws_lead.title = "K.00 Lead Sheet"
    ws_lead["C40"] = "科目名称"
    ws_lead["E40"] = "期末账面数"
    ws_lead["C41"] = "原值"
    ws_lead["E41"] = 9999
    ws_lead["C42"] = "累计折旧"
    ws_lead["E42"] = 1
    ws_lead["C43"] = "减值准备"
    ws_lead["E43"] = 0
    ws_lead["C44"] = "净值"
    ws_lead["E44"] = 1
    ws_rf = wb.create_sheet("K.01 Agree SL to GL")
    ws_rf.append(["类别", "原值", "累计折旧", "净值"])
    ws_rf.append(["合计", 3000, 500, 2500])
    wb.save(path)
    wb.close()

    lead = load_lead_from_workbook(path)
    import openpyxl as ox

    wb2 = ox.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb2["K.01 Agree SL to GL"].iter_rows(values_only=True))
    wb2.close()
    rf = parse_rollforward_rows(rows, source_sheet="K.01 Agree SL to GL")
    issues = check_lead_rollforward_tb_reconciliation(lead, rf)
    assert any(i.severity == Severity.FAIL for i in issues)


def test_rollforward_reconciliation_prefers_k01_check_column(tmp_path: Path):
    path = tmp_path / "lead_rf_k01_check.xlsx"
    wb = openpyxl.Workbook()
    ws_lead = wb.active
    ws_lead.title = "K.00 Lead Sheet"
    ws_lead["C40"] = "科目名称"
    ws_lead["E40"] = "期末账面数"
    ws_lead["C41"] = "原值"
    ws_lead["E41"] = 696
    ws_lead["C42"] = "累计折旧"
    ws_lead["E42"] = 134
    ws_lead["C43"] = "减值准备"
    ws_lead["E43"] = 0
    ws_lead["C44"] = "净值"
    ws_lead["E44"] = 562
    wb.save(path)
    wb.close()

    lead = load_lead_from_workbook(path)
    rf = RollforwardSheetDataset(
        source_file=str(path),
        source_sheet="K.01 Agree SL to GL",
        header_row=None,
        mapped_fields=[],
        ending_totals={
            "original_value": Decimal("694"),
            "accumulated_depreciation": Decimal("134"),
            "impairment_provision": Decimal("0"),
            "net_value": Decimal("560"),
        },
        table1_check_values={
            "original_value": Decimal("-2"),
            "accumulated_depreciation": Decimal("0"),
            "impairment_provision": Decimal("0"),
            "net_value": Decimal("-2"),
        },
        table1_check_rows={
            "original_value": 18,
            "accumulated_depreciation": 24,
            "impairment_provision": 30,
            "net_value": 32,
        },
    )

    issues = check_lead_rollforward_tb_reconciliation(lead, rf)
    assert len(issues) == 1
    assert issues[0].field == "期末|original_value"
    assert "并导致净值差异" in issues[0].message
    assert {i.source_sheet for i in issues} == {"K.01 Agree SL to GL"}
    assert {i.procedure_code for i in issues} == {"K.01"}
    assert {i.source_row for i in issues} == {18}

    observation = build_lead_rollforward_tb_reconciliation_observation(lead, rf)
    assert set(observation) == {
        "checked_data",
        "check_logic",
        "expected_result",
        "actual_result",
        "result_summary",
    }
    checked = observation["checked_data"][0]
    assert checked["sheet"] == "K.01 Agree SL to GL"
    assert checked["identified_by"]["section"] == "table1_check_values"
    assert checked["values_read"][0]["label"] == "原值 Check"
    assert checked["values_read"][0]["row"] == 18


def test_rollforward_reconciliation_checks_opening_and_merges_derived_net():
    lead = LeadSheetDataset(
        source_file="test.xlsx",
        source_sheet="K.00 Lead Sheet",
        movement_rows=[
            LeadMovementRow("原值", "K.01", {"py_audited": "1000", "audited_ending": "1100"}, 10),
            LeadMovementRow("累计折旧", "K.01", {"py_audited": "300", "audited_ending": "350"}, 11),
            LeadMovementRow("减值准备", "K.01", {"py_audited": "0", "audited_ending": "0"}, 12),
            LeadMovementRow("净值", None, {"py_audited": "700", "audited_ending": "750"}, 13),
        ],
    )
    rf = RollforwardSheetDataset(
        source_file="test.xlsx",
        source_sheet="K.01 Agree SL to GL",
        header_row=1,
        mapped_fields=[],
        amount_column_bindings=[
            RollforwardColumnBinding(
                measure="original_value",
                period_role=RollforwardPeriodRole.OPENING,
                source_header="期初原值",
                column_index=2,
            )
        ],
        opening_totals={
            "original_value": Decimal("1000"),
            "accumulated_depreciation": Decimal("290"),
            "impairment_provision": Decimal("0"),
            "net_value": Decimal("710"),
        },
        ending_totals={
            "original_value": Decimal("1100"),
            "accumulated_depreciation": Decimal("350"),
            "impairment_provision": Decimal("0"),
            "net_value": Decimal("750"),
        },
    )

    issues = check_lead_rollforward_tb_reconciliation(lead, rf)

    assert len(issues) == 1
    assert issues[0].field == "期初|accumulated_depreciation"
    assert "并导致净值差异" in issues[0].message


def test_run_lead_rules_attaches_metadata(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    issues = attach_rule_metadata(run_lead_rules(lead))
    codes = {i.dict_rule_code for i in issues if i.dict_rule_code}
    assert "LEAD-003" in codes or "AE-001" in codes


def test_run_lead_rules_records_required_fields_evidence_how(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    recorder = RuleExecutionRecorder()
    run_lead_rules(lead, recorder=recorder)
    item = {
        ledger_item["rule_id"]: ledger_item
        for ledger_item in recorder.to_ledger()["items"]
    }["lead_required_fields"]

    observation = item["observation"]
    assert set(observation) == {
        "checked_data",
        "check_logic",
        "expected_result",
        "actual_result",
        "result_summary",
    }
    checked = observation["checked_data"][0]
    assert checked["sheet"] == "K.00 Lead Sheet"
    assert checked["section"] == "K.00 Lead Sheet 基础信息"
    assert checked["values_read"][0]["label"] == "客户名称"
    assert checked["values_read"][0]["cell"] == "C2"


def test_run_lead_rules_records_ingest_readability_how_when_paused(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    lead.usable_for_rules = False
    recorder = RuleExecutionRecorder()
    run_lead_rules(lead, recorder=recorder)
    items = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}

    observation = items["lead_ingest_readability"]["observation"]
    assert observation["checked_data"][0]["section"] == "K.00 Lead Sheet 资料识别质量"
    assert items["lead_analysis_date_after_period_end"]["status"] == "DATA_INSUFFICIENT"
def test_run_lead_rules_records_parameter_rules_evidence_how(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    recorder = RuleExecutionRecorder()
    run_lead_rules(lead, recorder=recorder)
    items = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}

    for rule_id in (
        "lead_analysis_date_after_period_end",
        "materiality_consistency",
        "risk_threshold_consistency",
        "lead_tt_overall_min",
        "lead_tt_gam_range",
        "lead_volatility_threshold_link",
    ):
        observation = items[rule_id]["observation"]
        assert set(observation) == {
            "checked_data",
            "check_logic",
            "expected_result",
            "actual_result",
            "result_summary",
        }
        assert observation["checked_data"][0]["sheet"] == "K.00 Lead Sheet"
        assert observation["checked_data"][0]["values_read"]


def test_run_lead_rules_records_parameter_data_insufficient_how(swp_lead_xlsx: Path):
    lead = load_lead_from_workbook(swp_lead_xlsx)
    lead.usable_for_rules = False
    recorder = RuleExecutionRecorder()
    run_lead_rules(lead, recorder=recorder)
    items = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}

    for rule_id in (
        "lead_analysis_date_after_period_end",
        "materiality_consistency",
        "risk_threshold_consistency",
        "lead_tt_overall_min",
        "lead_tt_gam_range",
        "lead_volatility_threshold_link",
    ):
        observation = items[rule_id]["observation"]
        assert items[rule_id]["status"] == "DATA_INSUFFICIENT"
        assert observation["checked_data"][0]["section"] == "K.00 Lead Sheet 参数类规则执行前置条件"
        assert observation["checked_data"][0]["values_read"] == []
        assert observation["checked_data"][0]["missing_data"]
