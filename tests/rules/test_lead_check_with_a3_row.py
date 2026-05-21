from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from ingest.lead_sheet import load_lead_from_workbook
from rules.lead_check_with_a3_row import check_lead_check_with_a3_row
from rules.models import Severity


@pytest.fixture
def lead_with_a3_check(tmp_path: Path) -> Path:
    path = tmp_path / "lead_a3.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["B5"] = "可容忍误差"
    ws["C5"] = 100000
    ws["C48"] = "科目名称"
    ws["E48"] = "期末账面数"
    ws["I48"] = "期末审定数"
    ws["C49"] = "原值"
    ws["E49"] = 1000
    ws["I49"] = 1000
    ws["C50"] = "累计折旧"
    ws["E50"] = 200
    ws["I50"] = 200
    ws["C51"] = "减值准备"
    ws["E51"] = 0
    ws["I51"] = 0
    ws["C52"] = "净值"
    ws["E52"] = 800
    ws["I52"] = 800
    ws["C53"] = "Check with A3"
    ws["E53"] = 1000
    ws["I53"] = 1000
    ws["C54"] = "Diff"
    ws["E54"] = 0
    ws["I54"] = 0
    ws["C55"] = "Notes:"
    ws["D55"] = "无差异"
    wb.save(path)
    wb.close()
    return path


def test_ingest_extracts_check_with_a3(lead_with_a3_check: Path):
    lead = load_lead_from_workbook(lead_with_a3_check)
    assert lead.check_with_a3 is not None
    cw = lead.check_with_a3
    assert cw.check_source_row == 53
    assert cw.diff_source_row == 54
    assert len(cw.lines) >= 1
    orig = next(ln for ln in cw.lines if "原值" in ln.account_label)
    assert orig.movement_value == "1000"
    assert orig.a3_value == "1000"
    assert orig.diff_value == "0"


def test_diff_zero_passes(lead_with_a3_check: Path):
    lead = load_lead_from_workbook(lead_with_a3_check)
    issues = check_lead_check_with_a3_row(lead)
    assert not any(i.severity == Severity.FAIL for i in issues)


def test_nonzero_diff_without_notes_fails(tmp_path: Path):
    path = tmp_path / "lead_a3_bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["C48"] = "科目名称"
    ws["E48"] = "期末账面数"
    ws["I48"] = "期末审定数"
    ws["C49"] = "原值"
    ws["E49"] = 1000
    ws["I49"] = 1000
    ws["C50"] = "累计折旧"
    ws["I50"] = 200
    ws["C51"] = "减值准备"
    ws["I51"] = 0
    ws["C52"] = "净值"
    ws["I52"] = 800
    ws["C53"] = "Check with A3"
    ws["I53"] = 1000
    ws["C54"] = "Diff"
    ws["I54"] = 100
    wb.save(path)
    wb.close()

    lead = load_lead_from_workbook(path)
    issues = check_lead_check_with_a3_row(lead)
    assert any(
        i.rule_id == "lead_check_with_a3_row" and i.severity == Severity.FAIL
        for i in issues
    )
    assert any(i.field == "notes" for i in issues)


def test_missing_a3_rows_warns(tmp_path: Path):
    path = tmp_path / "lead_no_a3.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["C48"] = "科目名称"
    ws["E48"] = "期末账面数"
    ws["I48"] = "期末审定数"
    ws["C49"] = "原值"
    ws["I49"] = 1000
    ws["C50"] = "累计折旧"
    ws["I50"] = 200
    ws["C51"] = "减值准备"
    ws["I51"] = 0
    ws["C52"] = "净值"
    ws["I52"] = 800
    wb.save(path)
    wb.close()

    lead = load_lead_from_workbook(path)
    issues = check_lead_check_with_a3_row(lead)
    assert any(i.severity == Severity.WARN and i.field == "check_with_a3" for i in issues)
