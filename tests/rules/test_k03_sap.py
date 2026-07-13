from __future__ import annotations

from pathlib import Path

import openpyxl

from ingest.k03_sheet import (
    EXECUTION_PATH_SAP_HIGH,
    EXECUTION_PATH_SAP_MEDIUM,
    EXECUTION_PATH_SAP_PLUS_TOD_SAMPLING,
    EXECUTION_PATH_TOD_SAMPLING,
    K03ComponentSheet,
    K03ExecutionProfile,
    K03LeadLinkage,
    K03SheetDataset,
    load_k03_sheets_from_workbook,
)
from ingest.lead_sheet import LeadBasicInfoField, LeadSheetDataset, MaterialityCapture
from rules.execution_recorder import (
    STATUS_DATA_INSUFFICIENT,
    STATUS_EXECUTED,
    STATUS_NOT_APPLICABLE,
    RuleExecutionRecorder,
)
from rules.k03_sap import run_k03_sap_rules
from rules.models import Severity


def _lead(cra: str = "Minimal", te: str = "100") -> LeadSheetDataset:
    return LeadSheetDataset(
        source_file="test.xlsx",
        source_sheet="K.00 Lead Sheet",
        basic_info_fields=[
            LeadBasicInfoField(field_key="cra", label="CRA", value=cra, source_row=3, source_col=2),
            LeadBasicInfoField(field_key="te", label="TE", value=te, source_row=4, source_col=2),
        ],
        materiality=[
            MaterialityCapture(
                field_key="te",
                label="TE",
                workpaper_value=te,
                source_row=4,
                source_col_workpaper=2,
            )
        ],
    )


def _save_sap_workbook(path: Path, sheet_name: str, *, cra: str = "Minimal", te: int = 100, over: str = "否"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["客户名称", "测试公司", "实体类型", "非复杂实体"])
    ws.append(["年度", "2025", "可容忍误差(TE)", te])
    ws.append(["记账本位币", "CNY", "CRA", cra])
    ws.append([])
    ws.append(["预期", "根据资产类别、原值、使用寿命和残值率形成折旧预期"])
    ws.append(["精确度", "中精度说明，同时模板说明中提到高精确度适用场景"])
    ws.append(["细分", "按资产类别分解"])
    ws.append([])
    ws.append(["偏差阈值", 10])
    ws.append(["实际折旧费用", 95])
    ws.append(["偏差金额", 5])
    ws.append(["偏差是否超过阈值？", over])
    ws.append(["结论", "偏差未超过阈值"])
    wb.save(path)
    wb.close()
    return path


def _profile(
    dataset: K03SheetDataset,
    *,
    cra: str | None,
    with_tod: bool = False,
) -> K03ExecutionProfile:
    role = "sap_high" if dataset.execution_path == EXECUTION_PATH_SAP_HIGH else "sap_medium"
    components = {
        role: [
            K03ComponentSheet(
                role=role,
                sheet_name=dataset.sheet_name,
                execution_path=dataset.execution_path,
                template_type=dataset.template_type,
            )
        ]
    }
    primary_path = dataset.execution_path
    if with_tod:
        components["tod_sampling"] = [
            K03ComponentSheet(
                role="tod_sampling",
                sheet_name="K.03.2 TOD sampling",
                execution_path=EXECUTION_PATH_TOD_SAMPLING,
                template_type="tod_sampling",
            )
        ]
        primary_path = EXECUTION_PATH_SAP_PLUS_TOD_SAMPLING
    return K03ExecutionProfile(
        primary_depreciation_path=primary_path,
        executed_depreciation_paths=[
            dataset.execution_path,
            *([EXECUTION_PATH_TOD_SAMPLING] if with_tod else []),
        ],
        component_sheets=components,
        lead_linkage=K03LeadLinkage(
            assertion="V/M",
            cra=cra,
            source_row=25,
            cra_cell="C25",
            source="lead_vm_assertion",
        ),
    )


def _ledger_items(recorder: RuleExecutionRecorder) -> dict[str, dict]:
    return {item["rule_id"]: item for item in recorder.to_ledger()["items"]}


def test_sap_medium_precision_detects_from_sheet_name_even_when_high_text_exists(tmp_path: Path):
    path = _save_sap_workbook(tmp_path / "sap_medium.xlsx", "K.03.1 SAP-中精确度")

    dataset = load_k03_sheets_from_workbook(path)[0]

    assert dataset.execution_path == EXECUTION_PATH_SAP_MEDIUM
    assert dataset.template_type == "sap_medium_precision"
    assert dataset.unsupported_or_later_phase is False


def test_sap_high_precision_detects_from_sheet_name(tmp_path: Path):
    path = _save_sap_workbook(tmp_path / "sap_high.xlsx", "K.03.1 SAP-高精确度", cra="Moderate")

    dataset = load_k03_sheets_from_workbook(path)[0]

    assert dataset.execution_path == EXECUTION_PATH_SAP_HIGH
    assert dataset.template_type == "sap_high_precision"


def test_sap_medium_precision_non_minimal_without_tod_needs_review(tmp_path: Path):
    path = _save_sap_workbook(tmp_path / "sap_medium_non_minimal.xlsx", "K.03.1 SAP-中精确度", cra="Moderate")
    dataset = load_k03_sheets_from_workbook(path)[0]

    issues = run_k03_sap_rules(
        dataset,
        lead=_lead(cra="Moderate"),
        k03_sheets=[dataset],
        k03_execution_profile=_profile(dataset, cra="Moderate"),
    )

    assert any(
        issue.rule_id == "sap_precision_selection"
        and issue.severity == Severity.NEED_REVIEW
        for issue in issues
    )


def test_sap_medium_uses_lead_vm_cra_not_template_minimal(tmp_path: Path):
    path = _save_sap_workbook(
        tmp_path / "sap_medium_preset_minimal.xlsx",
        "K.03.1 SAP-中精确度",
        cra="Minimal",
    )
    dataset = load_k03_sheets_from_workbook(path)[0]

    issues = run_k03_sap_rules(
        dataset,
        lead=_lead(cra="Moderate"),
        k03_execution_profile=_profile(dataset, cra="Moderate"),
    )

    assert any(issue.rule_id == "sap_precision_selection" for issue in issues)
    assert not any(issue.rule_id == "sap_high_cra_consistency" for issue in issues)


def test_sap_medium_non_minimal_with_tod_is_allowed(tmp_path: Path):
    path = _save_sap_workbook(
        tmp_path / "sap_medium_with_tod.xlsx",
        "K.03.1 SAP-中精确度",
        cra="Minimal",
    )
    dataset = load_k03_sheets_from_workbook(path)[0]

    issues = run_k03_sap_rules(
        dataset,
        lead=_lead(cra="Moderate"),
        k03_execution_profile=_profile(dataset, cra="Moderate", with_tod=True),
    )

    assert not any(issue.rule_id == "sap_precision_selection" for issue in issues)


def test_sap_te_consistency_has_independent_rule_and_ledger_status(tmp_path: Path):
    path = _save_sap_workbook(
        tmp_path / "sap_bad_te.xlsx",
        "K.03.1 SAP-高精确度",
        cra="Moderate",
        te=90,
    )
    dataset = load_k03_sheets_from_workbook(path)[0]
    recorder = RuleExecutionRecorder()

    issues = run_k03_sap_rules(
        dataset,
        lead=_lead(cra="Moderate", te="100"),
        k03_execution_profile=_profile(dataset, cra="Moderate"),
        recorder=recorder,
    )

    assert any(
        issue.rule_id == "sap_te_consistency" and issue.severity == Severity.FAIL
        for issue in issues
    )
    assert _ledger_items(recorder)["sap_te_consistency"]["status"] == STATUS_EXECUTED


def test_sap_high_cra_consistency_compares_lead_vm_cra(tmp_path: Path):
    path = _save_sap_workbook(
        tmp_path / "sap_bad_cra.xlsx",
        "K.03.1 SAP-高精确度",
        cra="Moderate",
    )
    dataset = load_k03_sheets_from_workbook(path)[0]

    issues = run_k03_sap_rules(
        dataset,
        lead=_lead(cra="Low"),
        k03_execution_profile=_profile(dataset, cra="Low"),
    )

    assert any(
        issue.rule_id == "sap_high_cra_consistency" and issue.severity == Severity.FAIL
        for issue in issues
    )


def test_sap_missing_inputs_are_not_recorded_as_executed(tmp_path: Path):
    path = _save_sap_workbook(
        tmp_path / "sap_missing_inputs.xlsx",
        "K.03.1 SAP-高精确度",
        cra="Moderate",
    )
    dataset = load_k03_sheets_from_workbook(path)[0]
    dataset.summary.pop("sap_te", None)
    dataset.summary.pop("sap_cra", None)
    recorder = RuleExecutionRecorder()

    run_k03_sap_rules(
        dataset,
        lead=_lead(cra="Moderate"),
        k03_execution_profile=_profile(dataset, cra=None),
        recorder=recorder,
    )

    ledger = _ledger_items(recorder)
    assert ledger["sap_precision_selection"]["status"] == STATUS_DATA_INSUFFICIENT
    assert ledger["sap_te_consistency"]["status"] == STATUS_DATA_INSUFFICIENT
    assert ledger["sap_high_cra_consistency"]["status"] == STATUS_DATA_INSUFFICIENT


def test_sap_high_cra_rule_is_not_applicable_to_medium_sap(tmp_path: Path):
    path = _save_sap_workbook(
        tmp_path / "sap_medium_na.xlsx",
        "K.03.1 SAP-中精确度",
    )
    dataset = load_k03_sheets_from_workbook(path)[0]
    recorder = RuleExecutionRecorder()

    run_k03_sap_rules(
        dataset,
        lead=_lead(),
        k03_execution_profile=_profile(dataset, cra="Minimal"),
        recorder=recorder,
    )

    assert (
        _ledger_items(recorder)["sap_high_cra_consistency"]["status"]
        == STATUS_NOT_APPLICABLE
    )


def test_sap_deviation_over_threshold_without_note_fails(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.1 SAP-高精确度"
    ws.append(["可容忍误差(TE)", 100, "CRA", "Moderate"])
    ws.append(["预期", "根据资产信息形成折旧预期"])
    ws.append(["偏差阈值", 10])
    ws.append(["实际折旧费用", 95])
    ws.append(["偏差金额", 20])
    ws.append(["偏差是否超过阈值？", "是"])
    path = tmp_path / "sap_over.xlsx"
    wb.save(path)
    wb.close()
    dataset = load_k03_sheets_from_workbook(path)[0]

    issues = run_k03_sap_rules(
        dataset,
        lead=_lead(cra="Moderate"),
        k03_sheets=[dataset],
        k03_execution_profile=_profile(dataset, cra="Moderate"),
    )

    assert any(
        issue.rule_id == "sap_depreciation_difference"
        and issue.severity == Severity.FAIL
        for issue in issues
    )


def _save_medium_category_workbook(
    path: Path, *, include_total_note: bool = True, machine_note: str = "机器设备本期新增较多，已完成进一步量化。"
) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.1 SAP-中精确度"
    ws.append(["CRA", "Minimal"])
    ws.append(["预期", "按类别建立折旧预期"])
    ws.append([])
    ws.append(["资产类别", "房屋及建筑物", "机器设备", "total"])
    ws.append([])
    ws.append(["偏差阈值 c", 10, 10, 20])
    ws.append([])
    ws.append([])
    ws.append(["偏差金额 h=g-f", 5, -15, 25])
    ws.append([])
    ws.append(["偏差是否超过阈值 h vs c", "否", "是", "是"])
    ws.append([None, None, "NB1", "NB10"])
    ws.append([])
    ws.append(["Notes"])
    ws.append(["NB1", machine_note])
    if include_total_note:
        ws.append(["NB10", "合计偏差主要来自机器设备，已完成进一步量化。"])
    wb.save(path)
    wb.close()
    return path


def test_sap_medium_category_and_total_require_their_own_notes(tmp_path: Path):
    dataset = load_k03_sheets_from_workbook(
        _save_medium_category_workbook(tmp_path / "sap_medium_categories.xlsx")
    )[0]

    issues = run_k03_sap_rules(
        dataset,
        lead=_lead(),
        k03_execution_profile=_profile(dataset, cra="Minimal"),
    )

    category_issues = [
        issue for issue in issues if issue.rule_id == "sap_medium_category_deviation_explanation"
    ]
    assert len(category_issues) == 2
    assert all(issue.severity == Severity.NEED_REVIEW for issue in category_issues)
    assert any("机器设备" in issue.message for issue in category_issues)
    assert any("total" in issue.message for issue in category_issues)


def test_sap_medium_total_without_its_own_note_fails(tmp_path: Path):
    dataset = load_k03_sheets_from_workbook(
        _save_medium_category_workbook(tmp_path / "sap_medium_missing_total_note.xlsx", include_total_note=False)
    )[0]

    issues = run_k03_sap_rules(
        dataset,
        lead=_lead(),
        k03_execution_profile=_profile(dataset, cra="Minimal"),
    )

    assert any(
        issue.rule_id == "sap_medium_category_deviation_explanation"
        and issue.severity == Severity.FAIL
        and "total" in issue.message
        for issue in issues
    )


def test_sap_medium_placeholder_note_is_not_an_explanation(tmp_path: Path):
    dataset = load_k03_sheets_from_workbook(
        _save_medium_category_workbook(
            tmp_path / "sap_medium_placeholder_note.xlsx", machine_note="待补"
        )
    )[0]
    issues = run_k03_sap_rules(
        dataset, lead=_lead(), k03_execution_profile=_profile(dataset, cra="Minimal")
    )

    assert any(
        issue.rule_id == "sap_medium_category_deviation_explanation"
        and issue.severity == Severity.FAIL
        and "机器设备" in issue.message
        for issue in issues
    )


def test_sap_high_note_references_match_exact_marker_not_prefix(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.1 SAP-高精确度"
    ws.append(["CRA", "Moderate"])
    ws.append(["预期", "按类别建立折旧预期"])
    ws.append([])
    ws.append(["资产类别", "差异", "已分配偏差阈值", "差异是否超过已分配偏差阈值", "说明索引"])
    ws.append(["机器设备", -15, 10, "是", "NB1"])
    ws.append(["电子设备", 20, 10, "是", "NB10"])
    ws.append(["Notes"])
    ws.append(["NB1", "机器设备差异已完成进一步量化。"])
    ws.append(["NB10", "电子设备差异已完成进一步量化。"])
    path = tmp_path / "sap_high_markers.xlsx"
    wb.save(path)
    wb.close()
    dataset = load_k03_sheets_from_workbook(path)[0]
    items = dataset.summary["sap_high_deviation_items"]
    assert items[0]["note_reference"] == "NB1"
    assert items[0]["matched_note"]["marker"] == "NB1"
    assert items[1]["note_reference"] == "NB10"
    assert items[1]["matched_note"]["marker"] == "NB10"

    issues = run_k03_sap_rules(
        dataset,
        lead=_lead(cra="Moderate"),
        k03_execution_profile=_profile(dataset, cra="Moderate"),
    )

    category_issues = [
        issue for issue in issues if issue.rule_id == "sap_high_category_deviation_explanation"
    ]
    assert len(category_issues) == 2
    assert all(issue.severity == Severity.NEED_REVIEW for issue in category_issues)


def test_sap_high_missing_threshold_is_data_insufficient(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.1 SAP-高精确度"
    ws.append(["CRA", "Moderate"])
    ws.append(["预期", "按类别建立折旧预期"])
    ws.append(["资产类别", "差异", "差异是否超过已分配偏差阈值"])
    ws.append(["机器设备", "=20-5", "是"])
    path = tmp_path / "sap_high_missing_threshold.xlsx"
    wb.save(path)
    wb.close()
    dataset = load_k03_sheets_from_workbook(path)[0]
    recorder = RuleExecutionRecorder()

    run_k03_sap_rules(
        dataset,
        lead=_lead(cra="Moderate"),
        k03_execution_profile=_profile(dataset, cra="Moderate"),
        recorder=recorder,
    )

    assert _ledger_items(recorder)["sap_high_category_deviation_explanation"]["status"] == STATUS_DATA_INSUFFICIENT


def test_sap_high_uncalculated_formula_threshold_is_data_insufficient(tmp_path: Path):
    """openpyxl data_only reads an uncalculated Excel formula as None, never as zero."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.1 SAP-高精确度"
    ws.append(["CRA", "Moderate"])
    ws.append(["预期", "按类别建立折旧预期"])
    ws.append(["资产类别", "差异", "已分配偏差阈值", "差异是否超过已分配偏差阈值"])
    ws.append(["机器设备", 20, "=10", "是"])
    path = tmp_path / "sap_high_uncalculated_formula.xlsx"
    wb.save(path)
    wb.close()
    dataset = load_k03_sheets_from_workbook(path)[0]
    recorder = RuleExecutionRecorder()

    run_k03_sap_rules(
        dataset,
        lead=_lead(cra="Moderate"),
        k03_execution_profile=_profile(dataset, cra="Moderate"),
        recorder=recorder,
    )

    assert _ledger_items(recorder)["sap_high_category_deviation_explanation"]["status"] == STATUS_DATA_INSUFFICIENT


def test_sap_high_blank_over_threshold_flag_is_data_insufficient(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.1 SAP-高精确度"
    ws.append(["CRA", "Moderate"])
    ws.append(["预期", "按类别建立折旧预期"])
    ws.append(["资产类别", "差异", "已分配偏差阈值", "差异是否超过已分配偏差阈值"])
    ws.append(["机器设备", 20, 10, None])
    path = tmp_path / "sap_high_blank_flag.xlsx"
    wb.save(path)
    wb.close()
    dataset = load_k03_sheets_from_workbook(path)[0]
    recorder = RuleExecutionRecorder()

    issues = run_k03_sap_rules(
        dataset,
        lead=_lead(cra="Moderate"),
        k03_execution_profile=_profile(dataset, cra="Moderate"),
        recorder=recorder,
    )

    assert not any(issue.rule_id == "sap_high_category_deviation_explanation" for issue in issues)
    assert _ledger_items(recorder)["sap_high_category_deviation_explanation"]["status"] == STATUS_DATA_INSUFFICIENT
