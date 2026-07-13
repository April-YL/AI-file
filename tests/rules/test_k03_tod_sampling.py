from __future__ import annotations

from pathlib import Path
from decimal import Decimal
from types import SimpleNamespace

import openpyxl

from ingest.k03_sheet import EXECUTION_PATH_TOD_SAMPLING, load_k03_sheets_from_workbook
from ingest.lead_sheet import LeadBasicInfoField, LeadSheetDataset, MaterialityCapture
from rules.execution_recorder import RuleExecutionRecorder, STATUS_DATA_INSUFFICIENT
from rules.k03_tod_sampling import RULE_IDS, run_k03_tod_sampling_rules
from rules.models import Severity


def _lead(te: str = "100", sad: str = "5") -> LeadSheetDataset:
    return LeadSheetDataset(
        source_file="test.xlsx",
        source_sheet="K.00 Lead Sheet",
        basic_info_fields=[
            LeadBasicInfoField(field_key="te", label="TE", value=te, source_row=3, source_col=2),
            LeadBasicInfoField(field_key="sad", label="SAD", value=sad, source_row=4, source_col=2),
        ],
        materiality=[
            MaterialityCapture(
                field_key="te",
                label="TE",
                workpaper_value=te,
                source_row=3,
                source_col_workpaper=2,
            )
        ],
    )


def _save_sampling_workbook(path: Path, *, conclusion: str | None = "样本测试未见重大差异") -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.2 折旧测试TOD-抽样"
    ws.append(["折旧费用总体", 1000])
    ws.append(["Breakdown中折旧计提金额", 1000])
    ws.append(["减：测试的关键项目", 100])
    ws.append(["剩余总体", 900])
    ws.append(["关键项目", "选择大额资产作为关键项目"])
    ws.append([])
    ws.append(["样本数量", "样本类型", "固定资产类别", "固定资产编号", "固定资产名称", "原值", "本期计提折旧", "重新计算折旧费用", "差异", "获得的证据/支持的描述", "1", "2", "Notes"])
    ws.append([1, "关键项", "设备", "FA-TEST-001", "设备A", 1000, 100, 100, 0, "折旧政策", "Y", "Y", None])
    if conclusion is not None:
        ws.append(["结论", conclusion])
    out = wb.create_sheet("K.03.2a 折旧选样输出")
    out.append(["可容忍误差", 100])
    out.append(["抽样货币单元", "本期计提折旧"])
    out.append(["总体金额", 1000])
    out.append(["关键项目数量", 1])
    out.append(["关键项目金额", 100])
    out.append(["双重目的", "否"])
    out.append(["高估", "是"])
    out.append(["保证水平", "几乎没有"])
    out.append(["样本池", 900])
    out.append(["预期错报", "否"])
    out.append(["抽样方法", "随机抽样 (Random)"])
    out.append([])
    out.append(["样本类型", "固定资产编号", "固定资产名称", "本期计提折旧"])
    out.append(["关键项", "FA-TEST-001", "设备A", 100])
    wb.save(path)
    wb.close()
    return path


def _datasets(path: Path):
    datasets = load_k03_sheets_from_workbook(path)
    return (
        next(item for item in datasets if item.template_type == "tod_sampling"),
        next(item for item in datasets if item.template_type == "tod_sampling_output"),
    )


def _rollforward(amount: str):
    return SimpleNamespace(
        movement_transactions=[SimpleNamespace(
            transaction_key="depreciation",
            measure="accumulated_depreciation",
            amount=Decimal(amount),
            source_row=20,
        )],
        table4_rollforward_depreciation=None,
        table4_rollforward_depreciation_row=None,
    )


def test_tod_sampling_content_identifies_main_and_output_roles(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_sampling.xlsx")

    datasets = load_k03_sheets_from_workbook(path)
    main = next(item for item in datasets if item.sheet_name == "K.03.2 折旧测试TOD-抽样")

    assert main.execution_path == EXECUTION_PATH_TOD_SAMPLING
    assert main.template_type == "tod_sampling"


def test_tod_sampling_wrong_sampling_currency_fails(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_bad_currency.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["K.03.2a 折旧选样输出"]
    ws["B2"] = "原值"
    wb.save(path)
    wb.close()
    datasets = load_k03_sheets_from_workbook(path)
    main = next(item for item in datasets if item.template_type == "tod_sampling")
    output = next(item for item in datasets if item.template_type == "tod_sampling_output")

    issues = run_k03_tod_sampling_rules(main, sample_output=output, lead=_lead())

    assert any(
        issue.rule_id == "k03_tod_sampling_currency"
        and issue.severity == Severity.FAIL
        for issue in issues
    )


def test_tod_sampling_missing_conclusion_needs_review(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_no_conclusion.xlsx", conclusion=None)
    datasets = load_k03_sheets_from_workbook(path)
    main = next(item for item in datasets if item.template_type == "tod_sampling")
    output = next(item for item in datasets if item.template_type == "tod_sampling_output")

    issues = run_k03_tod_sampling_rules(main, sample_output=output, lead=_lead())

    assert any(
        issue.rule_id == "k03_tod_sampling_documentation"
        and issue.severity == Severity.NEED_REVIEW
        for issue in issues
    )


def test_tod_sampling_dynamic_shift_reorder_and_sheet_rename(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_dynamic.xlsx")
    wb = openpyxl.load_workbook(path)
    main = wb["K.03.2 折旧测试TOD-抽样"]
    output = wb["K.03.2a 折旧选样输出"]
    main.insert_rows(1, 4)
    main.insert_cols(1, 3)
    output.insert_rows(1, 6)
    output.insert_cols(1, 2)
    header_row = next(
        row for row in range(1, output.max_row + 1)
        if any(output.cell(row, col).value == "样本类型" for col in range(1, output.max_column + 1))
    )
    asset_col = next(col for col in range(1, output.max_column + 1) if output.cell(header_row, col).value == "固定资产编号")
    depreciation_col = next(col for col in range(1, output.max_column + 1) if output.cell(header_row, col).value == "本期计提折旧")
    for row in range(header_row, output.max_row + 1):
        output.cell(row, asset_col).value, output.cell(row, depreciation_col).value = (
            output.cell(row, depreciation_col).value,
            output.cell(row, asset_col).value,
        )
    main.title = "折旧详细测试_项目自定义"
    output.title = "抽样工具结果_项目自定义"
    wb.save(path)
    wb.close()

    main_ds, output_ds = _datasets(path)

    assert main_ds.summary["tod_sampling_detail_state"] == "FOUND"
    assert output_ds.summary["sample_output_detail_state"] == "FOUND"
    assert output_ds.summary["sample_output_rows"][0]["asset_id"] == "FA-TEST-001"


def test_tod_sampling_duplicate_asset_id_fails(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_duplicate.xlsx")
    wb = openpyxl.load_workbook(path)
    output = wb["K.03.2a 折旧选样输出"]
    output.append(["代表性样本", "FA-TEST-001", "设备A", 100])
    wb.save(path)
    wb.close()
    main, output = _datasets(path)

    issues = run_k03_tod_sampling_rules(main, sample_output=output, lead=_lead())

    assert any(issue.rule_id == "k03_tod_sampling_identity_consistency" and issue.severity == Severity.FAIL for issue in issues)


def test_tod_sampling_missing_sad_is_data_insufficient(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_missing_sad.xlsx")
    main, output = _datasets(path)
    recorder = RuleExecutionRecorder()
    lead = _lead(sad="")

    run_k03_tod_sampling_rules(main, sample_output=output, lead=lead, recorder=recorder)

    ledger = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}
    assert ledger["k03_tod_sampling_population_reconciliation"]["status"] == STATUS_DATA_INSUFFICIENT


def test_tod_sampling_te_mismatch_fails(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_te_mismatch.xlsx")
    main, output = _datasets(path)

    issues = run_k03_tod_sampling_rules(main, sample_output=output, lead=_lead(te="200"))

    assert any(issue.rule_id == "k03_tod_sampling_te_consistency" and issue.severity == Severity.FAIL for issue in issues)


def test_tod_sampling_population_over_sad_without_followup_fails(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_population_fail.xlsx", conclusion=None)
    main, output = _datasets(path)

    issues = run_k03_tod_sampling_rules(
        main, sample_output=output, lead=_lead(sad="5"), rollforward=_rollforward("990")
    )

    assert any(issue.rule_id == "k03_tod_sampling_population_reconciliation" and issue.severity == Severity.FAIL for issue in issues)


def test_tod_sampling_nonzero_difference_requires_row_note(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_difference_fail.xlsx", conclusion=None)
    wb = openpyxl.load_workbook(path)
    ws = wb["K.03.2 折旧测试TOD-抽样"]
    header_row = next(row for row in range(1, ws.max_row + 1) if any(ws.cell(row, col).value == "固定资产编号" for col in range(1, ws.max_column + 1)))
    difference_col = next(col for col in range(1, ws.max_column + 1) if ws.cell(header_row, col).value == "差异")
    ws.cell(header_row + 1, difference_col).value = 10
    wb.save(path)
    wb.close()
    main, output = _datasets(path)

    issues = run_k03_tod_sampling_rules(main, sample_output=output, lead=_lead())

    assert any(issue.rule_id == "k03_tod_sampling_difference_followup" and issue.severity == Severity.FAIL for issue in issues)


def test_tod_sampling_nonzero_difference_with_row_note_needs_review(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_difference_review.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["K.03.2 折旧测试TOD-抽样"]
    header_row = next(row for row in range(1, ws.max_row + 1) if any(ws.cell(row, col).value == "固定资产编号" for col in range(1, ws.max_column + 1)))
    difference_col = next(col for col in range(1, ws.max_column + 1) if ws.cell(header_row, col).value == "差异")
    note_col = next(col for col in range(1, ws.max_column + 1) if ws.cell(header_row, col).value == "Notes")
    ws.cell(header_row + 1, difference_col).value = 10
    ws.cell(header_row + 1, note_col).value = "客户折旧月份差异，已进一步调查"
    wb.save(path)
    wb.close()
    main, output = _datasets(path)

    issues = run_k03_tod_sampling_rules(main, sample_output=output, lead=_lead())

    assert any(issue.rule_id == "k03_tod_sampling_difference_followup" and issue.severity == Severity.NEED_REVIEW for issue in issues)


def test_tod_sampling_replacement_sample_is_traced_for_review(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_replacement.xlsx")
    wb = openpyxl.load_workbook(path)
    main = wb["K.03.2 折旧测试TOD-抽样"]
    output = wb["K.03.2a 折旧选样输出"]
    main_header = next(row for row in range(1, main.max_row + 1) if any(main.cell(row, col).value == "固定资产编号" for col in range(1, main.max_column + 1)))
    main_asset_col = next(col for col in range(1, main.max_column + 1) if main.cell(main_header, col).value == "固定资产编号")
    main.cell(main_header + 1, main_asset_col).value = "FA-TEST-R01"
    output.append(["替换样本", "FA-TEST-R01", "设备R", 100])
    wb.save(path)
    wb.close()
    main_ds, output_ds = _datasets(path)

    issues = run_k03_tod_sampling_rules(main_ds, sample_output=output_ds, lead=_lead())

    replacement_issues = [issue for issue in issues if issue.rule_id == "k03_tod_sampling_identity_consistency"]
    assert any(issue.severity == Severity.NEED_REVIEW for issue in replacement_issues)
    assert not any(issue.severity == Severity.FAIL for issue in replacement_issues)


def test_tod_sampling_observation_contains_main_and_output_sheets(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_observation.xlsx")
    main, output = _datasets(path)
    recorder = RuleExecutionRecorder()

    run_k03_tod_sampling_rules(
        main,
        sample_output=output,
        lead=_lead(),
        rollforward=_rollforward("1000"),
        recorder=recorder,
    )

    ledger = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}
    checked = ledger["k03_tod_sampling_identity_consistency"]["observation"]["checked_data"]
    assert {item["sheet"] for item in checked} == {main.sheet_name, output.sheet_name}
    output_values = next(item for item in checked if item["sheet"] == output.sheet_name)["values_read"]
    assert {
        "选样输出遗漏编号",
        "主测试额外编号",
        "重复资产编号",
        "实际使用的替换样本",
    }.issubset({item["label"] for item in output_values})


def test_tod_sampling_runner_records_every_rule_in_execution_ledger(tmp_path: Path):
    path = _save_sampling_workbook(tmp_path / "tod_ledger.xlsx")
    main, output = _datasets(path)
    recorder = RuleExecutionRecorder()

    run_k03_tod_sampling_rules(
        main,
        sample_output=output,
        lead=_lead(),
        rollforward=_rollforward("1000"),
        recorder=recorder,
    )

    ledger = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}
    assert set(ledger) == set(RULE_IDS)
    for rule_id in RULE_IDS[2:]:
        assert ledger[rule_id]["status"] == "EXECUTED"
        assert ledger[rule_id]["observation"]["checked_data"]
