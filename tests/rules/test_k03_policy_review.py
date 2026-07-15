from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl

from ingest.k03_sheet import (
    EXECUTION_PATH_POLICY_REVIEW,
    INGEST_DEPTH_LIGHTWEIGHT,
    K03Area,
    K03Column,
    K03PolicyRow,
    K03PolicyTable,
    K03SheetDataset,
    K03_BRANCH_POLICY_REVIEW,
    RULE_STATUS_LATER_PHASE,
    load_k03_sheets_from_workbook,
)
from ingest.models import AssetRecord
from ingest.records import FaListDataset
from rules.k03_policy_review import (
    RULE_IDS,
    parse_life_range,
    parse_rate,
    parse_rate_range,
    run_k03_policy_review_rules,
)
from rules.execution_recorder import (
    STATUS_DATA_INSUFFICIENT,
    STATUS_EXECUTED,
    RuleExecutionRecorder,
    validate_execution_ledger,
)
from rules.models import Severity


def _policy_dataset(
    rows: list[K03PolicyRow],
    *,
    notes: str | None = "政策未见异常。",
    column_map: dict[str, K03Column] | None = None,
) -> K03SheetDataset:
    if column_map is None:
        column_map = {
            "asset_category": K03Column("固定资产类别", 1, "A", "asset_category"),
            "current_useful_life": K03Column("本期使用寿命", 5, "E", "current_useful_life"),
            "current_salvage_rate": K03Column("本期残值率", 6, "F", "current_salvage_rate"),
            "prior_useful_life": K03Column("上期使用寿命", 9, "I", "prior_useful_life"),
            "prior_salvage_rate": K03Column("上期残值率", 10, "J", "prior_salvage_rate"),
            "useful_life_same_marker": K03Column("使用寿命", 12, "L", "useful_life_same_marker"),
            "salvage_rate_same_marker": K03Column("残值率", 13, "M", "salvage_rate_same_marker"),
            "difference_explanation": K03Column("差异说明", 14, "N", "difference_explanation"),
        }
    return K03SheetDataset(
        workbook_name="policy.xlsx",
        source_file="policy.xlsx",
        sheet_name="K.03.3 折旧政策复核",
        k03_branch=K03_BRANCH_POLICY_REVIEW,
        execution_path=EXECUTION_PATH_POLICY_REVIEW,
        template_type="policy_review",
        ingest_depth=INGEST_DEPTH_LIGHTWEIGHT,
        rule_status=RULE_STATUS_LATER_PHASE,
        policy_table=K03PolicyTable(
            range=K03Area(start_row=4, end_row=4 + len(rows), start_col=1, end_col=14),
            header_row=4,
            column_map=column_map,
            rows=rows,
        ),
        note_area=(
            K03Area(start_row=10, end_row=10, start_col=1, end_col=2, text=notes)
            if notes is not None
            else None
        ),
    )


def _row(
    *,
    category: str = "机器设备",
    current_life: str = "5-10年",
    prior_life: str = "5-10年",
    current_salvage: str = "5%",
    prior_salvage: str = "5%",
    life_marker: str = "TRUE",
    salvage_marker: str = "TRUE",
    explanation: str | None = None,
) -> K03PolicyRow:
    return K03PolicyRow(
        source_row=5,
        asset_category=category,
        current_method="年限平均法",
        current_useful_life=current_life,
        current_salvage_rate=current_salvage,
        current_annual_rate="10%",
        prior_method="年限平均法",
        prior_useful_life=prior_life,
        prior_salvage_rate=prior_salvage,
        prior_annual_rate="10%",
        useful_life_same_marker=life_marker,
        salvage_rate_same_marker=salvage_marker,
        difference_explanation=explanation,
        cell_refs={"current_useful_life": "E5", "difference_explanation": "N5"},
    )


def _fa_list(records: list[AssetRecord]) -> FaListDataset:
    return FaListDataset(
        source_file="policy.xlsx",
        source_sheet="FA list",
        mapped_fields=[],
        records=records,
    )


def test_parse_policy_life_range_and_rate_formats():
    assert parse_life_range("5-10年", assume_number_unit=None).min_months == Decimal("60")
    assert parse_life_range("10年-30年", assume_number_unit=None).max_months == Decimal("360")
    assert parse_life_range("60月", assume_number_unit=None).min_months == Decimal("60")
    assert parse_life_range("60", assume_number_unit="month").min_months == Decimal("60")
    assert parse_life_range("60", assume_number_unit=None) is None

    assert parse_rate("5%") == Decimal("0.05")
    assert parse_rate("0.05") == Decimal("0.05")
    assert parse_rate("5") == Decimal("0.05")
    assert parse_rate("105%") is None

    assert parse_rate_range("5%").min_rate == Decimal("0.05")
    assert parse_rate_range("0%-5%").max_rate == Decimal("0.05")
    assert parse_rate_range("2%-4%").min_rate == Decimal("0.02")


def test_policy_review_ingest_reads_sop_context_and_note_body_without_right_side_override(
    tmp_path: Path,
):
    path = tmp_path / "policy_sop.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.3 折旧政策复核"
    ws["B2"] = "K.03.3 折旧政策复核"
    ws["B9"] = "表1"
    ws["B10"] = "公司折旧政策"
    ws["C10"] = "2025-12-31"
    ws["F10"] = "2024-12-31"
    ws["I10"] = "差异"
    ws["B11"] = "折旧方法"
    ws["C11"] = "年限平均法"
    ws["F11"] = "年限平均法"
    ws["I11"] = True
    ws["S11"] = "右侧操作指引：使用寿命差异说明，不属于主表。"
    ws.merge_cells("S11:S23")
    ws["B12"] = "折旧政策"
    ws["C12"] = "使用寿命"
    ws["D12"] = "残值率"
    ws["E12"] = "年折旧率"
    ws["F12"] = "使用寿命"
    ws["G12"] = "残值率"
    ws["H12"] = "年折旧率"
    ws["I12"] = "使用寿命"
    ws["J12"] = "残值率"
    ws["K12"] = "差异说明"
    ws["B13"] = "机器设备"
    ws["C13"] = "5-10年"
    ws["D13"] = "0%-5%"
    ws["E13"] = "9.5%-19%"
    ws["F13"] = "5-10年"
    ws["G13"] = "0%-5%"
    ws["H13"] = "9.5%-19%"
    ws["I13"] = True
    ws["J13"] = True
    ws["B24"] = "Notes:"
    ws.merge_cells("B25:H25")
    ws["B25"] = "政策正文完整读取。"
    wb.save(path)
    wb.close()

    ds = load_k03_sheets_from_workbook(path)[0]
    table = ds.policy_table

    assert table is not None
    assert table.header_row == 12
    assert table.column_map["asset_category"].column_letter == "B"
    assert table.column_map["difference_explanation"].column_letter == "K"
    assert table.column_map["useful_life_same_marker"].column_letter == "I"
    assert table.column_map["salvage_rate_same_marker"].column_letter == "J"
    assert table.range is not None and table.range.end_col == 11
    assert all(column.column_index <= 11 for column in table.column_map.values())
    assert table.current_policy_date == "2025-12-31"
    assert table.prior_policy_date == "2024-12-31"
    assert table.current_method == "年限平均法"
    assert table.prior_method == "年限平均法"
    assert table.method_same_marker is True
    assert ds.note_area is not None
    assert ds.note_area.start_row == 24
    assert ds.note_area.end_row == 25
    assert ds.note_area.text == "政策正文完整读取。"


def test_policy_review_formula_without_cached_marker_is_data_insufficient_not_false_difference(
    tmp_path: Path,
):
    path = tmp_path / "policy_formula_marker.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.3 折旧政策复核"
    ws.append(["固定资产类别", "本期使用寿命", "本期残值率", "上期使用寿命", "上期残值率", "使用寿命差异", "残值率差异", "差异说明"])
    ws.append(["机器设备", "5-10年", "5%", "5-10年", "5%", "=B2=D2", "=C2=E2", None])
    ws.append([])
    ws.append(["Notes:"])
    ws.append(["政策未发生变化。"])
    wb.save(path)
    wb.close()

    ds = load_k03_sheets_from_workbook(path)[0]
    recorder = RuleExecutionRecorder()
    issues = run_k03_policy_review_rules(ds, fa_list=_fa_list([]), recorder=recorder)
    ledger = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}

    assert "k03_policy_period_consistency" not in {issue.rule_id for issue in issues}
    assert ledger["k03_policy_period_consistency"]["status"] == STATUS_DATA_INSUFFICIENT
    assert ledger["k03_policy_period_consistency"]["observation"]["result_summary"] == "资料不足，规则未完整执行。"
    validate_execution_ledger(recorder.to_ledger(), issues)


def test_policy_review_context_detection_is_relative_not_fixed_to_sop_rows_or_columns(
    tmp_path: Path,
):
    path = tmp_path / "policy_shifted.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.3 折旧政策复核"
    ws["E6"] = "2025-06-30"
    ws["H6"] = "2024-06-30"
    ws["D7"] = "折旧方法"
    ws["E7"] = "年限平均法"
    ws["H7"] = "年限平均法"
    ws["K7"] = True
    for cell, value in {
        "D8": "折旧政策",
        "E8": "使用寿命",
        "F8": "残值率",
        "G8": "年折旧率",
        "H8": "使用寿命",
        "I8": "残值率",
        "J8": "年折旧率",
        "K8": "使用寿命",
        "L8": "残值率",
        "M8": "差异说明",
    }.items():
        ws[cell] = value
    for cell, value in {
        "D9": "机器设备",
        "E9": "60月-120月",
        "F9": "5%",
        "G9": "9.5%-19%",
        "H9": "60月-120月",
        "I9": "5%",
        "J9": "9.5%-19%",
        "K9": True,
        "L9": True,
    }.items():
        ws[cell] = value
    ws["D12"] = "Notes:"
    ws.merge_cells("D13:M13")
    ws["D13"] = "第一段正文。"
    ws.merge_cells("D14:M14")
    ws["D14"] = "第二段正文。"
    wb.save(path)
    wb.close()

    dataset = load_k03_sheets_from_workbook(path)[0]
    table = dataset.policy_table

    assert table is not None
    assert table.header_row == 8
    assert table.column_map["asset_category"].column_letter == "D"
    assert table.current_policy_date == "2025-06-30"
    assert table.prior_policy_date == "2024-06-30"
    assert table.current_method == "年限平均法"
    assert dataset.note_area is not None
    assert dataset.note_area.text == "第一段正文。\n第二段正文。"


def test_policy_rules_record_fa_dependencies_as_data_insufficient_when_fa_list_missing():
    ds = _policy_dataset([_row()])
    recorder = RuleExecutionRecorder()

    issues = run_k03_policy_review_rules(ds, fa_list=None, recorder=recorder)
    ledger = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}

    assert not any(issue.rule_id.startswith("k03_policy_fa_") for issue in issues)
    for rule_id in (
        "k03_policy_fa_category_coverage",
        "k03_policy_fa_life_exception_followup",
        "k03_policy_fa_salvage_exception_followup",
    ):
        assert ledger[rule_id]["status"] == STATUS_DATA_INSUFFICIENT
        assert ledger[rule_id]["observation"]["result_summary"] == "资料不足，规则未完整执行。"
    assert ledger["k03_policy_period_consistency"]["status"] == STATUS_DATA_INSUFFICIENT
    validate_execution_ledger(recorder.to_ledger(), issues)


def test_policy_unreadable_table_executes_preconditions_and_marks_dependent_rules_insufficient():
    ds = _policy_dataset([])
    ds.policy_table = None
    recorder = RuleExecutionRecorder()

    issues = run_k03_policy_review_rules(ds, recorder=recorder)
    ledger = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}

    assert issues == []
    for rule_id in RULE_IDS:
        assert ledger[rule_id]["status"] == STATUS_DATA_INSUFFICIENT
    validate_execution_ledger(recorder.to_ledger(), issues)


def test_policy_missing_core_columns_marks_dependent_rules_data_insufficient():
    row = _row()
    row.current_method = None
    row.prior_method = None
    row.current_annual_rate = None
    row.prior_annual_rate = None
    column_map = {
        "asset_category": K03Column("固定资产类别", 1, "A", "asset_category"),
        "current_useful_life": K03Column("本期使用寿命", 2, "B", "current_useful_life"),
        "current_salvage_rate": K03Column("本期残值率", 3, "C", "current_salvage_rate"),
    }
    ds = _policy_dataset([row], column_map=column_map)
    recorder = RuleExecutionRecorder()

    issues = run_k03_policy_review_rules(ds, fa_list=_fa_list([]), recorder=recorder)
    ledger = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}

    assert "k03_policy_three_elements_complete" in {issue.rule_id for issue in issues}
    assert ledger["k03_policy_three_elements_complete"]["status"] == STATUS_EXECUTED
    assert ledger["k03_policy_method_change_consistency"]["status"] == STATUS_DATA_INSUFFICIENT
    assert ledger["k03_policy_annual_rate_recalculation"]["status"] == STATUS_DATA_INSUFFICIENT
    assert ledger["k03_policy_period_consistency"]["status"] == STATUS_DATA_INSUFFICIENT
    validate_execution_ledger(recorder.to_ledger(), issues)


def test_policy_unitless_life_is_inferred_only_from_consistent_explicit_policy_units():
    inferable = _policy_dataset(
        [
            _row(category="机器设备", current_life="5-10年", prior_life="5-10年"),
            _row(category="电子设备", current_life="3-8", prior_life="3-8"),
        ]
    )
    fa = _fa_list(
        [AssetRecord(source_row=8, asset_id="FA-TEST-001", asset_category="电子设备", useful_life_months="60", salvage_rate="5%")]
    )
    recorder = RuleExecutionRecorder()
    issues = run_k03_policy_review_rules(inferable, fa_list=fa, recorder=recorder)
    assert not any(issue.rule_id == "k03_policy_fa_life_exception_followup" for issue in issues)

    ambiguous = _policy_dataset([_row(current_life="5-10", prior_life="5-10")])
    ambiguous_fa = _fa_list(
        [AssetRecord(source_row=8, asset_id="FA-TEST-002", asset_category="机器设备", useful_life_months="60", salvage_rate="5%")]
    )
    recorder = RuleExecutionRecorder()
    issues = run_k03_policy_review_rules(ambiguous, fa_list=ambiguous_fa, recorder=recorder)
    ledger = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}
    assert not any(issue.rule_id == "k03_policy_fa_life_exception_followup" for issue in issues)
    assert ledger["k03_policy_fa_life_exception_followup"]["status"] == STATUS_DATA_INSUFFICIENT


def test_policy_unitless_life_is_uniquely_inferred_from_annual_rate_relationship():
    month_row = _row(
        current_life="36-60",
        prior_life="36-60",
        current_salvage="4%-5%",
        prior_salvage="4%-5%",
    )
    month_row.current_annual_rate = "19%-32%"
    month_row.prior_annual_rate = "19%-32%"
    month_recorder = RuleExecutionRecorder()
    month_issues = run_k03_policy_review_rules(
        _policy_dataset([month_row]),
        fa_list=_fa_list(
            [
                AssetRecord(
                    source_row=8,
                    asset_id="FA-TEST-003",
                    asset_category=month_row.asset_category,
                    useful_life_months="48",
                    salvage_rate="5%",
                )
            ]
        ),
        recorder=month_recorder,
    )

    year_row = _row(
        current_life="2-5",
        prior_life="2-5",
        current_salvage="0%",
        prior_salvage="0%",
    )
    year_row.current_annual_rate = "20%-50%"
    year_row.prior_annual_rate = "20%-50%"
    year_recorder = RuleExecutionRecorder()
    year_issues = run_k03_policy_review_rules(
        _policy_dataset([year_row]),
        fa_list=_fa_list(
            [
                AssetRecord(
                    source_row=9,
                    asset_id="FA-TEST-004",
                    asset_category=year_row.asset_category,
                    useful_life_months="36",
                    salvage_rate="0%",
                )
            ]
        ),
        recorder=year_recorder,
    )

    for issues, recorder in (
        (month_issues, month_recorder),
        (year_issues, year_recorder),
    ):
        assert not any(
            issue.rule_id == "k03_policy_three_elements_complete"
            and issue.field == "current_useful_life"
            for issue in issues
        )
        assert not any(issue.rule_id == "k03_policy_fa_life_exception_followup" for issue in issues)
        ledger = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}
        assert ledger["k03_policy_fa_life_exception_followup"]["status"] == STATUS_EXECUTED
        validate_execution_ledger(recorder.to_ledger(), issues)


def test_policy_category_review_ignores_unrelated_footer_and_repeated_header_rows():
    row = _row()
    ds = _policy_dataset([row])
    fa = _fa_list(
        [
            AssetRecord(source_row=20, asset_category="NB1 footer note"),
            AssetRecord(
                source_row=21,
                asset_id="asset_id",
                asset_name="asset_name",
                asset_category="asset_category",
                useful_life_months="useful_life_months",
                salvage_rate="salvage_rate",
            ),
            AssetRecord(
                source_row=22,
                asset_id="FA-TEST-005",
                asset_category="unmapped real category",
                useful_life_months="60",
                salvage_rate="5%",
            ),
        ]
    )

    issues = run_k03_policy_review_rules(ds, fa_list=fa, recorder=RuleExecutionRecorder())
    category_issues = [
        issue
        for issue in issues
        if issue.rule_id == "k03_policy_fa_category_coverage"
        and issue.field == "asset_category"
    ]

    assert len(category_issues) == 1


def test_policy_salvage_rate_range_includes_boundaries_and_reports_only_outside_values():
    ds = _policy_dataset([_row(current_salvage="0%-5%", prior_salvage="0%-5%")])
    records = [
        AssetRecord(source_row=idx, asset_id=f"FA-TEST-{idx:03d}", asset_category="机器设备", useful_life_months="60", salvage_rate=rate)
        for idx, rate in enumerate(("0%", "3%", "5%", "5.01%", "6%"), start=1)
    ]

    issues = run_k03_policy_review_rules(ds, fa_list=_fa_list(records))
    mismatch_rows = {
        issue.source_row
        for issue in issues
        if issue.rule_id == "k03_policy_fa_salvage_exception_followup" and issue.source_row is not None
    }

    assert mismatch_rows == {4, 5}


def test_complete_policy_review_does_not_report():
    ds = _policy_dataset([_row()])
    fa = _fa_list(
        [
            AssetRecord(
                source_row=8,
                asset_id="FA-TEST-001",
                asset_category="机器设备",
                useful_life_months="60",
                salvage_rate="0.05",
            )
        ]
    )

    issues = run_k03_policy_review_rules(ds, fa_list=fa)

    assert issues == []


def test_policy_change_with_true_marker_and_no_explanation_reports():
    ds = _policy_dataset(
        [_row(current_life="8年", prior_life="5年", life_marker="TRUE")],
        notes=None,
    )

    recorder = RuleExecutionRecorder()
    issues = run_k03_policy_review_rules(ds, fa_list=_fa_list([]), recorder=recorder)

    ids = {issue.rule_id for issue in issues}
    assert "k03_policy_period_consistency" in ids
    assert "k03_policy_change_field_explanation" in ids
    assert any(issue.severity == Severity.FAIL for issue in issues)
    ledger = {item["rule_id"]: item for item in recorder.to_ledger()["items"]}
    assert ledger["k03_policy_period_consistency"]["status"] == STATUS_EXECUTED
    assert ledger["k03_policy_period_consistency"]["finding_count"] == 1


def test_policy_change_with_explanation_does_not_fail():
    ds = _policy_dataset(
        [
            _row(
                current_life="8年",
                prior_life="5年",
                life_marker="FALSE",
                explanation="本期更新会计估计，已取得审批。",
            )
        ],
        notes=None,
    )

    issues = run_k03_policy_review_rules(ds, fa_list=_fa_list([]))

    assert "k03_policy_change_field_explanation" not in {issue.rule_id for issue in issues}


def test_fa_list_life_and_salvage_mismatch_report_with_top_n_summary():
    ds = _policy_dataset([_row(current_life="5-10年", current_salvage="5%")])
    records = [
        AssetRecord(
            source_row=idx,
            asset_id=f"FA-TEST-{idx:03d}",
            asset_category="机器设备",
            useful_life_months="24",
            salvage_rate="10%",
        )
        for idx in range(1, 8)
    ]

    issues = run_k03_policy_review_rules(ds, fa_list=_fa_list(records))

    life_issues = [issue for issue in issues if issue.rule_id == "k03_policy_fa_life_exception_followup"]
    salvage_issues = [issue for issue in issues if issue.rule_id == "k03_policy_fa_salvage_exception_followup"]
    assert len([issue for issue in life_issues if issue.source_row]) == 5
    assert len([issue for issue in salvage_issues if issue.source_row]) == 5
    assert any(issue.source_row is None for issue in life_issues)
    assert any(issue.source_row is None for issue in salvage_issues)


def test_unreadable_or_missing_policy_sheet_needs_review():
    missing_recorder = RuleExecutionRecorder()
    missing = run_k03_policy_review_rules(None, recorder=missing_recorder)
    assert missing == []
    assert all(
        item["status"] == STATUS_DATA_INSUFFICIENT
        for item in missing_recorder.to_ledger()["items"]
    )

    unreadable = _policy_dataset([])
    unreadable.policy_table = None
    unreadable_recorder = RuleExecutionRecorder()
    issues = run_k03_policy_review_rules(unreadable, recorder=unreadable_recorder)
    assert issues == []
    assert all(
        item["status"] == STATUS_DATA_INSUFFICIENT
        for item in unreadable_recorder.to_ledger()["items"]
    )


def test_policy_method_change_consistency_positive_negative_and_missing_marker():
    consistent = _policy_dataset([_row()])
    consistent.policy_table.current_method = "straight-line"
    consistent.policy_table.prior_method = "straight-line"
    consistent.policy_table.method_same_marker = True
    consistent.policy_table.current_policy_date = "2025-12-31"
    consistent.policy_table.prior_policy_date = "2024-12-31"
    assert not any(
        issue.rule_id == "k03_policy_method_change_consistency"
        for issue in run_k03_policy_review_rules(consistent, fa_list=_fa_list([]))
    )

    inconsistent = _policy_dataset([_row()])
    inconsistent.policy_table.current_method = "straight-line"
    inconsistent.policy_table.prior_method = "straight-line"
    inconsistent.policy_table.method_same_marker = False
    issues = run_k03_policy_review_rules(inconsistent, fa_list=_fa_list([]))
    assert any(issue.rule_id == "k03_policy_method_change_consistency" for issue in issues)

    missing = _policy_dataset([_row()])
    missing.policy_table.current_method = "straight-line"
    missing.policy_table.prior_method = "straight-line"
    missing_recorder = RuleExecutionRecorder()
    run_k03_policy_review_rules(missing, fa_list=_fa_list([]), recorder=missing_recorder)
    ledger = {item["rule_id"]: item for item in missing_recorder.to_ledger()["items"]}
    assert ledger["k03_policy_method_change_consistency"]["status"] == STATUS_DATA_INSUFFICIENT


def test_policy_annual_rate_recalculation_boundary_and_mismatch():
    boundary_row = _row(current_life="5年", prior_life="5年", current_salvage="5%", prior_salvage="5%")
    boundary_row.current_annual_rate = "19.5%"
    boundary_row.prior_annual_rate = "19%"
    boundary = _policy_dataset([boundary_row])
    assert not any(
        issue.rule_id == "k03_policy_annual_rate_recalculation"
        for issue in run_k03_policy_review_rules(boundary, fa_list=_fa_list([]))
    )

    mismatch_row = _row(current_life="5年", prior_life="5年", current_salvage="5%", prior_salvage="5%")
    mismatch_row.current_annual_rate = "20%"
    mismatch_row.prior_annual_rate = "19%"
    mismatch = _policy_dataset([mismatch_row])
    issues = run_k03_policy_review_rules(mismatch, fa_list=_fa_list([]))
    assert any(issue.rule_id == "k03_policy_annual_rate_recalculation" for issue in issues)


def test_policy_period_and_conclusion_consistency_positive_negative_and_missing():
    positive = _policy_dataset([_row()], notes="折旧政策无变化。")
    positive.policy_table.current_policy_date = "2025-12-31"
    positive.policy_table.prior_policy_date = "2024-12-31"
    positive_issues = run_k03_policy_review_rules(positive, fa_list=_fa_list([]))
    assert not any(
        issue.rule_id in {"k03_policy_period_consistency", "k03_policy_conclusion_consistency"}
        for issue in positive_issues
    )

    changed = _policy_dataset(
        [_row(current_life="8年", prior_life="5年", life_marker=False, explanation="approved")],
        notes="折旧政策无变化。",
    )
    changed.policy_table.current_policy_date = "2024-12-31"
    changed.policy_table.prior_policy_date = "2024-12-31"
    changed_issues = run_k03_policy_review_rules(changed, fa_list=_fa_list([]))
    changed_ids = {issue.rule_id for issue in changed_issues}
    assert "k03_policy_period_consistency" in changed_ids
    assert "k03_policy_conclusion_consistency" in changed_ids

    missing = _policy_dataset([_row()], notes=None)
    missing.policy_table.current_policy_date = "2025-12-31"
    missing.policy_table.prior_policy_date = "2024-12-31"
    missing_recorder = RuleExecutionRecorder()
    run_k03_policy_review_rules(missing, fa_list=_fa_list([]), recorder=missing_recorder)
    ledger = {item["rule_id"]: item for item in missing_recorder.to_ledger()["items"]}
    assert ledger["k03_policy_conclusion_consistency"]["status"] == STATUS_DATA_INSUFFICIENT


def test_policy_review_ingest_extracts_table_and_notes(tmp_path: Path):
    path = tmp_path / "policy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.3 折旧政策复核"
    ws.append(["程序说明", "复核折旧政策是否与上期一致"])
    ws.append([])
    ws.append(["表1：公司折旧政策对比表"])
    ws.append(
        [
            "固定资产类别",
            "本期日期",
            "上期日期",
            "本期折旧方法",
            "本期使用寿命",
            "本期残值率",
            "本期年折旧率",
            "上期折旧方法",
            "上期使用寿命",
            "上期残值率",
            "上期年折旧率",
            "使用寿命 TRUE/FALSE",
            "残值率 TRUE/FALSE",
            "差异说明",
        ]
    )
    ws.append(["机器设备", "2025/12/31", "2024/12/31", "年限平均法", "5-10年", "5%", "10%", "年限平均法", "5-10年", "5%", "10%", "TRUE", "TRUE", None])
    ws.append([])
    ws.append(["Notes", "政策未见变化"])
    wb.save(path)
    wb.close()

    ds = load_k03_sheets_from_workbook(path)[0]

    assert ds.execution_path == EXECUTION_PATH_POLICY_REVIEW
    assert ds.policy_table is not None
    assert ds.policy_table.range is not None
    assert ds.policy_table.rows, ds.policy_table.to_dict()
    assert ds.policy_table.rows[0].asset_category == "机器设备"
    assert ds.policy_table.column_map["current_useful_life"].column_letter == "E"
    assert ds.note_area is not None
    assert ds.llm_candidate_context["policy_table_summary"]["row_count"] == 1
    assert "rows" not in ds.llm_candidate_context["policy_table_summary"]
