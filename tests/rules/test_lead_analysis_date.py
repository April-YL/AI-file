from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from ingest.lead_sheet import load_lead_from_workbook
from rules.lead_analysis_date_after_period_end import (
    check_lead_analysis_date_after_period_end,
)
from rules.models import Severity


@pytest.fixture
def lead_dates_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "lead_dates.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["B2"] = "客户名称"
    ws["C2"] = "A"
    ws["B3"] = "期末"
    ws["C3"] = datetime(2025, 12, 31)
    ws["B4"] = "分析日期"
    ws["C4"] = datetime(2026, 3, 1)
    ws["B5"] = "可容忍误差"
    ws["C5"] = 1
    ws["B6"] = "名义金额"
    ws["C6"] = 1
    ws["B7"] = "适用会计准则"
    ws["C7"] = "CAS"
    ws["B8"] = "记账本位币"
    ws["C8"] = "CNY"
    wb.save(path)
    wb.close()
    return path


def test_analysis_after_period_end_passes(lead_dates_xlsx: Path):
    lead = load_lead_from_workbook(lead_dates_xlsx)
    assert check_lead_analysis_date_after_period_end(lead) == []


def test_analysis_on_period_end_fails(lead_dates_xlsx: Path):
    lead = load_lead_from_workbook(lead_dates_xlsx)
    for f in lead.basic_info_fields:
        if f.field_key == "analysis_date":
            f.value = "2025-12-31"
    issues = check_lead_analysis_date_after_period_end(lead)
    assert len(issues) == 1
    assert issues[0].severity == Severity.FAIL


def test_unparseable_dates_warn(lead_dates_xlsx: Path):
    lead = load_lead_from_workbook(lead_dates_xlsx)
    for f in lead.basic_info_fields:
        if f.field_key == "analysis_date":
            f.value = "TBD"
    issues = check_lead_analysis_date_after_period_end(lead)
    assert len(issues) == 1
    assert issues[0].severity == Severity.WARN
