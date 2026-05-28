from openpyxl import Workbook

from ingest.summary_sheet import PspProgramRow, SummarySheetDataset
from rules.models import Severity
from rules.psp_completion import WaiverSemanticReview, check_psp_completion


def _dataset(programs: list[PspProgramRow]) -> SummarySheetDataset:
    return SummarySheetDataset(
        source_file="test.xlsx",
        source_sheet="汇总",
        header_row=1,
        programs=programs,
    )


def test_waiver_missing_is_fail():
    rows = [
        PspProgramRow(
            procedure_name="PSP-折旧测试",
            sheet_ref="K.03.1",
            execution_status="否",
            waiver_reason=None,
            notes=None,
            source_row=2,
            is_psp=True,
        ),
    ]
    issues = check_psp_completion(_dataset(rows))
    assert any(i.severity == Severity.FAIL for i in issues)
    assert any(i.field == "waiver_reason" for i in issues)


def test_waiver_present_ok():
    rows = [
        PspProgramRow(
            procedure_name="PSP-新增",
            sheet_ref="K.02.1",
            execution_status="不适用",
            waiver_reason="客户已提供外部专家报告，范围已覆盖",
            notes=None,
            source_row=3,
            is_psp=True,
        ),
    ]
    issues = check_psp_completion(_dataset(rows))
    assert not any(i.severity == Severity.FAIL for i in issues)
    assert any(i.field == "waiver_reason" and i.severity == Severity.NEED_REVIEW for i in issues)


def test_empty_execution_need_review():
    rows = [
        PspProgramRow(
            procedure_name="K.01 后推",
            sheet_ref="K.01",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=2,
            is_psp=False,
        ),
    ]
    issues = check_psp_completion(_dataset(rows))
    assert any(i.severity == Severity.NEED_REVIEW for i in issues)


def test_no_programs_sheet_level_need_review():
    issues = check_psp_completion(
        SummarySheetDataset("t.xlsx", "", None, []),
    )
    assert len(issues) == 1
    assert issues[0].severity == Severity.NEED_REVIEW


def test_yes_without_workbook_skips_sheet_cross_check():
    rows = [
        PspProgramRow(
            procedure_name="K.01 后推",
            sheet_ref="K.01",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=2,
            is_psp=True,
        ),
    ]
    issues = check_psp_completion(_dataset(rows))
    assert not issues


def test_yes_matching_sheet_ok():
    rows = [
        PspProgramRow(
            procedure_name="K.01 后推",
            sheet_ref="K.01",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=2,
            is_psp=True,
        ),
    ]
    issues = check_psp_completion(
        _dataset(rows),
        workbook_sheet_titles=["Lead", "K.01 后推-2024"],
    )
    assert not issues


def test_yes_missing_sheet_fail():
    rows = [
        PspProgramRow(
            procedure_name="测试",
            sheet_ref="K.99.9",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=2,
            is_psp=True,
        ),
    ]
    issues = check_psp_completion(
        _dataset(rows),
        workbook_sheet_titles=["K.01"],
    )
    assert any(i.severity == Severity.FAIL for i in issues)


def test_yes_fuzzy_match_need_review():
    rows = [
        PspProgramRow(
            procedure_name="细节",
            sheet_ref="K.02.1 细节测试",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=2,
            is_psp=True,
        ),
    ]
    issues = check_psp_completion(
        _dataset(rows),
        workbook_sheet_titles=["K.02 新增"],
    )
    assert any(i.severity == Severity.NEED_REVIEW for i in issues)
    assert not any(i.severity == Severity.FAIL for i in issues)


def test_yes_sheet_ref_fallback_from_procedure_name():
    rows = [
        PspProgramRow(
            procedure_name="执行 K.01 后推程序",
            sheet_ref="",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=2,
            is_psp=True,
        ),
    ]
    issues = check_psp_completion(
        _dataset(rows),
        workbook_sheet_titles=["K.01 底稿"],
    )
    assert not issues


def test_yes_no_ref_need_review():
    rows = [
        PspProgramRow(
            procedure_name="仅名称无编号",
            sheet_ref="",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=2,
            is_psp=True,
        ),
    ]
    issues = check_psp_completion(
        _dataset(rows),
        workbook_sheet_titles=["某页"],
    )
    assert any(i.field == "sheet_ref" and i.severity == Severity.NEED_REVIEW for i in issues)


def test_yes_sparse_sheet_warns(tmp_path):
    wb_path = tmp_path / "psp_sparse.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "K.01 后推"
    ws["A1"] = "hdr"
    wb.save(wb_path)

    rows = [
        PspProgramRow(
            procedure_name="K.01",
            sheet_ref="K.01",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=2,
            is_psp=True,
        ),
    ]
    issues = check_psp_completion(
        _dataset(rows),
        workbook_sheet_titles=["K.01 后推"],
        workbook_path=str(wb_path),
    )
    assert any(i.field == "sheet_substance" and i.severity == Severity.WARN for i in issues)


def test_template_completeness_warns_when_missing_programs():
    rows = [
        PspProgramRow(
            procedure_name="K.00 Lead Sheet",
            sheet_ref="K.00 Lead Sheet",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=6,
            is_psp=False,
        )
    ]
    ds = _dataset(rows)
    ds.layout = "swp"
    issues = check_psp_completion(
        ds,
        workbook_sheet_titles=["K.00 Lead Sheet", "K.01 Agree SL to GL"],
        enforce_template_completeness=True,
    )
    assert any(i.field == "program_completeness" for i in issues)
    assert any(i.severity == Severity.FAIL for i in issues)


def test_template_completeness_missing_row_and_sheet_is_fail():
    rows = [
        PspProgramRow(
            procedure_name="K.00 Lead Sheet",
            sheet_ref="K.00 Lead Sheet",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=6,
            is_psp=False,
        )
    ]
    ds = _dataset(rows)
    ds.layout = "swp"
    issues = check_psp_completion(
        ds,
        workbook_sheet_titles=["K.00 Lead Sheet"],
        enforce_template_completeness=True,
    )
    assert any(
        i.field == "program_completeness"
        and i.severity == Severity.FAIL
        and "工作簿中也未发现相关程序页" in i.message
        for i in issues
    )


def test_waiver_semantic_reviewer_controls_severity():
    rows = [
        PspProgramRow(
            procedure_name="K.03.2 折旧测试TOD",
            sheet_ref="K.03.2 折旧测试TOD",
            execution_status="否",
            waiver_reason="金额小。",
            notes=None,
            source_row=19,
            is_psp=False,
        ),
    ]

    def reviewer(_row: PspProgramRow) -> WaiverSemanticReview:
        return WaiverSemanticReview(
            adequacy="insufficient",
            rationale="仅描述金额，未说明替代程序和风险应对",
            suggested_action="补充TE/TT判断、替代程序和证据来源。",
        )

    issues = check_psp_completion(
        _dataset(rows),
        waiver_reason_reviewer=reviewer,
    )
    assert any(i.field == "waiver_reason" and i.severity == Severity.WARN for i in issues)


def test_template_completeness_allows_sap_or_tod():
    rows = [
        PspProgramRow(
            procedure_name="K.00 Lead Sheet",
            sheet_ref="K.00 Lead Sheet",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=6,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.01 Agree SL to GL",
            sheet_ref="K.01 Agree SL to GL",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=8,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.02.1 新增测试",
            sheet_ref="K.02.1 新增测试",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=10,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.02.1a 新增选样输出",
            sheet_ref="K.02.1a 新增选样输出",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=12,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="新增清单",
            sheet_ref="新增清单",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=11,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.02.2 处置测试",
            sheet_ref="K.02.2 处置测试",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=14,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.02.2a 处置选样输出",
            sheet_ref="K.02.2a 处置选样输出",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=16,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="处置清单",
            sheet_ref="处置清单",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=15,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.03.1 SAP",
            sheet_ref="K.03.1 SAP",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=18,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.03.3 折旧政策复核",
            sheet_ref="K.03.3 折旧政策复核",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=20,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.04 固定资产减值",
            sheet_ref="K.04 固定资产减值",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=22,
            is_psp=False,
        ),
    ]
    ds = _dataset(rows)
    ds.layout = "swp"
    issues = check_psp_completion(
        ds,
        workbook_sheet_titles=[
            "K.00 Lead Sheet",
            "K.01 Agree SL to GL",
            "K.02.1 新增测试",
            "K.02.1a 新增选样输出",
            "新增清单",
            "K.02.2 处置测试",
            "K.02.2a 处置选样输出",
            "处置清单",
            "K.03.1 SAP",
            "K.03.3 折旧政策复核",
            "K.04 固定资产减值",
        ],
        enforce_template_completeness=True,
    )
    dep_alt_issues = [
        i
        for i in issues
        if i.field == "program_completeness" and "K.03.1 SAP / K.03.2 折旧测试TOD" in i.message
    ]
    assert not dep_alt_issues


def test_template_completeness_not_limited_to_psp_rows():
    """回归：当仅部分行被标记 is_psp=True 时，完整性检查仍应基于全部程序行。"""
    rows = [
        PspProgramRow(
            procedure_name="K.00 Lead Sheet",
            sheet_ref="K.00 Lead Sheet",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=6,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.01 Agree SL to GL",
            sheet_ref="K.01 Agree SL to GL",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=8,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.02.1 新增测试",
            sheet_ref="K.02.1 新增测试",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=10,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.02.1a 新增选样输出",
            sheet_ref="K.02.1a 新增选样输出",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=12,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="新增清单",
            sheet_ref="新增清单",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=11,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.02.2 处置测试",
            sheet_ref="K.02.2 处置测试",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=14,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.02.2a 处置选样输出",
            sheet_ref="K.02.2a 处置选样输出",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=16,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="处置清单",
            sheet_ref="处置清单",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=15,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.03.1 SAP",
            sheet_ref="K.03.1 SAP",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=18,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.03.3 折旧政策复核",
            sheet_ref="K.03.3 折旧政策复核",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=20,
            is_psp=False,
        ),
        # 仅此行标记为 PSP，不应导致完整性检查只看这一行
        PspProgramRow(
            procedure_name="K.04 固定资产减值",
            sheet_ref="K.04 固定资产减值",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=22,
            is_psp=True,
        ),
    ]
    ds = _dataset(rows)
    ds.layout = "swp"
    issues = check_psp_completion(
        ds,
        workbook_sheet_titles=[
            "K.00 Lead Sheet",
            "K.01 Agree SL to GL",
            "K.02.1 新增测试",
            "K.02.1a 新增选样输出",
            "新增清单",
            "K.02.2 处置测试",
            "K.02.2a 处置选样输出",
            "处置清单",
            "K.03.1 SAP",
            "K.03.3 折旧政策复核",
            "K.04 固定资产减值",
        ],
        enforce_template_completeness=True,
    )
    assert not any(i.field == "program_completeness" for i in issues)


def test_semantic_review_applies_to_non_psp_rows():
    """即便存在 is_psp=True 行，也应覆盖普通程序行的不执行语义判断。"""
    rows = [
        PspProgramRow(
            procedure_name="K.03.2 折旧测试TOD",
            sheet_ref="K.03.2 折旧测试TOD",
            execution_status="否",
            waiver_reason="实际处置金额小于TT，不执行本次测试",
            notes=None,
            source_row=19,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.04 固定资产减值",
            sheet_ref="项目组自行填写底稿索引",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=22,
            is_psp=True,
        ),
    ]

    def reviewer(_row: PspProgramRow) -> WaiverSemanticReview:
        return WaiverSemanticReview(
            adequacy="insufficient",
            rationale="仅提及阈值，未说明替代程序和风险应对",
            suggested_action="补充风险评估与替代程序依据。",
        )

    issues = check_psp_completion(
        _dataset(rows),
        waiver_reason_reviewer=reviewer,
    )
    assert any(i.field == "waiver_reason" and i.severity == Severity.WARN for i in issues)


def test_semantic_review_failure_is_need_review():
    rows = [
        PspProgramRow(
            procedure_name="K.02.2 处置测试",
            sheet_ref="K.02.2 处置测试",
            execution_status="否",
            waiver_reason="实际处置金额小于TT，不执行本次测试",
            notes=None,
            source_row=15,
            is_psp=False,
        ),
    ]

    def reviewer(_row: PspProgramRow) -> WaiverSemanticReview | None:
        return None

    issues = check_psp_completion(
        _dataset(rows),
        waiver_reason_reviewer=reviewer,
    )
    assert any(
        i.field == "waiver_reason"
        and i.severity == Severity.NEED_REVIEW
        and "语义复核未返回有效结果" in i.message
        for i in issues
    )


def test_merged_execution_status_inherits_for_k021a_and_k022a():
    rows = [
        PspProgramRow(
            procedure_name="K.02.1 新增测试",
            sheet_ref="K.02.1 新增测试",
            execution_status="是",
            waiver_reason=None,
            notes=None,
            source_row=11,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.02.1a 新增选样输出",
            sheet_ref="K.02.1a 新增选样输出",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=12,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.02.2 处置测试",
            sheet_ref="K.02.2 处置测试",
            execution_status="否",
            waiver_reason="处置总体金额很低，执行替代程序",
            notes=None,
            source_row=15,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.02.2a 处置选样输出",
            sheet_ref="K.02.2a 处置选样输出",
            execution_status="",
            waiver_reason=None,
            notes=None,
            source_row=16,
            is_psp=False,
        ),
    ]
    issues = check_psp_completion(_dataset(rows))
    assert not any(i.source_row == 12 and i.field == "execution_status" for i in issues)
    assert not any(i.source_row == 16 and i.field == "execution_status" for i in issues)
    assert not any(i.source_row == 16 and i.field == "waiver_reason" for i in issues)


def test_dep_sap_tod_either_or_allows_tod_evidence_by_number_semantic_content(tmp_path):
    wb_path = tmp_path / "dep_tod_by_item.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "K.03.2 固定资产折旧逐项重算工作底稿"
    ws["A1"] = "资产编号"
    ws["B1"] = "原值"
    ws["C1"] = "使用寿命"
    ws["D1"] = "残值率"
    ws["E1"] = "账面折旧"
    ws["F1"] = "重算折旧"
    ws["G1"] = "差异"
    ws["A2"] = "FA-TEST-001"
    ws["B2"] = 120000
    ws["C2"] = 60
    ws["D2"] = 0.05
    ws["E2"] = 1800
    ws["F2"] = 1750
    ws["G2"] = -50
    wb.save(wb_path)

    rows = [
        PspProgramRow(
            procedure_name="K.03.1 SAP",
            sheet_ref="K.03.1 SAP",
            execution_status="否",
            waiver_reason="未执行",
            notes=None,
            source_row=18,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.03.2 折旧测试TOD",
            sheet_ref="K.03.2 折旧测试TOD",
            execution_status="否",
            waiver_reason="未执行",
            notes=None,
            source_row=19,
            is_psp=False,
        ),
    ]

    def reviewer(_row: PspProgramRow) -> WaiverSemanticReview:
        return WaiverSemanticReview(
            adequacy="insufficient",
            rationale="理由不足",
            suggested_action="补充说明",
        )

    issues = check_psp_completion(
        _dataset(rows),
        workbook_sheet_titles=["K.03.2 固定资产折旧逐项重算工作底稿"],
        workbook_path=str(wb_path),
        waiver_reason_reviewer=reviewer,
    )
    assert not any(i.field == "waiver_reason" and i.source_row in (18, 19) for i in issues)
    assert any(
        i.field == "execution_status_consistency"
        and i.severity == Severity.NEED_REVIEW
        and i.source_row == 19
        for i in issues
    )


def test_dep_tod_name_only_without_content_does_not_count_as_executed(tmp_path):
    wb_path = tmp_path / "dep_tod_name_only.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "K.03.2 固定资产折旧逐项重算工作底稿"
    ws["A1"] = "模板"
    wb.save(wb_path)

    rows = [
        PspProgramRow(
            procedure_name="K.03.1 SAP",
            sheet_ref="K.03.1 SAP",
            execution_status="否",
            waiver_reason="金额小",
            notes=None,
            source_row=18,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.03.2 折旧测试TOD",
            sheet_ref="K.03.2 折旧测试TOD",
            execution_status="否",
            waiver_reason="金额小",
            notes=None,
            source_row=19,
            is_psp=False,
        ),
    ]

    def reviewer(_row: PspProgramRow) -> WaiverSemanticReview:
        return WaiverSemanticReview(
            adequacy="insufficient",
            rationale="理由不足",
            suggested_action="补充说明",
        )

    issues = check_psp_completion(
        _dataset(rows),
        workbook_sheet_titles=["K.03.2 固定资产折旧逐项重算工作底稿"],
        workbook_path=str(wb_path),
        waiver_reason_reviewer=reviewer,
    )
    assert any(i.field == "waiver_reason" and i.source_row in (18, 19) for i in issues)
    assert not any(i.field == "execution_status_consistency" for i in issues)
