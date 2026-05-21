from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from ingest.lead_sheet import load_lead_from_workbook
from rules.lead_required_fields import (
    LEAD_REQUIRED_FIELD_KEYS,
    check_lead_required_fields,
)
from rules.models import Severity


@pytest.fixture
def complete_lead_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "lead_complete.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["B2"] = "客户名称"
    ws["C2"] = "XYZ公司"
    ws["B3"] = "期末"
    ws["C3"] = datetime(2025, 12, 31)
    ws["B4"] = "分析日期"
    ws["C4"] = datetime(2026, 3, 1)
    ws["B5"] = "可容忍误差（TE）"
    ws["C5"] = 500000
    ws["B6"] = "名义金额（SAD）"
    ws["C6"] = 50000
    ws["B7"] = "适用会计准则（GAAP）"
    ws["C7"] = "CAS"
    ws["B8"] = "记账本位币"
    ws["C8"] = "CNY"
    wb.save(path)
    wb.close()
    return path


def test_complete_lead_passes_required_fields(complete_lead_xlsx: Path):
    lead = load_lead_from_workbook(complete_lead_xlsx)
    issues = check_lead_required_fields(lead)
    assert issues == []


def test_missing_client_name_fails(complete_lead_xlsx: Path):
    lead = load_lead_from_workbook(complete_lead_xlsx)
    lead.basic_info_fields = [f for f in lead.basic_info_fields if f.field_key != "client_name"]
    issues = check_lead_required_fields(lead)
    assert any(i.field == "client_name" and i.severity == Severity.FAIL for i in issues)


def test_no_lead_sheet_fails():
    issues = check_lead_required_fields(None)
    assert len(issues) == 1
    assert issues[0].severity == Severity.FAIL


def test_required_keys_match_planning():
    assert "pm" not in LEAD_REQUIRED_FIELD_KEYS
    assert "te" in LEAD_REQUIRED_FIELD_KEYS
