from pathlib import Path

import openpyxl

from ingest.fa_list_amount_basis import resolve_fa_list_amount_basis, resolve_unique_header_basis
from ingest.models import FaListAmountBasisSource, FaListAmountBasisStatus
from ingest.records import load_fa_list_from_workbook


def _build_formula_workbook(
    path: Path,
    *,
    include_k01: bool = True,
    include_summary: bool = True,
    summary_net_column: str = "I",
) -> list[tuple]:
    wb = openpyxl.Workbook()
    fa = wb.active
    fa.title = "FA list"

    if include_summary:
        fa.append(("固定资产类别", "原值", "累计折旧", "减值准备", "净值"))
        fa.append(
            (
                "机器设备",
                "=SUMIFS($F$6:$F$8,$D$6:$D$8,$A2)",
                "=SUMIFS($H$6:$H$8,$D$6:$D$8,$A2)",
                "=SUMIFS($K$6:$K$8,$D$6:$D$8,$A2)",
                f"=SUMIFS(${summary_net_column}$6:${summary_net_column}$8,$D$6:$D$8,$A2)",
            )
        )
        fa.append(())
        fa.append(())
    else:
        for _ in range(4):
            fa.append(())

    fa.append((None, None, None, "固定资产类别", "固定资产编号", "原值", None, "累计折旧", "净值", None, "减值准备"))
    fa.append((None, None, None, "机器设备", "FA-001", 100, None, 20, 80, None, 0))
    fa.append((None, None, None, "机器设备", "FA-002", 200, None, 40, 160, None, 0))
    fa.append((None, None, None, "机器设备", None, None, None, None, 0.01, None, -0.01))

    if include_k01:
        k01 = wb.create_sheet("K.01 Agree SL to GL")
        k01["A1"] = "表2"
        k01["H1"] = "表3"
        for column, value in enumerate(("固定资产类别", "原值", "累计折旧", "减值准备", "净值"), 1):
            k01.cell(2, column, value)
        k01["A3"] = "机器设备"
        k01["B3"] = "=SUMIFS('FA list'!$F$6:$F$8,'FA list'!$D$6:$D$8,$A3)"
        k01["C3"] = "=SUMIFS('FA list'!$H$6:$H$8,'FA list'!$D$6:$D$8,$A3)"
        k01["D3"] = "=SUMIFS('FA list'!$K$6:$K$8,'FA list'!$D$6:$D$8,$A3)"
        k01["E3"] = "=SUMIFS('FA list'!$I$6:$I$8,'FA list'!$D$6:$D$8,$A3)"
        k01["A4"] = "合计"
        k01["B4"] = "=SUM(B3:B3)"
        k01["C4"] = "=SUM(C3:C3)"
        k01["D4"] = "=SUM(D3:D3)"
        k01["E4"] = "=SUM(E3:E3)"

    wb.save(path)
    wb.close()

    check = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return [tuple(cell.value for cell in row) for row in check["FA list"].iter_rows()]
    finally:
        check.close()


def test_k01_table2_formulas_are_highest_confidence_and_cross_checked(tmp_path: Path):
    path = tmp_path / "formula_basis.xlsx"
    rows = _build_formula_workbook(path)

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=rows,
    )

    assert basis.status == FaListAmountBasisStatus.CONFIRMED
    assert basis.source == FaListAmountBasisSource.K01_FORMULA
    assert basis.bindings == {
        "original_value": 6,
        "accumulated_depreciation": 8,
        "impairment_provision": 11,
        "net_value": 9,
    }
    assert basis.category_column == 4
    assert (basis.data_start_row, basis.data_end_row) == (6, 8)


def test_unbounded_k01_bindings_are_completed_by_matching_bounded_fa_summary(tmp_path: Path):
    path = tmp_path / "unbounded_k01_with_bounded_summary.xlsx"
    _build_formula_workbook(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        k01 = wb["K.01 Agree SL to GL"]
        for cell, amount_column in (("B3", "F"), ("C3", "H"), ("D3", "K"), ("E3", "I")):
            k01[cell] = f"=SUMIFS('FA list'!{amount_column}:{amount_column},'FA list'!D:D,$A3)"
        wb.save(path)
    finally:
        wb.close()

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=_read_fa_rows(path),
    )

    assert basis.status == FaListAmountBasisStatus.CONFIRMED
    assert basis.source == FaListAmountBasisSource.K01_FORMULA
    assert basis.bindings == {
        "original_value": 6,
        "accumulated_depreciation": 8,
        "impairment_provision": 11,
        "net_value": 9,
    }
    assert basis.category_column == 4
    assert (basis.data_start_row, basis.data_end_row) == (6, 8)
    assert any("completed with matching bounded" in item for item in basis.evidence)


def test_unbounded_k01_without_matching_summary_keeps_bindings_but_stops(tmp_path: Path):
    path = tmp_path / "unbounded_k01_without_summary.xlsx"
    _build_formula_workbook(path, include_summary=False)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        k01 = wb["K.01 Agree SL to GL"]
        for cell, amount_column in (("B3", "F"), ("C3", "H"), ("D3", "K"), ("E3", "I")):
            k01[cell] = f"=SUMIFS('FA list'!{amount_column}:{amount_column},'FA list'!D:D,$A3)"
        wb.save(path)
    finally:
        wb.close()

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=_read_fa_rows(path),
    )

    assert basis.status == FaListAmountBasisStatus.INCOMPLETE
    assert basis.source == FaListAmountBasisSource.K01_FORMULA
    assert basis.bindings["original_value"] == 6
    assert basis.category_column == 4
    assert basis.data_start_row is None
    assert any("bounded FA detail range" in item for item in basis.conflicts)


def test_unbounded_k01_and_summary_with_different_currency_values_stop(tmp_path: Path):
    path = tmp_path / "unbounded_k01_currency_conflict.xlsx"
    _build_formula_workbook(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        fa = wb["FA list"]
        fa["L5"] = "币种"
        for row in range(6, 9):
            fa.cell(row, 12, "USD")
        for column, amount_column in zip(range(2, 6), ("F", "H", "K", "I")):
            fa.cell(
                2,
                column,
                f'=SUMIFS(${amount_column}$6:${amount_column}$8,$D$6:$D$8,$A2,$L$6:$L$8,"USD")',
            )
        k01 = wb["K.01 Agree SL to GL"]
        for cell, amount_column in (("B3", "F"), ("C3", "H"), ("D3", "K"), ("E3", "I")):
            k01[cell] = (
                f'=SUMIFS(\'FA list\'!{amount_column}:{amount_column},'
                f'\'FA list\'!D:D,$A3,\'FA list\'!L:L,"CNY")'
            )
        wb.save(path)
    finally:
        wb.close()

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=_read_fa_rows(path),
    )

    assert basis.status == FaListAmountBasisStatus.AMBIGUOUS
    assert any("currency_criteria_values" in item for item in basis.conflicts)


def test_k01_exact_bindings_ignore_unrelated_summary_amount_columns(tmp_path: Path):
    path = tmp_path / "formula_basis_with_opening_column.xlsx"
    _build_formula_workbook(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        fa = wb["FA list"]
        fa["F1"] = "期初累计折旧"
        fa["F2"] = "=SUMIFS($R$6:$R$8,$D$6:$D$8,$A2)"
        wb.save(path)
    finally:
        wb.close()

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=_read_fa_rows(path),
    )

    assert basis.status == FaListAmountBasisStatus.CONFIRMED
    assert basis.bindings["accumulated_depreciation"] == 8
    assert (basis.data_start_row, basis.data_end_row) == (6, 8)


def test_k01_and_fa_summary_formula_conflict_is_ambiguous(tmp_path: Path):
    path = tmp_path / "formula_conflict.xlsx"
    rows = _build_formula_workbook(path, summary_net_column="J")

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=rows,
    )

    assert basis.status == FaListAmountBasisStatus.AMBIGUOUS
    assert any("disagree" in conflict for conflict in basis.conflicts)


def test_k01_and_fa_summary_data_range_conflict_is_ambiguous(tmp_path: Path):
    path = tmp_path / "range_conflict.xlsx"
    _build_formula_workbook(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        fa = wb["FA list"]
        for column, amount_column in zip(range(2, 6), ("F", "H", "K", "I")):
            fa.cell(2, column, f"=SUMIFS(${amount_column}$6:${amount_column}$9,$D$6:$D$9,$A2)")
        wb.save(path)
    finally:
        wb.close()
    rows = _read_fa_rows(path)

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=rows,
    )

    assert basis.status == FaListAmountBasisStatus.AMBIGUOUS
    assert any("data_range" in conflict for conflict in basis.conflicts)


def test_fa_summary_formulas_are_used_when_k01_is_absent(tmp_path: Path):
    path = tmp_path / "summary_basis.xlsx"
    rows = _build_formula_workbook(path, include_k01=False)

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet=None,
        data_rows=rows,
    )

    assert basis.status == FaListAmountBasisStatus.CONFIRMED
    assert basis.source == FaListAmountBasisSource.FA_SUMMARY_FORMULA
    assert basis.bindings["net_value"] == 9


def test_cross_sheet_fa_summary_formulas_are_rejected(tmp_path: Path):
    path = tmp_path / "cross_sheet_summary.xlsx"
    _build_formula_workbook(path, include_k01=False)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        other = wb.create_sheet("Other")
        other.append(("类别", None, None, None, None, 100, None, 20, 80, None, 0))
        fa = wb["FA list"]
        for column, amount_column in zip(range(2, 6), ("F", "H", "K", "I")):
            fa.cell(2, column, f"=SUMIFS(Other!${amount_column}$1:${amount_column}$1,Other!$D$1:$D$1,$A2)")
        wb.save(path)
    finally:
        wb.close()

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet=None,
        data_rows=_read_fa_rows(path),
    )

    assert basis.status == FaListAmountBasisStatus.AMBIGUOUS
    assert any("cross-sheet" in conflict for conflict in basis.conflicts)


def test_duplicate_k01_amount_headers_and_incomplete_category_rows_are_rejected(tmp_path: Path):
    duplicate_path = tmp_path / "duplicate_k01.xlsx"
    _build_formula_workbook(duplicate_path)
    wb = openpyxl.load_workbook(duplicate_path, data_only=False)
    try:
        wb["K.01 Agree SL to GL"]["F2"] = "净值"
        wb.save(duplicate_path)
    finally:
        wb.close()
    duplicate_basis = resolve_fa_list_amount_basis(
        duplicate_path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=_read_fa_rows(duplicate_path),
    )

    incomplete_path = tmp_path / "incomplete_k01.xlsx"
    _build_formula_workbook(incomplete_path)
    wb = openpyxl.load_workbook(incomplete_path, data_only=False)
    try:
        k01 = wb["K.01 Agree SL to GL"]
        k01["A4"] = "房屋建筑物"
        k01["B4"] = "=SUMIFS('FA list'!$F$6:$F$8,'FA list'!$D$6:$D$8,$A4)"
        k01["C4"] = "=SUMIFS('FA list'!$H$6:$H$8,'FA list'!$D$6:$D$8,$A4)"
        k01["D4"] = "=SUMIFS('FA list'!$K$6:$K$8,'FA list'!$D$6:$D$8,$A4)"
        wb.save(incomplete_path)
    finally:
        wb.close()
    incomplete_basis = resolve_fa_list_amount_basis(
        incomplete_path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=_read_fa_rows(incomplete_path),
    )

    assert duplicate_basis.status == FaListAmountBasisStatus.AMBIGUOUS
    assert incomplete_basis.status == FaListAmountBasisStatus.AMBIGUOUS


def test_unique_headers_are_last_fallback_and_duplicates_stop_resolution():
    confirmed = resolve_unique_header_basis(
        [(1, "固定资产编号"), (2, "原值"), (3, "累计折旧"), (4, "净值")],
        header_row=1,
    )
    ambiguous = resolve_unique_header_basis(
        [(1, "原值"), (2, "期末原值"), (3, "累计折旧"), (4, "净值")],
        header_row=1,
    )

    assert confirmed.status == FaListAmountBasisStatus.CONFIRMED
    assert confirmed.source == FaListAmountBasisSource.UNIQUE_HEADERS
    assert "impairment_provision" not in confirmed.bindings
    assert ambiguous.status == FaListAmountBasisStatus.AMBIGUOUS


def test_formula_range_retains_unlabelled_amount_row_for_identity_diagnosis(tmp_path: Path):
    path = tmp_path / "asset_rows.xlsx"
    _build_formula_workbook(path)

    dataset = load_fa_list_from_workbook(
        path,
        sheet_name="FA list",
        k01_sheet_name="K.01 Agree SL to GL",
    )

    assert dataset.amount_basis is not None
    assert dataset.amount_basis.status == FaListAmountBasisStatus.CONFIRMED
    assert [record.source_row for record in dataset.records] == [6, 7]
    assert dataset.fa_profile is not None
    assert [record.source_row for record in dataset.fa_profile.population.identity_incomplete_records] == [8]
    row_8 = next(item for item in dataset.fa_profile.population.classified_rows if item.record.source_row == 8)
    assert row_8.role.value == "identity_incomplete_detail"


def test_inconsistent_sumifs_criteria_dimensions_are_ambiguous(tmp_path: Path):
    path = tmp_path / "criteria_conflict.xlsx"
    _build_formula_workbook(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        fa = wb["FA list"]
        fa["L5"] = "币种"
        for row, currency in zip(range(6, 9), ("CNY", "CNY", "CNY")):
            fa.cell(row, 12, currency)
        wb["K.01 Agree SL to GL"]["B3"] = (
            "=SUMIFS('FA list'!$F$6:$F$8,'FA list'!$D$6:$D$8,$A3,"
            "'FA list'!$L$6:$L$8,$F$1)"
        )
        wb.save(path)
    finally:
        wb.close()

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=_read_fa_rows(path),
    )

    assert basis.status == FaListAmountBasisStatus.AMBIGUOUS
    assert any("criteria" in item for item in basis.conflicts)


def test_multiple_row_currencies_without_formula_currency_criterion_are_incomplete(tmp_path: Path):
    path = tmp_path / "currency_scope.xlsx"
    _build_formula_workbook(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        fa = wb["FA list"]
        fa["L5"] = "币种"
        for row, currency in zip(range(6, 9), ("CNY", "USD", "CNY")):
            fa.cell(row, 12, currency)
        wb.save(path)
    finally:
        wb.close()

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=_read_fa_rows(path),
    )

    assert basis.status == FaListAmountBasisStatus.INCOMPLETE
    assert any("currencies" in item for item in basis.conflicts)


def test_same_criteria_columns_with_different_currency_values_are_ambiguous(tmp_path: Path):
    path = tmp_path / "criteria_value_conflict.xlsx"
    _build_formula_workbook(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        fa = wb["FA list"]
        fa["L5"] = "币种"
        for row, currency in zip(range(6, 9), ("CNY", "USD", "CNY")):
            fa.cell(row, 12, currency)
        k01 = wb["K.01 Agree SL to GL"]
        for cell, amount_column, currency in (
            ("B3", "F", "CNY"),
            ("C3", "H", "USD"),
            ("D3", "K", "CNY"),
            ("E3", "I", "CNY"),
        ):
            k01[cell] = (
                f'=SUMIFS(\'FA list\'!${amount_column}$6:${amount_column}$8,'
                f'\'FA list\'!$D$6:$D$8,$A3,\'FA list\'!$L$6:$L$8,"{currency}")'
            )
        wb.save(path)
    finally:
        wb.close()

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=_read_fa_rows(path),
    )

    assert basis.status == FaListAmountBasisStatus.AMBIGUOUS
    assert any("criteria values" in item or "currency criteria" in item for item in basis.conflicts)


def test_category_criterion_need_not_be_first_sumifs_pair(tmp_path: Path):
    path = tmp_path / "category_second.xlsx"
    _build_formula_workbook(path, include_summary=False)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        fa = wb["FA list"]
        fa["L5"] = "币种"
        for row in range(6, 9):
            fa.cell(row, 12, "CNY")
        k01 = wb["K.01 Agree SL to GL"]
        for cell, amount_column in (("B3", "F"), ("C3", "H"), ("D3", "K"), ("E3", "I")):
            k01[cell] = (
                f'=SUMIFS(\'FA list\'!${amount_column}$6:${amount_column}$8,'
                f'\'FA list\'!$L$6:$L$8,"CNY",\'FA list\'!$D$6:$D$8,$A3)'
            )
        wb.save(path)
    finally:
        wb.close()

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=_read_fa_rows(path),
    )

    assert basis.status == FaListAmountBasisStatus.CONFIRMED
    assert basis.category_column == 4


def test_partial_period_labels_do_not_define_the_whole_amount_group(tmp_path: Path):
    path = tmp_path / "partial_period.xlsx"
    _build_formula_workbook(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        wb["FA list"]["F5"] = "期末原值"
        wb.save(path)
    finally:
        wb.close()

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=_read_fa_rows(path),
    )

    assert basis.status == FaListAmountBasisStatus.INCOMPLETE
    assert any("period semantics" in item for item in basis.conflicts)


def test_fa_summary_category_criterion_need_not_be_first(tmp_path: Path):
    path = tmp_path / "fa_summary_category_second.xlsx"
    _build_formula_workbook(path, include_k01=False)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        fa = wb["FA list"]
        fa["L5"] = "币种"
        for row in range(6, 9):
            fa.cell(row, 12, "CNY")
        for column, amount_column in zip(range(2, 6), ("F", "H", "K", "I")):
            fa.cell(2, column, (
                f'=SUMIFS(${amount_column}$6:${amount_column}$8,'
                '$L$6:$L$8,"CNY",$D$6:$D$8,$A2)'
            ))
        wb.save(path)
    finally:
        wb.close()

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet=None,
        data_rows=_read_fa_rows(path),
    )

    assert basis.status == FaListAmountBasisStatus.CONFIRMED
    assert basis.category_column == 4


def test_composite_sumifs_formula_is_rejected_instead_of_partially_parsed(tmp_path: Path):
    path = tmp_path / "composite_formula.xlsx"
    _build_formula_workbook(path)
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        cell = wb["K.01 Agree SL to GL"]["B3"]
        cell.value = str(cell.value) + "+1"
        wb.save(path)
    finally:
        wb.close()

    basis = resolve_fa_list_amount_basis(
        path,
        fa_sheet="FA list",
        k01_sheet="K.01 Agree SL to GL",
        data_rows=_read_fa_rows(path),
    )

    assert basis.status == FaListAmountBasisStatus.AMBIGUOUS
    assert any("unsupported" in item for item in basis.conflicts)


def _read_fa_rows(path: Path) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return [tuple(cell.value for cell in row) for row in wb["FA list"].iter_rows()]
    finally:
        wb.close()
