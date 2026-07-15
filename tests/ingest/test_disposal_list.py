"""处置清单 ingest：字段映射、表头识别与行解析。"""

from pathlib import Path

import openpyxl

from ingest.records import build_disposal_list_summary
from ingest.field_mapping import check_required_fields, map_headers
from ingest.header_detection import scan_rows_for_headers
from ingest.models import AmountGroupStatus, SheetKind
from ingest.records import parse_fa_list_rows
from ingest.sheet_loader import load_asset_sheet_from_workbook

from ingest.workbook_ingest import load_workbook_ingest


def test_parse_disposal_list_keeps_disposal_fields():
    rows = [
        (
            "固定资产类别",
            "卡片编码",
            "固定资产名称",
            "处置原值",
            "处置累计折旧",
            "处置净值",
            "业务日期",
            "减少方式",
        ),
        ("机器设备", "FA-D-001", "旧设备A", 10000, -8000, 2000, "2025-06-01", "出售"),
    ]
    dataset = parse_fa_list_rows(
        rows,
        source_file="dummy.xlsx",
        source_sheet="处置清单",
        sheet_kind=SheetKind.DISPOSAL_LIST,
    )
    record = dataset.records[0]
    assert record.asset_id == "FA-D-001"
    assert record.disposal_date == "2025-06-01"
    assert record.disposal_method == "出售"
    assert record.original_value == "10000"
    assert record.net_value == "2000"
    mapped = {m.standard_field for m in dataset.mapped_fields}
    assert "disposal_date" in mapped
    assert "disposal_method" in mapped


def test_disposal_amount_group_does_not_mix_ending_and_disposal_columns():
    rows = [
        (
            "固定资产编号",
            "原值",
            "累计折旧",
            "减值准备",
            "净值",
            "处置原值-CNY",
            "处置累计折旧-CNY",
            "处置净值-CNY",
            "处置时间",
            "新增/处置",
        ),
        ("FA-D-001", 0, 0, 0, 0, 100, 80, 20, "2025-06-30", "处置或报废"),
    ]

    dataset = parse_fa_list_rows(
        rows,
        source_sheet="处置清单",
        sheet_kind=SheetKind.DISPOSAL_LIST,
    )

    selected = next(
        group for group in dataset.amount_groups
        if group.group_id == dataset.selected_amount_group_id
    )
    assert selected.status == AmountGroupStatus.INCOMPLETE
    assert selected.members["original_value"].source_header == "处置原值-CNY"
    assert selected.members["net_value"].source_header == "处置净值-CNY"
    assert "impairment_provision" in selected.missing_measures
    assert dataset.records[0].original_value == "100"
    assert dataset.records[0].net_value == "20"
    assert dataset.records[0].disposal_date == "2025-06-30"
    assert dataset.records[0].disposal_method == "处置或报废"


def test_disposal_amount_group_marks_duplicate_semantic_blocks_as_conflicted():
    headers = [
        (1, "处置原值-CNY"), (2, "处置累计折旧-CNY"),
        (3, "处置减值准备-CNY"), (4, "处置净值-CNY"),
        (5, "本期减少原值-CNY"), (6, "本期减少累计折旧-CNY"),
        (7, "本期减少减值准备-CNY"), (8, "本期减少净值-CNY"),
    ]

    dataset = parse_fa_list_rows(
        [tuple(header for _, header in headers), tuple(range(1, 9))],
        source_sheet="处置清单",
        sheet_kind=SheetKind.DISPOSAL_LIST,
    )
    selected = next(
        group for group in dataset.amount_groups
        if group.group_id == dataset.selected_amount_group_id
    )

    assert selected.status == AmountGroupStatus.CONFLICTED


def test_disposal_amount_group_prefers_complete_original_currency_group():
    rows = [
        (
            "固定资产编号",
            "处置原值", "处置累计折旧", "处置减值准备", "处置净值",
            "处置原值-CNY", "处置累计折旧-CNY", "处置净值-CNY",
        ),
        ("FA-D-001", 100, 80, 0, 20, 700, 560, 140),
    ]

    dataset = parse_fa_list_rows(
        rows,
        source_sheet="处置清单",
        sheet_kind=SheetKind.DISPOSAL_LIST,
    )
    selected = next(
        group for group in dataset.amount_groups
        if group.group_id == dataset.selected_amount_group_id
    )

    assert selected.status == AmountGroupStatus.CONFIRMED
    assert selected.currency_role.value == "original"
    assert dataset.records[0].original_value == "100"


def test_disposal_list_prefers_disposal_original_value_over_opening():
    mapped, _ = map_headers(
        [
            (1, "编码"),
            (2, "名称"),
            (3, "期初原值"),
            (4, "处置原值"),
            (5, "减少方式"),
            (6, "业务日期"),
            (7, "净值"),
        ],
        SheetKind.DISPOSAL_LIST,
    )
    by_field = {m.standard_field: m for m in mapped}
    assert by_field["original_value"].source_header == "处置原值"
    assert "期初原值" not in {m.source_header for m in mapped}


def test_disposal_list_sap_style_headers():
    rows = [
        (
            "企业类别名称",
            "卡片编码",
            "附加资产描述",
            "未税成本",
            "累计折旧",
            "净值",
            "业务日期",
            "处置情况",
        ),
        ("生产设备", "CARD-99", "冲床", 50000, -40000, 10000, "2025-03-15", "报废"),
    ]
    dataset = parse_fa_list_rows(
        rows,
        source_sheet="K.02.2b 减少清单",
        sheet_kind=SheetKind.DISPOSAL_LIST,
    )
    record = dataset.records[0]
    assert record.asset_id == "CARD-99"
    assert record.asset_name == "冲床"
    assert record.disposal_method == "报废"
    assert record.disposal_date == "2025-03-15"


def test_disposal_header_skips_sop_prose_row():
    rows = [
        ("获取当期按资产类别划分的处置清单，并与 K.01 后推明细表核对差异",),
        (
            "固定资产编号",
            "资产名称",
            "原值",
            "累计折旧",
            "净值",
            "业务日期",
            "减少方式",
        ),
        ("FA-D-002", "设备B", 100, -80, 20, "2025-01-01", "出售"),
    ]
    header_row, cells, _ = scan_rows_for_headers(rows, sheet_kind=SheetKind.DISPOSAL_LIST)
    assert header_row == 2
    assert any(text == "减少方式" for _, text in cells)
    dataset = parse_fa_list_rows(
        rows,
        source_sheet="处置清单",
        sheet_kind=SheetKind.DISPOSAL_LIST,
    )
    assert len(dataset.records) == 1
    assert dataset.records[0].asset_id == "FA-D-002"


def test_disposal_list_skips_subtotal_rows():
    rows = [
        ("减少方式", "固定资产编号", "固定资产名称", "原值", "累计折旧", "净值", "业务日期"),
        ("出售", "FA-D-001", "设备A", 100, -80, 20, "2025-01-01"),
        ("出售-小计", None, None, 100, -80, 20, None),
        ("报废", "FA-D-002", "设备B", 200, -150, 50, "2025-02-01"),
    ]
    dataset = parse_fa_list_rows(
        rows,
        source_sheet="处置清单",
        sheet_kind=SheetKind.DISPOSAL_LIST,
    )
    assert [r.asset_id for r in dataset.records] == ["FA-D-001", "FA-D-002"]


def test_disposal_required_fields_check():
    mapped, _ = map_headers(
        [
            (1, "卡片编码"),
            (2, "资产名称"),
            (3, "原值"),
            (4, "累计折旧"),
            (5, "净值"),
            (6, "业务日期"),
            (7, "减少方式"),
        ],
        SheetKind.DISPOSAL_LIST,
    )
    missing_req, _ = check_required_fields(mapped, SheetKind.DISPOSAL_LIST)
    assert "disposal_date" not in missing_req
    assert "disposal_method" not in missing_req


def test_load_disposal_list_from_workbook(tmp_path: Path):
    xlsx = tmp_path / "disposal.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "处置清单"
    ws.append(
        [
            "固定资产类别",
            "卡片编码",
            "固定资产名称",
            "原值",
            "累计折旧",
            "净值",
            "业务日期",
            "处置情况",
        ]
    )
    ws.append(["机器设备", "FA-D-010", "设备X", 1000, -700, 300, "2025-05-01", "出售"])
    wb.save(xlsx)
    wb.close()

    dataset = load_asset_sheet_from_workbook(
        xlsx,
        SheetKind.DISPOSAL_LIST,
    )
    assert dataset.source_sheet == "处置清单"
    assert dataset.records[0].disposal_method == "出售"
    assert dataset.records[0].disposal_date == "2025-05-01"


def test_disposal_list_summary_classifies_methods_and_totals():
    rows = [
        (
            "固定资产类别",
            "固定资产编号",
            "固定资产名称",
            "原值",
            "累计折旧",
            "减值",
            "净值",
            "处置日期",
            "处置/报废",
        ),
        ("仪器设备", "FA-D-001", "设备A", 1000, 700, 0, 300, "2025-06-01", "出售"),
        ("仪器设备", "FA-D-002", "设备B", 2000, 1500, 0, 500, "2025-06-02", "报废"),
        ("仪器设备", "FA-D-003", "设备C", 3000, 2500, 0, 500, "2025-06-03", "处置及报废"),
        ("仪器设备", "FA-D-004", "设备D", 4000, 3500, 0, 500, "2025-06-04", "调拨"),
        ("仪器设备", "FA-D-005", "设备E", 5000, 4500, 0, 500, "2025-06-05", None),
    ]
    dataset = parse_fa_list_rows(
        rows,
        source_file="dummy.xlsx",
        source_sheet="处置清单",
        sheet_kind=SheetKind.DISPOSAL_LIST,
    )
    summary = build_disposal_list_summary(dataset)
    assert summary is not None
    assert summary.record_count == 5
    assert summary.sale_net_value == "300"
    assert summary.scrap_net_value == "500"
    assert summary.sale_scrap_net_value == "1300"
    assert summary.other_reduction_net_value == "500"
    assert summary.unclassified_net_value == "500"
    assert any(bucket.bucket_key == "sale" for bucket in summary.buckets)
    assert any(bucket.bucket_key == "scrap" for bucket in summary.buckets)
    assert any(bucket.bucket_key == "sale_scrap" for bucket in summary.buckets)
    assert any(bucket.bucket_key == "other" for bucket in summary.buckets)
    assert any(bucket.bucket_key == "unknown" for bucket in summary.buckets)


def test_workbook_ingest_exposes_disposal_list_summary(tmp_path: Path):
    xlsx = tmp_path / "disposal_summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(["程序", "工作表", "是否执行"])
    ws.append(["K.02.2 处置测试", "K.02.2 处置测试", "是"])
    ws_list = wb.create_sheet("处置清单")
    ws_list.append(
        [
            "固定资产类别",
            "卡片编码",
            "固定资产名称",
            "原值",
            "累计折旧",
            "减值",
            "净值",
            "业务日期",
            "处置/报废",
        ]
    )
    ws_list.append(["仪器设备", "FA-D-001", "设备A", 1000, 700, 0, 300, "2025-06-01", "处置及报废"])
    wb.save(xlsx)
    wb.close()

    ctx = load_workbook_ingest(xlsx)
    assert ctx.disposal_list_summary is not None
    assert ctx.disposal_list_summary.sale_scrap_net_value == "300"
