from __future__ import annotations

from pathlib import Path

import openpyxl

from ingest.lead_sheet import load_lead_from_workbook


def test_check_with_a3_notes_multiline(tmp_path: Path):
    path = tmp_path / "lead_a3_notes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["C48"] = "科目名称"
    ws["E48"] = "期末账面数"
    ws["I48"] = "期末审定数"
    ws["C49"] = "原值"
    ws["I49"] = 500
    ws["C50"] = "累计折旧"
    ws["I50"] = 100
    ws["C51"] = "减值准备"
    ws["I51"] = 0
    ws["C52"] = "净值"
    ws["I52"] = 400
    ws["C53"] = "check with A3"
    ws["I53"] = 500
    ws["C54"] = "Diff"
    ws["I54"] = 50
    ws["C55"] = "Notes:"
    ws["D55"] = "折旧增加主要系原值增加"
    wb.save(path)
    wb.close()

    lead = load_lead_from_workbook(path)
    assert lead.check_with_a3 is not None
    assert lead.check_with_a3.notes_text
    assert "折旧" in lead.check_with_a3.notes_text
