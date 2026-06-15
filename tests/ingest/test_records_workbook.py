from pathlib import Path

import openpyxl
import pytest

from ingest.models import SheetKind
from ingest.records import (
    find_fa_list_sheets,
    load_fa_list_csv,
    load_fa_list_from_workbook,
    parse_fa_list_rows,
)
from ingest.sheet_loader import load_asset_sheet_from_workbook

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures"


@pytest.fixture
def fa_list_xlsx(tmp_path: Path) -> Path:
    """由 fa_list_mixed.csv 生成最小 Excel 底稿（FA list sheet）。"""
    csv_path = FIXTURES / "fa_list_mixed.csv"
    xlsx_path = tmp_path / "fa_list_mixed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FA list"
    with csv_path.open(encoding="utf-8-sig") as f:
        for line in f:
            ws.append(line.strip().split(","))
    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


def test_find_fa_list_sheets(fa_list_xlsx: Path):
    candidates = find_fa_list_sheets(fa_list_xlsx)
    assert len(candidates) >= 1
    assert candidates[0].sheet_name == "FA list"
    assert candidates[0].confidence > 0.5


def test_load_fa_list_from_workbook_matches_csv(fa_list_xlsx: Path):
    from_csv = load_fa_list_csv(FIXTURES / "fa_list_mixed.csv")
    from_xlsx = load_fa_list_from_workbook(fa_list_xlsx)

    assert from_xlsx.source_sheet == "FA list"
    assert len(from_xlsx.records) == len(from_csv.records)
    assert from_xlsx.records[0].asset_id == "FA-TEST-001"
    assert from_xlsx.mapped_fields


def test_load_fa_list_from_workbook_explicit_sheet(fa_list_xlsx: Path):
    dataset = load_fa_list_from_workbook(fa_list_xlsx, sheet_name="FA list")
    assert len(dataset.records) == 6


def test_load_fa_list_from_workbook_no_fa_list(tmp_path: Path):
    xlsx = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(["项目", "金额"])
    ws.append(["合计", 100])
    wb.save(xlsx)
    wb.close()

    dataset = load_fa_list_from_workbook(xlsx)
    assert dataset.records == []
    assert dataset.source_sheet == ""


def test_parse_fa_list_rows_skips_reclassification_summary_rows():
    rows = [
        ("资产编号", "资产名称", "原值", "累计折旧", "净值"),
        ("FA-TEST-001", "设备A", 100, -10, 90),
        ("资产类别重分类", None, 0, 0, 0),
    ]
    dataset = parse_fa_list_rows(rows)
    assert [r.asset_id for r in dataset.records] == ["FA-TEST-001"]


def test_parse_addition_list_keeps_addition_method():
    rows = [
        (
            "固定资产类别",
            "固定资产编号",
            "固定资产名称",
            "入账开始日期",
            "原值",
            "新增方式",
        ),
        ("机器设备", "FA-TEST-001", "设备A", "2024-01-01", 1000, "在建工程转入"),
    ]
    dataset = parse_fa_list_rows(
        rows,
        source_file="dummy.xlsx",
        source_sheet="新增清单",
        sheet_kind=SheetKind.ADDITION_LIST,
    )
    assert dataset.records[0].addition_method == "在建工程转入"
    assert {m.standard_field for m in dataset.mapped_fields} >= {
        "asset_category",
        "asset_id",
        "asset_name",
        "start_date",
        "original_value",
        "addition_method",
    }


def test_parse_addition_list_uses_ending_original_value_not_opening():
    rows = [
        ("编码", "名称", "期初原值", "期末原值", "新增方式"),
        ("FA-TEST-001", "设备A", 0, 500, "购置"),
    ]
    dataset = parse_fa_list_rows(
        rows,
        source_file="dummy.xlsx",
        source_sheet="新增清单",
        sheet_kind=SheetKind.ADDITION_LIST,
    )
    assert dataset.records[0].original_value == "500"


def test_parse_addition_list_uses_added_original_value_not_opening():
    rows = [
        ("编码", "名称", "原值原币", "原值本币", "期初原值", "新增原值", "变动方式"),
        ("FA-TEST-001", "设备A", 1000, 1000, 900, 300, "购置"),
    ]
    dataset = parse_fa_list_rows(
        rows,
        source_file="dummy.xlsx",
        source_sheet="新增清单",
        sheet_kind=SheetKind.ADDITION_LIST,
    )
    assert dataset.records[0].original_value == "300"
    assert dataset.records[0].addition_method == "购置"


def test_parse_addition_list_skips_subtotal_and_total_rows():
    rows = [
        ("新增方式", "固定资产编号", "固定资产名称", "资产类别", "原值"),
        ("外购", "FA-TEST-001", "设备A", "机器设备", 100),
        ("外购-小计", None, None, None, 100),
        ("在建转固", "FA-TEST-002", "设备B", "机器设备", 200),
        (None, None, None, "合计", 300),
        (None, None, None, None, 300),
    ]
    dataset = parse_fa_list_rows(
        rows,
        source_file="dummy.xlsx",
        source_sheet="新增清单",
        sheet_kind=SheetKind.ADDITION_LIST,
    )
    assert [r.asset_id for r in dataset.records] == ["FA-TEST-001", "FA-TEST-002"]


def test_load_addition_list_keeps_rows_after_preview_limit(tmp_path: Path):
    xlsx = tmp_path / "addition_after_preview_limit.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "新增清单"
    ws.append(("新增方式", "固定资产编号", "固定资产名称", "固定资产类别", "原值"))
    for index in range(1, 105):
        ws.append(("在建工程转入", f"FA-WIP-{index:03d}", "设备A", "机器设备", 100))
    ws.append(("购置", "FA-BUY-001", "设备B", "机器设备", 386061.06))
    wb.save(xlsx)
    wb.close()

    dataset = load_asset_sheet_from_workbook(xlsx, SheetKind.ADDITION_LIST, max_rows=100)

    purchase_records = [r for r in dataset.records if r.addition_method == "购置"]
    assert len(dataset.records) == 105
    assert purchase_records[0].asset_id == "FA-BUY-001"
    assert purchase_records[0].original_value == "386061.06"
