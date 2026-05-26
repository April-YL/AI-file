"""K.00 Lead 分块识别与扩展 parse 单测。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from typing import Any

import openpyxl
import pytest

from ingest.lead_sheet import load_lead_from_workbook, parse_lead_sheet_rows
from ingest.lead_sheet_blocks import LeadBlockKind, detect_lead_blocks

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures"
TEMPLATE_GLOB = "K1 SWP*XYZ*.xlsx"


@pytest.fixture
def lead_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "lead_wb.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["A1"] = "项目"
    ws["B1"] = "底稿值"
    ws["C1"] = "Canvas最终"
    ws["A2"] = "可容忍误差"
    ws["B2"] = "100000"
    ws["C2"] = "100000"
    ws["A3"] = "名义金额"
    ws["B3"] = "50000"
    ws["C3"] = "50000"
    ws["A4"] = "计划重要性"
    ws["B4"] = "200000"
    ws["C4"] = "200000"
    ws["A6"] = "认定"
    ws["B6"] = "CRA"
    ws["C6"] = "TT"
    ws["A7"] = "固定资产"
    ws["B7"] = "High"
    ws["C7"] = "150000"
    ws["A8"] = "存货"
    ws["B8"] = "Moderate"
    ws["C8"] = "80000"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def swp_lead_xlsx(tmp_path: Path) -> Path:
    """近似标准 K1 SWP Lead 版式（行序可漂移）。"""
    path = tmp_path / "swp_lead.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["B2"] = "客户名称"
    ws["C2"] = "XYZ公司"
    ws["B3"] = "期末"
    ws["C3"] = datetime(2025, 12, 31)
    ws["B4"] = "分析日期"
    ws["C4"] = datetime(2026, 3, 1)
    ws["B5"] = "可容忍误差（TE）"
    ws["C5"] = 500000
    ws["B6"] = "名义金额（SAD）"
    ws["C6"] = 50000
    ws["B7"] = "适用会计准则（GAAP）"
    ws["C7"] = "企业会计准则"
    ws["B8"] = "记账本位币"
    ws["C8"] = "CNY"
    ws["B14"] = "认定"
    ws["C14"] = "CRA"
    ws["D14"] = "各项认定1"
    ws["E14"] = "所有相关认定 2"
    ws["B15"] = "完整性（C）"
    ws["C15"] = "Minimal"
    ws["D15"] = 100000
    ws["E15"] = 100000
    ws["B16"] = "计价/计量（V/M）"
    ws["C16"] = "Low"
    ws["D16"] = 80000
    ws["B27"] = "账户变更"
    ws["C27"] = "预期及额外考虑"
    ws["B28"] = "新增"
    ws["C28"] = "无重大新增"
    ws["B29"] = "减少"
    ws["C29"] = "无处置"
    ws["B37"] = "波动范围"
    ws["B38"] = "波动幅度 ():"
    ws["C38"] = 0
    ws["B39"] = "波动幅度 (%):"
    ws["C39"] = 0.1
    ws["B48"] = "总账科目编码"
    ws["C48"] = "科目名称"
    ws["D48"] = "索引号"
    ws["E48"] = "期末账面数"
    ws["G48"] = "期末未审数"
    ws["I48"] = "期末审定数"
    ws["J48"] = "上期末审定数"
    ws["C49"] = "原值"
    ws["D49"] = "K.01"
    ws["E49"] = 1000
    ws["G49"] = 1000
    ws["I49"] = 1000
    ws["J49"] = 900
    ws["C50"] = "累计折旧"
    ws["D50"] = "K.01"
    ws["E50"] = 200
    ws["I50"] = 200
    ws["J50"] = 150
    ws["C53"] = "净值"
    ws["I53"] = 800
    ws["J53"] = 750
    ws["B58"] = "波动说明"
    ws["B59"] = "无异常波动"
    ws["B64"] = "调整汇总表（如不适用请删除）"
    ws["B65"] = "调整类型"
    ws["B66"] = "审计调整"
    ws["C66"] = "示例"
    wb.save(path)
    wb.close()
    return path


def test_parse_materiality_and_cra(lead_xlsx: Path):
    ds = load_lead_from_workbook(lead_xlsx)
    assert ds.source_sheet == "K.00 Lead Sheet"
    keys = {m.field_key: m for m in ds.materiality}
    assert keys["te"].workpaper_value == "100000"
    assert keys["te"].canvas_value == "100000"
    assert len(ds.cra_rows) >= 1
    fa_row = next(r for r in ds.cra_rows if "固定资产" in r.assertion)
    assert fa_row.cra == "High"
    assert fa_row.tt == "150000"


def test_swp_layout_blocks_and_modules(swp_lead_xlsx: Path):
    ds = load_lead_from_workbook(swp_lead_xlsx)
    kinds = {b.kind for b in ds.blocks}
    assert LeadBlockKind.BASIC_INFO in kinds
    assert LeadBlockKind.CRA_THRESHOLD in kinds
    assert LeadBlockKind.EXPECTATION in kinds
    assert LeadBlockKind.MOVEMENT_TABLE in kinds
    assert LeadBlockKind.FLUCTUATION_NOTES in kinds
    assert LeadBlockKind.ADJUSTMENT_SUMMARY in kinds

    by_key = {f.field_key: f for f in ds.basic_info_fields}
    assert by_key["client_name"].value == "XYZ公司"
    assert by_key["period_end"].value == "2025-12-31"
    assert by_key["te"].value == "500000"
    assert by_key["gaap"].value == "企业会计准则"
    assert by_key["currency"].value == "CNY"

    vm = next(r for r in ds.cra_rows if "计价" in r.assertion)
    assert vm.cra == "Low"
    assert vm.tt == "80000"

    c_row = next(r for r in ds.cra_rows if "完整性" in r.assertion)
    assert c_row.tt_overall == "100000"

    assert len(ds.expectations) >= 2
    assert any("新增" in e.account_change for e in ds.expectations)
    assert ds.volatility is not None
    assert ds.volatility.percent == "0.1"

    labels = {r.account_label for r in ds.movement_rows}
    assert "原值" in labels
    assert "累计折旧" in labels
    assert "净值" in labels
    roles = {b.role for b in ds.movement_bindings}
    assert "py_audited" in roles
    orig = next(r for r in ds.movement_rows if r.account_label == "原值")
    assert orig.sheet_ref == "K.01"
    assert orig.values.get("sheet_ref") == "K.01"
    assert orig.values.get("audited_ending") == "1000"
    assert orig.values.get("py_audited") == "900"

    assert ds.fluctuation_notes and "无异常波动" in ds.fluctuation_notes
    assert len(ds.adjustment_rows) >= 1


def test_blocks_shift_when_extra_rows_prepended(swp_lead_xlsx: Path):
    wb = openpyxl.load_workbook(swp_lead_xlsx)
    ws = wb["K.00 Lead Sheet"]
    ws.insert_rows(1, 5)
    for i in range(5):
        ws.cell(row=i + 1, column=2, value=f"说明行{i}")
    path = swp_lead_xlsx.parent / "shifted_lead.xlsx"
    wb.save(path)
    wb.close()

    ds = load_lead_from_workbook(path)
    assert ds.block(LeadBlockKind.BASIC_INFO) is not None
    assert ds.block(LeadBlockKind.BASIC_INFO).anchor_row > 5
    assert len(ds.movement_rows) >= 1
    assert any(f.field_key == "client_name" for f in ds.basic_info_fields)


def test_match_movement_role_py_audited_header():
    from ingest.lead_sheet import _match_movement_role

    assert _match_movement_role("上期末审定数") == "py_audited"
    assert _match_movement_role("期末审定数") == "audited_ending"


def test_hyphen_cell_does_not_match_anchors():
    from ingest.lead_sheet_blocks import _label_in_cell, _EXPECTATION_ANCHORS

    assert not _label_in_cell("-", _EXPECTATION_ANCHORS)
    assert not _label_in_cell("N/A", ("期末账面数", "科目名称"))


def test_detect_blocks_on_template_rows():
    rows: list[tuple[Any, ...]] = []
    for _ in range(50):
        rows.append(tuple([None] * 6))
    rows[1] = (None, "客户名称", "A", None, None, None)
    rows[13] = (None, "认定", "CRA", "TT", None, None)
    rows[26] = (None, "账户变更", "预期及额外考虑", None, None, None)
    rows[47] = (None, "总账科目编码", "科目名称", "期末账面数", None, None)
    blocks = detect_lead_blocks(rows)
    kinds = {b.kind for b in blocks}
    assert LeadBlockKind.BASIC_INFO in kinds
    assert LeadBlockKind.CRA_THRESHOLD in kinds
    assert LeadBlockKind.EXPECTATION in kinds
    assert LeadBlockKind.MOVEMENT_TABLE in kinds


def test_no_cra_layout_volatility_from_te(tmp_path: Path):
    """简版 Lead：无 CRA 区，波动幅度金额 = TE（案例 A 口径）。"""
    path = tmp_path / "lead_no_cra.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["B2"] = "客户名称"
    ws["C2"] = "A公司"
    ws["B5"] = "可容忍误差（TE）"
    ws["C5"] = 1148000000
    ws["B26"] = "账户变更"
    ws["C26"] = "预期及额外考虑"
    ws["B27"] = "新增"
    ws["B36"] = "波动范围"
    ws["B37"] = "波动幅度 (CNY):"
    ws["C37"] = 1148000000
    ws["B38"] = "波动幅度 (%):"
    ws["C38"] = 0.1
    ws["B47"] = "总账科目编码"
    ws["C47"] = "科目名称"
    ws["E47"] = "期末账面数"
    ws["C48"] = "原值"
    ws["E48"] = 100
    wb.save(path)
    wb.close()

    ds = load_lead_from_workbook(path)
    assert ds.layout_variant == "no_cra_te_volatility"
    assert ds.volatility is not None
    assert ds.volatility.amount_source == "te"
    assert not ds.cra_rows


@pytest.mark.skipif(
    not any(Path(r"D:\AI file\固定资产质检agent\资料库").glob(TEMPLATE_GLOB)),
    reason="标准模板不在资料库路径",
)
def test_xyz_template_from_materials_library():
    lib = Path(r"D:\AI file\固定资产质检agent\资料库")
    path = next(p for p in lib.glob(TEMPLATE_GLOB) if "By item" not in p.name)
    ds = load_lead_from_workbook(path)
    assert ds.source_sheet == "K.00 Lead Sheet"
    kinds = {b.kind for b in ds.blocks}
    assert LeadBlockKind.BASIC_INFO in kinds
    assert LeadBlockKind.CRA_THRESHOLD in kinds
    assert LeadBlockKind.MOVEMENT_TABLE in kinds
    assert len(ds.cra_rows) >= 4
    assert len(ds.movement_rows) >= 3
