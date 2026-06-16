from pathlib import Path

import openpyxl

from ingest.workbook_ingest import load_workbook_ingest


def _base_workbook(path: Path, *, summary_status: str = "是") -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "汇总"
    ws.append(["程序", "工作表", "是否执行", "不执行的原因"])
    ws.append(["K.02.2 处置测试", "K.02.2 处置测试", summary_status, ""])

    ws_list = wb.create_sheet("处置清单")
    ws_list.append(
        [
            "固定资产类别",
            "卡片编码",
            "固定资产名称",
            "处置原值",
            "处置累计折旧",
            "处置净值",
            "业务日期",
            "减少方式",
        ]
    )
    ws_list.append(["机器设备", "FA-D-001", "旧设备A", 1000, -700, 300, "2025-06-01", "出售"])

    ws_test = wb.create_sheet("K.02.2 处置测试")
    ws_test.append(["记录我们详细测试所涵盖的样本总体"])
    ws_test.append(["处置/报废总金额", "", "", "", 300])
    ws_test.append(["Breakdown中处置/报废金额", "", "", "", 300])
    ws_test.append(["差异", "", "", "", 0])
    ws_test.append(["出售净值", "", "", "", 300])
    ws_test.append(["报废净值", "", "", "", 0])
    ws_test.append(["其他减少净值", "", "", "", 0])
    ws_test.append(["定量关键项目金额", "", "", "", 100])
    ws_test.append(["代表性抽样的剩余总体", "", "", "", 200])
    ws_test.append(["无异常情况"])
    ws_test.append([])
    ws_test.append(
        [
            "样本类型",
            "固定资产类别",
            "固定资产编号",
            "固定资产名称",
            "资产原价",
            "累计折旧",
            "减值准备",
            "净值",
            "处置/报废日",
            "出售价格",
            "处置交易凭证号",
            "处置损益",
            "出售价格（通过审计证据/支持性文件取得）",
            "出售价格差异",
            "获得的证据/支持的描述",
            "金额差异",
            "1",
            "2",
            "3",
            "4",
        ]
    )
    ws_test.append(
        [
            "代表性样本",
            "机器设备",
            "FA-D-001",
            "旧设备A",
            1000,
            -700,
            0,
            300,
            "2025-06-01",
            300,
            "12-记511",
            0,
            300,
            0,
            "合同；发票；收款记录",
            0,
            "Y",
            "Y",
            "Y",
            "Y",
        ]
    )

    ws_sample = wb.create_sheet("K.02.2a 处置选样输出")
    ws_sample.append(["源数据汇总"])
    ws_sample.append(["已上传数据", "", "", "", 300])
    ws_sample.append(["必要的数据排除项", "", "", "", 0])
    ws_sample.append(["样本池总体金额", "", "", "", 300])
    ws_sample.append(["代表性总体价值", "", "", "", 200])
    ws_sample.append(["总金额", "", "", "", 300])
    ws_sample.append(["会计记录金额", "", "", "", 300])
    ws_sample.append(["差额", "", "", "", 0])
    ws_sample.append(["关键项数量", "", "", "", 1])
    ws_sample.append(["定量关键项金额", "", "", "", 100])
    ws_sample.append(["代表性样本量", "", "", "", 1])
    ws_sample.append(["样本选择方法", "", "", "", "随机抽样"])
    ws_sample.append(["已选取样本"])
    ws_sample.append(
        [
            "源样本#",
            "抽样ID",
            "样本类型",
            "固定资产类别",
            "固定资产编号",
            "固定资产名称",
            "资产原价",
            "累计折旧",
            "净值",
            "处置日期",
            "减少方式",
        ]
    )
    ws_sample.append([1, 260, "代表性样本", "机器设备", "FA-D-001", "旧设备A", 1000, -700, 300, "2025-06-01", "出售"])

    wb.save(path)
    return wb


def test_disposal_execution_path_complete(tmp_path: Path):
    path = tmp_path / "disposal_complete.xlsx"
    wb = _base_workbook(path)
    wb.close()

    ctx = load_workbook_ingest(path)

    assert ctx.disposal_test is not None
    assert ctx.disposal_test.source_sheet == "K.02.2 处置测试"
    assert ctx.disposal_test.amounts["sale_scrap_net_value"].amount == "300"
    assert ctx.disposal_test.amounts["rollforward_disposal_net_value"].amount == "300"
    assert ctx.disposal_test.amounts["difference_amount"].amount == "0"
    assert len(ctx.disposal_test.tested_samples) == 1
    assert ctx.disposal_test.tested_samples[0].asset_id == "FA-D-001"
    assert ctx.disposal_test.tested_samples[0].sale_price == "300"
    assert ctx.disposal_test.tested_samples[0].net_value == "300"
    assert ctx.disposal_test.tested_samples[0].disposal_date == "2025-06-01"
    assert ctx.disposal_test.tested_samples[0].evidence_amount == "300"

    assert ctx.disposal_sample_output is not None
    assert ctx.disposal_sample_output.source_sheet == "K.02.2a 处置选样输出"
    assert ctx.disposal_sample_output.amounts["sample_pool_amount"].amount == "300"
    assert len(ctx.disposal_sample_output.selected_samples) == 1
    assert ctx.disposal_sample_output.selected_samples[0].asset_id == "FA-D-001"
    assert ctx.disposal_sample_output.selected_samples[0].net_value == "300"
    assert ctx.disposal_sample_output.selected_samples[0].disposal_date == "2025-06-01"

    assert ctx.disposal_execution_path is not None
    assert ctx.disposal_execution_path.path_kind == "executed_package_complete"
    assert ctx.disposal_execution_path.missing_components == []


def test_disposal_execution_path_summary_waived_does_not_require_later_sheets(tmp_path: Path):
    path = tmp_path / "disposal_summary_waived.xlsx"
    wb = _base_workbook(path, summary_status="否")
    ws = wb["汇总"]
    ws["D2"] = "本期处置资产净值小于TE，且无性质异常，未执行处置测试。"
    del wb["处置清单"]
    del wb["K.02.2 处置测试"]
    del wb["K.02.2a 处置选样输出"]
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path)

    assert ctx.disposal_test is None
    assert ctx.disposal_sample_output is None
    assert ctx.disposal_execution_path is not None
    assert ctx.disposal_execution_path.path_kind == "summary_waived"
    assert ctx.disposal_execution_path.missing_components == []


def test_disposal_execution_path_test_sheet_waiver_note(tmp_path: Path):
    path = tmp_path / "disposal_test_sheet_waiver.xlsx"
    wb = _base_workbook(path)
    ws = wb["K.02.2 处置测试"]
    ws["A10"] = "本期处置资产净值小于TE，且各单项处置净值小于TT，无性质异常，未执行抽样测试。"
    del wb["K.02.2a 处置选样输出"]
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path)

    assert ctx.disposal_test is not None
    assert ctx.disposal_test.waiver_note_text is not None
    assert ctx.disposal_execution_path is not None
    assert ctx.disposal_execution_path.path_kind == "test_sheet_waiver_note"
    assert ctx.disposal_execution_path.missing_components == []


def test_disposal_reconciliation_matrix_uses_dynamic_layout_and_formula_sources(tmp_path: Path):
    path = tmp_path / "disposal_dynamic_matrix.xlsx"
    wb = _base_workbook(path)
    ws = wb["K.02.2 处置测试"]
    ws.insert_rows(1, 20)
    for row in ws.iter_rows(min_row=33, max_row=37, min_col=1, max_col=12):
        for cell in row:
            cell.value = None
    ws["A5"] = "基础操作指引：处置/报废总金额、Breakdown中处置/报废金额及差异。"
    ws["F27"] = "账面净值"
    ws["H27"] = "原值"
    ws["J27"] = "累计折旧"
    ws["L27"] = "减值准备"
    ws["C28"] = "处置/报废总金额"
    ws["F28"] = "=H28-J28-L28"
    ws["H28"] = "='处置清单'!E50"
    ws["J28"] = "='处置清单'!F50"
    ws["L28"] = "='处置清单'!G50"
    ws["C29"] = "Breakdown中处置/报废金额"
    ws["F29"] = "=H29-J29-L29"
    ws["H29"] = "=-'K.01 Agree SL to GL'!AI15"
    ws["J29"] = "=-'K.01 Agree SL to GL'!AI21"
    ws["L29"] = "=-'K.01 Agree SL to GL'!AI27"
    ws["C30"] = "差异"
    ws["F30"] = "=F28-F29"
    ws["H30"] = "=H28-H29"
    ws["J30"] = "=J28-J29"
    ws["L30"] = "=L28-L29"
    ws["C31"] = "差异是否需要进一步调查？"
    ws["F31"] = '=IF(ABS(F30)>100,"是","否")'
    ws["H31"] = '=IF(ABS(H30)>100,"是","否")'
    ws["J31"] = '=IF(ABS(J30)>100,"是","否")'
    ws["L31"] = '=IF(ABS(L30)>100,"是","否")'
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path)
    matrix = ctx.disposal_test.reconciliation_matrix

    assert matrix is not None
    assert matrix.header_row == 27
    assert matrix.measure_columns == {
        "net_value": 6,
        "original_value": 8,
        "accumulated_depreciation": 10,
        "impairment_provision": 12,
    }
    assert matrix.rows["disposal_list"].source_row == 28
    assert matrix.rows["rollforward"].source_row == 29
    assert matrix.rows["disposal_list"].measures["original_value"].formula == "='处置清单'!E50"
    assert "K.01 Agree SL to GL" in matrix.rows["rollforward"].measures["original_value"].formula
    assert "disposal_list_formula_source=recognized" in matrix.recognition_evidence
    assert "rollforward_formula_source=recognized" in matrix.recognition_evidence
    assert "net_formula_relationship=recognized" in matrix.recognition_evidence
    assert matrix.ambiguous_candidates == []
    assert matrix.usable_for_rules is True
    assert ctx.disposal_test.usable_for_rules is True


def test_disposal_reconciliation_matrix_blocks_wrong_formula_source(tmp_path: Path):
    path = tmp_path / "disposal_wrong_source.xlsx"
    wb = _base_workbook(path)
    ws = wb["K.02.2 处置测试"]
    ws["E13"] = "账面净值"
    ws["G13"] = "原值"
    ws["I13"] = "累计折旧"
    ws["K13"] = "减值准备"
    ws["B14"] = "处置/报废总金额"
    ws["E14"] = "=G14-I14-K14"
    ws["G14"] = "=1000"
    ws["I14"] = "=700"
    ws["K14"] = "=0"
    ws["B15"] = "Breakdown中处置/报废金额"
    ws["E15"] = "=G15-I15-K15"
    ws["G15"] = "=1000"
    ws["I15"] = "=700"
    ws["K15"] = "=0"
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path)
    matrix = ctx.disposal_test.reconciliation_matrix

    assert matrix is not None
    assert matrix.rows["disposal_list"].source_row == 14
    assert "disposal_list_formula_source=unconfirmed" in matrix.recognition_evidence
    assert "rollforward_formula_source=unconfirmed" in matrix.recognition_evidence
    assert matrix.usable_for_rules is False
    assert ctx.disposal_test.usable_for_rules is False


def test_disposal_modules_and_sample_output_confidence_are_exposed(tmp_path: Path):
    path = tmp_path / "disposal_modules.xlsx"
    wb = _base_workbook(path)
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path)
    test_modules = {m.module_key: m for m in ctx.disposal_test.module_assessments}
    sample_modules = {m.module_key: m for m in ctx.disposal_sample_output.module_assessments}

    assert set(test_modules) == {
        "population_definition",
        "amount_reconciliation",
        "key_item_representation",
        "test_attributes",
        "sample_table",
        "exception_summary",
    }
    assert test_modules["sample_table"].status == "recognized"
    assert test_modules["amount_reconciliation"].status == "partial"
    assert set(sample_modules) == {
        "sampling_prerequisites",
        "source_data_summary",
        "sampling_strategy",
        "accounting_reconciliation",
        "selected_samples",
    }
    assert sample_modules["selected_samples"].status == "recognized"
    assert ctx.disposal_sample_output.usable_for_rules is True
