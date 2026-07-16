"""整本底稿结构、后推解析与勾稽单测。"""

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from ingest.models import SheetClassification, SheetKind
from ingest.reconciliation import ReconciliationStatus, run_workbook_reconciliations
from ingest.rollforward_sheet import parse_rollforward_rows
from ingest.workbook_ingest import _select_rollforward_sheet, load_workbook_ingest
from ingest.workbook_structure import WorkbookStructure, StructureIssueCode, analyze_workbook_structure

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures"


def _rollforward_structure(path: Path, names: list[str]) -> WorkbookStructure:
    return WorkbookStructure(
        source_file=str(path),
        sheets_by_kind={
            SheetKind.ROLLFORWARD.value: [
                SheetClassification(
                    sheet_name=name,
                    kind=SheetKind.ROLLFORWARD,
                    confidence=0.9,
                    name_score=0.9,
                    content_score=0.9,
                )
                for name in names
            ]
        },
    )


def test_k01_route_prefers_unique_summary_candidate(tmp_path: Path):
    path = tmp_path / "k01_route_summary.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "FA list-汇总"
    wb.create_sheet("K.01-A公司")
    wb.create_sheet("K.01-汇总")
    wb.save(path)
    wb.close()

    selected, reason = _select_rollforward_sheet(
        path,
        _rollforward_structure(path, ["K.01-A公司", "K.01-汇总"]),
        fa_sheet="FA list-汇总",
        explicit_name=None,
    )

    assert selected == "K.01-汇总"
    assert reason is None


def test_k01_route_uses_unique_formula_reference_or_stops(tmp_path: Path):
    path = tmp_path / "k01_route_formula.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "FA list-汇总"
    k01_a = wb.create_sheet("K.01-A公司")
    k01_b = wb.create_sheet("K.01-B公司")
    k01_b["A1"] = "='FA list-汇总'!A1"
    wb.save(path)
    wb.close()
    structure = _rollforward_structure(path, ["K.01-A公司", "K.01-B公司"])

    selected, reason = _select_rollforward_sheet(
        path,
        structure,
        fa_sheet="FA list-汇总",
        explicit_name=None,
    )
    assert selected == "K.01-B公司"
    assert reason is None

    wb = openpyxl.load_workbook(path)
    wb["K.01-B公司"]["A1"] = None
    wb.save(path)
    wb.close()
    selected, reason = _select_rollforward_sheet(
        path,
        structure,
        fa_sheet="FA list-汇总",
        explicit_name=None,
    )
    assert selected is None
    assert "multiple peer" in str(reason)


@pytest.fixture
def reconciliation_workbook(tmp_path: Path) -> Path:
    """FA list 与 K.01 后推期末合计一致。"""
    path = tmp_path / "recon_demo.xlsx"
    wb = openpyxl.Workbook()

    ws_sum = wb.active
    ws_sum.title = "汇总"
    ws_sum.append(["程序", "工作表", "是否执行"])
    ws_sum.append(["K.01", "K.01", "是"])

    ws_lead = wb.create_sheet("K.00 Lead Sheet")
    ws_lead["A1"] = "客户名称"
    ws_lead["B1"] = "测试客户"
    ws_lead["A2"] = "可容忍误差"
    ws_lead["B2"] = 100000

    ws_fa = wb.create_sheet("FA list")
    ws_fa.append(
        [
            "固定资产编号",
            "固定资产名称",
            "原值",
            "累计折旧",
            "净值",
            "币种",
        ]
    )
    ws_fa.append(["FA-TEST-001", "设备A", 1000, 200, 800, "CNY"])
    ws_fa.append(["FA-TEST-002", "设备B", 2000, 300, 1700, "CNY"])

    ws_rf = wb.create_sheet("K.01 Agree SL to GL")
    ws_rf.append(["固定资产类别", "原值", "累计折旧", "净值"])
    ws_rf.append(["机器设备", 1500, 250, 1250])
    ws_rf.append(["电子设备", 1500, 250, 1250])
    ws_rf.append(["合计", 3000, 500, 2500])

    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def mismatch_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "recon_mismatch.xlsx"
    wb = openpyxl.Workbook()
    ws_fa = wb.active
    ws_fa.title = "FA list"
    ws_fa.append(["固定资产编号", "原值", "累计折旧", "净值", "币种"])
    ws_fa.append(["FA-TEST-001", 1000, 100, 900, "CNY"])
    ws_rf = wb.create_sheet("K.01 Agree SL to GL")
    ws_rf.append(["类别", "原值", "累计折旧", "净值"])
    ws_rf.append(["合计", 1000, 100, 800])
    wb.save(path)
    wb.close()
    return path


def test_workbook_structure_core_sheets(reconciliation_workbook: Path):
    structure = analyze_workbook_structure(reconciliation_workbook)
    kinds = set(structure.sheets_by_kind.keys())
    assert "summary" in kinds
    assert "lead" in kinds
    assert "fa_list" in kinds
    assert "rollforward" in kinds
    assert len(structure.program_flow) >= 4
    missing_core = [
        i for i in structure.issues if i.code == StructureIssueCode.MISSING_CORE_SHEET
    ]
    assert not missing_core
    assert structure.sheet_resolutions
    for sheets in structure.sheets_by_kind.values():
        for sheet in sheets:
            assert sheet.resolution_decision is not None
            assert sheet.resolution_decision.status.value == "RESOLVED"


def test_workbook_structure_skips_internal_storage_sheet(tmp_path: Path):
    path = tmp_path / "internal_skip.xlsx"
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "汇总"
    ws_sum.append(["程序", "工作表", "是否执行"])
    ws_sum.append(["K.01", "K.01", "是"])

    ws_internal = wb.create_sheet("DS_INTERNAL_DOCUMENT_STORAGE")
    ws_internal.append(["固定资产编号", "原值", "累计折旧", "净值", "处置日期", "处置方式"])
    ws_internal.append(["FA-TEST-001", 100, 10, 90, "2025-01-01", "报废"])
    wb.save(path)
    wb.close()

    structure = analyze_workbook_structure(path)

    assert "disposal_list" not in structure.sheets_by_kind
    assert not any(
        item["sheet_name"] == "DS_INTERNAL_DOCUMENT_STORAGE"
        for item in structure.program_flow
    )


def test_rollforward_total_row():
    rows = [
        ("固定资产类别", "原值", "累计折旧", "净值"),
        ("A", 1000, 100, 900),
        ("合计", 1000, 100, 900),
    ]
    rf = parse_rollforward_rows(rows, source_sheet="K.01")
    assert rf.total_row == 3
    assert rf.ending_totals.get("net_value") == Decimal("900")
    assert len(rf.amount_column_bindings) == 3
    assert all(b.period_role.value == "unknown" for b in rf.amount_column_bindings)


def test_rollforward_dual_period_audit2_audit3():
    rows: list[tuple] = [()] * 56
    rows[51] = ("", "审2", "", "", "", "", "", "审3", "", "", "", "")
    rows[53] = (
        "",
        "固定资产类别",
        "原值",
        "累计折旧",
        "减值准备",
        "净值",
        "",
        "原值",
        "累计折旧",
        "减值准备",
        "净值",
    )
    rows[54] = ("", "办公设备", 100, 10, 0, 90, "", 110, 12, 0, 98)
    rows[55] = ("", "合计", 100, 10, 0, 90, "", 110, 12, 0, 98)
    rows[32] = ("", "变动", "原值变动金额", "本年VS上年", 0, 0, 0, 10)

    rf = parse_rollforward_rows(rows, source_sheet="K.01")
    from ingest.models import RollforwardPeriodRole

    assert any(b.period_role == RollforwardPeriodRole.OPENING for b in rf.amount_column_bindings)
    assert any(b.period_role == RollforwardPeriodRole.ENDING for b in rf.amount_column_bindings)
    assert rf.has_movement_rows
    assert rf.opening_totals.get("original_value") == Decimal("100")
    assert rf.ending_totals.get("original_value") == Decimal("110")


def test_rollforward_opening_ending_bindings_and_totals():
    rows = [
        ("类别", "期初原值", "期末原值", "期初累计折旧", "期末累计折旧", "期末净值"),
        ("机器", 1000, 1200, 200, 240, 960),
        ("合计", 1000, 1200, 200, 240, 960),
    ]
    rf = parse_rollforward_rows(rows, source_sheet="K.01")
    roles = {b.period_role.value for b in rf.amount_column_bindings}
    assert "opening" in roles
    assert "ending" in roles
    assert rf.opening_totals.get("original_value") == Decimal("1000")
    assert rf.ending_totals.get("original_value") == Decimal("1200")
    assert rf.ending_totals.get("net_value") == Decimal("960")


def test_workbook_ingest_does_not_truncate_addition_list(tmp_path: Path):
    path = tmp_path / "long_addition.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "新增清单"
    ws.append(["固定资产编号", "原值", "新增方式"])
    for idx in range(300):
        ws.append([f"FA-TEST-{idx:03d}", 1, "购置"])
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path, max_rows=150)

    assert ctx.addition_list is not None
    assert len(ctx.addition_list.records) == 300


def test_workbook_ingest_does_not_truncate_fa_list(tmp_path: Path):
    path = tmp_path / "long_fa_list.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FA list"
    ws.append(["固定资产编号", "固定资产名称", "原值", "累计折旧", "净值"])
    for idx in range(260):
        ws.append([f"FA-TEST-{idx:03d}", f"设备{idx}", 100, 10, 90])
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path, max_rows=50)

    assert ctx.fa_list is not None
    assert len(ctx.fa_list.records) == 260


def test_workbook_ingest_does_not_truncate_disposal_list(tmp_path: Path):
    path = tmp_path / "long_disposal.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "处置清单"
    ws.append(
        [
            "固定资产编号",
            "固定资产名称",
            "原值",
            "累计折旧",
            "净值",
            "处置日期",
            "处置方式",
        ]
    )
    for idx in range(240):
        ws.append([f"FA-D-{idx:03d}", f"旧设备{idx}", 100, 10, 90, "2025-01-01", "报废"])
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path, max_rows=50)

    assert ctx.disposal_list is not None
    assert len(ctx.disposal_list.records) == 240


def test_rollforward_detects_k01_sections_six_blocks():
    rows = [
        ("K.01 Agree SL to GL",),
        ("表1", "固定资产类别", "账面数", "审定数", "年初余额", "年末余额"),
        ("机器设备", 1000, 1000, 1000, 1000, 1000),
        ("原值变动金额", "TB-原值", "差异"),
        ("表2", "固定资产清单", "分类汇总"),
        ("表3", "表2 check with 表1"),
        ("表4", "折旧费用与利润表科目核对"),
        ("Notes", "超过SAD差异调查", "超过TE转K.02/K.03", "拒绝执行原因"),
    ]
    rf = parse_rollforward_rows(rows, source_sheet="K.01 Agree SL to GL")
    presence = rf.section_presence
    assert presence["b1_bkd_main_table"] is True
    assert presence["b2_movement_tb_reconciliation"] is True
    assert presence["b3_table2_fa_summary"] is True
    assert presence["b4_table3_check_with_table1"] is True
    assert presence["b5_table4_depreciation_pl"] is True
    assert presence["b6_notes_investigation_routing"] is True
    assert "k01_sections_detected:6/6" in rf.notes
    assert "表1" in rf.section_evidence["b1_bkd_main_table"]


def test_rollforward_detects_partial_sections_when_missing_blocks():
    rows = [
        ("K.01 Agree SL to GL",),
        ("固定资产类别", "原值", "累计折旧", "净值"),
        ("设备A", 100, 10, 90),
        ("合计", 100, 10, 90),
    ]
    rf = parse_rollforward_rows(rows, source_sheet="K.01")
    # 仅有类别简表：行锚点可识别 b1，但无表2–表6
    assert rf.section_presence["b1_bkd_main_table"] is True
    assert "b1_bkd_main_table" in rf.section_regions
    assert rf.section_presence["b2_movement_tb_reconciliation"] is False
    assert rf.section_presence["b3_table2_fa_summary"] is False
    assert rf.section_presence["b6_notes_investigation_routing"] is False
    assert rf.recognition_confidence < 0.65
    assert "k01_recognition_needs_review" in rf.notes


def test_rollforward_b1_total_prefers_region_over_insert_block():
    """插入表与标准字段重叠时，合计行应取自 b1 区块而非上方插入表。"""
    rows = [
        ("项目特殊说明",),
        ("原值", "累计折旧", "净值"),
        ("插入合计", 9999, 1000, 8999),
        ("表1", "固定资产类别", "原值", "累计折旧", "净值"),
        ("", "设备A", 100, 10, 90),
        ("", "合计", 100, 10, 90),
    ]
    rf = parse_rollforward_rows(rows, source_sheet="K.01")
    assert rf.total_row == 6
    assert rf.ending_totals.get("net_value") == Decimal("90")
    assert any(c.startswith("ambiguous_total_rows") for c in rf.section_conflicts) is False


def test_rollforward_section_regions_and_conflicts_on_six_blocks():
    rows = [
        ("表1", "固定资产类别", "原值", "净值"),
        ("原值变动金额", "TB-原值", "差异"),
        ("表2", "固定资产清单"),
        ("表3", "表2 check with 表1"),
        ("表4", "折旧费用与利润表科目核对"),
        ("Notes", "SAD调查"),
    ]
    rf = parse_rollforward_rows(rows, source_sheet="K.01")
    assert len(rf.section_regions) >= 5
    assert rf.section_regions["b3_table2_fa_summary"].anchor_row == 3
    assert rf.recognition_confidence >= 0.65


def test_reconciliation_match(reconciliation_workbook: Path):
    ctx = load_workbook_ingest(reconciliation_workbook)
    assert ctx.fa_list is not None
    assert ctx.rollforward is not None
    assert ctx.fa_list.amount_basis is not None
    assert ctx.fa_list.amount_basis.status.value == "confirmed", ctx.fa_list.amount_basis
    assert ctx.fa_list.amount_basis.currency_role.value != "unknown", ctx.fa_list.amount_basis
    assert ctx.fa_list.fa_profile is not None
    assert ctx.fa_list.fa_profile.population.status.value == "ready", ctx.fa_list.fa_profile.population
    net_checks = [c for c in ctx.reconciliations if c.link_id == "fa_list_rollforward_net"]
    assert len(net_checks) == 1
    assert net_checks[0].status == ReconciliationStatus.MATCH


def test_reconciliation_mismatch(mismatch_workbook: Path):
    ctx = load_workbook_ingest(mismatch_workbook)
    net_checks = [c for c in ctx.reconciliations if c.link_id == "fa_list_rollforward_net"]
    assert net_checks[0].status == ReconciliationStatus.MISMATCH


def test_load_workbook_ingest_on_fixture():
    path = FIXTURES / "workbook_with_lead.xlsx"
    if not path.is_file():
        pytest.skip("fixture missing")
    ctx = load_workbook_ingest(path)
    assert ctx.structure.sheets_by_kind
    assert ctx.fa_list is not None or ctx.fa_list_sheets


def test_rollforward_summary_contains_section_outputs():
    rows = [
        ("K.01 Agree SL to GL",),
        ("表1", "固定资产类别", "账面数", "审定数"),
        ("原值变动金额", "TB-原值", "差异"),
        ("表2", "固定资产清单"),
        ("表3", "表2 check with 表1"),
        ("表4", "折旧费用与利润表科目核对"),
        ("Notes", "超过SAD差异调查", "超过TE转K.02/K.03"),
    ]
    rf = parse_rollforward_rows(rows, source_sheet="K.01")
    from ingest.workbook_ingest import _rollforward_summary

    data = _rollforward_summary(rf)
    assert data is not None
    assert "section_presence" in data
    assert "section_evidence" in data
    assert data["section_presence"]["b1_bkd_main_table"] is True
    assert isinstance(data["section_evidence"]["b6_notes_investigation_routing"], list)
