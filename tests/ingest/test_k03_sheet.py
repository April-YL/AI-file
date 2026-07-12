from pathlib import Path

import openpyxl
import pytest

from ingest.k03_sheet import (
    EXECUTION_PATH_POLICY_REVIEW,
    EXECUTION_PATH_SAP_HIGH,
    EXECUTION_PATH_SAP_MEDIUM,
    EXECUTION_PATH_SAP_PLUS_TOD_SAMPLING,
    EXECUTION_PATH_TOD_BY_ITEM,
    EXECUTION_PATH_TOD_SAMPLING,
    EXECUTION_PATH_UNKNOWN,
    INGEST_DEPTH_DETAILED,
    INGEST_DEPTH_LIGHTWEIGHT,
    K03_BRANCH_DEPRECIATION_TEST,
    K03_BRANCH_POLICY_REVIEW,
    RULE_STATUS_READY_FOR_LATER_RULES,
    RULE_STATUS_LATER_PHASE,
    load_k03_sheets_from_workbook,
)
from ingest.workbook_ingest import load_workbook_ingest

_ROOT = Path(__file__).resolve().parents[2]


def _save(wb: openpyxl.Workbook, path: Path) -> Path:
    wb.save(path)
    wb.close()
    return path


def _append_by_item_sheet(wb: openpyxl.Workbook, title: str = "K.03.2 折旧测试"):
    ws = wb.create_sheet(title)
    ws.append(["说明", "TOD-by item 全量折旧测试"])
    ws.append([])
    ws.append(
        [
            "资产编号",
            "资产名称",
            "资产类别",
            "原值",
            "残值率",
            "折旧年限",
            "折旧起始日期",
            "管理层计算折旧",
            "审计重新计算折旧",
            "差异",
            "结论",
            "备注字段",
        ]
    )
    ws.append(
        [
            "FA-TEST-001",
            "设备A",
            "机器设备",
            1200,
            "5%",
            60,
            "2025-01-01",
            228,
            228,
            0,
            "通过",
            "extra",
        ]
    )
    ws.append(["合计", "", "", 1200, "", "", "", 228, 228, 0, "", ""])
    ws.append(["结论", "未见重大差异"])
    return ws


def _append_tod_sampling_sheet(wb: openpyxl.Workbook, title: str = "K.03.2 折旧测试TOD-抽样"):
    ws = wb.create_sheet(title)
    ws.append(["说明", "抽样 TOD 折旧测试"])
    ws.append([])
    ws.append(["折旧费用总体", "", "", 1000])
    ws.append(["Breakdown中折旧计提金额", "", "", 1000])
    ws.append(["减：测试的关键项目", "", "", 200])
    ws.append(["在下方描述选择关键项目的理由", "选取本期计提折旧超过 TT 的资产"])
    ws.append([])
    ws.append(
        [
            "样本数量",
            "样本类型",
            "固定资产类别",
            "固定资产编号",
            "固定资产名称",
            "原值",
            "资本开始折旧的日期",
            "使用寿命(月)",
            "折旧方法",
            "账面计提折旧费用",
            "重新计算折旧费用",
            "差异",
            "获得的证据/支持的描述",
            "1",
            "2",
        ]
    )
    ws.append(
        [
            1,
            "关键项",
            "机器设备",
            "FA-TEST-001",
            "设备A",
            1200,
            "2025-01-01",
            60,
            "直线法",
            228,
            228,
            0,
            "公司折旧政策",
            "Y",
            "Y",
        ]
    )
    ws.append(["结论", "样本测试未见异常"])
    return ws


def _append_tod_sampling_output_sheet(
    wb: openpyxl.Workbook,
    title: str = "K.03.2a 折旧选样输出",
):
    ws = wb.create_sheet(title)
    ws.append(["Skywind-TOD", "抽样策略工作底稿"])
    ws.append(["客户名称", "", "", "", "ABC公司"])
    ws.append(["可容忍误差 （TE）", "", "", "", 1000])
    ws.append(["综合风险评估", "", "", "", "最低"])
    ws.append(["抽样货币单元", "", "", "", "本期计提折旧"])
    ws.append(["总体金额", "", "", "", 1000])
    ws.append(["关键项目数量", "", "", "", 1])
    ws.append(["关键项目金额", "", "", "", 200])
    ws.append(["双重目的", "", "", "", "否"])
    ws.append(["高估", "", "", "", "是"])
    ws.append(["保证水平", "", "", "", "几乎没有"])
    ws.append(["样本池", "", "", "", 800])
    ws.append(["预期错报", "", "", "", 0])
    ws.append(["抽样方法", "", "", "", "随机抽样 (Random)"])
    ws.append([])
    ws.append(["样本", "抽样ID", "样本类型", "固定资产类别", "固定资产编号", "固定资产名称", "本期计提折旧"])
    ws.append([1, 10, "关键项", "机器设备", "FA-TEST-001", "设备A", 228])
    return ws


def _append_current_depreciation_sheet(wb: openpyxl.Workbook, title: str = "本期计提"):
    ws = wb.create_sheet(title)
    ws.append(
        [
            "固定资产类别",
            "固定资产编号",
            "固定资产名称",
            "入账开始日期",
            "使用寿命(月)",
            "残值率",
            "原值",
            "累计折旧",
            "减值准备",
            "净值",
            "本期计提折旧",
        ]
    )
    ws.append(["机器设备", "FA-TEST-001", "设备A", "2025-01-01", 60, 0.05, 1200, 100, 0, 1100, 228])
    ws.append(["合计", "", "", "", "", "", 1200, 100, 0, 1100, 228])
    return ws


def _append_lead_sheet_with_vm_cra(
    wb: openpyxl.Workbook,
    *,
    cra: str = "Moderate",
    tt: int = 500,
    title: str = "K.00 Lead Sheet",
):
    ws = wb.create_sheet(title)
    ws["B2"] = "客户名称"
    ws["C2"] = "ABC公司"
    ws["B5"] = "可容忍误差（TE）"
    ws["C5"] = 1000
    ws["B6"] = "名义金额（SAD）"
    ws["C6"] = 100
    ws["B14"] = "认定"
    ws["C14"] = "CRA"
    ws["D14"] = "各项认定1"
    ws["E14"] = "所有相关认定2"
    ws["B15"] = "完整性（C）"
    ws["C15"] = "Minimal"
    ws["D15"] = 1000
    ws["E15"] = 1000
    ws["B16"] = "计价/计量（V/M）"
    ws["C16"] = cra
    ws["D16"] = tt
    return ws


def _append_lead_sheet_without_vm_cra(
    wb: openpyxl.Workbook,
    *,
    title: str = "K.00 Lead Sheet",
):
    ws = wb.create_sheet(title)
    ws["B2"] = "Client"
    ws["C2"] = "ABC Company"
    ws["B5"] = "TE"
    ws["C5"] = 1000
    ws["B14"] = "Assertion"
    ws["C14"] = "CRA"
    ws["D14"] = "TT"
    ws["E14"] = "All relevant"
    ws["B15"] = "Completeness (C)"
    ws["C15"] = "Minimal"
    ws["D15"] = 1000
    ws["E15"] = 1000
    ws["B16"] = "Existence (E)"
    ws["C16"] = "Low"
    ws["D16"] = 750
    ws["E16"] = 1000
    return ws


def test_loads_independent_tod_by_item_dataset(tmp_path: Path):
    path = tmp_path / "K1 SWP 固定资产 202YMMDD XYZ公司（By item折旧测试）.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _append_by_item_sheet(wb)
    _save(wb, path)

    datasets = load_k03_sheets_from_workbook(path)

    assert len(datasets) == 1
    ds = datasets[0]
    assert ds.k03_branch == K03_BRANCH_DEPRECIATION_TEST
    assert ds.execution_path == EXECUTION_PATH_TOD_BY_ITEM
    assert ds.ingest_depth == INGEST_DEPTH_DETAILED
    assert ds.rule_status == RULE_STATUS_READY_FOR_LATER_RULES
    assert ds.sheet_name == "K.03.2 折旧测试"
    assert ds.header_rows == [3]
    assert ds.detail_table_range is not None
    assert ds.detail_table_range.start_row == 4
    assert ds.total_rows == [5]
    assert ds.conclusion_area is not None
    assert ds.row_count == 1
    assert ds.column_count == 12
    assert ds.detail_table_ref is not None
    assert ds.detail_table_ref.sheet_name == "K.03.2 折旧测试"
    assert "original_value" in ds.normalized_column_map
    assert "management_depreciation" in ds.normalized_column_map
    assert "audit_recalculated_depreciation" in ds.normalized_column_map
    assert "depreciation_difference" in ds.normalized_column_map
    assert "备注字段" in ds.unmapped_columns
    assert "original_value" in ds.amount_columns
    assert "depreciation_start_date" in ds.date_columns
    assert ds.preview_rows[0]["cell_refs"]["asset_id"] == "A4"
    data = ds.to_dict()
    assert "detail_rows" not in data
    assert data["detail_table_ref"]["start_row"] == 4
    assert data["summary"]["mapped_field_count"] >= 8
    assert "preview_rows" not in data["llm_candidate_context"]


def test_standard_k032_tod_uses_structure_not_name_only(tmp_path: Path):
    path = tmp_path / "standard.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("K.03.2 折旧测试TOD")
    ws.append(["程序说明", "抽样 TOD 测试"])
    ws.append(["样本编号", "资产编号", "原值", "本期折旧", "检查程序", "结论"])
    ws.append([1, "FA-TEST-001", 1000, 100, "检查折旧凭证", "通过"])
    ws.append(["结论", "样本测试未见异常"])
    _save(wb, path)

    ds = load_k03_sheets_from_workbook(path)[0]

    assert ds.k03_branch == K03_BRANCH_DEPRECIATION_TEST
    assert ds.execution_path == EXECUTION_PATH_TOD_SAMPLING
    assert ds.ingest_depth == INGEST_DEPTH_LIGHTWEIGHT
    assert ds.rule_status == RULE_STATUS_READY_FOR_LATER_RULES
    assert ds.conclusion_area is not None


def test_standard_k032_can_be_classified_as_tod_by_item_by_structure(tmp_path: Path):
    path = tmp_path / "standard_by_item.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _append_by_item_sheet(wb, title="K.03.2 折旧测试TOD")
    _save(wb, path)

    ds = load_k03_sheets_from_workbook(path)[0]

    assert ds.execution_path == EXECUTION_PATH_TOD_BY_ITEM
    assert ds.ingest_depth == INGEST_DEPTH_DETAILED


def test_unknown_tod_writes_warning_instead_of_forcing_path(tmp_path: Path):
    path = tmp_path / "unknown_tod.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("K.03.2 折旧测试TOD")
    ws.append(["项目", "说明"])
    ws.append(["折旧测试", "结构不足，无法判断执行路径"])
    _save(wb, path)

    ds = load_k03_sheets_from_workbook(path)[0]

    assert ds.execution_path == EXECUTION_PATH_UNKNOWN
    assert "k03_tod_execution_path_not_identified" in ds.warnings


def test_sap_medium_and_high_templates_are_recognized_as_ready_paths(tmp_path: Path):
    path = tmp_path / "sap.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.1 SAP-中精确度"
    ws.append(["SAP 中精确度测试"])
    ws2 = wb.create_sheet("K.03.1 SAP-高精确度")
    ws2.append(["SAP 高精确度测试"])
    _save(wb, path)

    datasets = load_k03_sheets_from_workbook(path)
    paths = {ds.sheet_name: ds.execution_path for ds in datasets}

    assert paths["K.03.1 SAP-中精确度"] == EXECUTION_PATH_SAP_MEDIUM
    assert paths["K.03.1 SAP-高精确度"] == EXECUTION_PATH_SAP_HIGH
    assert all(ds.ingest_depth == INGEST_DEPTH_LIGHTWEIGHT for ds in datasets)
    assert all(ds.rule_status == RULE_STATUS_READY_FOR_LATER_RULES for ds in datasets)
    assert all(not ds.unsupported_or_later_phase for ds in datasets)


def test_generic_k031_sap_is_identified_without_forcing_precision(tmp_path: Path):
    path = tmp_path / "generic_sap.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.1 SAP"
    ws.append(["K.03.1 实质性分析程序"])
    ws.append(["实体类型", "非复杂实体"])
    ws.append(["TE", 1000])
    ws.append(["CRA", "Moderate"])
    _save(wb, path)

    ds = load_k03_sheets_from_workbook(path)[0]

    assert ds.k03_branch == K03_BRANCH_DEPRECIATION_TEST
    assert ds.execution_path == EXECUTION_PATH_UNKNOWN
    assert ds.template_type == "sap"
    assert "k03_sap_precision_not_identified" in ds.warnings
    assert ds.summary["sap_entity_type"] == "非复杂实体"
    assert ds.summary["sap_te"] == 1000
    assert ds.summary["sap_cra"] == "Moderate"


def test_k032a_depreciation_sampling_output_is_auxiliary_to_tod_sampling_path(tmp_path: Path):
    path = tmp_path / "tod_sampling_output.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _append_tod_sampling_output_sheet(wb)
    _save(wb, path)

    ds = load_k03_sheets_from_workbook(path)[0]

    assert ds.k03_branch == K03_BRANCH_DEPRECIATION_TEST
    assert ds.execution_path == EXECUTION_PATH_TOD_SAMPLING
    assert ds.template_type == "tod_sampling_output"
    assert ds.ingest_depth == INGEST_DEPTH_LIGHTWEIGHT
    assert ds.summary["sample_output_sampling_currency"] == "本期计提折旧"
    assert ds.summary["sample_output_sampling_method"] == "随机抽样 (Random)"


def test_sap_plus_tod_structure_keeps_component_paths_for_later_strategy_review(tmp_path: Path):
    path = tmp_path / "sap_plus_tod.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.1 SAP-中精确度"
    ws.append(["K.03.1 SAP-中精确度"])
    ws.append(["CRA", "Moderate"])
    _append_tod_sampling_sheet(wb)
    _append_tod_sampling_output_sheet(wb)
    _save(wb, path)

    datasets = load_k03_sheets_from_workbook(path)
    by_template = {ds.template_type: ds for ds in datasets}

    assert len(datasets) == 3
    assert by_template["sap_medium_precision"].execution_path == EXECUTION_PATH_SAP_MEDIUM
    assert by_template["tod_sampling"].execution_path == EXECUTION_PATH_TOD_SAMPLING
    assert by_template["tod_sampling_output"].execution_path == EXECUTION_PATH_TOD_SAMPLING
    # SAP+TOD is a strategy combination prepared from component sheets, not a duplicate dataset.
    assert all(ds.execution_path != EXECUTION_PATH_SAP_PLUS_TOD_SAMPLING for ds in datasets)


def test_workbook_profile_prefers_high_precision_sap_for_non_minimal_cra(tmp_path: Path):
    path = tmp_path / "k03_profile_multi_path.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _append_lead_sheet_with_vm_cra(wb, cra="Moderate", tt=500)
    ws_medium = wb.create_sheet("K.03.1 SAP-中精确度")
    ws_medium.append(["实体类型", "非复杂实体"])
    ws_medium.append(["TE", 1000])
    ws_medium.append(["CRA", "Moderate"])
    ws_high = wb.create_sheet("K.03.1 SAP-高精确度")
    ws_high.append(["实体类型", "非复杂实体"])
    ws_high.append(["TE", 1000])
    ws_high.append(["CRA", "Moderate"])
    _append_tod_sampling_sheet(wb)
    _append_tod_sampling_output_sheet(wb)
    ws_policy = wb.create_sheet("K.03.3 折旧政策复核")
    ws_policy.append(["结论", "折旧政策未见变化"])
    _append_current_depreciation_sheet(wb)
    _save(wb, path)

    ctx = load_workbook_ingest(path)
    profile = ctx.k03_execution_profile

    assert profile is not None
    assert profile.primary_depreciation_path == EXECUTION_PATH_SAP_HIGH
    assert profile.lead_linkage is not None
    assert profile.lead_linkage.cra == "Moderate"
    assert "计价" in (profile.lead_linkage.assertion or "")
    assert "sap_medium" in profile.component_sheets
    assert "sap_high" in profile.component_sheets
    assert "tod_sampling" in profile.component_sheets
    assert "tod_sampling_output" in profile.component_sheets
    assert "policy_review" in profile.component_sheets
    assert "auxiliary_current_depreciation" in profile.component_sheets
    assert profile.evidence_completeness["policy_review"]["is_independent_required_procedure"] is True
    assert profile.evidence_completeness["auxiliary_current_depreciation"]["is_required_procedure_page"] is False
    assert "k03_multiple_depreciation_test_components_detected" in profile.warnings
    data = ctx.to_dict()
    assert data["k03_execution_profile"]["primary_depreciation_path"] == EXECUTION_PATH_SAP_HIGH


def test_workbook_profile_identifies_sap_plus_tod_when_medium_sap_supplemented_by_sampling(tmp_path: Path):
    path = tmp_path / "k03_profile_sap_plus_tod.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _append_lead_sheet_with_vm_cra(wb, cra="Low", tt=750)
    ws_medium = wb.create_sheet("K.03.1 SAP-中精确度")
    ws_medium.append(["实体类型", "非复杂实体"])
    ws_medium.append(["TE", 1000])
    ws_medium.append(["CRA", "Low"])
    _append_tod_sampling_sheet(wb)
    _append_tod_sampling_output_sheet(wb)
    _save(wb, path)

    profile = load_workbook_ingest(path).k03_execution_profile

    assert profile is not None
    assert profile.primary_depreciation_path == EXECUTION_PATH_SAP_PLUS_TOD_SAMPLING
    assert "sap_medium" in profile.component_sheets
    assert "tod_sampling" in profile.component_sheets
    assert "tod_sampling_output" in profile.component_sheets


def test_workbook_profile_does_not_force_first_lead_cra_row_without_vm(tmp_path: Path):
    path = tmp_path / "k03_profile_no_vm_cra.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _append_lead_sheet_without_vm_cra(wb)
    ws_medium = wb.create_sheet("K.03.1 SAP-中精确度")
    ws_medium.append(["实体类型", "非复杂实体"])
    ws_medium.append(["TE", 1000])
    ws_medium.append(["CRA", "Low"])
    _save(wb, path)

    profile = load_workbook_ingest(path).k03_execution_profile

    assert profile is not None
    assert profile.lead_linkage is None
    assert "k03_lead_depreciation_cra_tt_not_identified" in profile.warnings
    assert "k03_sap_medium_without_high_or_tod_for_non_minimal_cra" in profile.warnings


def test_current_depreciation_sheet_is_not_required_k03_procedure_page(tmp_path: Path):
    path = tmp_path / "current_depreciation_only.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _append_current_depreciation_sheet(wb)
    _save(wb, path)

    datasets = load_k03_sheets_from_workbook(path)

    assert datasets == []


def test_policy_review_is_lightweight_k03_branch(tmp_path: Path):
    path = tmp_path / "policy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.03.3 折旧政策复核"
    ws.append(["说明", "复核折旧政策是否与上期一致"])
    ws.append(["政策描述", "机器设备使用年限 5 年，残值率 5%"])
    ws.append(["结论", "政策未见变化"])
    _save(wb, path)

    ds = load_k03_sheets_from_workbook(path)[0]

    assert ds.k03_branch == K03_BRANCH_POLICY_REVIEW
    assert ds.execution_path == EXECUTION_PATH_POLICY_REVIEW
    assert ds.ingest_depth == INGEST_DEPTH_LIGHTWEIGHT
    assert ds.rule_status == RULE_STATUS_LATER_PHASE
    assert ds.unsupported_or_later_phase is False
    assert ds.llm_candidate_context["candidate_for"] == "depreciation_policy_semantic_review"


def test_workbook_ingest_exposes_k03_sheets(tmp_path: Path):
    path = tmp_path / "workbook.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _append_by_item_sheet(wb)
    ws = wb.create_sheet("K.03.3 折旧政策复核")
    ws.append(["结论", "政策未见变化"])
    _save(wb, path)

    ctx = load_workbook_ingest(path)

    assert len(ctx.k03_sheets) == 2
    branches = {ds.k03_branch for ds in ctx.k03_sheets}
    assert branches == {K03_BRANCH_DEPRECIATION_TEST, K03_BRANCH_POLICY_REVIEW}
    data = ctx.to_dict()
    assert len(data["k03_sheets"]) == 2
    assert data["k03_execution_profile"]["component_sheets"]["tod_by_item"][0]["sheet_name"] == "K.03.2 折旧测试"
    assert "detail_rows" not in data["k03_sheets"][0]


def test_workbook_ingest_does_not_truncate_tod_by_item_detail_rows(tmp_path: Path):
    path = tmp_path / "long_k03_by_item.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("K.03.2 折旧测试")
    ws.append(
        [
            "资产编号",
            "资产名称",
            "原值",
            "残值率",
            "折旧年限",
            "管理层计算折旧",
            "审计重新计算折旧",
            "差异",
        ]
    )
    for idx in range(250):
        ws.append([f"FA-TEST-{idx:03d}", f"设备{idx}", 1200, "5%", 60, 228, 228, 0])
    ws.append(["合计", "", 300000, "", "", 57000, 57000, 0])
    _save(wb, path)

    ctx = load_workbook_ingest(path, max_rows=50)

    assert ctx.k03_sheets[0].execution_path == EXECUTION_PATH_TOD_BY_ITEM
    assert ctx.k03_sheets[0].row_count == 250
    assert ctx.k03_sheets[0].column_count == 8
    assert ctx.k03_sheets[0].total_rows == [252]
    assert ctx.k03_sheets[0].detail_table_ref is not None
    assert ctx.k03_sheets[0].detail_table_ref.end_row == 252
    assert len(ctx.k03_sheets[0].preview_rows) <= 5
    assert "preview_rows" not in ctx.k03_sheets[0].llm_candidate_context
    assert "detail_rows" not in ctx.k03_sheets[0].to_dict()


def test_schema_reserves_sap_plus_tod_sampling_path():
    assert EXECUTION_PATH_SAP_PLUS_TOD_SAMPLING == "sap_plus_tod_sampling"


def test_local_sop_workbook_profile_identifies_k03_components_when_available():
    path = _ROOT / "固定资产质检agent" / "资料库" / "FY26_SOP K1 SWP 固定资产.xlsx"
    if not path.is_file():
        pytest.skip("SOP Excel workbook is not available in the local materials library")

    ctx = load_workbook_ingest(path, max_rows=120)
    profile = ctx.k03_execution_profile

    assert profile is not None
    assert "sap_medium" in profile.component_sheets
    assert "sap_high" in profile.component_sheets
    assert "tod_by_item" in profile.component_sheets
    assert "tod_sampling" in profile.component_sheets
    assert "tod_sampling_output" in profile.component_sheets
    assert "policy_review" in profile.component_sheets
    assert "auxiliary_current_depreciation" in profile.component_sheets
    assert profile.evidence_completeness["policy_review"]["is_independent_required_procedure"] is True
    assert profile.evidence_completeness["auxiliary_current_depreciation"]["is_required_procedure_page"] is False


def test_local_j_copy_profile_keeps_multi_path_components_without_requiring_all_methods():
    path = _ROOT / "固定资产质检agent" / "案例库" / "K1 固定资产 20251231 J有限公司 - 副本.xlsx"
    if not path.is_file():
        pytest.skip("J copy workbook is not available in the local case library")

    ctx = load_workbook_ingest(path, max_rows=120)
    profile = ctx.k03_execution_profile

    assert profile is not None
    assert "sap_medium" in profile.component_sheets
    assert "sap_high" in profile.component_sheets
    assert "tod_by_item" in profile.component_sheets
    assert "tod_sampling" in profile.component_sheets
    assert "tod_sampling_output" in profile.component_sheets
    assert "policy_review" in profile.component_sheets
    assert profile.primary_depreciation_path in {
        EXECUTION_PATH_SAP_MEDIUM,
        EXECUTION_PATH_SAP_HIGH,
        EXECUTION_PATH_SAP_PLUS_TOD_SAMPLING,
        EXECUTION_PATH_TOD_BY_ITEM,
        EXECUTION_PATH_TOD_SAMPLING,
        EXECUTION_PATH_UNKNOWN,
    }
    assert "k03_multiple_depreciation_test_components_detected" in profile.warnings
