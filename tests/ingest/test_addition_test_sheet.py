from pathlib import Path

import openpyxl

from ingest.workbook_ingest import load_workbook_ingest


def _base_workbook(path: Path, *, summary_status: str = "是") -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "汇总"
    ws.append(["程序", "工作表", "是否执行", "不执行的原因"])
    ws.append(["K.02.1 新增测试", "K.02.1 新增测试", summary_status, ""])

    ws_add = wb.create_sheet("新增清单")
    ws_add.append(["固定资产类别", "固定资产编号", "固定资产名称", "入账开始日期", "原值", "新增方式"])
    ws_add.append(["机器设备", "FA-TEST-001", "设备A", "2025-01-01", 1000, "购置"])

    ws_test = wb.create_sheet("K.02.1 新增测试")
    ws_test.append(["记录我们的详细测试所涵盖的总体"])
    ws_test.append(["购置总金额", "", "", "", 1000])
    ws_test.append(["Breakdown中购置金额", "", "", "", 1000])
    ws_test.append(["差异", "", "", "", 0, "Rx", "差异是否需要进一步调查？", "", "否"])
    ws_test.append(["减：测试的关键项目", "", "", "", 200])
    ws_test.append(["进行代表性抽样的剩余总体", "", "", "", 800])
    ws_test.append(["关键项数量", "", "", "", 1])
    ws_test.append(["定量关键项金额", "", "", "", 200])
    ws_test.append(["代表性样本量", "", "", "", 1])
    ws_test.append(["样本选择方法", "", "", "", "随机抽样"])
    ws_test.append(["在下方描述选择关键项目的理由", "", "", "", "", "TT", "定量关键项示例"])
    ws_test.append(["无异常情况"])
    ws_test.append([])
    ws_test.append(
        [
            "样本类型",
            "固定资产类别",
            "总账账户代码",
            "固定资产编号",
            "固定资产名称",
            "资产原价",
            "资本化日期",
            "使用寿命(月)",
            "残值率",
            "折旧方法",
            "资产原价（通过审计证据/支持性文件取得）",
            "获得的证据/支持的描述（例如，我们检查的控制权转移时点的单据/销售发票/合同/订单等）",
            "资产原价差异",
            "1",
            "2",
            "3",
            "4",
        ]
    )
    ws_test.append(
        [
            "代表性样本",
            "机器设备",
            "2240",
            "FA-TEST-001",
            "设备A",
            1000,
            "2025-01-01",
            120,
            0.05,
            "直线法",
            1000,
            "验收单；发票；合同",
            0,
            "Y",
            "Y",
            "Y",
            "Y",
        ]
    )

    ws_sample = wb.create_sheet("K.02.1a 新增选样输出")
    ws_sample.append(["源数据汇总"])
    ws_sample.append(["已上传数据", "", "", "", 1000])
    ws_sample.append(["必要的数据排除项", "", "", "", 0])
    ws_sample.append(["样本池总体金额", "", "", "", 1000])
    ws_sample.append(["代表性总体价值", "", "", "", 800])
    ws_sample.append(["总体与会计记录进行核对"])
    ws_sample.append(["总金额", "", "", "", 1000])
    ws_sample.append(["会计记录的重大账户余额或活动", "", "", "", 1000])
    ws_sample.append(["差额", "", "", "", 0])
    ws_sample.append(["抽样结论汇总"])
    ws_sample.append(["关键项数量", "", "", "", 1])
    ws_sample.append(["定量关键项金额", "", "", "", 200])
    ws_sample.append(["代表性样本量", "", "", "", 1])
    ws_sample.append(["代表性样本与关键项数量合计", "", "", "", 2])
    ws_sample.append(["样本选择方法", "", "", "", "随机抽样 (Random)"])
    ws_sample.append(["已选取样本"])
    ws_sample.append(
        [
            "源样本#",
            "抽样ID",
            "样本类型",
            "固定资产类别",
            "固定资产编号",
            "固定资产名称",
            "入账开始日期",
            "使用寿命(月)",
            "残值率",
            "原值",
            "新增方式",
        ]
    )
    ws_sample.append([1, 140, "代表性样本", "机器设备", "FA-TEST-001", "设备A", "2025-01-01", 120, 0.05, 1000, "购置"])

    wb.save(path)
    return wb


def test_addition_execution_path_complete(tmp_path: Path):
    path = tmp_path / "addition_complete.xlsx"
    wb = _base_workbook(path)
    wb.close()

    ctx = load_workbook_ingest(path)

    assert ctx.addition_test is not None
    assert ctx.addition_test.source_sheet == "K.02.1 新增测试"
    assert ctx.addition_test.amounts["purchase_population_amount"].amount == "1000"
    assert ctx.addition_test.amounts["rollforward_purchase_amount"].amount == "1000"
    assert ctx.addition_test.amounts["difference_amount"].amount == "0"
    assert len(ctx.addition_test.tested_samples) == 1
    assert ctx.addition_test.tested_samples[0].asset_id == "FA-TEST-001"
    assert ctx.addition_test.tested_samples[0].original_value == "1000"
    assert ctx.addition_test.tested_samples[0].gl_account_code == "2240"
    assert ctx.addition_test.tested_samples[0].useful_life_months == "120"
    assert ctx.addition_test.tested_samples[0].salvage_rate == "0.05"
    assert ctx.addition_test.tested_samples[0].depreciation_method == "直线法"
    assert ctx.addition_test.module_assessments
    modules = {m.module_key: m for m in ctx.addition_test.module_assessments}
    assert set(modules) == {
        "execution_path",
        "population_definition",
        "amount_reconciliation",
        "key_item_representation",
        "sample_table",
        "exception_summary",
    }
    assert modules["execution_path"].status == "recognized"
    assert modules["population_definition"].status == "recognized"
    assert modules["amount_reconciliation"].status == "recognized"
    assert modules["key_item_representation"].status == "recognized"
    assert modules["sample_table"].status == "recognized"
    assert modules["exception_summary"].status == "recognized"

    assert ctx.addition_sample_output is not None
    assert ctx.addition_sample_output.source_sheet == "K.02.1a 新增选样输出"
    assert ctx.addition_sample_output.amounts["uploaded_data_amount"].amount == "1000"
    assert ctx.addition_sample_output.amounts["sample_pool_amount"].amount == "1000"
    assert ctx.addition_sample_output.amounts["sample_method"].amount == "随机抽样 (Random)"
    assert len(ctx.addition_sample_output.selected_samples) == 1
    assert ctx.addition_sample_output.selected_samples[0].asset_id == "FA-TEST-001"
    assert ctx.addition_sample_output.selected_samples[0].addition_method == "购置"
    assert ctx.addition_sample_output.selected_samples[0].useful_life_months == "120"
    assert ctx.addition_sample_output.selected_samples[0].salvage_rate == "0.05"
    sample_modules = {m.module_key: m for m in ctx.addition_sample_output.module_assessments}
    assert set(sample_modules) == {
        "source_data_summary",
        "sampling_strategy",
        "accounting_reconciliation",
        "selected_samples",
    }
    assert sample_modules["source_data_summary"].status == "recognized"
    assert sample_modules["sampling_strategy"].status == "recognized"
    assert sample_modules["accounting_reconciliation"].status == "recognized"
    assert sample_modules["selected_samples"].status == "recognized"
    assert ctx.addition_execution_path is not None
    assert ctx.addition_execution_path.path_kind == "executed_package_complete"
    assert ctx.addition_execution_path.missing_components == []


def test_addition_test_ignores_right_side_guidance_amount_labels(tmp_path: Path):
    path = tmp_path / "addition_guidance_labels.xlsx"
    wb = _base_workbook(path)
    ws = wb["K.02.1 新增测试"]
    ws.insert_rows(2)
    ws.cell(row=2, column=22, value="差异")
    ws.cell(row=2, column=23, value="样本总体与固定资产清单存在差异，但未进行调查或说明")
    ws.cell(row=2, column=24, value="进行代表性抽样的剩余总体")
    ws.cell(row=2, column=25, value="抽样策略记录：关键项目测试的选择标准和理由未恰当记录")
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path)

    assert ctx.addition_test is not None
    assert ctx.addition_test.amounts["difference_amount"].amount == "0"
    assert ctx.addition_test.amounts["remaining_population_amount"].amount == "800"


def test_addition_sample_output_ignores_right_side_guidance_strategy(tmp_path: Path):
    path = tmp_path / "addition_sample_guidance.xlsx"
    wb = _base_workbook(path)
    ws = wb["K.02.1a 新增选样输出"]
    ws.insert_rows(2)
    ws.cell(row=2, column=18, value="抽样策略：\n在抽样工具中，对抽样策略进行选择。")
    ws.cell(row=2, column=19, value="抽样策略记录：从其他程序取得的证据水平填写错误。")
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path)

    assert ctx.addition_sample_output is not None
    assert ctx.addition_sample_output.amounts["sample_method"].amount == "随机抽样 (Random)"


def test_addition_execution_path_test_sheet_waiver_note(tmp_path: Path):
    path = tmp_path / "addition_waiver_note.xlsx"
    wb = _base_workbook(path)
    ws = wb["K.02.1 新增测试"]
    ws["A11"] = "本期新增固定资产金额小于 TE，且各单项资产均小于 TT，无性质异常，无须执行详细测试。"
    del wb["K.02.1a 新增选样输出"]
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path)

    assert ctx.addition_execution_path is not None
    assert ctx.addition_execution_path.path_kind == "test_sheet_waiver_note"
    assert "K.02.1a 新增选样输出" in ctx.addition_execution_path.missing_components
    assert ctx.addition_execution_path.test_sheet_waiver_note


def test_addition_execution_path_summary_waived(tmp_path: Path):
    path = tmp_path / "addition_summary_waived.xlsx"
    wb = _base_workbook(path, summary_status="否")
    ws = wb["汇总"]
    ws["D2"] = "新增购置金额小于SAD，其他新增已经在在建工程底稿执行程序。"
    del wb["新增清单"]
    del wb["K.02.1 新增测试"]
    del wb["K.02.1a 新增选样输出"]
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path)

    assert ctx.addition_execution_path is not None
    assert ctx.addition_execution_path.path_kind == "summary_waived"
    assert ctx.addition_execution_path.summary_waiver_reason == "新增购置金额小于SAD，其他新增已经在在建工程底稿执行程序。"


def test_addition_execution_path_test_sheet_waiver_note_sad_tod(tmp_path: Path):
    path = tmp_path / "addition_test_sheet_sad_tod.xlsx"
    wb = _base_workbook(path)
    ws = wb["K.02.1 新增测试"]
    ws["A11"] = "本期购置金额小于SAD，不在执行TOD测试；关键项目选择理由已在新增测试中说明。"
    wb.save(path)
    wb.close()

    ctx = load_workbook_ingest(path)

    assert ctx.addition_test is not None
    assert ctx.addition_test.waiver_note_text is not None
    assert ctx.addition_execution_path is not None
    assert ctx.addition_execution_path.path_kind == "test_sheet_waiver_note"
    assert ctx.addition_test.module_assessments
    assert {m.module_key: m.status for m in ctx.addition_test.module_assessments}["execution_path"] == "recognized"
