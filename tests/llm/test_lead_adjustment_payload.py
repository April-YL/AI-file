"""调整汇总 LLM payload 与 LEAD-017 门控单测（不调用 API）。"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from ingest.lead_adjustment_grid import build_adjustment_grid, load_adjustment_grid_for_lead
from ingest.lead_sheet import (
    AdjustmentSummaryRow,
    LeadMovementRow,
    LeadSheetDataset,
    load_lead_from_workbook,
    parse_lead_sheet_rows,
)
from ingest.lead_sheet_blocks import LeadBlock, LeadBlockKind
from llm.config import LlmConfig
from llm.lead_adjustment_review import (
    _LEAD017_ADDITIONAL_SYSTEM,
    RULE_LAYOUT,
    RULE_SEMANTIC,
    build_adjustment_review_payload,
    build_guidance_adjustments,
    build_lead_adjustment_issues,
    extract_layout_and_rows_for_gating,
    run_lead_adjustment_llm_review,
    should_review_adjustments,
)
from rules.lead_adjustment_gating import is_direct_ppe_account, should_run_strict_total_check
from rules.lead_adjustment_internal_consistency import check_lead_adjustment_internal_consistency
from rules.models import Severity

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures"
FIXTURE_EN = FIXTURES / "lead_adjustment_en_debit_credit.xlsx"
FIXTURE_CROSS = FIXTURES / "lead_adjustment_cross_account_aa.xlsx"


def _config(*, enabled: bool = True) -> LlmConfig:
    return LlmConfig(
        enabled=enabled,
        base_url="https://api.example.com/v1",
        api_key="sk-test",
        model="gpt-4o-mini",
    )


def _minimal_lead_with_adjustment_block() -> LeadSheetDataset:
    return LeadSheetDataset(
        source_file="test.xlsx",
        source_sheet="K.00 Lead Sheet",
        blocks=[
            LeadBlock(
                kind=LeadBlockKind.ADJUSTMENT_SUMMARY,
                anchor_row=64,
                start_row=64,
                end_row=80,
                confidence=0.9,
                anchor_text="调整汇总表",
            ),
            LeadBlock(
                kind=LeadBlockKind.MOVEMENT_TABLE,
                anchor_row=40,
                start_row=40,
                end_row=60,
                confidence=0.9,
                anchor_text="引导主表",
            ),
        ],
        movement_rows=[
            LeadMovementRow(
                account_label="原值",
                sheet_ref="K.01",
                values={"audit_adjustment": "100", "book_adjustment": None},
                source_row=49,
            ),
        ],
        adjustment_rows=[
            AdjustmentSummaryRow(
                adjustment_type="未更正审计调整",
                source_row=66,
                raw_cells=["未更正审计调整", "AA1", "原值", "100"],
            ),
        ],
    )


def test_should_review_when_adjustment_rows_present():
    lead = _minimal_lead_with_adjustment_block()
    assert should_review_adjustments(lead) is True


def test_should_review_when_main_has_adjustment_only():
    lead = _minimal_lead_with_adjustment_block()
    lead.adjustment_rows = []
    assert should_review_adjustments(lead) is True


def test_should_review_false_without_adjustment_block():
    lead = LeadSheetDataset(
        source_file="test.xlsx",
        source_sheet="K.00 Lead Sheet",
        movement_rows=[
            LeadMovementRow(
                account_label="原值",
                sheet_ref="K.01",
                values={"audit_adjustment": "100"},
                source_row=49,
            ),
        ],
    )
    assert should_review_adjustments(lead) is False


def test_build_guidance_adjustments():
    lead = _minimal_lead_with_adjustment_block()
    guidance = build_guidance_adjustments(lead)
    assert guidance[0]["account_label"] == "原值"
    assert guidance[0]["audit_adjustment"] == "100"


def test_build_adjustment_review_payload_includes_grid_and_policy():
    lead = _minimal_lead_with_adjustment_block()
    payload = build_adjustment_review_payload(
        lead,
        adjustment_grid={"grid": [["调整类型", "金额"], ["审计调整", "100"]]},
        deterministic_hints=[{"rule_id": "lead_adjustment_internal_consistency"}],
    )
    assert payload["cross_account_policy"] == "flag_not_fail"
    assert payload["adjustment_grid"] == [["调整类型", "金额"], ["审计调整", "100"]]
    assert len(payload["guidance_adjustments"]) == 1
    assert "原值" in payload["ppe_direct_aliases"]


def test_lead017_prompt_prevents_whole_journal_net_comparison():
    prompt = _LEAD017_ADDITIONAL_SYSTEM
    compact = " ".join(prompt.split())
    assert "Do not compare the whole journal-entry net total" in prompt
    assert "Only direct PPE rows may enter direct_ppe_net_amount" in prompt
    assert "Counterparty accounts" in prompt
    assert "do not create a direct mismatch from uncertain evidence" in compact


def test_is_direct_ppe_account():
    assert is_direct_ppe_account("固定资产-原值") is True
    assert is_direct_ppe_account("PPE Cost") is True
    assert is_direct_ppe_account("管理费用") is False
    assert is_direct_ppe_account("SG&A expense") is False


def test_strict_total_gated_by_low_layout():
    lead = _minimal_lead_with_adjustment_block()
    assert should_run_strict_total_check(
        lead,
        layout_result={"confidence": "low", "amount_layout": "unknown"},
    ) is False


def test_strict_total_off_for_indirect_only_rows():
    lead = _minimal_lead_with_adjustment_block()
    assert should_run_strict_total_check(
        lead,
        layout_result={"confidence": "high", "amount_layout": "single_signed_column"},
        extracted_rows=[{"ppe_impact": "indirect", "account_label": "管理费用"}],
    ) is False


def test_strict_total_on_for_direct_rows():
    lead = _minimal_lead_with_adjustment_block()
    assert should_run_strict_total_check(
        lead,
        layout_result={"confidence": "high", "amount_layout": "single_signed_column"},
        extracted_rows=[{"ppe_impact": "direct", "account_label": "原值"}],
    ) is True


def test_llm_review_disabled_returns_empty():
    lead = _minimal_lead_with_adjustment_block()
    issues, review = run_lead_adjustment_llm_review(lead, _config(enabled=False))
    assert issues == []
    assert review is None


def test_llm_review_layout_low_produces_lead_018():
    lead = _minimal_lead_with_adjustment_block()
    mock_review = {
        "layout": {
            "amount_layout": "unknown",
            "confidence": "low",
            "layout_notes": "Dr/Cr 列未识别",
        },
        "rows": [],
        "assessment": "unclear",
    }
    with patch("llm.lead_adjustment_review.chat_completion_json", return_value=mock_review):
        issues, _ = run_lead_adjustment_llm_review(lead, _config())

    assert any(
        i.rule_id == RULE_LAYOUT
        and i.severity == Severity.NEED_REVIEW
        and "版式" in i.message
        for i in issues
    )


def test_llm_review_insufficient_direct_mismatch_warns():
    lead = _minimal_lead_with_adjustment_block()
    mock_review = {
        "layout": {"amount_layout": "single_signed_column", "confidence": "high"},
        "rows": [{"ppe_impact": "direct", "account_label": "原值", "signed_amount": "200"}],
        "assessment": "insufficient",
        "direct_amount_checks": [
            {
                "account_label": "原值",
                "match": False,
                "guidance_signed": "100",
                "summary_signed": "200",
            }
        ],
        "rationale": "direct 行金额与引导表不一致",
    }
    with patch("llm.lead_adjustment_review.chat_completion_json", return_value=mock_review):
        issues = build_lead_adjustment_issues(lead, _config())

    assert any(
        i.rule_id == RULE_SEMANTIC
        and i.severity == Severity.WARN
        and "语义上不足" in i.message
        for i in issues
    )


def test_llm_review_indirect_insufficient_need_review():
    lead = _minimal_lead_with_adjustment_block()
    mock_review = {
        "layout": {"amount_layout": "single_signed_column", "confidence": "high"},
        "rows": [{"ppe_impact": "indirect", "account_label": "管理费用", "adjustment_ref": "AA3"}],
        "assessment": "insufficient",
        "direct_amount_checks": [],
        "cross_account_flags": [{"adjustment_ref": "AA3", "issue": "indirect_adjustment_without_ppe_link_narrative"}],
        "rationale": "跨科目调整缺少对 PPE 影响的说明",
    }
    with patch("llm.lead_adjustment_review.chat_completion_json", return_value=mock_review):
        issues = build_lead_adjustment_issues(lead, _config())

    assert any(
        i.rule_id == RULE_SEMANTIC
        and i.severity == Severity.NEED_REVIEW
        and "跨科目提示" in i.message
        for i in issues
    )


def test_llm_review_sufficient_no_semantic_issue():
    lead = _minimal_lead_with_adjustment_block()
    mock_review = {
        "layout": {"amount_layout": "debit_credit_two_columns", "confidence": "high"},
        "rows": [{"ppe_impact": "direct", "account_label": "原值", "signed_amount": "100"}],
        "assessment": "sufficient",
        "direct_amount_checks": [{"account_label": "原值", "match": True}],
    }
    with patch("llm.lead_adjustment_review.chat_completion_json", return_value=mock_review):
        issues = build_lead_adjustment_issues(lead, _config())

    assert not any(i.rule_id == RULE_SEMANTIC for i in issues)


def test_extract_layout_and_rows_for_gating():
    review = {
        "layout": {"confidence": "high", "amount_layout": "debit_credit_two_columns"},
        "rows": [{"ppe_impact": "indirect"}, "skip-me"],
    }
    layout, rows = extract_layout_and_rows_for_gating(review)
    assert layout is not None
    assert layout["amount_layout"] == "debit_credit_two_columns"
    assert len(rows) == 1


@pytest.fixture
def lead_grid_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "lead_adj.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K.00 Lead Sheet"
    ws["B64"] = "调整汇总表（如不适用请删除）"
    ws["B65"] = "调整类型"
    ws["C65"] = "Ref"
    ws["D65"] = "Account"
    ws["E65"] = "Amount"
    ws["B66"] = "未更正审计调整"
    ws["C66"] = "AA1"
    ws["D66"] = "原值"
    ws["E66"] = 100
    wb.save(path)
    wb.close()
    return path


def test_build_adjustment_grid_from_workbook(lead_grid_xlsx: Path):
    rows = []
    wb = openpyxl.load_workbook(lead_grid_xlsx, data_only=True)
    ws = wb["K.00 Lead Sheet"]
    for row in ws.iter_rows(max_row=80, values_only=True):
        rows.append(row)
    wb.close()
    lead = parse_lead_sheet_rows(rows, source_file=str(lead_grid_xlsx))
    grid = build_adjustment_grid(rows, lead)
    assert grid is not None
    assert grid["grid_row_count"] >= 2
    assert any("调整类型" in (c or "") for row in grid["grid"] for c in row if c)


@pytest.mark.skipif(not FIXTURE_EN.is_file(), reason="run scripts/build_lead_adjustment_fixtures.py")
def test_fixture_en_debit_credit_grid_and_guidance():
    lead = load_lead_from_workbook(FIXTURE_EN)
    assert lead.block(LeadBlockKind.ADJUSTMENT_SUMMARY) is not None
    assert len(lead.adjustment_rows) >= 1

    grid = load_adjustment_grid_for_lead(FIXTURE_EN, lead)
    assert grid is not None
    grid_text = " ".join(c or "" for row in grid["grid"] for c in row)
    assert "Dr" in grid_text
    assert "Cr" in grid_text

    payload = build_adjustment_review_payload(lead, adjustment_grid=grid)
    assert payload["adjustment_grid"] == grid["grid"]
    guidance = payload["guidance_adjustments"]
    orig = next((g for g in guidance if g["account_label"] == "原值"), None)
    assert orig is not None
    assert orig.get("audit_adjustment") in ("500000", "500000.0", 500000)


@pytest.mark.skipif(not FIXTURE_CROSS.is_file(), reason="run scripts/build_lead_adjustment_fixtures.py")
def test_fixture_cross_account_aa_indirect_gating():
    lead = load_lead_from_workbook(FIXTURE_CROSS)
    assert len(lead.adjustment_rows) >= 1
    assert "AA3" in " ".join(str(c) for r in lead.adjustment_rows for c in r.raw_cells if c)

    grid = load_adjustment_grid_for_lead(FIXTURE_CROSS, lead)
    assert grid is not None
    grid_text = " ".join(c or "" for row in grid["grid"] for c in row)
    assert "AA3" in grid_text
    assert "管理费用" in grid_text or "SG&A" in grid_text

    extracted_rows = [{"ppe_impact": "indirect", "account_label": "SG&A expense / 管理费用"}]
    layout = {"confidence": "high", "amount_layout": "single_signed_column"}
    assert should_run_strict_total_check(
        lead,
        layout_result=layout,
        extracted_rows=extracted_rows,
    ) is False

    issues = check_lead_adjustment_internal_consistency(
        lead,
        layout_result=layout,
        extracted_rows=extracted_rows,
    )
    assert not any(i.field == "adjustment_amount" for i in issues)
