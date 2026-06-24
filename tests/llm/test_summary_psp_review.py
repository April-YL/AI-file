from unittest.mock import patch

from openpyxl import Workbook

from ingest.lead_sheet import LeadSheetDataset, MaterialityCapture
from ingest.models import AssetRecord, FieldMapping
from ingest.records import FaListDataset
from ingest.rollforward_sheet import RollforwardSheetDataset
from ingest.summary_sheet import PspProgramRow, SummarySheetDataset
from llm.config import LlmConfig
from llm.summary_psp_review import (
    build_sheet_semantic_issues,
    build_waiver_semantic_context,
    review_waiver_reasons_batch_with_llm,
    review_waiver_reason_with_llm,
)


def _config() -> LlmConfig:
    return LlmConfig(
        enabled=True,
        base_url="https://api.example.com/v1",
        api_key="sk-test",
        model="gpt-4o-mini",
    )


def test_review_waiver_reason_with_llm_parses_result():
    row = PspProgramRow(
        procedure_name="K.03.2 折旧测试TOD",
        sheet_ref="K.03.2 折旧测试TOD",
        execution_status="否",
        waiver_reason="本期金额不大。",
        notes=None,
        source_row=19,
        is_psp=False,
    )
    with patch(
        "llm.summary_psp_review.chat_completion_json",
        return_value={
            "adequacy": "insufficient",
            "rationale": "仅有金额描述，未说明风险应对",
            "suggested_action": "补充替代程序与证据来源",
        },
    ):
        res = review_waiver_reason_with_llm(row, _config())
    assert res is not None
    assert res.adequacy == "insufficient"


def test_review_waiver_reasons_batch_with_llm_parses_results():
    rows = [
        PspProgramRow(
            procedure_name="K.02.1 addition test",
            sheet_ref="K.02.1 addition test",
            execution_status="No",
            waiver_reason="amount is small",
            notes=None,
            source_row=12,
            is_psp=False,
        ),
        PspProgramRow(
            procedure_name="K.03.2 depreciation TOD",
            sheet_ref="K.03.2 depreciation TOD",
            execution_status="No",
            waiver_reason="not applicable",
            notes=None,
            source_row=19,
            is_psp=False,
        ),
    ]
    with patch(
        "llm.summary_psp_review.chat_completion_json",
        return_value={
            "reviews": [
                {
                    "row_id": 0,
                    "adequacy": "insufficient",
                    "rationale": "missing threshold basis",
                    "suggested_action": "add TE/TT/SAD basis",
                },
                {
                    "row_id": 1,
                    "adequacy": "unclear",
                    "rationale": "manual review needed",
                    "suggested_action": "check impairment indicator procedure",
                },
            ]
        },
    ) as mock_call:
        reviews = review_waiver_reasons_batch_with_llm(rows, _config())

    assert reviews[0].adequacy == "insufficient"
    assert reviews[1].adequacy == "unclear"
    user = mock_call.call_args.kwargs["user"]
    assert '"programs"' in user
    assert '"row_id": 0' in user
    assert '"row_id": 1' in user


def test_review_waiver_reason_prompt_uses_calibrated_psp_criteria():
    row = PspProgramRow(
        procedure_name="K.02.1 新增测试",
        sheet_ref="K.02.1 新增测试",
        execution_status="否",
        waiver_reason="N/A",
        notes=None,
        source_row=12,
        is_psp=False,
    )
    with patch(
        "llm.summary_psp_review.chat_completion_json",
        return_value={
            "adequacy": "insufficient",
            "rationale": "空泛理由",
            "suggested_action": "补充金额和性质风险判断",
        },
    ) as mock_call:
        res = review_waiver_reason_with_llm(row, _config())

    assert res is not None
    system = mock_call.call_args.kwargs["system"]
    user = mock_call.call_args.kwargs["user"]
    assert "包括但不限于" in system
    assert "N/A" in system and "NA" in system and "N/a" in system
    assert "TE、TT、SAD 不是并列标准" in system
    assert "第一层" in system and "小于 K.00 Lead Sheet 的 SAD" in system
    assert "第二层" in system and "小于 K.00 Lead Sheet 的 TT" in system
    assert "第三层" in system and "小于 K.00 Lead Sheet 的 TE" in system
    assert "不得仍要求补充 TT/TE" in system
    assert "K.00 Lead Sheet 中读取到的对应金额" in system
    assert "K.01 Agree SL to GL" in system
    assert "减值迹象" in system
    assert "不得编造" in system
    assert '"waiver_reason": "N/A"' in user


def test_review_waiver_reason_prompt_rejects_te_only_disposal_reason():
    row = PspProgramRow(
        procedure_name="K.02.2 处置测试",
        sheet_ref="K.02.2 处置测试",
        execution_status="否",
        waiver_reason="本期处置资产净值小于TE。",
        notes=None,
        source_row=18,
        is_psp=False,
    )
    with patch(
        "llm.summary_psp_review.chat_completion_json",
        return_value={
            "adequacy": "insufficient",
            "rationale": "仅说明小于TE，未说明单项TT和性质异常。",
            "suggested_action": "补充无单项大于TT且无性质异常项，或说明金额小于SAD。",
        },
    ) as mock_call:
        res = review_waiver_reason_with_llm(row, _config())

    assert res is not None
    assert res.adequacy == "insufficient"
    system = mock_call.call_args.kwargs["system"]
    user = mock_call.call_args.kwargs["user"]
    assert "如理由仅说明总体金额、处置资产净值或新增金额小于 K.00 Lead Sheet 的 TE" in system
    assert "还必须同时说明无单项金额大于 K.00 Lead Sheet 的 TT" in system
    assert "已明确小于 SAD 时，不得仍要求补充 TT/TE" in system
    assert "已明确小于 TT 且无性质异常项时，不得仍要求补充 TE" in system
    assert '"waiver_reason": "本期处置资产净值小于TE。"' in user


def test_review_waiver_reason_prompt_does_not_over_require_te_or_tt_when_below_sad():
    row = PspProgramRow(
        procedure_name="K.02.1 新增测试",
        sheet_ref="K.02.1 新增测试",
        execution_status="否",
        waiver_reason="本期新增金额小于SAD，未发现性质异常项。",
        notes=None,
        source_row=12,
        is_psp=False,
    )
    with patch(
        "llm.summary_psp_review.chat_completion_json",
        return_value={
            "adequacy": "sufficient",
            "rationale": "命中SAD层级。",
            "suggested_action": "",
        },
    ) as mock_call:
        res = review_waiver_reason_with_llm(row, _config())

    assert res is not None
    system = mock_call.call_args.kwargs["system"]
    assert "小于 K.00 Lead Sheet 的 SAD，则金额层面可接受" in system
    assert "除非输入摘录显示存在性质异常项，否则不得再要求补充 TE 或 TT" in system
    assert "已明确小于 SAD 时，不得仍要求补充 TT/TE" in system


def test_review_waiver_reason_prompt_does_not_require_te_when_below_tt_and_no_nature_exception():
    row = PspProgramRow(
        procedure_name="K.02.2 处置测试",
        sheet_ref="K.02.2 处置测试",
        execution_status="否",
        waiver_reason="本期处置金额小于TT，且无性质异常项。",
        notes=None,
        source_row=18,
        is_psp=False,
    )
    with patch(
        "llm.summary_psp_review.chat_completion_json",
        return_value={
            "adequacy": "sufficient",
            "rationale": "命中TT层级且无性质异常项。",
            "suggested_action": "",
        },
    ) as mock_call:
        res = review_waiver_reason_with_llm(row, _config())

    assert res is not None
    system = mock_call.call_args.kwargs["system"]
    assert "小于 K.00 Lead Sheet 的 TT，则金额层面可接受" in system
    assert "仍需说明或能从输入摘录判断不存在性质异常项" in system
    assert "不得再要求补充 TE" in system


def test_review_waiver_reason_prompt_includes_workbook_context():
    row = PspProgramRow(
        procedure_name="K.02.2 处置测试",
        sheet_ref="K.02.2 处置测试",
        execution_status="否",
        waiver_reason="本期处置资产净值小于TE。",
        notes=None,
        source_row=18,
        is_psp=True,
    )
    rollforward = RollforwardSheetDataset(
        source_file="case.xlsx",
        source_sheet="K.01 Agree SL to GL",
        header_row=None,
        mapped_fields=[],
        has_movement_rows=True,
        tb_difference_values=[],
        table4_difference=None,
    )
    disposal = FaListDataset(
        source_file="case.xlsx",
        source_sheet="K.02.2a 处置选样输出",
        mapped_fields=[
            FieldMapping("asset_id", "资产编号", 1),
            FieldMapping("net_value", "账面净值", 2),
        ],
        records=[
            AssetRecord(
                source_row=5,
                asset_id="FA-TEST-001",
                asset_name="处置设备",
                net_value="120000",
            )
        ],
    )
    lead = LeadSheetDataset(
        source_file="case.xlsx",
        source_sheet="K.00 Lead Sheet",
        materiality=[
            MaterialityCapture(field_key="te", label="TE", workpaper_value="1000000"),
            MaterialityCapture(field_key="sad", label="SAD", workpaper_value="50000"),
        ],
    )
    context = build_waiver_semantic_context(
        lead=lead,
        rollforward=rollforward,
        disposal_list=disposal,
        workbook_sheet_titles=["汇总", "K.01 Agree SL to GL", "K.02.2a 处置选样输出"],
    )

    with patch(
        "llm.summary_psp_review.chat_completion_json",
        return_value={
            "adequacy": "insufficient",
            "rationale": "处置清单已有记录，仅说明小于TE不足以支持不执行。",
            "suggested_action": "补充TT/SAD及性质风险判断，或执行处置测试。",
        },
    ) as mock_call:
        res = review_waiver_reason_with_llm(
            row,
            _config(),
            semantic_context=context,
        )

    assert res is not None
    user = mock_call.call_args.kwargs["user"]
    assert '"workbook_context"' in user
    assert '"has_movement_rows": true' in user
    assert "K.02.2a 处置选样输出" in user
    assert '"record_count": 1' in user
    assert '"field_key": "sad"' in user


def test_build_sheet_semantic_issues_for_weak_match(tmp_path):
    wb_path = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "K.02 新增"
    ws["A1"] = "新增资产清单"
    ws["A2"] = "资产名称"
    wb.save(wb_path)

    ds = SummarySheetDataset(
        source_file=str(wb_path),
        source_sheet="汇总",
        header_row=1,
        programs=[
            PspProgramRow(
                procedure_name="K.02.1 新增测试",
                sheet_ref="K.02.1 细节测试",
                execution_status="是",
                waiver_reason=None,
                notes=None,
                source_row=10,
                is_psp=False,
            )
        ],
    )

    with patch(
        "llm.summary_psp_review.chat_completion_json",
        return_value={
            "assessment": "match_supported",
            "chosen_sheet": "K.02 新增",
            "rationale": "候选页出现新增资产测试字段",
            "suggested_action": "确认并更新汇总页程序页引用",
        },
    ):
        issues = build_sheet_semantic_issues(
            ds,
            _config(),
            workbook_path=str(wb_path),
            workbook_sheet_titles=["K.02 新增"],
        )

    assert len(issues) == 1
    assert "更可能对应" in issues[0].message
    assert issues[0].review_source == "LLM辅助判断"
    assert issues[0].llm_review_type == "汇总页程序页语义匹配"
