from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import openpyxl

from ingest.addition_test_sheet import ModuleAssessment
from ingest.lead_sheet import LeadSheetDataset
from ingest.models import (
    AmountBusinessRole,
    AmountColumnCandidate,
    AmountCurrencyRole,
    AmountFieldGroup,
    AmountGroupStatus,
    AmountPeriodRole,
    EvidenceType,
    FieldCandidate,
    FieldEvidence,
    FieldResolutionDecision,
    ResolutionStatus,
    SheetKind,
    SheetResolutionDecision,
)
from ingest.records import FaListDataset
from ingest.disposal_test_sheet import (
    DisposalExecutionPathDataset,
    DisposalSampleOutputDataset,
    DisposalSampleRow,
    DisposalTestSheetDataset,
    DisposalTestedSampleRow,
)
from llm.config import LlmConfig
from llm.ingest_review import (
    EXPECTED_INGEST_OBJECTS,
    K01_PROFILE_HINT,
    K021_ADDITION_PROFILE_HINT,
    K022_DISPOSAL_PROFILE_HINT,
    LEAD_PROFILE_HINT,
    SYSTEM_PROMPT,
    SUMMARY_PROFILE_HINT,
    IngestReviewCandidatePreview,
    IngestReviewPayload,
    build_k022_disposal_ingest_review_payload,
    build_lead_ingest_review_payload,
    build_missing_k01_ingest_review_payload,
    build_ingest_review_user_prompt,
    parse_ingest_review_result,
    run_ingest_review,
    run_field_identification_fallback,
    run_amount_group_identification_fallback,
    run_sheet_identification_fallback,
    run_workbook_ingest_reviews,
    should_review_k022_disposal_ingest,
    should_review_lead_ingest,
)
from ingest.workbook_structure import WorkbookStructure


def _config(*, enabled: bool = True) -> LlmConfig:
    return LlmConfig(
        enabled=enabled,
        base_url="https://api.example.com/v1",
        api_key="sk-test",
        model="gpt-4o-mini",
    )


def _payload() -> IngestReviewPayload:
    return IngestReviewPayload(
        review_target="K.01 表3 check with 表1 漏读发现",
        review_type="missing_object_discovery",
        program_profile_hint=K01_PROFILE_HINT,
        coding_result={
            "classified_sheet": "K.01 Agree SL to GL",
            "recognized_sections": ["b1_bkd_main_table", "b2_movement_tb_reconciliation"],
            "missing_sections": ["b4_table3_check_with_table1"],
            "recognition_confidence": 0.58,
            "conflicts": ["duplicate_anchor:b4_table3_check_with_table1"],
        },
        expected_object={
            "procedure": "K.01",
            "object_type": "module",
            "object_name": "b4_table3_check_with_table1",
        },
        candidate_previews=[
            IngestReviewCandidatePreview(
                sheet_name="K.01 Agree SL to GL",
                name_score=0.9,
                content_score=0.7,
                preview_lines=[
                    {"row": 38, "text": "表2 固定资产清单分类汇总"},
                    {"row": 45, "text": "表2 check with 表1 差异 Notes"},
                    {"row": 82, "text": "表4 折旧费用与利润表科目核对 差异"},
                ],
                anchor_hits=[
                    {"row": 45, "anchors": ["表2 check with 表1", "差异"]},
                    {"row": 82, "anchors": ["表4", "折旧费用与利润表"]},
                ],
            )
        ],
        question="请判断 coding 是否可能漏识别 K.01 表3。",
    )


def test_ingest_review_prompt_sets_boundaries():
    assert "不是重新读取整本 Excel" in SYSTEM_PROMPT
    assert "不得直接给出金额勾稽结论" in SYSTEM_PROMPT
    assert "不得将低置信度读取直接改成高置信度" in SYSTEM_PROMPT
    assert "只提出候选 sheet" in SYSTEM_PROMPT
    assert "表3 check、TB check、表4折旧核对是不同专题" in K01_PROFILE_HINT


def test_build_user_prompt_includes_payload_and_k01_hint():
    prompt = build_ingest_review_user_prompt(_payload())
    assert "K.01 表3 check with 表1 漏读发现" in prompt
    assert "missing_object_discovery" in prompt
    assert "program_profile_hint" in prompt
    assert "表4折旧费用与利润表核对的差异不得被当作 TB 差异" in prompt


def test_program_profiles_are_attached_to_missing_discovery_objects():
    by_name = {obj.object_name: obj.profile_hint for obj in EXPECTED_INGEST_OBJECTS}

    assert by_name["汇总"] == SUMMARY_PROFILE_HINT
    assert by_name["K.00 Lead Sheet"] == LEAD_PROFILE_HINT
    assert by_name["新增清单"] == K021_ADDITION_PROFILE_HINT
    assert by_name["K.02.1 新增测试"] == K021_ADDITION_PROFILE_HINT
    assert by_name["K.02.1a 新增选样输出"] == K021_ADDITION_PROFILE_HINT
    assert by_name["处置清单"] == K022_DISPOSAL_PROFILE_HINT
    assert by_name["K.02.2 处置测试"] == K022_DISPOSAL_PROFILE_HINT
    assert by_name["K.02.2a 处置选样输出"] == K022_DISPOSAL_PROFILE_HINT


def test_program_profile_prompts_include_false_positive_guardrails():
    assert "sheet 名称存在尾随空格" in SUMMARY_PROFILE_HINT
    assert "no_cra_te_volatility" in LEAD_PROFILE_HINT
    assert "右侧 SOP/易错点说明区" in K021_ADDITION_PROFILE_HINT
    assert "汇总页已拒绝执行" in K022_DISPOSAL_PROFILE_HINT


def test_run_ingest_review_disabled_returns_none():
    result, raw = run_ingest_review(_config(enabled=False), _payload())
    assert result is None
    assert raw is None


def test_parse_valid_suspicious_result():
    raw = {
        "assessment": "suspicious",
        "risk_level": "high",
        "risk_area": "missing_module",
        "suspected_object": "b4_table3_check_with_table1",
        "candidate_sheet": "K.01 Agree SL to GL",
        "candidate_rows": [45],
        "evidence_anchors": ["表2 check with 表1", "差异"],
        "rationale": "第45行出现表2 check with 表1。",
        "suggested_action": "二次读取第45行附近。",
        "should_retry_deterministic_ingest": True,
        "manual_review_focus": "核对第45行附近。",
    }
    result = parse_ingest_review_result(raw, _payload())
    assert result is not None
    assert result.assessment == "suspicious"
    assert result.candidate_rows == [45]
    assert result.should_retry_deterministic_ingest is True


def test_parse_rejects_unknown_candidate_sheet():
    raw = {
        "assessment": "suspicious",
        "risk_level": "high",
        "risk_area": "missing_module",
        "candidate_sheet": "Invented Sheet",
        "candidate_rows": [45],
        "evidence_anchors": ["表2 check with 表1"],
    }
    assert parse_ingest_review_result(raw, _payload()) is None


def test_parse_rejects_invented_candidate_row():
    raw = {
        "assessment": "suspicious",
        "risk_level": "high",
        "risk_area": "missing_module",
        "candidate_sheet": "K.01 Agree SL to GL",
        "candidate_rows": [999],
        "evidence_anchors": ["表2 check with 表1"],
    }
    assert parse_ingest_review_result(raw, _payload()) is None


def test_parse_rejects_invented_anchor():
    raw = {
        "assessment": "suspicious",
        "risk_level": "high",
        "risk_area": "missing_module",
        "candidate_sheet": "K.01 Agree SL to GL",
        "candidate_rows": [45],
        "evidence_anchors": ["不存在的锚点"],
    }
    assert parse_ingest_review_result(raw, _payload()) is None


def test_parse_rejects_invalid_assessment():
    raw = {
        "assessment": "pass",
        "risk_level": "low",
        "risk_area": "other",
    }
    assert parse_ingest_review_result(raw, _payload()) is None


def test_parse_rejects_severity_override():
    raw = {
        "assessment": "suspicious",
        "risk_level": "high",
        "risk_area": "missing_module",
        "candidate_sheet": "K.01 Agree SL to GL",
        "candidate_rows": [45],
        "evidence_anchors": ["表2 check with 表1"],
        "severity": "FAIL",
    }
    assert parse_ingest_review_result(raw, _payload()) is None


def test_run_ingest_review_returns_validated_result_and_raw():
    raw = {
        "assessment": "likely_ok",
        "risk_level": "low",
        "risk_area": "section_boundary",
        "candidate_sheet": "K.01 Agree SL to GL",
        "candidate_rows": [],
        "evidence_anchors": [],
        "rationale": "未见明显错分证据。",
        "should_retry_deterministic_ingest": False,
    }
    with patch("llm.ingest_review.chat_completion_json", return_value=raw):
        result, returned_raw = run_ingest_review(_config(), _payload())

    assert returned_raw == raw
    assert result is not None
    assert result.assessment == "likely_ok"


def test_build_missing_k01_payload_from_candidate_workbook(tmp_path: Path):
    path = tmp_path / "missing_k01_candidate.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K01 SL-GL"
    ws.append(["表1", "固定资产类别", "年初余额", "年末余额", "审定数"])
    ws.append(["TB-原值", "差异", "Notes"])
    wb.save(path)
    wb.close()

    payload = build_missing_k01_ingest_review_payload(
        workbook_path=str(path),
        workbook_sheet_titles=["K01 SL-GL"],
    )

    assert payload is not None
    assert payload.review_type == "missing_object_discovery"
    assert payload.expected_object["object_type"] == "sheet"
    assert payload.candidate_previews[0].sheet_name == "K01 SL-GL"
    assert payload.candidate_previews[0].anchor_hits


def test_run_workbook_ingest_reviews_handles_missing_k01(tmp_path: Path):
    path = tmp_path / "missing_k01_candidate.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K01 SL-GL"
    ws.append(["表1", "固定资产类别", "年初余额", "年末余额", "审定数"])
    ws.append(["表2 check with 表1", "差异", "Notes"])
    wb.save(path)
    wb.close()

    raw = {
        "assessment": "suspicious",
        "risk_level": "high",
        "risk_area": "missing_sheet",
        "suspected_object": "K.01 Agree SL to GL",
        "candidate_sheet": "K01 SL-GL",
        "candidate_rows": [1],
        "evidence_anchors": ["表1", "固定资产类别"],
        "rationale": "候选 sheet 出现 K.01 后推锚点。",
        "suggested_action": "对候选 sheet 执行 deterministic ingest。",
        "should_retry_deterministic_ingest": True,
        "manual_review_focus": "打开 K01 SL-GL 核对是否为 K.01。",
    }
    with patch("llm.ingest_review.chat_completion_json", return_value=raw):
        results = run_workbook_ingest_reviews(
            _config(),
            rollforward=None,
            workbook_path=str(path),
            workbook_sheet_titles=["K01 SL-GL"],
        )

    assert len(results) == 1
    assert results[0].assessment == "suspicious"
    assert results[0].risk_area == "missing_sheet"


def test_run_workbook_ingest_reviews_covers_all_core_program_sheets(tmp_path: Path):
    path = tmp_path / "all_program_candidates.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "汇总"
    for title in (
        "K.00 Lead Sheet",
        "K.01 Agree SL to GL",
        "FA list",
        "新增清单",
        "K.02.1 新增测试",
        "K.02.1a 新增选样输出",
        "处置清单",
        "K.02.2 处置测试",
        "K.02.2a 处置选样输出",
        "K.03.1 SAP",
        "K.03.2 折旧测试TOD",
        "K.03.3 折旧政策复核",
    ):
        wb.create_sheet(title)
    wb.save(path)
    wb.close()

    def fake_chat_completion_json(*args, **kwargs):
        user = kwargs["user"]
        payload = user.split("输入数据：", 1)[1].strip()
        data = __import__("json").loads(payload)
        candidate = data["candidate_previews"][0]
        return {
            "assessment": "suspicious",
            "risk_level": "medium",
            "risk_area": "missing_sheet",
            "suspected_object": data["expected_object"]["object_name"],
            "candidate_sheet": candidate["sheet_name"],
            "candidate_rows": [],
            "evidence_anchors": [],
            "rationale": "mock project-level missing sheet review",
            "suggested_action": "人工核对候选 sheet。",
            "should_retry_deterministic_ingest": False,
            "manual_review_focus": "确认 sheet 类型。",
        }

    with patch("llm.ingest_review.chat_completion_json", side_effect=fake_chat_completion_json):
        results = run_workbook_ingest_reviews(
            _config(),
            rollforward=None,
            workbook_path=str(path),
            workbook_sheet_titles=wb.sheetnames,
            recognized_sheet_kinds={},
        )

    assert len(results) == len(EXPECTED_INGEST_OBJECTS)
    assert {r.procedure_code for r in results} >= {"SUMMARY", "K.00", "K.01", "K.02.1", "K.02.2", "K.03.1", "K.03.2", "K.03.3"}


def _disposal_test_sheet() -> DisposalTestSheetDataset:
    return DisposalTestSheetDataset(
        source_file="wb.xlsx",
        source_sheet="K.02.2 处置测试",
        module_assessments=[
            ModuleAssessment(
                module_key="population_reconciliation",
                module_name="处置/报废总金额核对",
                status="weak",
                confidence=0.45,
                evidence=["处置测试", "净值"],
            )
        ],
        recognition_confidence=0.52,
        usable_for_rules=False,
        tested_samples=[
            DisposalTestedSampleRow(
                source_row=30,
                sample_type="关键项",
                asset_id="FA-TEST-001",
                asset_name="设备A",
                net_value="100",
                disposal_method="出售",
                evidence_description="合同及收款单",
            )
        ],
        notes=["disposal_reconciliation_matrix_low_confidence"],
    )


def _disposal_sample_output() -> DisposalSampleOutputDataset:
    return DisposalSampleOutputDataset(
        source_file="wb.xlsx",
        source_sheet="K.02.2a 处置选样输出",
        module_assessments=[
            ModuleAssessment(
                module_key="selected_samples",
                module_name="已选取样本",
                status="recognized",
                confidence=0.9,
                evidence=["处置选样输出", "样本池"],
            )
        ],
        recognition_confidence=0.91,
        usable_for_rules=True,
        selected_samples=[
            DisposalSampleRow(
                source_row=22,
                sample_type="代表性样本",
                asset_id="FA-TEST-001",
                asset_name="设备A",
                net_value="100",
                disposal_method="出售",
            )
        ],
    )


def _disposal_execution_path() -> DisposalExecutionPathDataset:
    return DisposalExecutionPathDataset(
        path_kind="full_expected",
        recognition_confidence=0.8,
        disposal_test_sheet="K.02.2 处置测试",
        disposal_sample_output_sheet="K.02.2a 处置选样输出",
        missing_components=[],
    )


def test_should_review_k022_disposal_ingest_when_low_confidence_or_unusable():
    assert (
        should_review_k022_disposal_ingest(
            disposal_test=_disposal_test_sheet(),
            disposal_sample_output=_disposal_sample_output(),
            disposal_execution_path=_disposal_execution_path(),
        )
        is True
    )


def test_lead_ingest_review_runs_only_when_deterministic_result_is_unusable():
    lead = LeadSheetDataset(
        source_file="wb.xlsx",
        source_sheet="K.00 Lead Sheet",
        usable_for_rules=False,
    )

    assert should_review_lead_ingest(lead) is True
    payload = build_lead_ingest_review_payload(lead)
    assert payload.expected_object["procedure"] == "K.00"
    assert payload.coding_result["usable_for_rules"] is False

    lead.usable_for_rules = True
    assert should_review_lead_ingest(lead) is False


def test_build_k022_disposal_payload_uses_profile_and_read_result_review():
    payload = build_k022_disposal_ingest_review_payload(
        disposal_test=_disposal_test_sheet(),
        disposal_sample_output=_disposal_sample_output(),
        disposal_execution_path=_disposal_execution_path(),
    )

    assert payload.review_type == "read_result_review"
    assert payload.program_profile_hint == K022_DISPOSAL_PROFILE_HINT
    assert payload.expected_object["procedure"] == "K.02.2"
    assert payload.coding_result["recognition_confidence"] == 0.52
    assert "disposal_test:population_reconciliation:weak" in payload.coding_result["missing_sections"]
    assert payload.candidate_previews[0].sheet_name == "K.02.2 处置测试"
    assert payload.candidate_previews[0].anchor_hits


def test_run_workbook_ingest_reviews_handles_k022_read_result_review():
    raw = {
        "assessment": "suspicious",
        "risk_level": "medium",
        "risk_area": "section_boundary",
        "suspected_object": "K.02.2 disposal ingest result",
        "candidate_sheet": "K.02.2 处置测试",
        "candidate_rows": [30],
        "evidence_anchors": ["处置测试", "净值", "出售"],
        "rationale": "K.02.2 测试页低置信且样本行含处置净值锚点，建议人工核对读取区域。",
        "suggested_action": "核对 K.02.2 处置测试页第30行附近是否为实测样本区域。",
        "should_retry_deterministic_ingest": False,
        "manual_review_focus": "确认处置测试样本区域和处置净值字段是否读对。",
    }

    with patch("llm.ingest_review.chat_completion_json", return_value=raw):
        results = run_workbook_ingest_reviews(
            _config(),
            disposal_test=_disposal_test_sheet(),
            disposal_sample_output=_disposal_sample_output(),
            disposal_execution_path=_disposal_execution_path(),
            recognized_sheet_kinds={
                "disposal_test": True,
                "disposal_sample_output": True,
            },
        )

    assert len(results) == 1
    assert results[0].procedure_code == "K.02.2"
    assert results[0].review_type == "k022_ingest_review"
    assert results[0].risk_area == "section_boundary"


def test_identification_fallback_can_only_select_existing_supported_candidate():
    first = FieldCandidate(
        "asset_id",
        "资产编号",
        1,
        evidence=[
            FieldEvidence(EvidenceType.HEADER_SEMANTIC, "header"),
            FieldEvidence(EvidenceType.VALUE_DISTRIBUTION, "distinct values"),
        ],
    )
    second = FieldCandidate(
        "asset_id",
        "备用编号",
        2,
        evidence=[FieldEvidence(EvidenceType.HEADER_SEMANTIC, "header")],
    )
    dataset = FaListDataset(
        source_file="demo.xlsx",
        source_sheet="FA list",
        mapped_fields=[],
        records=[],
        field_resolutions={
            "asset_id": FieldResolutionDecision(
                standard_field="asset_id",
                candidates=[first, second],
                status=ResolutionStatus.AMBIGUOUS,
            )
        },
    )
    config = LlmConfig(
        enabled=True,
        base_url="https://api.example.com/v1",
        api_key="sk-test",
        model="test",
        identification_enabled=True,
    )
    raw = {
        "selections": [
            {"standard_field": "asset_id", "column": 1, "confidence": 0.9},
            {"standard_field": "invented", "column": 99, "confidence": 1},
        ]
    }

    with patch("llm.ingest_review.chat_completion_json", return_value=raw):
        selections = run_field_identification_fallback(config, dataset)

    assert selections == {"asset_id": 1}


def test_identification_disabled_does_not_call_llm():
    dataset = FaListDataset(
        source_file="demo.xlsx",
        source_sheet="FA list",
        mapped_fields=[],
        records=[],
    )
    with patch("llm.ingest_review.chat_completion_json") as client:
        assert run_field_identification_fallback(_config(), dataset) == {}
    client.assert_not_called()


def test_sheet_identification_can_only_select_verified_candidate():
    evidence = [
        FieldEvidence(EvidenceType.HEADER_SEMANTIC, "name"),
        FieldEvidence(EvidenceType.STRUCTURAL_CONTEXT, "headers"),
    ]
    structure = WorkbookStructure(
        source_file="demo.xlsx",
        sheet_resolutions={
            "新增清单": SheetResolutionDecision(
                sheet_name="新增清单",
                candidates=[(SheetKind.ADDITION_LIST, 0.72)],
                status=ResolutionStatus.AMBIGUOUS,
                evidence=evidence,
            )
        },
    )
    config = LlmConfig(
        enabled=True,
        base_url="https://api.example.com/v1",
        api_key="sk-test",
        model="test",
        identification_enabled=True,
    )

    with patch(
        "llm.ingest_review.chat_completion_json",
        return_value={"sheet": "新增清单", "confidence": 0.9},
    ):
        result = run_sheet_identification_fallback(
            config,
            structure,
            SheetKind.ADDITION_LIST,
        )

    assert result is not None
    assert result.selected_kind == SheetKind.ADDITION_LIST
    assert result.status == ResolutionStatus.RESOLVED
    assert result.reorganization_count == 1


def test_sheet_identification_rejects_invented_candidate():
    structure = WorkbookStructure(
        source_file="demo.xlsx",
        sheet_resolutions={
            "新增清单": SheetResolutionDecision(
                sheet_name="新增清单",
                candidates=[(SheetKind.ADDITION_LIST, 0.8)],
                status=ResolutionStatus.AMBIGUOUS,
                evidence=[
                    FieldEvidence(EvidenceType.HEADER_SEMANTIC, "name"),
                    FieldEvidence(EvidenceType.STRUCTURAL_CONTEXT, "headers"),
                ],
            )
        },
    )
    config = LlmConfig(
        enabled=True,
        base_url="https://api.example.com/v1",
        api_key="sk-test",
        model="test",
        identification_enabled=True,
    )

    with patch(
        "llm.ingest_review.chat_completion_json",
        return_value={"sheet": "不存在", "confidence": 1.0},
    ):
        result = run_sheet_identification_fallback(
            config,
            structure,
            SheetKind.ADDITION_LIST,
        )

    assert result is None


def test_amount_group_identification_selects_whole_existing_group():
    def group(group_id: str, start_col: int, status: AmountGroupStatus):
        return AmountFieldGroup(
            group_id=group_id,
            members={
                "original_value": AmountColumnCandidate(
                    measure="original_value",
                    source_header="本期新增原值",
                    column_index=start_col,
                    period_role=AmountPeriodRole.CURRENT_PERIOD,
                    currency_role=AmountCurrencyRole.REPORTING,
                    business_role=AmountBusinessRole.ADDITION,
                    evidence=("header", "shared semantic group"),
                ),
                "net_value": AmountColumnCandidate(
                    measure="net_value",
                    source_header="本期新增净值",
                    column_index=start_col + 1,
                    period_role=AmountPeriodRole.CURRENT_PERIOD,
                    currency_role=AmountCurrencyRole.REPORTING,
                    business_role=AmountBusinessRole.ADDITION,
                    evidence=("header", "shared semantic group"),
                ),
            },
            period_role=AmountPeriodRole.CURRENT_PERIOD,
            currency_role=AmountCurrencyRole.REPORTING,
            business_role=AmountBusinessRole.ADDITION,
            status=status,
            confidence=0.9,
        )

    dataset = FaListDataset(
        source_file="demo.xlsx",
        source_sheet="新增清单",
        mapped_fields=[],
        records=[],
        amount_groups=[
            group("addition_list:amount:1", 3, AmountGroupStatus.AMBIGUOUS),
            group("addition_list:amount:2", 8, AmountGroupStatus.CONFIRMED),
        ],
        selected_amount_group_id="addition_list:amount:1",
    )
    config = LlmConfig(
        enabled=True,
        base_url="https://api.example.com/v1",
        api_key="sk-test",
        model="test",
        identification_enabled=True,
    )

    with patch(
        "llm.ingest_review.chat_completion_json",
        return_value={"group_id": "addition_list:amount:2", "confidence": 0.9},
    ):
        result = run_amount_group_identification_fallback(
            config,
            dataset,
            SheetKind.ADDITION_LIST,
        )

    assert result == "addition_list:amount:2"
